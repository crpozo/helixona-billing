"""
Helixona Billing Agent — Main Entry Point
Polls SQS for task messages and dispatches to the appropriate handler.

Task Types:
  - nightly_bulk_extract     → Stage 1: ECW claim creation
  - claims_ecw / test_ecw_login → Stage 0: ECW claim extraction
  - generate_hcfa            → Stage 0C: HCFA form PDF generation from ECW
  - capture_blueshield_claim → Stage 3: BS portal claim # capture
  - verify_medical_record    → Stage 4: AI IV prescription verification
  - generate_cover_letter    → Stage 5: Cover letter PDF generation
  - symplisend_submission    → Stage 6: Full SympliSend submission flow
  - process_adjudication     → Stage 8: EOB/ERA processing
"""

import time
import json
import uuid
import random
import os
import subprocess
from src.utils.logger import get_logger
from src.aws.clients import AWSClient
from src.ecw.browser import BrowserManager
from src.ecw.login import perform_ecw_login
from src.ecw.claims import process_nightly_bulk_claims
from src.blueshield.portal import BlueShieldPortal
from src.ai.verification import IVVerificationService
from src.documents.cover_letter import generate_cover_letter_pdf
from src.symplisend.submission import SympliSendSubmission
from src.rules.engine import RulesEngine
from src.adjudication.processor import process_adjudication
from src.audit.submission_log import (
    record_submission,
    describe_documents,
    METHOD_SYMPLISEND,
    OUTCOME_SUBMITTED,
    OUTCOME_FAILED,
    OUTCOME_BLOCKED,
)

logger = get_logger(__name__)
rules = RulesEngine()


def _dismiss_ecw_data_error(page, label=''):
    """
    Auto-dismiss the recurring eClinicalWorks 'data loading error' popup
    (and similar transient OK-only popups) by clicking OK.
    Safe to call repeatedly — does nothing if no popup is open.
    """
    try:
        dismissed = page.evaluate('''() => {
            const targets = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a'));
            const text = (document.body && document.body.innerText) || '';
            const isError = /data loading error|eClinicalWorks/i.test(text);
            for (const t of targets) {
                const label = ((t.innerText || t.value || '') + '').trim();
                if (label === 'OK' || label === 'Ok') {
                    // Only click if visible
                    const r = t.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0) {
                        t.click();
                        return label;
                    }
                }
            }
            return null;
        }''')
        if dismissed:
            logger.info(f"🟦 Dismissed ECW popup via {dismissed}{(' @' + label) if label else ''}")
            time.sleep(0.6)
            return True
    except Exception:
        pass
    return False


def _dismiss_claim_alert(page):
    """Dismiss the small eCW validation alerts that pop on top of the claim detail
    popup (e.g. 'Master/Default fee schedule is selected.', 'data loading error').
    Clicks the OK inside the alert dialog only — scoped by the alert's text so it
    won't hit the status-picker OK or the claim popup's Save/OK. Returns True if one
    was dismissed. Safe to call repeatedly."""
    import time as _t
    ALERT_RE = "fee schedule is selected|data loading error|already open|cannot be|is required|please select"
    for ctx in [page] + list(page.frames):
        try:
            clicked = ctx.evaluate('''(reSrc) => {
                const re = new RegExp(reSrc, 'i');
                const btns = document.querySelectorAll('button, input[type="button"], input[type="submit"]');
                // 1) Prefer an OK whose enclosing dialog text matches a known alert
                for (const b of btns) {
                    const t = (b.value || b.textContent || '').trim();
                    if ((t === 'OK' || t === 'Ok') && b.offsetWidth > 0) {
                        const dlg = b.closest('.modal, .modal-content, .ui-dialog, [role="dialog"], .bootbox, .alert');
                        if (dlg && re.test(dlg.textContent || '')) { b.click(); return true; }
                    }
                }
                // 2) Fallback: if the page clearly shows the alert text, click first visible OK
                if (re.test((document.body && document.body.innerText) || '')) {
                    for (const b of btns) {
                        const t = (b.value || b.textContent || '').trim();
                        if ((t === 'OK' || t === 'Ok') && b.offsetWidth > 0) { b.click(); return true; }
                    }
                }
                return false;
            }''', ALERT_RE)
            if clicked:
                logger.info("  🟦 Dismissed eCW claim alert (fee schedule / validation popup)")
                _t.sleep(0.6)
                return True
        except Exception:
            continue
    return False


def _set_claim_status_in_ecw(page, claim_id, target_label='Claim sent via Symplisend'):
    """Set a claim's Claim Status in the ECW claim popup and verify it persisted.

    Assumes ECW is already logged in and on the Billing → Claims lookup screen.
    Flow (mirrors the manual one): open the claim via lookup → dismiss the recurring
    'Master/Default fee schedule is selected' alert → open the Claim Status Code picker
    (#billingClaimBtn83) → select the target status row → click the picker's OK button
    (saveClaimStatusCodes(true); the × close calls saveClaimStatusCodes(false)) → save
    the claim (saveAllData) → re-open and read the status field to confirm.

    The status is rendered as text in a <span>/<td> next to the '...' button — NOT an
    <input> — and the in-popup value does not refresh after the picker save, so the only
    authoritative check is re-opening the claim. Returns True only if the re-opened claim
    shows target_label. Leaves nothing changed in our datastore on failure (caller decides).
    """
    target_lc = target_label.strip().lower()

    READ_STATUS_JS = '''() => {
        const btn = document.querySelector('#billingClaimBtn83')
                 || document.querySelector('button[data-ng-click^="selectClaimStatusCode"]');
        if (btn) {
            let p = btn.parentElement;
            for (let i = 0; i < 4 && p; i++) {
                const inp = p.querySelector('input');
                if (inp && (inp.value || '').trim()) return inp.value.trim();
                const txt = ((p.innerText || p.textContent || '').replace(/\\.\\.\\./g, '').trim());
                if (txt && txt.length < 80) return txt;
                p = p.parentElement;
            }
        }
        // Fallback (the robust path): a span/td showing a known claim status verbatim
        for (const el of document.querySelectorAll('span, td')) {
            const t = (el.textContent || '').trim();
            const lt = t.toLowerCase();
            if ((lt === 'claim sent via symplisend'
                 || lt === 'ready to submit to symplisend') && t.length < 60) {
                return t;
            }
        }
        return null;
    }'''

    def _read_status():
        for frm in page.frames:
            try:
                v = frm.evaluate(READ_STATUS_JS)
                if v is not None:
                    return v
            except Exception:
                continue
        return None

    # STEP 1 — open the claim
    opened, _f = _open_claim_popup_via_lookup(page, claim_id)
    if not opened:
        logger.warning(f"  Could not find claim {claim_id} in ECW")
        return False
    logger.info(f"  ✅ STEP 1 — Opened claim {claim_id} modal")
    time.sleep(3)
    _dismiss_claim_alert(page)

    logger.info(f"  STEP 2 — Claim Status BEFORE = {_read_status()!r}")

    # STEP 3 — open the Claim Status Code picker
    status_btn_clicked = False
    for frm in page.frames:
        try:
            which = frm.evaluate('''() => {
                const b = document.querySelector('button#billingClaimBtn83');
                if (b && b.offsetWidth > 0) { b.click(); return 'billingClaimBtn83'; }
                for (const x of document.querySelectorAll('button[data-ng-click="selectClaimStatusCode()"]')) {
                    if (x.offsetWidth > 0) { x.click(); return 'selectClaimStatusCode'; }
                }
                return null;
            }''')
            if which:
                status_btn_clicked = True
                logger.info(f"  ✅ STEP 3 — Opened status picker (via {which})")
                break
        except Exception:
            continue
    if not status_btn_clicked:
        logger.warning(f"  ❌ STEP 3 — status '...' button not found for claim {claim_id}")
        _close_claim_popup(page)
        return False
    time.sleep(2)

    # STEP 4 — select the target status row
    found = False
    for frm in page.frames:
        try:
            res = frm.evaluate('''(targetLc) => {
                for (const row of document.querySelectorAll('tr')) {
                    for (const cell of row.querySelectorAll('td')) {
                        if ((cell.textContent || '').trim().toLowerCase() === targetLc) {
                            row.click();
                            return true;
                        }
                    }
                }
                return false;
            }''', target_lc)
            if res:
                found = True
                break
        except Exception:
            continue
    if not found:
        logger.warning(f"  ❌ STEP 4 — status row {target_label!r} not found for claim {claim_id}")
        _close_claim_popup(page)
        return False
    logger.info(f"  ✅ STEP 4 — Selected status row {target_label!r}")
    time.sleep(1)

    # STEP 5 — click the picker's OK (saveClaimStatusCodes(true))
    confirm = None
    for frm in page.frames:
        try:
            confirm = frm.evaluate('''() => {
                const btns = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"]'))
                    .filter(b => b.offsetWidth > 0);
                const ngOf = b => (b.getAttribute('ng-click') || b.getAttribute('data-ng-click') || '');
                let ok = btns.find(b => /saveClaimStatusCodes\\(\\s*true\\s*\\)/.test(ngOf(b)));
                if (!ok) ok = btns.find(b => /saveClaimStatusCodes/.test(ngOf(b)) && !/\\(\\s*false\\s*\\)/.test(ngOf(b)));
                if (!ok) ok = btns.find(b => ['OK', 'Ok'].includes((b.value || b.textContent || '').trim()));
                if (ok) { ok.click(); return ngOf(ok) || 'OK'; }
                return null;
            }''')
            if confirm:
                break
        except Exception:
            continue
    if confirm:
        logger.info(f"  ✅ STEP 5 — Clicked picker OK ({confirm})")
    else:
        logger.warning(f"  ⚠️ STEP 5 — picker OK (saveClaimStatusCodes(true)) not found for claim {claim_id}")
    time.sleep(1.5)
    _dismiss_claim_alert(page)

    # STEP 6 — save the claim popup (saveAllData)
    saved_via = None
    for frm in page.frames:
        try:
            saved_via = frm.evaluate('''() => {
                const okBtn = document.querySelector('button[ng-click="saveAllData()"]');
                if (okBtn && okBtn.offsetWidth > 0) { okBtn.click(); return 'saveAllData'; }
                for (const b of document.querySelectorAll('button[id^="claimScreenOkBtn"]')) {
                    if (b.offsetWidth > 0) { b.click(); return 'claimScreenOkBtn'; }
                }
                const all = document.querySelectorAll('button, input[type="button"]');
                for (const b of all) {
                    const t = (b.value || b.textContent || '').trim();
                    if (t === 'OK' && b.offsetWidth > 0) {
                        const dlg = b.closest('.modal, .ui-dialog, [role="dialog"]');
                        if (dlg && /Print HCFA|Prog\\.? Notes|Claim No/i.test(dlg.textContent || '')) { b.click(); return 'claimPopupOK'; }
                    }
                }
                for (const b of all) {
                    const t = (b.value || b.textContent || '').trim();
                    if (t === 'OK' && b.offsetWidth > 0) { b.click(); return 'anyOK'; }
                }
                return null;
            }''')
            if saved_via:
                logger.info(f"  ✅ STEP 6 — Saved claim via {saved_via}")
                break
        except Exception:
            continue
    if not saved_via:
        logger.warning(f"  ❌ STEP 6 — Save/OK button not found for claim {claim_id}")
    time.sleep(3)

    # STEP 7 (verify) — re-open the claim and confirm the status persisted
    verified = False
    try:
        reopened, _vf = _open_claim_popup_via_lookup(page, claim_id, wait_seconds=3)
        if reopened:
            _dismiss_claim_alert(page)
            now = _read_status()
            logger.info(f"  STEP 7 (verify) — re-opened claim {claim_id}, status = {now!r}")
            verified = (now or '').strip().lower() == target_lc
        else:
            logger.warning(f"  STEP 7 — could not re-open claim {claim_id} to verify")
        _close_claim_popup(page)
    except Exception as ve:
        logger.warning(f"  STEP 7 — verify failed for {claim_id}: {ve}")

    if not verified:
        logger.warning(f"  ❌ Claim {claim_id} status NOT persisted in ECW — leaving unmarked for retry")
    return verified


def _combine_excels(paths, output_path):
    """Combine multiple .xlsx files into one (single sheet, shared header)."""
    import openpyxl
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Combined'
    header_written = False
    for path in paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        if not header_written:
            ws_out.append(rows[0])
            header_written = True
        for row in rows[1:]:
            ws_out.append(row)
    wb_out.save(output_path)
    return output_path


def _run_fix_coding_ivs_stage2(page, body, aws_client, ecw_url):
    """
    Fix Coding IVs — Stage 2: Remove test claims.
      1. Navigate to Billing → Claims.
      2. For each status in ['Pending With Errors', 'Pending']:
         a. Set status filter + date range 07/01/2025 → today
         b. Click #btnclaimlookup
         c. Find rows where the patient name (LastName, FirstName) contains 'test'
         d. For each match: select row checkbox → Claims dropdown (#claimLookupBtn10)
            → 'Delete Claim' → confirm Yes
         e. Re-find after every deletion since the table refreshes.

    body flags:
      - testing_mode: stop after deleting 1 test claim total (verification only).
    """
    from datetime import datetime as _dt
    testing_mode = bool(body.get('testing_mode'))
    today_str = _dt.now().strftime('%m/%d/%Y')

    logger.info("═══ Fix Coding IVs — Stage 2: Remove test claims ═══")
    if testing_mode:
        logger.info("🧪 Test mode ON — will stop after deleting 1 test claim total")

    # Ensure we're on Billing → Claims (Stage 1 may have left us elsewhere)
    logger.info("Navigating to Billing → Claims...")
    try:
        page.evaluate("window.location.hash = '/mobiledoc/jsp/webemr/webpm/claimLookup.jsp'")
        time.sleep(4)
        try:
            page.wait_for_load_state('networkidle', timeout=10000)
        except Exception:
            pass
        _dismiss_ecw_data_error(page, 'stage2-billing-nav')
        page.screenshot(path='/tmp/fix_ivs_stage2_01_billing.png')
    except Exception as e:
        logger.error(f"Stage 2 navigation failed: {e}")

    total_deleted = 0
    MAX_DELETIONS_PER_STATUS = 200  # safety cap

    for status_label in ['Pending With Errors', 'Pending']:
        if testing_mode and total_deleted >= 1:
            logger.info("🧪 Test mode — limit reached, skipping remaining statuses")
            break

        logger.info(f"━━ Status: {status_label} ━━")

        # Set date range 07/01/2025 → today
        try:
            inputs = page.locator('input[type="text"]:visible').all()
            n = 0
            for inp in inputs:
                try:
                    val = inp.input_value()
                    if val and '/' in val and len(val) == 10:
                        if n == 0:
                            inp.click(click_count=3); inp.fill('07/01/2025')
                            inp.dispatch_event('change'); inp.dispatch_event('blur')
                        elif n == 1:
                            inp.click(click_count=3); inp.fill(today_str)
                            inp.dispatch_event('change'); inp.dispatch_event('blur')
                            break
                        n += 1
                except Exception:
                    continue
            page.keyboard.press('Escape'); time.sleep(0.5)
        except Exception as e:
            logger.warning(f"Date set failed: {e}")

        # Set status filter
        try:
            for sel in page.locator('select:visible').all():
                try:
                    target = sel.evaluate(
                        '''(el, label) => {
                            const want = label.trim().toLowerCase();
                            for (const o of el.options) {
                                if ((o.text || '').trim().toLowerCase() === want) return o.value;
                            }
                            for (const o of el.options) {
                                if ((o.text || '').trim().toLowerCase().includes(want)) return o.value;
                            }
                            return null;
                        }''',
                        status_label
                    )
                    if target is not None:
                        sel.select_option(value=str(target))
                        sel.dispatch_event('change')
                        logger.info(f"✅ Status='{status_label}'")
                        break
                except Exception:
                    continue
        except Exception as e:
            logger.warning(f"Status filter failed: {e}")
        time.sleep(0.5)

        # Apply filter
        try:
            lookup_btn = page.query_selector('#btnclaimlookup')
            if lookup_btn:
                lookup_btn.click()
                logger.info("✅ Clicked #btnclaimlookup")
                time.sleep(3)
                try:
                    page.wait_for_load_state('networkidle', timeout=15000)
                except Exception:
                    pass
                for _ in range(3):
                    if not _dismiss_ecw_data_error(page, f'stage2-after-filter-{status_label}'):
                        break
                    time.sleep(0.4)
        except Exception as e:
            logger.warning(f"Filter click failed: {e}")

        page.screenshot(path=f'/tmp/fix_ivs_stage2_02_{status_label.replace(" ", "_").lower()}.png')

        # Loop: scan current page → if test row, delete it; else try next page; stop when
        # both current page is clean AND no next page. Selectors come from the existing
        # ECW claims scraper (tr[ng-repeat*="lstClaimReport"] + #nextBtn).
        deleted_for_this_status = 0
        no_op_iters = 0  # safety: stop if too many consecutive iterations do nothing
        for iteration in range(MAX_DELETIONS_PER_STATUS):
            if testing_mode and total_deleted >= 1:
                break

            test_row_info = page.evaluate('''() => {
                const rows = document.querySelectorAll('tr[ng-repeat*="lstClaimReport"]');
                for (let i = 0; i < rows.length; i++) {
                    const row = rows[i];
                    const candidates = row.querySelectorAll('a, span.ng-binding');
                    for (const el of candidates) {
                        const t = ((el.innerText || el.textContent || '') + '').trim();
                        // Patient names look like "LastName, FirstName"
                        if (t.includes(',') && /test/i.test(t)) {
                            return { index: i, patientName: t };
                        }
                    }
                }
                return null;
            }''')

            if not test_row_info:
                # No test on current page — try next page if available
                page_state = page.evaluate('''() => {
                    const btn = document.querySelector('#nextBtn');
                    if (!btn) return { hasNext: false };
                    const disabled = btn.disabled || btn.classList.contains('disabled');
                    return { hasNext: !disabled };
                }''') or {}
                if not page_state.get('hasNext'):
                    logger.info(f"  ✓ No more test claims for '{status_label}' (deleted {deleted_for_this_status}, all pages scanned)")
                    break
                logger.info("  ↪ Current page clean — clicking Next page")
                try:
                    page.evaluate('document.querySelector("#nextBtn")?.click()')
                    time.sleep(2)
                    try:
                        page.wait_for_load_state('networkidle', timeout=8000)
                    except Exception:
                        pass
                    _dismiss_ecw_data_error(page, f'stage2-next-page-{status_label}')
                except Exception as e:
                    logger.warning(f"  ⚠️ Next page click failed: {e}")
                    break
                no_op_iters += 1
                if no_op_iters > 30:  # safety: at most 30 pages
                    logger.warning("  ⚠️ Too many pages, stopping")
                    break
                continue
            no_op_iters = 0

            patient_name = test_row_info.get('patientName', '?')
            row_idx = test_row_info.get('index', 0)
            logger.info(f"  🎯 Deleting test claim — patient='{patient_name}' (row {row_idx})")

            # ── Clean state: uncheck every already-selected checkbox (rows + master).
            # ECW's AngularJS sometimes keeps a stale selection from a previous deletion,
            # which disables Delete Claim when we add the next row's selection.
            try:
                # Master select-all (in thead) — uncheck if it's checked
                try:
                    master_loc = page.locator('thead input[type="checkbox"]').first
                    if master_loc.count() > 0 and master_loc.is_checked():
                        master_loc.uncheck(timeout=2000)
                except Exception:
                    pass
                # Individual row checkboxes — uncheck any currently checked one
                checked_locs = page.locator('tr[ng-repeat*="lstClaimReport"] input[type="checkbox"]:checked')
                for i in range(checked_locs.count()):
                    try:
                        checked_locs.nth(i).uncheck(timeout=1500)
                    except Exception:
                        try:
                            checked_locs.nth(i).click(timeout=1500)
                        except Exception:
                            pass
                time.sleep(0.3)
            except Exception:
                pass

            # Select the row's checkbox using Playwright's locator (.check() dispatches
            # the real change events AngularJS needs to enable Delete Claim).
            try:
                row_loc = page.locator('tr[ng-repeat*="lstClaimReport"]').nth(row_idx)
                cb_loc = row_loc.locator('input[type="checkbox"]:not([disabled])').first
                try:
                    cb_loc.scroll_into_view_if_needed(timeout=3000)
                except Exception:
                    pass
                cb_loc.check(timeout=5000)
                time.sleep(0.4)

                # Verify the row is actually selected (mnuClaimDelete becomes true).
                is_selected = page.evaluate('''(idx) => {
                    const rows = document.querySelectorAll('tr[ng-repeat*="lstClaimReport"]');
                    const row = rows[idx];
                    if (!row) return false;
                    const cb = row.querySelector('input[type="checkbox"]');
                    return !!(cb && cb.checked);
                }''', row_idx)
                if not is_selected:
                    logger.warning(f"  ⚠️ Row {row_idx} checkbox didn't register as checked — retrying with click()")
                    try:
                        cb_loc.click(timeout=3000)
                        time.sleep(0.3)
                    except Exception:
                        pass
            except Exception as e:
                logger.warning(f"  ⚠️ Row selection failed: {e}")
                break

            # Open Claims dropdown and click Delete Claim. The dropdown may be residually
            # open from the previous iteration, so close any open menus first, then click.
            # Verify Delete Claim becomes visible+enabled before clicking; retry up to 2x.
            delete_clicked_ok = False
            for attempt in range(2):
                try:
                    # Close any open dropdown/menu
                    try:
                        page.keyboard.press('Escape')
                        time.sleep(0.2)
                        page.click('body', position={'x': 5, 'y': 5}, force=True)
                        time.sleep(0.2)
                    except Exception:
                        pass

                    # Open the Claims dropdown
                    page.locator('#claimLookupBtn10').click(timeout=5000)
                    time.sleep(0.5)

                    # Wait for Delete Claim to become visible AND enabled (not via :not([disabled])
                    # which doesn't catch ng-disabled in all cases — read attributes directly).
                    delete_ready = False
                    for _ in range(20):  # up to ~4s
                        state = page.evaluate('''() => {
                            const el = document.querySelector('#claimLookupBtn14');
                            if (!el) return null;
                            const r = el.getBoundingClientRect();
                            const disabled = el.disabled || el.hasAttribute('disabled') || el.classList.contains('disabled');
                            return { visible: r.width > 0 && r.height > 0, disabled };
                        }''')
                        if state and state.get('visible') and not state.get('disabled'):
                            delete_ready = True
                            break
                        time.sleep(0.2)

                    if not delete_ready:
                        logger.warning(f"  ⚠️ Delete Claim not ready on attempt {attempt+1} — retrying dropdown")
                        continue

                    page.locator('#claimLookupBtn14').click(timeout=5000)
                    logger.info(f"  ✅ Clicked 'Delete Claim' (#claimLookupBtn14, attempt {attempt+1})")
                    delete_clicked_ok = True
                    time.sleep(1.0)
                    break
                except Exception as e:
                    logger.warning(f"  ⚠️ Attempt {attempt+1} failed: {e}")
                    continue

            if not delete_clicked_ok:
                try:
                    page.screenshot(path=f'/tmp/fix_ivs_stage2_delete_failed_{patient_name.replace(",", "").replace(" ", "_")}.png', full_page=True)
                except Exception:
                    pass
                logger.warning("  ⚠️ Could not click Delete Claim (screenshot saved)")
                break

            # Confirm with YES — the modal "Claim Lookup → Are you sure..." uses:
            #   <input type="button" value="Yes" ng-click="handleYesClaimLookupConfirmation();">
            # The ng-click attribute is unique (the id `deleteBtn` is duplicated for Yes/No).
            # Search every frame, poll up to ~15s, with multiple fallbacks.
            def _try_click_yes_in_context(ctx):
                try:
                    return ctx.evaluate('''() => {
                        // Primary: unique ng-click handler for Yes
                        const ngYes = document.querySelector('input[ng-click*="handleYesClaimLookupConfirmation"]');
                        if (ngYes) {
                            const r = ngYes.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) {
                                ngYes.click();
                                return 'ng-click';
                            }
                        }
                        // Secondary: input id=deleteBtn with value=Yes
                        for (const el of document.querySelectorAll('input#deleteBtn[type="button"]')) {
                            if ((el.value || '').trim().toLowerCase() === 'yes') {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) { el.click(); return 'id-value'; }
                            }
                        }
                        // Tertiary: generic Yes button/input/a
                        const candidates = Array.from(document.querySelectorAll(
                            'button, input[type="button"], input[type="submit"], a, [role="button"]'
                        ));
                        for (const el of candidates) {
                            const txt = ((el.innerText || el.textContent || el.value || '') + '').trim().toLowerCase();
                            if (txt === 'yes') {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) { el.click(); return 'text'; }
                            }
                        }
                        // Final fallback: invoke the AngularJS handler directly
                        if (window.angular) {
                            const stub = document.querySelector('[ng-controller], [ng-app], .modal-content');
                            if (stub) {
                                try {
                                    const s = angular.element(stub).scope();
                                    if (s && typeof s.handleYesClaimLookupConfirmation === 'function') {
                                        s.handleYesClaimLookupConfirmation();
                                        s.$apply && s.$apply();
                                        return 'angular-scope';
                                    }
                                } catch (e) { }
                            }
                        }
                        return null;
                    }''')
                except Exception:
                    return None

            yes_method = None
            try:
                for poll in range(30):  # up to ~15s
                    result = _try_click_yes_in_context(page)
                    if result:
                        yes_method = result
                        break
                    for fr in page.frames:
                        if fr == page.main_frame:
                            continue
                        result = _try_click_yes_in_context(fr)
                        if result:
                            yes_method = result
                            break
                    if yes_method:
                        break
                    time.sleep(0.5)
            except Exception as e:
                logger.warning(f"  ⚠️ Yes search errored: {e}")

            if not yes_method:
                try:
                    page.screenshot(path=f'/tmp/fix_ivs_stage2_yes_missing_{patient_name.replace(",", "").replace(" ", "_")}.png', full_page=True)
                except Exception:
                    pass
                logger.warning("  ⚠️ Yes confirm button not found (screenshot saved)")
                break
            logger.info(f"  ✅ Confirmed deletion (Yes via {yes_method})")

            # Wait for the "Are you sure..." modal to fully close before the next iteration
            for _ in range(20):  # up to 4s
                modal_open = page.evaluate('''() => {
                    const els = document.querySelectorAll('.modal-content, .modal.in, .modal[style*="display: block"]');
                    for (const el of els) {
                        const r = el.getBoundingClientRect();
                        if (r.width > 0 && r.height > 0) {
                            const txt = (el.innerText || '').toLowerCase();
                            if (txt.includes('are you sure')) return true;
                        }
                    }
                    return false;
                }''')
                if not modal_open:
                    break
                time.sleep(0.2)

            # Wait for table data to refresh
            time.sleep(1.5)
            try:
                page.wait_for_load_state('networkidle', timeout=10000)
            except Exception:
                pass
            _dismiss_ecw_data_error(page, f'stage2-after-delete-{status_label}')
            # Click body to dismiss any lingering dropdowns/popovers between iterations
            try:
                page.click('body', position={'x': 5, 'y': 5}, force=True)
            except Exception:
                pass

            # Re-apply the filter to fully reset AngularJS state (selection,
            # dropdown state, pagination). This is slower but eliminates the
            # "Delete Claim not ready" / sticky selection issue we hit otherwise.
            try:
                page.locator('#btnclaimlookup').click(timeout=5000)
                time.sleep(2)
                try:
                    page.wait_for_load_state('networkidle', timeout=10000)
                except Exception:
                    pass
                _dismiss_ecw_data_error(page, f'stage2-after-refilter-{status_label}')
                logger.info("  ↻ Re-applied filter to reset state")
            except Exception as e:
                logger.warning(f"  ⚠️ Re-filter failed (continuing anyway): {e}")

            deleted_for_this_status += 1
            total_deleted += 1

        logger.info(f"━━ '{status_label}': {deleted_for_this_status} test claim(s) deleted ━━")

    page.screenshot(path='/tmp/fix_ivs_stage2_99_final.png')
    logger.info(f"═══ Stage 2 complete — total test claims deleted: {total_deleted} ═══")


def _run_fix_coding_ivs_stage1(page, body, aws_client, ecw_url):
    """
    Fix Coding IVs — Stage 1: Create encounter claims.
      1. (login already done by caller)
      2. Encounters → 07/01/2025-today, "Progress Notes Done/Locked" filter, click Filter.
      3. Click "Claims IPE All" → confirm Yes → close result modal.
         (If "No encounters selected" appears, IPE already ran — skip.)

    body flags:
      - testing_mode: no-op for this stage (kept for symmetry with other stages).
    """
    from datetime import datetime as _dt
    today_str = _dt.now().strftime('%m/%d/%Y')

    logger.info("═══ Fix Coding IVs — Stage 1: Create encounter claims ═══")

    # ── Cleanup previous run artifacts so /tmp doesn't grow unbounded ──
    import glob as _glob
    removed = 0
    for prev in _glob.glob('/tmp/fix_ivs_*'):
        try:
            os.remove(prev)
            removed += 1
        except Exception:
            pass
    if removed:
        logger.info(f"🧹 Cleared {removed} file(s) from previous fix_coding_ivs runs")

    # ── Step A: Encounters page (under Billing → Encounters) ──
    logger.info("┌─ A. Navigate to Billing → Encounters")
    try:
        page.evaluate("window.location.hash = '/mobiledoc/jsp/webemr/webpm/billing/encounter.jsp'")
        time.sleep(4)
        try:
            page.wait_for_load_state('networkidle', timeout=10000)
        except Exception:
            pass
        _dismiss_ecw_data_error(page, 'after-encounters-nav')
        page.screenshot(path='/tmp/fix_ivs_01_encounters.png', full_page=True)
        logger.info("📸 /tmp/fix_ivs_01_encounters.png")
    except Exception as e:
        logger.error(f"Encounters navigation failed: {e}")

    logger.info("Setting Encounters date range 07/01/2025 → today...")
    try:
        date_inputs = page.locator('input[type="text"]:visible').all()
        set_count = 0
        for inp in date_inputs:
            try:
                val = inp.input_value()
                if val and '/' in val and len(val) == 10:
                    if set_count == 0:
                        inp.click(click_count=3); inp.fill('07/01/2025')
                        inp.dispatch_event('change'); inp.dispatch_event('blur')
                        logger.info("✅ FROM 07/01/2025")
                    elif set_count == 1:
                        inp.click(click_count=3); inp.fill(today_str)
                        inp.dispatch_event('change'); inp.dispatch_event('blur')
                        logger.info(f"✅ TO {today_str}")
                        break
                    set_count += 1
            except Exception:
                continue
        page.keyboard.press('Escape'); time.sleep(0.5)
    except Exception as e:
        logger.warning(f"Encounters date set failed: {e}")

    logger.info("Filtering for 'Progress Notes Done/Locked'...")
    try:
        selected = False
        for sel in page.locator('select:visible').all():
            try:
                target = sel.evaluate(
                    '''(el) => {
                        for (const o of el.options) {
                            const t = (o.text || '').trim().toLowerCase();
                            if (t.includes('progress notes') && t.includes('done') && t.includes('locked')) {
                                return o.value;
                            }
                        }
                        return null;
                    }'''
                )
                if target is not None:
                    sel.select_option(value=str(target))
                    sel.dispatch_event('change')
                    logger.info("✅ Select set to 'Progress Notes Done/Locked'")
                    selected = True
                    break
            except Exception:
                continue
        if not selected:
            logger.warning("⚠️ 'Progress Notes Done/Locked' option not found in any visible <select>")
    except Exception as e:
        logger.warning(f"Encounters select filter failed: {e}")

    # Click the Filter button to apply
    logger.info("Clicking Filter button (#btnFilter)...")
    try:
        filter_btn = page.query_selector('#btnFilter')
        if filter_btn:
            filter_btn.click()
            logger.info("✅ Clicked Filter")
            time.sleep(2)
            try:
                page.wait_for_load_state('networkidle', timeout=10000)
            except Exception:
                pass
            for _ in range(3):
                if not _dismiss_ecw_data_error(page, 'after-encounters-filter'):
                    break
                time.sleep(0.4)
        else:
            logger.warning("⚠️ #btnFilter not found")
    except Exception as e:
        logger.warning(f"Filter click failed: {e}")
    page.screenshot(path='/tmp/fix_ivs_02_enc_filters.png')

    logger.info("Clicking 'Claims IPE All' (#btnClaimIPEAll)...")
    skip_ipe_modal = False
    try:
        btn = page.query_selector('#btnClaimIPEAll')
        if btn:
            btn.click()
            logger.info("✅ Clicked Claims IPE All")
            time.sleep(1.5)

            # Detect whichever dialog rendered: "No encounters selected..." or the Yes/No confirm.
            decision = None
            for attempt in range(10):  # up to ~8s for a dialog to render
                try:
                    decision = page.evaluate('''() => {
                        const body = (document.body && document.body.innerText) || '';
                        const noEnc = /no encounters selected/i.test(body);
                        // Find a visible OK button (for the "No encounters" alert)
                        const okBtn = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a'))
                            .find(el => {
                                const t = ((el.innerText || el.value || '') + '').trim();
                                if (t !== 'OK' && t !== 'Ok') return false;
                                const r = el.getBoundingClientRect();
                                return r.width > 0 && r.height > 0;
                            });
                        // Find a visible Yes button (for the confirm dialog)
                        const yesBtn = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a'))
                            .find(el => {
                                const t = ((el.innerText || el.value || '') + '').trim();
                                if (t !== 'Yes') return false;
                                const r = el.getBoundingClientRect();
                                return r.width > 0 && r.height > 0;
                            });
                        if (noEnc && okBtn) { okBtn.click(); return 'no-encounters'; }
                        if (yesBtn) { yesBtn.click(); return 'yes'; }
                        return null;
                    }''')
                    if decision:
                        break
                except Exception:
                    pass
                time.sleep(0.8)

            if decision == 'no-encounters':
                logger.info("ℹ️ 'No encounters selected to run IPE' — IPE already ran. Skipping modal wait.")
                skip_ipe_modal = True
            elif decision == 'yes':
                logger.info("✅ Clicked YES on 'Run IPE for all encounters?' dialog")
            else:
                logger.warning("⚠️ Neither YES nor 'No encounters' dialog detected")

            # Settle any follow-up popups
            time.sleep(2)
            for _ in range(3):
                if not _dismiss_ecw_data_error(page, 'after-IPE-action'):
                    break
                time.sleep(0.5)

            if not skip_ipe_modal:
                logger.info("Waiting for Claims IPE result modal to finish generating...")
                close_clicked = False
                for poll in range(60):
                    try:
                        x_btn = page.query_selector('#ClaimIPEBtn1')
                        if x_btn and x_btn.is_visible():
                            x_btn.click()
                            logger.info("✅ Closed Claims IPE modal via X (#ClaimIPEBtn1)")
                            close_clicked = True
                            time.sleep(1)
                            break
                    except Exception:
                        pass
                    _dismiss_ecw_data_error(page, f'waiting-ipe-{poll}')
                    time.sleep(1)
                if not close_clicked:
                    logger.warning("⚠️ #ClaimIPEBtn1 (close X) not found within 60s — proceeding anyway")
        else:
            logger.warning("⚠️ #btnClaimIPEAll not found — verify on screenshot")
    except Exception as e:
        logger.error(f"Claims IPE All step failed: {e}")
    page.screenshot(path='/tmp/fix_ivs_03_after_ipe.png')

    logger.info("═══ Fix Coding IVs — Stage 1 complete ═══")


def _run_fix_coding_ivs_stage3(page, body, aws_client, ecw_url):
    """
    Fix Coding IVs — Stage 3: Generate Excel documentation.
      1. Navigate to Billing → Claims.
      2. For each status in ['Pending with Errors', 'Pending']:
         a. Set status filter + date range 07/01/2025 → today
         b. Click #btnclaimlookup
         c. Billing menu → View Claims Report → EXCEL (download)
      3. Combine the two Excels and upload to S3 (+ DynamoDB metadata).

    body flags:
      - testing_mode: if True, only download "Pending with Errors" and skip combine.
    """
    from datetime import datetime as _dt
    testing_mode = bool(body.get('testing_mode'))
    today_str = _dt.now().strftime('%m/%d/%Y')
    timestamp = _dt.now().strftime('%Y%m%d_%H%M%S')

    logger.info("═══ Fix Coding IVs — Stage 3: Generate Excel documentation ═══")
    if testing_mode:
        logger.info("🧪 Test mode ON — only 'Pending with Errors' will be downloaded; combine step skipped.")

    # ── Navigate to Billing → Claims ──
    logger.info("┌─ Navigate to Billing → Claims")
    page.evaluate("window.location.hash = '/mobiledoc/jsp/webemr/webpm/claimLookup.jsp'")
    time.sleep(4)
    try:
        page.wait_for_load_state('networkidle', timeout=10000)
    except Exception:
        pass
    _dismiss_ecw_data_error(page, 'stage3-after-billing-nav')
    page.screenshot(path='/tmp/fix_ivs_04_billing_claims.png')

    statuses = ['Pending with Errors']
    if not testing_mode:
        statuses.append('Pending')

    downloaded = []
    for status_label in statuses:
        logger.info(f"━━ Processing status: {status_label} ━━")

        logger.info("Setting Claims date range 07/01/2025 → today...")
        try:
            inputs = page.locator('input[type="text"]:visible').all()
            n = 0
            for inp in inputs:
                try:
                    val = inp.input_value()
                    if val and '/' in val and len(val) == 10:
                        if n == 0:
                            inp.click(click_count=3); inp.fill('07/01/2025')
                            inp.dispatch_event('change'); inp.dispatch_event('blur')
                        elif n == 1:
                            inp.click(click_count=3); inp.fill(today_str)
                            inp.dispatch_event('change'); inp.dispatch_event('blur')
                            break
                        n += 1
                except Exception:
                    continue
            page.keyboard.press('Escape'); time.sleep(0.5)
        except Exception as e:
            logger.warning(f"Claims date set failed: {e}")

        logger.info(f"Setting status filter to '{status_label}'...")
        try:
            status_set = False
            for sel in page.locator('select:visible').all():
                try:
                    target = sel.evaluate(
                        '''(el, label) => {
                            const want = label.trim().toLowerCase();
                            for (const o of el.options) {
                                if ((o.text || '').trim().toLowerCase() === want) return o.value;
                            }
                            for (const o of el.options) {
                                if ((o.text || '').trim().toLowerCase().includes(want)) return o.value;
                            }
                            return null;
                        }''',
                        status_label
                    )
                    if target is not None:
                        sel.select_option(value=str(target))
                        sel.dispatch_event('change')
                        logger.info(f"✅ Status='{status_label}'")
                        status_set = True
                        break
                except Exception:
                    continue
            if not status_set:
                logger.warning(f"⚠️ Status '{status_label}' not found in any visible <select>")
        except Exception as e:
            logger.warning(f"Status filter failed: {e}")
        time.sleep(0.5)

        page.screenshot(path=f'/tmp/fix_ivs_05a_filters_{status_label.replace(" ", "_").lower()}.png')

        logger.info("Applying filters (#btnclaimlookup → fallbacks)...")
        try:
            applied = False
            lookup_btn = page.query_selector('#btnclaimlookup')
            if lookup_btn and lookup_btn.is_visible():
                lookup_btn.click()
                logger.info("✅ Clicked #btnclaimlookup (Lookup)")
                applied = True
            else:
                fallback = page.query_selector(
                    '#btnFilter, '
                    'button:has-text("Lookup"), input[value="Lookup"], '
                    'button:has-text("Filter"), input[value="Filter"]'
                )
                if fallback:
                    fallback.click()
                    logger.info("✅ Clicked Lookup/Filter fallback")
                    applied = True
                else:
                    logger.warning("⚠️ No filter/lookup button found")

            if applied:
                time.sleep(3)
                try:
                    page.wait_for_load_state('networkidle', timeout=15000)
                except Exception:
                    pass
                for _ in range(3):
                    if not _dismiss_ecw_data_error(page, f'stage3-after-filter-{status_label}'):
                        break
                    time.sleep(0.4)

                # Wait for the claims table to actually reflect the filter (rows
                # populated by tr[ng-repeat*="lstClaimReport"]). Otherwise EXCEL
                # export may include stale data from the previous filter.
                row_count = 0
                for poll in range(15):
                    row_count = page.evaluate(
                        'document.querySelectorAll("tr[ng-repeat*=\\"lstClaimReport\\"]").length'
                    ) or 0
                    if row_count > 0:
                        break
                    time.sleep(0.5)
                logger.info(f"📊 Filtered claims table: {row_count} row(s) visible for '{status_label}'")
        except Exception as e:
            logger.warning(f"Filter apply failed: {e}")

        page.screenshot(path=f'/tmp/fix_ivs_05b_after_filter_{status_label.replace(" ", "_").lower()}.png')

        _dismiss_ecw_data_error(page, f'stage3-before-billing-menu-{status_label}')

        logger.info("Opening Billing menu → View Claims Report → EXCEL...")
        try:
            billing_btn = page.query_selector('#claimLookupBtn17')
            if billing_btn:
                billing_btn.click()
                time.sleep(0.8)
                logger.info("✅ Opened Billing dropdown")
            else:
                logger.warning("⚠️ #claimLookupBtn17 not found")

            view_report = page.query_selector('#claimLookupBtn71')
            if view_report:
                try:
                    view_report.hover()
                    time.sleep(0.4)
                except Exception:
                    pass
                try:
                    view_report.click()
                except Exception:
                    pass
                page.evaluate('''() => {
                    const li = document.querySelector('#claimLookupBtn71')?.closest('li.dropdown-submenu');
                    if (li) li.classList.add('open');
                    const ul = document.querySelector('#claimLookupUl1000');
                    if (ul) ul.style.display = 'block';
                }''')
                time.sleep(0.5)
                logger.info("✅ View Claims Report submenu opened")
            else:
                logger.warning("⚠️ #claimLookupBtn71 (View Claims Report) not found")

            page.screenshot(path=f'/tmp/fix_ivs_06_menu_{status_label.replace(" ", "_").lower()}.png')
            with page.expect_download(timeout=120000) as dl_info:
                excel_btn = page.query_selector(
                    'button[ng-click*="mnuViewReport_Click"][ng-click*="EXCEL"]'
                )
                if excel_btn and excel_btn.is_visible():
                    excel_btn.click()
                    logger.info("✅ Clicked EXCEL (direct button)")
                else:
                    logger.warning("EXCEL button not visible via selector — invoking AngularJS handler")
                    page.evaluate('''() => {
                        const btn = Array.from(document.querySelectorAll('button'))
                            .find(b => (b.getAttribute('ng-click') || '').includes("EXCEL"));
                        if (btn) {
                            if (window.angular) {
                                angular.element(btn).triggerHandler('click');
                            } else {
                                btn.click();
                            }
                        }
                    }''')
            dl = dl_info.value
            fname = f'/tmp/fix_ivs_{timestamp}_{status_label.replace(" ", "_").lower()}.xlsx'
            dl.save_as(fname)
            downloaded.append((status_label, fname))
            logger.info(f"✅ Downloaded: {fname}")
        except Exception as e:
            logger.error(f"Excel download for '{status_label}' failed: {e}")

    # ── Combine + upload ──
    if not downloaded:
        logger.warning("No Excel files were downloaded — nothing to upload.")
        logger.info("═══ Fix Coding IVs — Stage 3 complete ═══")
        return

    s3_results = []
    if len(downloaded) == 2:
        try:
            combined = _combine_excels(
                [p for _, p in downloaded],
                f'/tmp/fix_ivs_{timestamp}_combined.xlsx'
            )
            # Enrich combined.xlsx with CLAIM_TYPE (IV vs Other) by opening each
            # claim in ECW and reading its Resource field. Skippable via body flag.
            enrich_limit = body.get('enrich_limit')  # None = all rows
            if body.get('skip_enrichment'):
                logger.info("⏭️ Skipping CLAIM_TYPE enrichment (skip_enrichment=true)")
            else:
                logger.info(f"🔍 Enriching combined.xlsx with CLAIM_TYPE (limit={enrich_limit or 'all'})")
                _enrich_combined_with_claim_type(page, combined, max_claims=enrich_limit)
            s3_key = f"fix_coding_ivs/{timestamp}/combined.xlsx"
            s3_url = aws_client.upload_to_s3(combined, s3_key)
            s3_results.append(('combined', s3_url))
            logger.info(f"✅ Combined uploaded: {s3_url}")
        except Exception as e:
            logger.error(f"Combine/upload failed: {e}")

    for status, p in downloaded:
        try:
            tag = status.replace(' ', '_').lower()
            s3_url = aws_client.upload_to_s3(p, f"fix_coding_ivs/{timestamp}/{tag}.xlsx")
            s3_results.append((tag, s3_url))
        except Exception as e:
            logger.error(f"Upload {status} failed: {e}")

    try:
        meta = aws_client.dynamodb.Table('helixona-fix-coding-ivs-runs')
        meta.put_item(Item={
            'run_id': timestamp,
            'created_at': _dt.utcnow().isoformat() + 'Z',
            'testing_mode': testing_mode,
            'statuses': statuses,
            's3_results': [{'tag': t, 'url': u} for t, u in s3_results],
        })
        logger.info("✅ Run metadata saved to helixona-fix-coding-ivs-runs")
    except Exception as e:
        logger.warning(f"Metadata save skipped ({e}) — table may not exist yet")

    logger.info("═══ Fix Coding IVs — Stage 3 complete ═══")


def _read_claim_resource(page):
    """Returns the value of the 'Resource' field shown in the claim detail popup,
    or '' if not found. Used by Stage 3 to tag IV Therapy claims (Resource='IV-1',
    'IV-2', etc.) vs other claim types."""
    for ctx in [page] + list(page.frames):
        try:
            val = ctx.evaluate('''() => {
                // Strategy 1: label "Resource" + next input/text element
                const all = Array.from(document.querySelectorAll('label, td, span, div'));
                for (const el of all) {
                    const t = (el.innerText || el.textContent || '').trim();
                    if (t === 'Resource' || t === 'Resource:' || t === 'Resource :') {
                        // Find next input — search forward in DOM and within parent siblings
                        let cur = el;
                        for (let i = 0; i < 8 && cur; i++) {
                            const inp = cur.querySelector && cur.querySelector('input, select');
                            if (inp && inp.value) return inp.value.trim();
                            cur = cur.parentElement;
                        }
                        // Look at siblings of label
                        let sib = el.nextElementSibling;
                        for (let i = 0; i < 6 && sib; i++) {
                            const inp = sib.tagName === 'INPUT' || sib.tagName === 'SELECT'
                                       ? sib : sib.querySelector && sib.querySelector('input, select');
                            if (inp && inp.value) return inp.value.trim();
                            sib = sib.nextElementSibling;
                        }
                    }
                }
                // Strategy 2: input whose ng-model / id mentions resource
                const inps = document.querySelectorAll('input, select');
                for (const inp of inps) {
                    const attrs = (inp.id || '') + ' ' + (inp.name || '') + ' '
                                + (inp.getAttribute('ng-model') || '');
                    if (/resource/i.test(attrs) && inp.value) return inp.value.trim();
                }
                return '';
            }''')
            if val:
                return val
        except Exception:
            continue
    return ''


def _close_claim_popup(page):
    """Closes the claim detail popup by clicking Cancel. Best-effort — returns
    True if a Cancel/OK was clicked, False otherwise."""
    import time as _time
    for ctx in [page] + list(page.frames):
        try:
            clicked = ctx.evaluate('''() => {
                const btns = document.querySelectorAll('button, input[type="button"]');
                for (const b of btns) {
                    const t = (b.value || b.textContent || '').trim();
                    if (t === 'Cancel') {
                        if (b.offsetWidth > 0 && b.offsetHeight > 0) { b.click(); return true; }
                    }
                }
                return false;
            }''')
            if clicked:
                _time.sleep(0.4)
                return True
        except Exception:
            continue
    return False


def _enrich_combined_with_claim_type(page, combined_path, max_claims=None):
    """Opens each claim in the combined Excel, reads the Resource field, appends a
    CLAIM_TYPE column (IV / Other / Unknown). Saves back to the same path.

    Returns dict {iv:N, other:N, unknown:N, processed:N}.
    """
    from openpyxl import load_workbook
    wb = load_workbook(combined_path)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    # Find CLAIMS# column (1-indexed in openpyxl)
    claims_col = None
    for i, h in enumerate(headers):
        if h == 'CLAIMS#':
            claims_col = i + 1
            break
    if claims_col is None:
        logger.warning("No CLAIMS# column in combined.xlsx — skipping enrichment")
        return {'iv': 0, 'other': 0, 'unknown': 0, 'processed': 0}

    type_col = len(headers) + 1
    ws.cell(row=1, column=type_col, value='CLAIM_TYPE')

    counts = {'iv': 0, 'other': 0, 'unknown': 0, 'processed': 0}
    total_rows = ws.max_row - 1
    for r in range(2, ws.max_row + 1):
        if max_claims and counts['processed'] >= max_claims:
            logger.info(f"Reached enrich_limit={max_claims} — stopping enrichment")
            break
        claim_no = ws.cell(row=r, column=claims_col).value
        if not claim_no:
            continue
        claim_id = str(claim_no).strip()
        try:
            popup_found, _frame = _open_claim_popup_via_lookup(page, claim_id, wait_seconds=2)
            if not popup_found:
                ws.cell(row=r, column=type_col, value='Unknown')
                counts['unknown'] += 1
            else:
                resource = (_read_claim_resource(page) or '').strip()
                # Any Resource value whose token starts with "IV" is an IV Therapy:
                # matches "IV", "IV-1", "IV2", "IV Room", etc. Excludes substrings
                # like "PRIVATE" or "DIVERSE" by requiring a leading word boundary.
                import re as _re
                is_iv = bool(_re.match(r'^\s*IV\b', resource, _re.IGNORECASE)) \
                        or bool(_re.match(r'^\s*IV\d', resource, _re.IGNORECASE))
                if is_iv:
                    ws.cell(row=r, column=type_col, value='IV')
                    counts['iv'] += 1
                else:
                    ws.cell(row=r, column=type_col, value='Other')
                    counts['other'] += 1
                _close_claim_popup(page)
        except Exception as e:
            logger.warning(f"Enrich failed for claim {claim_id}: {e}")
            ws.cell(row=r, column=type_col, value='Unknown')
            counts['unknown'] += 1
        counts['processed'] += 1
        if counts['processed'] % 20 == 0:
            logger.info(f"🔍 Enriched {counts['processed']}/{total_rows} (IV={counts['iv']} Other={counts['other']} Unknown={counts['unknown']})")

    wb.save(combined_path)
    logger.info(f"✅ Enrichment done: IV={counts['iv']} Other={counts['other']} Unknown={counts['unknown']} (of {counts['processed']} processed)")
    return counts


def _perform_ecw_login(page, creds, aws_client) -> bool:
    """Full ECW login flow with Cloudflare Turnstile solving (via 2Captcha).

    Faithful extraction of the proven login used by claims_ecw / fix_coding_ivs.
    Handles: two-step (username → Next → password) form, hidden field population,
    Turnstile auto-solve + 2Captcha fallback, error=6 retry, error=24 security-image
    re-auth, and the V12 plugin popup. Returns True if login appears successful.

    Use this instead of a bare username/password fill — ECW gates login behind a
    Turnstile captcha, so a naive login silently lands back on the login page.
    """
    import requests as req

    ecw_url = creds.get('url', 'https://zjdvbjliaam2udi33vapp.ecwcloud.com/mobiledoc/jsp/webemr/login/newLogin.jsp')
    turnstile_sitekey = "0x4AAAAAAAEasHHO0y-ORWXs"

    logger.info(f"Navigating to ECW: {ecw_url}")
    page.goto(ecw_url, wait_until="domcontentloaded", timeout=30000)
    try:
        page.wait_for_load_state('networkidle', timeout=15000)
    except Exception:
        pass
    time.sleep(1)

    # Dismiss "Restore pages?" dialog if present
    try:
        dismiss_btn = page.query_selector('button:has-text("Dismiss"), button:has-text("Don\'t restore")')
        if dismiss_btn:
            dismiss_btn.click()
            time.sleep(1)
        else:
            for btn in page.query_selector_all('button[aria-label="Close"], .infobar-close'):
                try:
                    btn.click()
                except Exception:
                    pass
            page.keyboard.press('Escape')
            time.sleep(0.5)
    except Exception:
        pass

    current_url = page.url
    logger.info(f"Landed on: {page.title()} | {current_url[:120]}")

    # ECW may use iframes — discover the login form context
    login_frame = page
    for f in page.frames:
        try:
            if f.query_selector('input[name="userName"], input[name="username"], input[id="userName"], input[type="text"]'):
                login_frame = f
                logger.info(f"Login form found in frame: {f.url[:80]}")
                break
        except Exception:
            continue

    is_login_page = 'login' in current_url.lower() or 'Login' in (page.title() or '')
    if not is_login_page:
        logger.info("🎉 Already logged in via persistent session!")
        return True

    logger.info("Login form detected — ECW two-step flow...")

    # STEP 1: Enter username in #doctorID
    username_input = login_frame.query_selector('#doctorID')
    if username_input:
        username_input.click()
        username_input.fill(creds['username'])
        logger.info("✅ Entered username")
    else:
        logger.warning("Could not find #doctorID field")

    time.sleep(0.5)

    next_btn = login_frame.query_selector('#nextStep')
    if next_btn:
        next_btn.click()
        logger.info("✅ Clicked Next Step")
    else:
        submit = login_frame.query_selector('input[type="submit"]')
        if submit:
            submit.click()
            logger.info("✅ Clicked submit button")

    try:
        page.wait_for_load_state('networkidle', timeout=10000)
    except Exception:
        pass
    time.sleep(1)

    # Re-discover the login frame (page may have reloaded)
    login_frame = page
    for f in page.frames:
        try:
            if f.query_selector('#passwordField, input[type="password"]'):
                login_frame = f
                logger.info(f"Password form found in frame: {f.url[:80]}")
                break
        except Exception:
            continue

    # STEP 2: Enter password (#passwordField on the password page)
    password_input = (login_frame.query_selector('#passwordField')
                      or login_frame.query_selector('input[type="password"]:visible')
                      or login_frame.query_selector('input[type="password"]'))
    if password_input:
        password_input.click()
        password_input.fill(creds['password'])
        logger.info("✅ Entered password")
        try:
            login_frame.evaluate(f'''
                const usernameHidden = document.querySelector('#doctorIDVal');
                const passwordHidden = document.querySelector('#password');
                if (usernameHidden) usernameHidden.value = "{creds['username']}";
                if (passwordHidden) passwordHidden.value = document.querySelector('#passwordField')?.value || "";
            ''')
            logger.info("✅ Populated hidden username + password fields")
        except Exception as e:
            logger.warning(f"Could not set hidden fields: {e}")
    else:
        logger.warning("Could not find password field after Step 1")

    time.sleep(0.5)

    # Scroll down to make the Turnstile checkbox visible
    try:
        login_frame.evaluate('window.scrollTo(0, document.body.scrollHeight)')
        time.sleep(1)
    except Exception:
        pass

    # STEP 3: Handle Cloudflare Turnstile
    logger.info("STEP 3: Solving Cloudflare Turnstile...")
    token_found = False
    try:
        token = login_frame.evaluate('document.querySelector("[name=\'cf-turnstile-response\']")?.value || ""')
        if token:
            logger.info(f"✅ Turnstile auto-solved! ({len(token)} chars)")
            token_found = True
    except Exception:
        pass

    if not token_found:
        captcha_key = None
        try:
            captcha_key = aws_client.get_secret("captcha_api_key").get('api_key', '')
        except Exception:
            pass

        if captcha_key:
            try:
                turnstile_page_url = login_frame.url or page.url
                logger.info(f"Sending Turnstile to 2Captcha (url: {turnstile_page_url[:60]})")
                resp = req.post("https://2captcha.com/in.php", data={
                    'key': captcha_key,
                    'method': 'turnstile',
                    'sitekey': turnstile_sitekey,
                    'pageurl': turnstile_page_url,
                    'json': 1,
                }, timeout=15)
                task_data = resp.json()
                logger.info(f"2Captcha response: {task_data}")
                if task_data.get('status') == 1:
                    task_id = task_data['request']
                    logger.info(f"2Captcha task: {task_id} — polling...")
                    for poll in range(40):  # up to ~120s
                        time.sleep(3)
                        result = req.get(
                            f"https://2captcha.com/res.php?key={captcha_key}&action=get&id={task_id}&json=1",
                            timeout=15
                        ).json()
                        if result.get('status') == 1:
                            solved_token = result['request']
                            logger.info(f"✅ 2Captcha solved! Token: ({len(solved_token)} chars)")
                            login_frame.evaluate('''
                                (token) => {
                                    const cfResp = document.querySelector("[name='cf-turnstile-response']");
                                    const tsResp = document.querySelector("#turnstileResponse");
                                    if (cfResp) cfResp.value = token;
                                    if (tsResp) tsResp.value = token;
                                    if (window.turnstile) {
                                        try {
                                            const containers = document.querySelectorAll('[id*="cf-turnstile"], .cf-turnstile');
                                            containers.forEach(c => {
                                                const cbName = c.getAttribute('data-callback');
                                                if (cbName && window[cbName]) { window[cbName](token); }
                                            });
                                        } catch(e) {}
                                    }
                                }
                            ''', solved_token)
                            logger.info("Token injected + callback triggered")
                            token_found = True
                            break
                        elif result.get('request') == 'CAPCHA_NOT_READY':
                            if poll % 4 == 0:
                                logger.info(f"  Waiting... {(poll+1)*5}s")
                        else:
                            logger.warning(f"2Captcha error: {result}")
                            break
                else:
                    logger.warning(f"2Captcha submit failed: {task_data}")
            except Exception as e:
                logger.error(f"2Captcha failed: {e}")
        else:
            logger.warning("No 2Captcha API key configured — set 'captcha_api_key' secret")

    if not token_found:
        logger.warning("⚠️ Turnstile not solved — login will likely fail")

    # STEP 4: Click the "Log In" button
    login_btn = login_frame.query_selector(
        '#Login, input[value="Log In"], input[type="submit"], button:has-text("Log In")'
    )
    if login_btn:
        login_btn.click()
        logger.info("✅ Clicked Log In button")
    else:
        page.keyboard.press('Enter')
        logger.info("✅ Pressed Enter to submit")

    time.sleep(random.uniform(2, 3))
    try:
        page.context.on("dialog", lambda dialog: dialog.dismiss())
    except Exception:
        pass

    post_url = page.url
    logger.info(f"After login: {page.title()} | {post_url[:120]}")

    # Handle error=6 — retry (first attempt sometimes fails)
    if 'error=6' in post_url:
        logger.info("Got error=6, waiting/retrying...")
        time.sleep(2)
        post_url = page.url
        logger.info(f"After wait: {page.title()} | {post_url[:120]}")

    # Handle error=24 — "VERIFY YOUR SECURITY IMAGE" re-authentication
    if 'error=24' in post_url or 'getPwdPage' in post_url:
        logger.info("🔐 Security image re-authentication detected (error=24)")
        time.sleep(1)
        reauth_frame = page
        for f in page.frames:
            if f.query_selector('#passwordField, input[type="password"]'):
                reauth_frame = f
                break
        try:
            pwd_field = reauth_frame.query_selector('#passwordField, input[type="password"]')
            if pwd_field:
                pwd_field.fill(creds['password'])
                logger.info("✅ Re-entered password on re-auth page")
                try:
                    reauth_frame.evaluate('''(pwd) => {
                        const hidden = document.querySelector('#password');
                        if (hidden) hidden.value = pwd;
                    }''', creds['password'])
                except Exception:
                    pass
                # Turnstile on re-auth page
                try:
                    ts_field = reauth_frame.query_selector('[name="cf-turnstile-response"]')
                    ts_value = reauth_frame.evaluate('document.querySelector("[name=\'cf-turnstile-response\']")?.value || ""') if ts_field else ''
                    if ts_field and not ts_value:
                        logger.info("Turnstile found on re-auth page — solving...")
                        captcha_key = ''
                        try:
                            captcha_key = aws_client.get_secret("captcha_api_key").get('api_key', '')
                        except Exception:
                            pass
                        if captcha_key:
                            reauth_url = reauth_frame.url or page.url
                            task_data = req.post("https://2captcha.com/in.php", data={
                                'key': captcha_key, 'method': 'turnstile',
                                'sitekey': turnstile_sitekey, 'pageurl': reauth_url, 'json': 1,
                            }, timeout=15).json()
                            if task_data.get('status') == 1:
                                task_id = task_data['request']
                                for poll in range(40):
                                    time.sleep(3)
                                    result = req.get(
                                        f"https://2captcha.com/res.php?key={captcha_key}&action=get&id={task_id}&json=1",
                                        timeout=15
                                    ).json()
                                    if result.get('status') == 1:
                                        reauth_frame.evaluate('''(token) => {
                                            const cfResp = document.querySelector("[name='cf-turnstile-response']");
                                            const tsResp = document.querySelector("#turnstileResponse");
                                            if (cfResp) cfResp.value = token;
                                            if (tsResp) tsResp.value = token;
                                        }''', result['request'])
                                        logger.info("✅ Turnstile solved on re-auth page")
                                        break
                                    elif result.get('request') != 'CAPCHA_NOT_READY':
                                        break
                    elif ts_value:
                        logger.info("✅ Turnstile already solved on re-auth page")
                except Exception as e:
                    logger.warning(f"Turnstile handling on re-auth page: {e}")

                time.sleep(0.5)
                submit_btn = reauth_frame.query_selector(
                    '#Login, input[type="submit"], button:has-text("Log In"), '
                    'button:has-text("Continue"), input[value="Log In"]'
                )
                if submit_btn:
                    submit_btn.click()
                    logger.info("✅ Clicked submit on re-auth page")
                else:
                    reauth_frame.evaluate('document.querySelector("form")?.submit()')
                    logger.info("✅ Submitted re-auth form programmatically")
                time.sleep(random.uniform(3, 5))
                post_url = page.url
                logger.info(f"After re-auth: {page.title()} | {post_url[:120]}")
            else:
                logger.warning("No password field found on re-auth page")
        except Exception as e:
            logger.error(f"Re-auth handling failed: {e}")

    # Handle V12 Plugin popup — click "Ignore and continue"
    try:
        for f in page.frames:
            ignore_btn = f.query_selector('text="Ignore and continue"')
            if ignore_btn:
                ignore_btn.click()
                logger.info("✅ Dismissed V12 Plugin popup")
                time.sleep(1)
                break
    except Exception:
        pass

    post_url = page.url
    if 'login' not in post_url.lower():
        logger.info("🎉 ECW LOGIN SUCCESS!")
        # Wait for the SPA shell to finish "Building your user experience"
        for _wait_i in range(30):
            try:
                if 'Building your user experience' not in page.inner_text('body'):
                    break
            except Exception:
                pass
            time.sleep(1)
        time.sleep(2)
        return True

    logger.warning("Login may have failed — still on login page")
    return False


def _open_claim_popup_via_lookup(page, claim_id, wait_seconds=3):
    """Open (or re-open) the ECW claim detail popup via the Claim Lookup input.
    Returns (popup_found, best_frame) where best_frame is the iframe whose body
    contains the claim popup (Cancel/Prog. Notes buttons), or None.
    """
    import time as _time

    # 1. Set claim ID in the lookup input
    lookup_set = False
    for ctx in [page] + list(page.frames):
        try:
            result = ctx.evaluate('''(claimId) => {
                const input = document.querySelector('input[id^="claimLookupIpt"]')
                           || document.querySelector('input[ng-model="_InvId"]');
                if (!input) return { ok: false };
                input.value = '';
                input.focus();
                const ngEl = window.angular && window.angular.element(input);
                if (ngEl && ngEl.controller) {
                    const scope = ngEl.scope();
                    if (scope) { scope._InvId = claimId; scope.$apply(); }
                }
                input.value = claimId;
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                return { ok: true };
            }''', str(claim_id))
            if result and result.get('ok'):
                _time.sleep(0.3)
                clicked = ctx.evaluate('''() => {
                    const btns = document.querySelectorAll('button, input[type="button"]');
                    for (const btn of btns) {
                        const text = (btn.textContent || btn.value || '').trim();
                        if (text === 'Lookup' || text === 'Look Up') { btn.click(); return true; }
                    }
                    return false;
                }''')
                if not clicked:
                    try:
                        input_el = ctx.query_selector('input[id^="claimLookupIpt"]') or ctx.query_selector('input[ng-model="_InvId"]')
                        if input_el:
                            input_el.press('Enter')
                    except Exception:
                        pass
                lookup_set = True
                break
        except Exception:
            continue

    if not lookup_set:
        return (False, None)

    _time.sleep(wait_seconds)

    # 2. Verify popup opened — look for Cancel/OK + Prog. Notes
    popup_found = False
    for f_ctx in page.frames:
        try:
            has_popup = f_ctx.evaluate('''() => {
                const btns = document.querySelectorAll('button');
                let hasCancel = false, hasProgNotes = false, hasOK = false;
                for (const btn of btns) {
                    const t = (btn.value || btn.textContent || '').trim();
                    if (t === 'Cancel') hasCancel = true;
                    if (t.includes('Prog') && t.includes('Note')) hasProgNotes = true;
                    if (t === 'OK') hasOK = true;
                }
                return (hasCancel || hasOK) && hasProgNotes;
            }''')
            if has_popup:
                popup_found = True
                break
        except Exception:
            continue
    if not popup_found:
        try:
            popup_found = page.evaluate('''() => {
                const bodyText = document.body.innerText || '';
                return bodyText.includes('Prog. Notes') && bodyText.includes('Cancel');
            }''')
        except Exception:
            pass

    if not popup_found:
        return (False, None)

    # 3. Find best frame (the one with the claim popup buttons)
    best_frame = None
    best_score = 0
    for ctx in page.frames:
        try:
            info = ctx.evaluate('''() => {
                const text = document.body?.innerText || '';
                const hasHCFA = text.includes('Print HCFA') || text.includes('HCFA');
                const hasClaimNo = text.includes('Claim No');
                const hasProgNotes = text.includes('Prog. Notes') || text.includes('Prog Notes');
                let hasVisibleBtn = false;
                const btns = document.querySelectorAll('button');
                for (const btn of btns) {
                    const t = (btn.value || btn.textContent || '').trim();
                    if ((t === 'Cancel' || t.includes('Prog')) && btn.offsetWidth > 0) {
                        hasVisibleBtn = true; break;
                    }
                }
                return { hasHCFA, hasClaimNo, hasProgNotes, hasVisibleBtn, btnCount: btns.length };
            }''')
            score = 0
            if info.get('hasHCFA'): score += 1
            if info.get('hasClaimNo'): score += 1
            if info.get('hasProgNotes'): score += 2
            if info.get('hasVisibleBtn'): score += 3
            if info.get('btnCount', 0) > 5: score += 1
            if score > best_score:
                best_score = score
                best_frame = ctx
        except Exception:
            continue

    return (True, best_frame if best_score >= 2 else None)


def _close_all_hub_popups(page, preserve_claim_popup=True):
    """Close nested PHM Hub / Patient Hub / Edit modal / alert popups.

    When preserve_claim_popup=True (default), the underlying claim detail popup
    is left intact — detected by markers like "Print HCFA", "Prog. Notes", or
    "Adjustments" in the modal body. Without this guard, a generic .close click
    would also dismiss the claim detail popup, breaking subsequent HCFA / Prog
    Notes / Encounter-file extraction.
    """
    import time as _time

    CLAIM_POPUP_MARKERS = ['Print HCFA', 'Prog. Notes', 'Prog Notes', 'view_hcfa', 'Adjustments']

    try:
        # Step 1: Close any alert dialogs — click OK (only known alert texts)
        for f_ctx in page.frames:
            try:
                f_ctx.evaluate('''() => {
                    const btns = document.querySelectorAll('button');
                    for (const btn of btns) {
                        const text = (btn.textContent || '').trim();
                        if (text === 'OK' && btn.offsetWidth > 0 && btn.offsetHeight > 0) {
                            const parent = btn.closest('.modal, .ui-dialog, [role="dialog"], .alert, .bootbox');
                            if (parent) {
                                const parentText = parent.textContent || '';
                                if (parentText.includes('already open') ||
                                    parentText.includes('Session Expiration') ||
                                    parentText.includes('Access to the facility')) {
                                    btn.click();
                                    return true;
                                }
                            }
                        }
                    }
                    return false;
                }''')
            except Exception:
                continue
        _time.sleep(0.3)

        # Step 2: Close Edit Goal/Intervention modal — Cancel button.
        # Constrain to modals that look like the Edit modal (not the claim popup).
        for f_ctx in page.frames:
            try:
                f_ctx.evaluate('''(claimMarkers) => {
                    const btns = document.querySelectorAll('button');
                    for (const btn of btns) {
                        const text = (btn.textContent || '').trim();
                        if (text === 'Cancel' && btn.offsetWidth > 0) {
                            const modal = btn.closest('.modal, .modal-content, [class*="edit"], [class*="Edit"]');
                            if (!modal) continue;
                            const modalText = modal.textContent || '';
                            // Skip the claim detail popup
                            if (claimMarkers.some(m => modalText.includes(m))) continue;
                            // Require Edit-Intervention/Goal text or single-date-picker (intervention edit)
                            const looksLikeEdit = modalText.includes('Edit Intervention') ||
                                                  modalText.includes('Edit Goal') ||
                                                  modal.querySelector('single-date-picker');
                            if (looksLikeEdit) {
                                btn.click();
                                return true;
                            }
                        }
                    }
                    return false;
                }''', CLAIM_POPUP_MARKERS)
            except Exception:
                continue
        _time.sleep(0.3)

        # Step 3 + 4: Close PHM Hub and Patient Hub modals — but NEVER the claim popup.
        # We loop a couple times to peel back nested hubs.
        for _round in range(3):
            closed_any = False
            for f_ctx in page.frames:
                try:
                    did_close = f_ctx.evaluate('''(claimMarkers) => {
                        const closeButtons = document.querySelectorAll(
                            '.close, button.close, [aria-label="Close"], .btn-close, i.icon-close'
                        );
                        for (const btn of closeButtons) {
                            if (!(btn.offsetWidth > 0 && btn.offsetHeight > 0)) continue;
                            // Climb to the containing modal/dialog and inspect its body
                            const modal = btn.closest('.modal, .ui-dialog, [role="dialog"], [class*="popup"], [class*="Popup"], [class*="hub"], [class*="Hub"]')
                                          || btn.parentElement;
                            if (!modal) continue;
                            const modalText = modal.textContent || '';
                            // Hard skip: claim detail popup
                            if (claimMarkers.some(m => modalText.includes(m))) continue;
                            btn.click();
                            return true;
                        }
                        return false;
                    }''', CLAIM_POPUP_MARKERS)
                    if did_close:
                        closed_any = True
                        break
                except Exception:
                    continue
            _time.sleep(0.4)
            if not closed_any:
                break

        # Step 5: Final OK dismissal for stray alerts (skip claim-popup OK).
        for f_ctx in page.frames:
            try:
                f_ctx.evaluate('''(claimMarkers) => {
                    const btns = document.querySelectorAll('button');
                    for (const btn of btns) {
                        const text = (btn.textContent || '').trim();
                        if (text === 'OK' && btn.offsetWidth > 0) {
                            const modal = btn.closest('.modal, .ui-dialog, [role="dialog"], .alert, .bootbox');
                            if (modal) {
                                const modalText = modal.textContent || '';
                                if (claimMarkers.some(m => modalText.includes(m))) continue;
                            }
                            btn.click();
                            return true;
                        }
                    }
                    return false;
                }''', CLAIM_POPUP_MARKERS)
            except Exception:
                continue
        _time.sleep(0.3)

        logger.info("✅ Closed nested Hub popups (claim popup preserved)")
    except Exception as e:
        logger.warning(f"Hub popup cleanup error: {e}")


# Office-visit E/M CPT codes: new patient 99201-99205, established 99211-99215.
# Claims billed with one of these need only HCFA + IV Note — no Progress Note
# (encounter file). Anything else is treated as IV-therapy (Progress Note required).
OFFICE_VISIT_CPTS = {
    '99201', '99202', '99203', '99204', '99205',
    '99211', '99212', '99213', '99214', '99215',
}


def is_office_visit(cpt_value) -> bool:
    """True if the claim's CPT contains an office-visit E/M code.

    `cpt_value` may be a single code, a comma/space separated list, or None.
    """
    if not cpt_value:
        return False
    import re as _re
    codes = _re.findall(r'\d{5}', str(cpt_value))
    return any(c in OFFICE_VISIT_CPTS for c in codes)


def _cpt_from_hcfa_pdf(pdf_path):
    """Parse CPT/HCPCS from a HCFA PDF (Box 24D). Returns office-visit codes
    if present, else any 9xxxx CPT / HCPCS codes, else None. Reused by the
    live HCFA parse and the backfill path when HCFA capture is skipped."""
    try:
        import pdfplumber
        import re as _re
        with pdfplumber.open(pdf_path) as _pdf:
            txt = '\n'.join((p.extract_text() or '') for p in _pdf.pages)
    except Exception:
        return None
    all5 = set(_re.findall(r'\b\d{5}\b', txt))
    hcpcs = set(_re.findall(r'\b[A-Z]\d{4}\b', txt))
    office = sorted(c for c in all5 if c in OFFICE_VISIT_CPTS)
    if office:
        return ', '.join(office)
    cpt_like = sorted(c for c in all5 if c and c[0] == '9')
    return ', '.join(cpt_like + sorted(hcpcs)) or None


def _subscriber_id_from_hcfa_pdf(pdf_path):
    """Extract the PRIMARY subscriber id from HCFA-1500 box 1a ("INSURED'S I.D.
    NUMBER") of the PDF that ECW generated for the claim.

    This is the RELIABLE source of the subscriber id. ECW renders box 1a from
    its own database, so it is independent of the flaky on-screen insurance-grid
    scrape (which reads stale popups / the wrong row and is non-deterministic).
    Verified across all captured claims: exactly one id-token sits in the box-1a
    region per form, the payer ZIP above it is never matched, and the same patient
    always yields the same id.

    Box 1a is the top-right of page 1: x in 0.55-0.92, y in 0.10-0.18 of the page.
    The value is an alphanumeric token >=8 chars with >=4 digits (covers R60056135,
    XED913505676, OCO903699558, FNN100011240352, WUE9404090GJ, ...). Returns the
    uppercased id, or None if the form layout doesn't match (caller keeps fallback).
    """
    import re as _re
    ID_TOKEN = _re.compile(r'^[A-Z0-9]{8,}$')
    BOX1A_X = (0.55, 0.92)
    BOX1A_Y = (0.10, 0.18)
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as _pdf:
            pg = _pdf.pages[0]
            W, H = pg.width, pg.height
            words = pg.extract_words()
    except Exception:
        return None
    cands = []
    for w in words:
        tok = w['text'].strip().upper()
        if not ID_TOKEN.match(tok) or sum(c.isdigit() for c in tok) < 4:
            continue
        fx, fy = w['x0'] / W, w['top'] / H
        if BOX1A_X[0] <= fx <= BOX1A_X[1] and BOX1A_Y[0] <= fy <= BOX1A_Y[1]:
            cands.append((fy, fx, tok))
    if not cands:
        return None
    cands.sort()  # topmost, then leftmost
    return cands[0][2]


def _sync_ecw_claim_visibility(claims_table, present_ids):
    """Reconcile DynamoDB with what is currently visible in ECW WITHOUT deleting.

    Claims that fall out of ECW's filtered claim-lookup (e.g. once submitted or
    out of the date window) are marked ecw_visible=False (archived) instead of
    being removed, so submitted/historical claims — their FLNs, subscriber ids,
    document paths and audit trail — are NEVER lost. Claims that reappear in ECW
    are restored to ecw_visible=True. The dashboard filters on ecw_visible.

    Replaces the old destructive delete_item sync that wiped historical claims.
    """
    present = {str(x) for x in (present_ids or set())}
    ts = time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())
    items, kw = [], {}
    while True:
        r = claims_table.scan(**kw)
        items += r.get('Items', [])
        if not r.get('LastEvaluatedKey'):
            break
        kw['ExclusiveStartKey'] = r['LastEvaluatedKey']
    archived = restored = 0
    for item in items:
        if item.get('source') != 'ECW':
            continue  # never touch non-ECW claims
        cid = item.get('claim_id')
        if not cid:
            continue
        is_present = str(cid) in present
        visible = item.get('ecw_visible')
        try:
            if not is_present and visible is not False:
                claims_table.update_item(
                    Key={'claim_id': cid},
                    UpdateExpression='SET ecw_visible = :f, ecw_removed_at = :t',
                    ExpressionAttributeValues={':f': False, ':t': ts},
                )
                archived += 1
                logger.info(f"📦 Archived claim {cid} ({item.get('patient_name', '?')}) — no longer in ECW (KEPT in DB)")
            elif is_present and visible is False:
                claims_table.update_item(
                    Key={'claim_id': cid},
                    UpdateExpression='SET ecw_visible = :t REMOVE ecw_removed_at',
                    ExpressionAttributeValues={':t': True},
                )
                restored += 1
        except Exception as _vis_e:
            logger.warning(f"Visibility update failed for claim {cid}: {_vis_e}")
    logger.info(f"📦 ECW visibility sync: archived={archived}, restored={restored} (0 deletions — claims are never removed)")
    return archived, restored


def _dump_insurance_grid(page, claim_id, aws_client):
    """READ-ONLY diagnostic: dump the ECW claim "Insurances & Payment" grid.

    Triggered by body flag dump_insurance_grid (testing only). Captures the real
    DOM/Angular structure we need to fix the Subscriber No extraction:
      - every span[ng-bind="insurance.SubscriberNo"] on the page (to detect the
        HCFA-section duplicate that the current Strategy-1 querySelector grabs),
      - the insurance grid headers,
      - per row: the Angular scope `.insurance` object (SubscriberNo, BillTo,
        IsPrimary, sequence, Name, InsId...), the ng-repeat expression, cell
        texts, whether a checkbox/radio is checked, and outerHTML,
      - the Claim No shown in the popup (to confirm we are NOT reading a stale
        popup from a prior claim),
      - what the CURRENT (buggy) extraction returns, for side-by-side comparison.

    Writes JSON to /tmp/insurance_grid_dump_<claim_id>.json, uploads it to S3, and
    logs a compact summary. Modifies nothing in ECW (only a best-effort tab click).
    """
    import json as _json
    DUMP_JS = r'''() => {
      const clean = s => (s || '').replace(/\s+/g, ' ').trim();
      const out = { url: location.href, found: false };

      // Best-effort: show the Insurances & Payment tab so the grid is rendered.
      try {
        for (const el of document.querySelectorAll('a,button,li,span,div')) {
          const t = clean(el.textContent).toLowerCase();
          if ((t === 'insurances & payment' || t === 'insurance & payment') && el.offsetParent !== null) {
            el.click();
            break;
          }
        }
      } catch (e) {}

      // Every SubscriberNo binding on the page — count + their containing table id.
      const binds = Array.from(document.querySelectorAll('[ng-bind="insurance.SubscriberNo"]'));
      out.subscriberno_binding_count = binds.length;
      out.subscriberno_binding_values = binds.map(b => clean(b.textContent));

      // What the current extraction (Strategy 1) returns right now.
      const s1 = document.querySelector('span[ng-bind="insurance.SubscriberNo"]');
      out.current_strategy1_value = s1 ? clean(s1.textContent) : null;

      // Locate the insurance grid.
      let grid = binds.length ? binds[0].closest('table') : null;
      if (!grid) {
        for (const tb of document.querySelectorAll('table')) {
          for (const h of tb.querySelectorAll('th, thead td')) {
            if (clean(h.textContent).toLowerCase().indexOf('subscriber') === 0) { grid = tb; break; }
          }
          if (grid) break;
        }
      }

      // Claim No shown in the popup (stale-popup detector).
      try {
        const m = clean(document.body.innerText).match(/Claim\s*(?:No|#|Number)[:\s]*([0-9]{3,})/i);
        out.popup_claim_no = m ? m[1] : null;
      } catch (e) {}

      if (!grid) return out;
      out.found = true;

      out.headers = Array.from(grid.querySelectorAll('thead th, thead td, tr:first-child th, tr:first-child td'))
        .map((h, i) => ({ i: i, cellIndex: h.cellIndex, text: clean(h.textContent) }));

      const safeScope = (el) => {
        try {
          if (!(window.angular && angular.element)) return null;
          const sc = angular.element(el).scope();
          const ins = sc && sc.insurance;
          if (!ins) return null;
          const o = {};
          for (const k in ins) {
            if (k.charAt(0) === '$') continue;
            const v = ins[k];
            if (v === null || typeof v === 'string' || typeof v === 'number' || typeof v === 'boolean') o[k] = v;
          }
          return o;
        } catch (e) { return { __error: String(e) }; }
      };

      out.rows = Array.from(grid.querySelectorAll('tr')).map((r, idx) => ({
        idx: idx,
        ng_repeat: r.getAttribute('ng-repeat') || r.getAttribute('ng-repeat-start') || null,
        hidden: r.offsetParent === null,
        has_th: !!r.querySelector('th'),
        checked_input: !!r.querySelector('input[type="radio"]:checked, input[type="checkbox"]:checked'),
        cells: Array.from(r.querySelectorAll('td,th')).map(c => clean(c.textContent)),
        scope_insurance: safeScope(r),
        outer_html: (r.outerHTML || '').slice(0, 1500),
      }));

      return out;
    }'''

    logger.info(f"═══ INSURANCE GRID DUMP (read-only) for claim {claim_id} ═══")

    # The grid lives in the claim-popup frame — try the page and every frame,
    # keep whichever actually contains the grid.
    best = None
    for ctx in [page] + list(page.frames):
        try:
            res = ctx.evaluate(DUMP_JS)
        except Exception as e:
            logger.warning(f"  dump eval failed in a frame: {e}")
            continue
        if not res:
            continue
        if res.get('found') and (best is None or not best.get('found')):
            best = res
        elif best is None:
            best = res  # keep something even if no grid, for diagnostics

    if not best:
        logger.warning("  ⚠️ Could not evaluate the dump JS in any frame.")
        return

    # Persist full JSON + upload to S3.
    out_path = f'/tmp/insurance_grid_dump_{claim_id}.json'
    try:
        with open(out_path, 'w') as f:
            _json.dump(best, f, indent=2, default=str)
        logger.info(f"  📄 Wrote {out_path}")
    except Exception as e:
        logger.warning(f"  Could not write dump file: {e}")
    try:
        s3_url = aws_client.upload_to_s3(out_path, f"diagnostics/insurance_grid/{claim_id}.json")
        logger.info(f"  ☁️  Uploaded: {s3_url}")
    except Exception as e:
        logger.warning(f"  S3 upload skipped: {e}")
    try:
        page.screenshot(path=f'/tmp/insurance_grid_dump_{claim_id}.png', full_page=True)
    except Exception:
        pass

    # Compact, human-readable summary in the logs.
    logger.info(f"  popup_claim_no={best.get('popup_claim_no')!r}  (requested claim_id={claim_id})")
    if str(best.get('popup_claim_no') or '') and str(best.get('popup_claim_no')) != str(claim_id):
        logger.warning("  ⚠️ popup_claim_no != requested claim_id — possible STALE POPUP being read!")
    logger.info(f"  SubscriberNo bindings on page: {best.get('subscriberno_binding_count')} "
                f"→ values={best.get('subscriberno_binding_values')}")
    logger.info(f"  current Strategy-1 would capture: {best.get('current_strategy1_value')!r}")
    logger.info(f"  grid found: {best.get('found')}")
    if best.get('headers'):
        logger.info(f"  headers: {[h['text'] for h in best['headers']]}")
    for row in best.get('rows', []):
        if row.get('has_th') and not row.get('cells'):
            continue
        logger.info(f"  row[{row['idx']}] hidden={row['hidden']} checked_input={row['checked_input']} "
                    f"ng_repeat={row.get('ng_repeat')!r}")
        logger.info(f"      cells={row.get('cells')}")
        logger.info(f"      scope.insurance={row.get('scope_insurance')}")
    logger.info("═══ END INSURANCE GRID DUMP ═══")


IV_CORRECTIONS_TASKS = {'fix_coding_ivs'}


def process_message(message: dict, aws_client: AWSClient):
    body = json.loads(message.get('Body', '{}'))
    task_type = body.get('task_type')

    logger.info(f"═══ Processing task: {task_type} ═══")

    # Defense in depth: queues are already separate by bot_role, but if a
    # message slips into the wrong queue, refuse it rather than run a workflow
    # the wrong bot was not deployed to handle.
    from src.config import settings as _settings
    if _settings.bot_role == 'iv_corrections' and task_type not in IV_CORRECTIONS_TASKS:
        logger.warning(f"BOT_ROLE=iv_corrections refused task_type={task_type}")
        return
    if _settings.bot_role == 'submissions' and task_type in IV_CORRECTIONS_TASKS:
        logger.warning(f"BOT_ROLE=submissions refused task_type={task_type}")
        return

    # ──────────────────────────────────────
    # STEP 0: ECW Claims Extraction
    # ──────────────────────────────────────
    if task_type in ('test_ecw_login', 'claims_ecw', 'fix_coding_ivs'):
        if task_type == 'fix_coding_ivs':
            logger.info("STEP 0: ECW Login (for Fix Coding IVs)...")
        else:
            logger.info("STEP 0: ECW Claims Extraction...")

        # ECW uses Turnstile captcha (not IP blocking) — skip proxy for faster loads
        # Proxy is only needed for Blue Shield portal which blocks datacenter IPs
        manager = BrowserManager().start(proxy_config=None)
        try:
            page = manager.new_page()
            creds = aws_client.get_secret("ecw_credentials")
            ecw_url = creds.get('url', 'https://zjdvbjliaam2udi33vapp.ecwcloud.com/mobiledoc/jsp/webemr/login/newLogin.jsp')

            logger.info(f"Navigating to: {ecw_url}")
            page.goto(ecw_url, wait_until="domcontentloaded", timeout=30000)
            try:
                page.wait_for_load_state('networkidle', timeout=15000)
            except Exception:
                pass
            time.sleep(1)

            # Dismiss "Restore pages?" dialog if present
            try:
                dismiss_btn = page.query_selector('button:has-text("Dismiss"), button:has-text("Don\'t restore")')
                if dismiss_btn:
                    dismiss_btn.click()
                    logger.info("Dismissed 'Restore pages' dialog")
                    time.sleep(1)
                else:
                    # Try closing via X button on the infobar
                    close_btns = page.query_selector_all('button[aria-label="Close"], .infobar-close')
                    for btn in close_btns:
                        try:
                            btn.click()
                            logger.info("Closed info bar")
                        except Exception:
                            pass
                    # Press Escape to dismiss any dialogs
                    page.keyboard.press('Escape')
                    time.sleep(0.5)
            except Exception:
                pass

            current_url = page.url
            logger.info(f"Landed on: {page.title()} | {current_url[:120]}")

            # ECW may use iframes — discover the login form context
            frames = page.frames
            logger.info(f"Found {len(frames)} frames on page")
            login_frame = page  # default to main page
            for f in frames:
                try:
                    ufield = f.query_selector('input[name="userName"], input[name="username"], input[id="userName"], input[type="text"]')
                    if ufield:
                        login_frame = f
                        logger.info(f"Login form found in frame: {f.url[:80]}")
                        break
                except Exception:
                    continue

            # Discover all inputs for debugging
            try:
                all_inputs = login_frame.query_selector_all('input')
                for i, inp in enumerate(all_inputs[:15]):
                    inp_type = inp.get_attribute('type') or ''
                    inp_name = inp.get_attribute('name') or ''
                    inp_id = inp.get_attribute('id') or ''
                    inp_ph = inp.get_attribute('placeholder') or ''
                    logger.info(f"  INPUT[{i}]: type={inp_type} name={inp_name} id={inp_id} placeholder={inp_ph}")
            except Exception as e:
                logger.warning(f"Input discovery failed: {e}")

            # Determine if login is needed
            is_login_page = 'login' in current_url.lower() or 'Login' in (page.title() or '')
            needs_login = is_login_page

            if needs_login:
                logger.info("Login form detected — ECW two-step flow...")

                # STEP 1: Enter username in #doctorID
                username_input = login_frame.query_selector('#doctorID')
                if username_input:
                    username_input.click()
                    username_input.fill(creds['username'])
                    logger.info(f"✅ Entered username")
                else:
                    logger.warning("Could not find #doctorID field")

                time.sleep(0.5)

                # Click "Next Step" button
                next_btn = login_frame.query_selector('#nextStep')
                if next_btn:
                    next_btn.click()
                    logger.info("✅ Clicked Next Step")
                else:
                    login_frame.query_selector('input[type="submit"]').click()
                    logger.info("✅ Clicked submit button")

                # Wait for password step to load (page may navigate)
                try:
                    page.wait_for_load_state('networkidle', timeout=10000)
                except Exception:
                    pass
                time.sleep(1)

                # Re-discover the login frame (page may have reloaded)
                frames = page.frames
                login_frame = page
                for f in frames:
                    try:
                        pw = f.query_selector('#passwordField, input[type="password"]')
                        if pw:
                            login_frame = f
                            logger.info(f"Password form found in frame: {f.url[:80]}")
                            break
                    except Exception:
                        continue

                # Re-discover inputs after step 1
                try:
                    all_inputs2 = login_frame.query_selector_all('input')
                    for i, inp in enumerate(all_inputs2[:15]):
                        inp_type = inp.get_attribute('type') or ''
                        inp_name = inp.get_attribute('name') or ''
                        inp_id = inp.get_attribute('id') or ''
                        vis = inp.is_visible()
                        logger.info(f"  STEP2 INPUT[{i}]: type={inp_type} name={inp_name} id={inp_id} visible={vis}")
                except Exception:
                    pass

                # STEP 2: Enter password (field is #passwordField on the password page)
                password_input = login_frame.query_selector('#passwordField')
                if not password_input:
                    password_input = login_frame.query_selector('input[type="password"]:visible')
                if not password_input:
                    password_input = login_frame.query_selector('input[type="password"]')

                if password_input:
                    password_input.click()
                    password_input.fill(creds['password'])
                    logger.info("✅ Entered password")

                    # ECW has hidden fields that must be populated for submit
                    # The visible field is #passwordField, but the form sends #password (hidden)
                    # Also populate #doctorIDVal (hidden username field)
                    try:
                        login_frame.evaluate(f'''
                            const usernameHidden = document.querySelector('#doctorIDVal');
                            const passwordHidden = document.querySelector('#password');
                            if (usernameHidden) usernameHidden.value = "{creds['username']}";
                            if (passwordHidden) passwordHidden.value = document.querySelector('#passwordField')?.value || "";
                        ''')
                        logger.info("✅ Populated hidden username + password fields")
                    except Exception as e:
                        logger.warning(f"Could not set hidden fields: {e}")
                else:
                    logger.warning("Could not find password field after Step 1")

                time.sleep(0.5)

                # Scroll down to make the Turnstile checkbox visible
                try:
                    login_frame.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                    time.sleep(1)
                    logger.info("Scrolled down in login frame")
                except Exception:
                    pass

                # STEP 3: Handle Cloudflare Turnstile
                logger.info("STEP 3: Solving Cloudflare Turnstile...")
                token_found = False
                turnstile_sitekey = "0x4AAAAAAAEasHHO0y-ORWXs"

                # Check if already solved (auto-solve)
                try:
                    token = login_frame.evaluate('document.querySelector("[name=\'cf-turnstile-response\']")?.value || ""')
                    if token:
                        logger.info(f"✅ Turnstile auto-solved! ({len(token)} chars)")
                        token_found = True
                except Exception:
                    pass

                # Use 2Captcha to solve the Turnstile challenge
                if not token_found:
                    captcha_key = None
                    try:
                        captcha_creds = aws_client.get_secret("captcha_api_key")
                        captcha_key = captcha_creds.get('api_key', '')
                    except Exception:
                        pass

                    if captcha_key:
                        import requests as req
                        try:
                            # Use the frame URL where Turnstile lives (not the outer page)
                            turnstile_page_url = login_frame.url or page.url
                            logger.info(f"Sending Turnstile to 2Captcha (sitekey: {turnstile_sitekey[:20]}..., url: {turnstile_page_url[:60]})")
                            resp = req.post("https://2captcha.com/in.php", data={
                                'key': captcha_key,
                                'method': 'turnstile',
                                'sitekey': turnstile_sitekey,
                                'pageurl': turnstile_page_url,
                                'json': 1,
                            }, timeout=15)
                            task_data = resp.json()
                            logger.info(f"2Captcha response: {task_data}")

                            if task_data.get('status') == 1:
                                task_id = task_data['request']
                                logger.info(f"2Captcha task: {task_id} — polling...")

                                for poll in range(40):  # up to ~120s
                                    time.sleep(3)
                                    res = req.get(
                                        f"https://2captcha.com/res.php?key={captcha_key}&action=get&id={task_id}&json=1",
                                        timeout=15
                                    )
                                    result = res.json()

                                    if result.get('status') == 1:
                                        solved_token = result['request']
                                        logger.info(f"✅ 2Captcha solved! Token: ({len(solved_token)} chars)")

                                        # Inject token into ALL relevant hidden fields AND trigger callback
                                        login_frame.evaluate('''
                                            (token) => {
                                                // Set all turnstile response fields
                                                const cfResp = document.querySelector("[name='cf-turnstile-response']");
                                                const tsResp = document.querySelector("#turnstileResponse");
                                                if (cfResp) cfResp.value = token;
                                                if (tsResp) tsResp.value = token;
                                                
                                                // Also trigger Turnstile callback if it exists
                                                if (window.turnstile) {
                                                    try {
                                                        // Find the widget container and trigger its callback
                                                        const containers = document.querySelectorAll('[id*="cf-turnstile"], .cf-turnstile');
                                                        containers.forEach(c => {
                                                            const cbName = c.getAttribute('data-callback');
                                                            if (cbName && window[cbName]) {
                                                                window[cbName](token);
                                                            }
                                                        });
                                                    } catch(e) {}
                                                }
                                            }
                                        ''', solved_token)
                                        logger.info("Token injected + callback triggered")
                                        token_found = True
                                        break
                                    elif result.get('request') == 'CAPCHA_NOT_READY':
                                        if poll % 4 == 0:
                                            logger.info(f"  Waiting... {(poll+1)*5}s")
                                    else:
                                        logger.warning(f"2Captcha error: {result}")
                                        break
                            else:
                                logger.warning(f"2Captcha submit failed: {task_data}")
                        except Exception as e:
                            logger.error(f"2Captcha failed: {e}")
                    else:
                        logger.warning("No 2Captcha API key configured — set 'captcha_api_key' secret")
                        logger.info("⏳ Waiting 60s for manual solve via noVNC...")
                        for wait in range(20):
                            time.sleep(3)
                            try:
                                token = login_frame.evaluate('document.querySelector("[name=\'cf-turnstile-response\']")?.value || ""')
                                if token:
                                    logger.info(f"✅ Turnstile solved manually! ({len(token)} chars)")
                                    token_found = True
                                    break
                            except Exception:
                                pass

                if not token_found:
                    logger.warning("⚠️ Turnstile not solved — login will likely fail")

                # DEBUG: Dump form values before submit
                try:
                    form_debug = login_frame.evaluate('''
                        (() => {
                            const fields = {};
                            const inputs = document.querySelectorAll('input');
                            inputs.forEach(inp => {
                                const name = inp.name || inp.id;
                                if (name) {
                                    let val = inp.value || '';
                                    if (inp.type === 'password') val = val ? '***(' + val.length + ' chars)' : 'EMPTY';
                                    else if (val.length > 40) val = val.substring(0, 40) + '...(' + val.length + ')';
                                    fields[name] = val;
                                }
                            });
                            return JSON.stringify(fields, null, 2);
                        })()
                    ''')
                    logger.info(f"Form values before submit:\n{form_debug}")
                except Exception as e:
                    logger.info(f"Form debug failed: {e}")

                # STEP 4: Click the "Log In" button
                login_btn = login_frame.query_selector(
                    '#Login, input[value="Log In"], input[type="submit"], '
                    'button:has-text("Log In")'
                )
                if login_btn:
                    login_btn.click()
                    logger.info("✅ Clicked Log In button")
                else:
                    page.keyboard.press('Enter')
                    logger.info("✅ Pressed Enter to submit")

                # Wait for login to process
                time.sleep(random.uniform(2, 3))

                # Handle Chrome's "Open xdg-open?" dialog
                try:
                    page.context.on("dialog", lambda dialog: dialog.dismiss())
                except Exception:
                    pass

                # Take screenshot of result
                try:
                    page.screenshot(path="/tmp/login_result.png", full_page=True)
                except Exception:
                    pass

                # Check if login succeeded
                post_url = page.url
                post_title = page.title()
                logger.info(f"After login: {post_title} | {post_url[:120]}")

                # Handle error=6 — retry (first attempt sometimes fails)
                if 'error=6' in post_url:
                    logger.info("Got error=6, retrying login...")
                    time.sleep(2)
                    # Re-check URL — sometimes ECW auto-redirects after plugin popup
                    post_url = page.url
                    post_title = page.title()
                    logger.info(f"After wait: {post_title} | {post_url[:120]}")

                # Handle error=24 — "VERIFY YOUR SECURITY IMAGE" re-authentication
                post_url = page.url
                if 'error=24' in post_url or 'getPwdPage' in post_url:
                    logger.info("🔐 Security image re-authentication detected (error=24)")
                    logger.info("Re-entering password on security verification page...")
                    time.sleep(1)

                    # This page may be in a frame or the main page
                    reauth_frame = page
                    for f in page.frames:
                        pwd_field = f.query_selector('#passwordField, input[type="password"]')
                        if pwd_field:
                            reauth_frame = f
                            break

                    # Enter password
                    try:
                        pwd_field = reauth_frame.query_selector('#passwordField, input[type="password"]')
                        if pwd_field:
                            pwd_field.fill(ecw_password)
                            logger.info("✅ Re-entered password on re-auth page")

                            # Also set the hidden password field if it exists
                            try:
                                reauth_frame.evaluate(f'''(pwd) => {{
                                    const hidden = document.querySelector('#password');
                                    if (hidden) hidden.value = pwd;
                                }}''', ecw_password)
                            except Exception:
                                pass

                            # Check for Turnstile on re-auth page
                            try:
                                ts_field = reauth_frame.query_selector('[name="cf-turnstile-response"]')
                                ts_value = reauth_frame.evaluate('document.querySelector("[name=\'cf-turnstile-response\']")?.value || ""') if ts_field else ''
                                if ts_field and not ts_value:
                                    logger.info("Turnstile found on re-auth page — solving...")
                                    captcha_key = None
                                    try:
                                        captcha_creds = aws_client.get_secret("captcha_api_key")
                                        captcha_key = captcha_creds.get('api_key', '')
                                    except Exception:
                                        pass
                                    if captcha_key:
                                        import requests as req
                                        reauth_url = reauth_frame.url or page.url
                                        resp = req.post("https://2captcha.com/in.php", data={
                                            'key': captcha_key,
                                            'method': 'turnstile',
                                            'sitekey': turnstile_sitekey,
                                            'pageurl': reauth_url,
                                            'json': 1,
                                        }, timeout=15)
                                        task_data = resp.json()
                                        if task_data.get('status') == 1:
                                            task_id = task_data['request']
                                            for poll in range(40):
                                                time.sleep(3)
                                                res = req.get(
                                                    f"https://2captcha.com/res.php?key={captcha_key}&action=get&id={task_id}&json=1",
                                                    timeout=15
                                                )
                                                result = res.json()
                                                if result.get('status') == 1:
                                                    solved_token = result['request']
                                                    reauth_frame.evaluate('''
                                                        (token) => {
                                                            const cfResp = document.querySelector("[name='cf-turnstile-response']");
                                                            const tsResp = document.querySelector("#turnstileResponse");
                                                            if (cfResp) cfResp.value = token;
                                                            if (tsResp) tsResp.value = token;
                                                        }
                                                    ''', solved_token)
                                                    logger.info("✅ Turnstile solved on re-auth page")
                                                    break
                                                elif result.get('request') != 'CAPCHA_NOT_READY':
                                                    break
                                elif ts_value:
                                    logger.info("✅ Turnstile already solved on re-auth page")
                            except Exception as e:
                                logger.warning(f"Turnstile handling on re-auth page: {e}")

                            # Click submit/login button
                            time.sleep(0.5)
                            submit_btn = reauth_frame.query_selector(
                                '#Login, input[type="submit"], button:has-text("Log In"), '
                                'button:has-text("Continue"), input[value="Log In"]'
                            )
                            if submit_btn:
                                submit_btn.click()
                                logger.info("✅ Clicked submit on re-auth page")
                            else:
                                reauth_frame.evaluate('document.querySelector("form")?.submit()')
                                logger.info("✅ Submitted re-auth form programmatically")

                            time.sleep(random.uniform(3, 5))
                            post_url = page.url
                            post_title = page.title()
                            logger.info(f"After re-auth: {post_title} | {post_url[:120]}")
                        else:
                            logger.warning("No password field found on re-auth page")
                    except Exception as e:
                        logger.error(f"Re-auth handling failed: {e}")

                # Handle V12 Plugin popup — click "Ignore and continue"
                try:
                    for f in page.frames:
                        ignore_btn = f.query_selector('text="Ignore and continue"')
                        if ignore_btn:
                            ignore_btn.click()
                            logger.info("✅ Dismissed V12 Plugin popup")
                            time.sleep(1)
                            break
                except Exception:
                    pass

                # Final URL check
                post_url = page.url
                post_title = page.title()

                if 'login' not in post_url.lower():
                    logger.info("🎉 ECW LOGIN SUCCESS!")
                else:
                    try:
                        error_text = login_frame.evaluate('document.body?.innerText?.substring(0, 500) || ""')
                        logger.info(f"Page text: {error_text[:200]}")
                    except Exception:
                        pass
                    logger.warning("Login may have failed — still on login page")
            else:
                logger.info("🎉 Already logged in via persistent session!")

            # ── Branch: Fix Coding IVs uses the ECW session for its own flow ──
            if task_type == 'fix_coding_ivs':
                # Wait for ECW SPA shell to finish loading before we start navigating
                for _wait_i in range(20):
                    try:
                        if 'Building your user experience' not in page.inner_text('body'):
                            break
                    except Exception:
                        pass
                    time.sleep(1)
                time.sleep(2)
                # Order: 1 (create encounter claims) → 2 (delete test claims) → 3 (generate Excel).
                stage = str(body.get('stage', 'all')).lower()
                if stage in ('all', '1', 'stage1', 'one'):
                    _run_fix_coding_ivs_stage1(page, body, aws_client, ecw_url)
                if stage in ('all', '2', 'stage2', 'two'):
                    _run_fix_coding_ivs_stage2(page, body, aws_client, ecw_url)
                if stage in ('all', '3', 'stage3', 'three', 'excel'):
                    _run_fix_coding_ivs_stage3(page, body, aws_client, ecw_url)
                return

            # ═══════════════════════════════════════════
            # CLAIMS EXTRACTION: Navigate & Filter
            # ═══════════════════════════════════════════
            claims_base = ecw_url.split('/mobiledoc/')[0]
            claims_page_url = claims_base + '/mobiledoc/jsp/webemr/index.jsp#/mobiledoc/jsp/webemr/webpm/claimLookup.jsp'

            # Wait for ECW SPA to finish "Building your user experience" loading
            logger.info("Waiting for ECW app to finish loading...")
            for wait_i in range(30):  # up to 30 seconds
                try:
                    body_text = page.inner_text('body')
                    if 'Building your user experience' not in body_text:
                        logger.info(f"✅ ECW app loaded after {wait_i}s")
                        break
                except Exception:
                    pass
                time.sleep(1)
            else:
                logger.warning("ECW app still loading after 30s — proceeding anyway")

            time.sleep(2)  # Extra settle time for AngularJS

            # Navigate to Claims page
            logger.info(f"Navigating to Claims...")

            # ECW is a SPA with hash routing — use hash navigation directly
            # (Avoid sidebar selector loops which cause erratic UI scrolling)
            nav_success = False

            # Strategy 1: Direct hash navigation (most reliable, no UI side-effects)
            logger.info("Using direct hash navigation to Claims...")
            try:
                page.evaluate('''() => {
                    window.location.hash = '/mobiledoc/jsp/webemr/webpm/claimLookup.jsp';
                }''')
                time.sleep(4)
                nav_success = True
                logger.info("✅ Hash navigation triggered")
            except Exception as e:
                logger.warning(f"Hash navigation failed: {e}")

            if not nav_success:
                # Strategy 2: Full page reload with claims URL
                logger.info("Hash failed, full page goto...")
                page.goto(claims_page_url, wait_until='domcontentloaded', timeout=60000)
                time.sleep(5)

            # Take a screenshot right after navigation to debug
            try:
                page.screenshot(path='/tmp/ecw_after_nav.png')
                logger.info("📸 Screenshot after navigation saved")
            except Exception:
                pass

            # ECW renders claims in the main AngularJS view (NOT a separate iframe)
            # Wait for the claims form to actually render by looking for visible UI elements
            logger.info("Waiting for claims UI to render...")
            for wait_i in range(30):
                try:
                    # Look for the "Service Dt(s)" label which is specific to the claims page
                    service_dt_label = page.query_selector('text=Service Dt')
                    claim_status_label = page.query_selector('text=Claim Status')
                    if service_dt_label or claim_status_label:
                        logger.info(f"✅ Claims form visible after {wait_i}s")
                        break
                except Exception:
                    pass
                if wait_i % 5 == 0:
                    logger.info(f"  Waiting for claims UI... {wait_i}s | URL hash: {page.url.split('#')[-1][:60]}")
                time.sleep(1)
            else:
                logger.warning("Claims form not detected after 30s — proceeding anyway")

            time.sleep(3)  # Extra time for AngularJS to fully initialize bindings

            # Take screenshot of claims page state
            try:
                page.screenshot(path='/tmp/ecw_claims_form.png')
                logger.info("📸 Screenshot of claims form saved")
            except Exception:
                pass

            # ── SET FILTERS using Playwright's native interaction (proper AngularJS binding) ──
            logger.info("Setting date filter: from 06/01/2025")

            # Find the Service Dt FROM input using visible label association
            date_set = False
            try:
                # Strategy 1: Find by AngularJS ng-model for the claims view
                from_selectors = [
                    'input[ng-model*="fromDate"]',
                    'input[ng-model*="serviceDateFrom"]',
                    'input[ng-model*="fromServiceDate"]',
                    'input[ng-model*="svcFromDate"]',
                    'input[ng-model*="fromDt"]',
                ]
                for sel in from_selectors:
                    try:
                        inp = page.query_selector(sel)
                        if inp and inp.is_visible():
                            # Use triple-click to select all, then type to replace
                            inp.click(click_count=3)
                            time.sleep(0.3)
                            inp.fill('06/01/2025')
                            inp.dispatch_event('change')
                            inp.dispatch_event('blur')
                            logger.info(f"✅ FROM date set via: {sel}")
                            date_set = True
                            break
                    except Exception:
                        continue

                if not date_set:
                    # Strategy 2: Find inputs near "Service Dt" label text
                    logger.info("Trying label-based date input discovery...")
                    all_visible_inputs = page.locator('input[type="text"]:visible').all()
                    logger.info(f"  Found {len(all_visible_inputs)} visible text inputs")
                    for i, inp in enumerate(all_visible_inputs[:15]):
                        try:
                            val = inp.input_value()
                            ng_model = inp.get_attribute('ng-model') or ''
                            inp_id = inp.get_attribute('id') or ''
                            logger.info(f"  INPUT[{i}]: id={inp_id} ng-model={ng_model} val={val[:20]}")
                            # Date inputs have values like MM/DD/YYYY
                            if val and '/' in val and len(val) == 10:
                                if not date_set:
                                    # First date input is FROM
                                    inp.click(click_count=3)
                                    time.sleep(0.3)
                                    inp.fill('06/01/2025')
                                    inp.dispatch_event('change')
                                    inp.dispatch_event('blur')
                                    logger.info(f"✅ FROM date set via visible input[{i}]: {inp_id}")
                                    date_set = True
                        except Exception as e:
                            logger.debug(f"  Input[{i}] check failed: {e}")
                            continue

            except Exception as e:
                logger.error(f"Date filter failed: {e}")

            # ── DISMISS DATEPICKER POPUP ──
            # The jQuery UI datepicker stays open after date entry, blocking other elements
            logger.info("Dismissing datepicker popup...")
            try:
                # Method 1: Press Escape to close any open datepicker
                page.keyboard.press('Escape')
                time.sleep(0.5)

                # Method 2: Click on the page body to dismiss any popups
                page.click('body', position={'x': 10, 'y': 10}, force=True)
                time.sleep(0.3)

                # Method 3: Force-hide the datepicker via JavaScript
                page.evaluate('''() => {
                    // jQuery UI datepicker
                    const dp = document.querySelector('#ui-datepicker-div, .ui-datepicker');
                    if (dp) {
                        dp.style.display = 'none';
                        dp.style.visibility = 'hidden';
                    }
                    // Also try jQuery API
                    if (typeof jQuery !== 'undefined' && jQuery.datepicker) {
                        jQuery.datepicker._hideDatepicker();
                    }
                }''')
                logger.info("✅ Datepicker dismissed")
            except Exception as e:
                logger.warning(f"Datepicker dismiss attempt: {e}")

            time.sleep(0.5)

            # ── VERIFY DATE WAS RETAINED ──
            if date_set:
                try:
                    # Re-read the FROM date input to confirm it stuck
                    for sel in from_selectors:
                        inp = page.query_selector(sel)
                        if inp and inp.is_visible():
                            actual_val = inp.input_value()
                            if actual_val == '06/01/2025':
                                logger.info(f"✅ FROM date verified: {actual_val}")
                            else:
                                logger.warning(f"⚠️ FROM date value changed to: {actual_val} — re-setting")
                                inp.click(click_count=3)
                                time.sleep(0.2)
                                inp.fill('06/01/2025')
                                inp.dispatch_event('change')
                                inp.dispatch_event('blur')
                                page.keyboard.press('Escape')
                                time.sleep(0.3)
                            break
                except Exception as e:
                    logger.debug(f"Date verification failed: {e}")

            time.sleep(0.5)

            # ── SET CLAIM STATUS FILTER ──
            logger.info("Setting claim status filter...")
            status_set = False
            try:
                # Find all visible select elements
                selects = page.locator('select:visible').all()
                logger.info(f"  Found {len(selects)} visible selects")

                for i, sel in enumerate(selects):
                    try:
                        sel_id = sel.get_attribute('id') or ''
                        ng_model = sel.get_attribute('ng-model') or ''
                        # Get all options to check if this is the claim status dropdown
                        options_text = sel.evaluate('''el => {
                            return Array.from(el.options).map(o => o.text.trim()).join('|');
                        }''')
                        if 'Ready to Submit' in options_text or 'Symplisend' in options_text or 'claimStatus' in ng_model.lower():
                            logger.info(f"  Found status dropdown: id={sel_id} ng-model={ng_model}")
                            # Find the matching option value
                            target_value = sel.evaluate('''el => {
                                for (const opt of el.options) {
                                    if (opt.text.includes('Ready to Submit to Symplisend')) {
                                        return opt.value;
                                    }
                                }
                                return null;
                            }''')
                            if target_value:
                                sel.select_option(value=target_value)
                                time.sleep(0.3)
                                sel.dispatch_event('change')
                                logger.info(f"✅ Status set to 'Ready to Submit to Symplisend' (value={target_value})")
                                status_set = True
                                break
                            else:
                                # Log available options for debugging
                                logger.info(f"  Options: {options_text[:300]}")
                    except Exception as e:
                        logger.debug(f"  Select[{i}] check failed: {e}")
                        continue

                if not status_set:
                    logger.warning("Status dropdown not found among visible selects")
            except Exception as e:
                logger.error(f"Status filter failed: {e}")

            time.sleep(0.5)

            # ── VERIFY STATUS DROPDOWN RETAINED VALUE ──
            if status_set:
                try:
                    selects = page.locator('select:visible').all()
                    for sel in selects:
                        ng_model = sel.get_attribute('ng-model') or ''
                        if 'claimStatus' in ng_model.lower() or 'status' in ng_model.lower():
                            selected_text = sel.evaluate('''el => {
                                return el.options[el.selectedIndex]?.text?.trim() || '';
                            }''')
                            if 'Ready to Submit' in selected_text:
                                logger.info(f"✅ Status dropdown verified: {selected_text}")
                            else:
                                logger.warning(f"⚠️ Status dropdown shows: '{selected_text}' — expected 'Ready to Submit to Symplisend'")
                            break
                except Exception as e:
                    logger.debug(f"Status verification failed: {e}")

            time.sleep(1)

            # ── PRE-LOOKUP VERIFICATION ──
            # Take screenshot AFTER all filters are applied, BEFORE clicking Lookup
            logger.info("=== PRE-LOOKUP FILTER VERIFICATION ===")
            try:
                page.screenshot(path='/tmp/ecw_after_filters.png')
                logger.info("📸 Screenshot after filters saved")
            except Exception:
                pass

            # Set up XHR interception BEFORE clicking Lookup
            # ECW loads claims via AJAX, so we capture the response JSON
            captured_claims_data = []
            def on_response(response):
                try:
                    url = response.url
                    if response.status == 200 and ('claim' in url.lower() or 'lookup' in url.lower() or 'grid' in url.lower() or 'list' in url.lower()):
                        content_type = response.headers.get('content-type', '')
                        if 'json' in content_type or 'javascript' in content_type:
                            try:
                                body = response.json()
                                if isinstance(body, list) and len(body) > 0:
                                    captured_claims_data.append({'url': url, 'data': body, 'type': 'list'})
                                    logger.info(f"📡 Captured XHR list: {url[:80]} ({len(body)} items)")
                                elif isinstance(body, dict):
                                    for key in body:
                                        val = body[key]
                                        if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                                            captured_claims_data.append({'url': url, 'data': val, 'key': key, 'type': 'dict_array'})
                                            logger.info(f"📡 Captured XHR dict: {url[:80]} key={key} ({len(val)} items)")
                            except Exception:
                                pass
                except Exception:
                    pass
            
            page.on('response', on_response)

            # ── CAPTURE CURRENT TABLE STATE before Lookup ──
            # The claims page auto-loads with default filters (usually "Encounters Without Claims")
            # We need to detect when the table CHANGES after clicking Lookup
            pre_lookup_state = {}
            try:
                pre_lookup_state = page.evaluate('''() => {
                    const table = document.querySelector('#this-table, table.table-bordered');
                    const rows = table ? table.querySelectorAll('tbody tr') : [];
                    const totalEl = document.body.innerText.match(/Total Counts\\s*:\\s*(\\d+)/);
                    const firstRowText = rows.length > 0 ? rows[0].innerText.substring(0, 80) : '';
                    return {
                        rowCount: rows.length,
                        totalCounts: totalEl ? totalEl[1] : '',
                        firstRowText: firstRowText
                    };
                }''')
                logger.info(f"📊 Pre-lookup state: {pre_lookup_state['rowCount']} rows, total={pre_lookup_state.get('totalCounts','?')}")
            except Exception as e:
                logger.warning(f"Could not capture pre-lookup state: {e}")

            # ── CLICK LOOKUP using Playwright locator (not JS click) ──
            # JS .click() may not trigger AngularJS ng-click handlers properly
            logger.info("Clicking Claims Lookup via Playwright locator...")
            lookup_clicked = False
            try:
                # Strategy 1: Find the Lookup button by its text, exclude "Patient Lookup"
                lookup_buttons = page.locator('button:has-text("Lookup"), input[type="button"][value*="Lookup"]').all()
                for btn in lookup_buttons:
                    btn_text = btn.inner_text() if btn.evaluate('el => el.tagName') == 'BUTTON' else btn.get_attribute('value') or ''
                    if 'Patient' in btn_text:
                        continue
                    if btn.is_visible():
                        btn.click()
                        lookup_clicked = True
                        logger.info(f"✅ Clicked Lookup button: '{btn_text.strip()}'")
                        break

                if not lookup_clicked:
                    # Strategy 2: Click any visible element with exact "Lookup" text
                    lookup_el = page.locator(':text-is("Lookup"):visible').last
                    if lookup_el.count() > 0:
                        lookup_el.click()
                        lookup_clicked = True
                        logger.info("✅ Clicked Lookup via text locator")
            except Exception as e:
                logger.warning(f"Playwright Lookup click failed: {e}")

            if not lookup_clicked:
                # Strategy 3: Fallback to JS click with Angular scope apply
                logger.info("Falling back to JS click with Angular trigger...")
                try:
                    clicked = page.evaluate('''() => {
                        const btns = document.querySelectorAll('button, input[type="button"], a, span');
                        for (const btn of btns) {
                            const text = (btn.innerText || btn.value || '').trim();
                            if (text.includes('Patient')) continue;
                            if (text === 'Lookup' || text === 'Look up') {
                                // Trigger via Angular if available
                                if (window.angular) {
                                    try {
                                        angular.element(btn).triggerHandler('click');
                                    } catch(e) {
                                        btn.click();
                                    }
                                } else {
                                    btn.click();
                                }
                                return 'Clicked: ' + text;
                            }
                        }
                        return null;
                    }''')
                    if clicked:
                        lookup_clicked = True
                        logger.info(f"✅ {clicked} (JS fallback)")
                except Exception as e:
                    logger.error(f"JS Lookup click also failed: {e}")

            # ── WAIT FOR TABLE TO CHANGE (not just exist) ──
            logger.info("Waiting for results to refresh after Lookup...")
            time.sleep(2)  # Initial wait for AJAX to start

            # Wait for network to settle
            try:
                page.wait_for_load_state('networkidle', timeout=15000)
                logger.info("Network idle reached")
            except Exception:
                logger.info("Network idle timeout — proceeding")

            # Wait for table content to actually change from pre-lookup state
            table_changed = False
            for wait_i in range(15):
                try:
                    post_state = page.evaluate('''() => {
                        const table = document.querySelector('#this-table, table.table-bordered');
                        const rows = table ? table.querySelectorAll('tbody tr') : [];
                        const totalEl = document.body.innerText.match(/Total Counts\\s*:\\s*(\\d+)/);
                        const firstRowText = rows.length > 0 ? rows[0].innerText.substring(0, 80) : '';
                        return {
                            rowCount: rows.length,
                            totalCounts: totalEl ? totalEl[1] : '',
                            firstRowText: firstRowText
                        };
                    }''')

                    # Check if the table content has changed
                    old_total = pre_lookup_state.get('totalCounts', '')
                    new_total = post_state.get('totalCounts', '')
                    old_first = pre_lookup_state.get('firstRowText', '')
                    new_first = post_state.get('firstRowText', '')
                    old_count = pre_lookup_state.get('rowCount', 0)
                    new_count = post_state.get('rowCount', 0)

                    if new_total != old_total or new_first != old_first or new_count != old_count:
                        logger.info(f"✅ Table changed after {wait_i}s: {old_count} → {new_count} rows, total: {old_total} → {new_total}")
                        table_changed = True
                        break
                    else:
                        if wait_i % 3 == 0:
                            logger.info(f"  Waiting for table to change... {wait_i}s (rows={new_count}, total={new_total})")
                except Exception as e:
                    logger.debug(f"Table change check failed: {e}")
                time.sleep(1)

            if not table_changed:
                logger.warning("⚠️ Table did not change after 15s — results may be stale")
                # Extra wait just in case
                time.sleep(3)

            # Additional settle time for Angular rendering
            time.sleep(2)

            # Take screenshot of results AFTER table has changed
            try:
                page.screenshot(path='/tmp/ecw_claims_result.png')
                logger.info("📸 Screenshot of claims results saved")
            except Exception:
                pass

            # Force Angular to render all data before extraction
            try:
                page.evaluate('''() => {
                    // Trigger Angular digest cycle
                    if (window.angular) {
                        try {
                            const scope = angular.element(document.body).scope();
                            if (scope && scope.$apply) {
                                scope.$apply();
                            }
                        } catch(e) {}
                        try {
                            const injector = angular.element(document.body).injector();
                            if (injector) {
                                const $rootScope = injector.get('$rootScope');
                                $rootScope.$digest();
                            }
                        } catch(e) {}
                    }
                    // Scroll the claims table to force rendering
                    const table = document.getElementById('this-table');
                    if (table) {
                        const container = table.closest('.table-responsive, [style*="overflow"], [style*="scroll"]');
                        if (container) {
                            container.scrollTop = 0;
                        }
                    }
                }''')
                time.sleep(1)
                logger.info("🔄 Triggered Angular digest and scroll")
            except Exception:
                pass

            # ── EXTRACT CLAIMS TABLE ──
            logger.info("Extracting claims data...")
            try:
                # First, discover the table structure
                table_debug = page.evaluate('''() => {
                    const tables = document.querySelectorAll('table');
                    const result = [];
                    for (let i = 0; i < tables.length; i++) {
                        const t = tables[i];
                        const rows = t.querySelectorAll('tr');
                        if (rows.length > 2) {
                            // Get header text
                            const headerRow = rows[0] || rows[1];
                            const headers = Array.from(headerRow.querySelectorAll('th, td')).map(c => c.innerText.trim());
                            result.push({
                                tableIdx: i,
                                rowCount: rows.length,
                                headers: headers.slice(0, 15),
                                visible: t.offsetParent !== null,
                                className: t.className || '',
                                id: t.id || ''
                            });
                        }
                    }
                    return result;
                }''')
                for td in table_debug:
                    logger.info(f"  TABLE[{td['tableIdx']}]: {td['rowCount']} rows, visible={td['visible']}, class={td['className'][:40]}, headers={td['headers'][:8]}")

                # Extract claims from the visible table with claim data columns
                # From screenshot: columns are COLL | CLAIM # | SERVICE DATE | PVDR | PATIENT | PAYER | STATUS | CHARGES | PMTS/ADJS | ADJUSTMENT | WITHHELD | BALANCE
                claims_data = page.evaluate('''() => {
                    const claims = [];
                    const tables = document.querySelectorAll('table');
                    let claimsTable = null;

                    // PRIORITY 1: Find the table with class containing "aqua-" 
                    // This is the actual claims data grid, NOT the filter form
                    for (const t of tables) {
                        const cls = t.className || '';
                        if (cls.includes('aqua-') || cls.includes('table-striped')) {
                            if (t.offsetParent !== null) {  // visible
                                const rows = t.querySelectorAll('tr');
                                // The data grid has many rows; filter form only has a few
                                if (rows.length > 3) {
                                    claimsTable = t;
                                    break;
                                }
                            }
                        }
                    }

                    // PRIORITY 2: Find table with ng-repeat rows (data grid pattern)
                    if (!claimsTable) {
                        const ngRows = document.querySelectorAll('tr[ng-repeat], tr[ng-click]');
                        if (ngRows.length > 0) {
                            claimsTable = ngRows[0].closest('table');
                        }
                    }

                    // PRIORITY 3: Find visible table with CLAIM # header but skip 
                    // tables that look like filter forms (few rows, many selects)
                    if (!claimsTable) {
                        for (const t of tables) {
                            if (t.offsetParent === null) continue;
                            const headerText = t.innerText || '';
                            if (headerText.includes('CLAIM #') || headerText.includes('SERVICE DATE')) {
                                // Skip if table has select/input elements (it's a filter form)
                                const selectCount = t.querySelectorAll('select, input[type="text"]').length;
                                if (selectCount > 3) continue;
                                claimsTable = t;
                                break;
                            }
                        }
                    }

                    if (!claimsTable) {
                        return { 
                            error: 'No claims table found', 
                            tableCount: tables.length,
                            hint: 'No table with aqua- class or ng-repeat rows found'
                        };
                    }

                    // Log which table we found for debugging
                    const tableInfo = {
                        className: claimsTable.className || '',
                        id: claimsTable.id || '',
                        rowCount: claimsTable.querySelectorAll('tr').length
                    };

                    const rows = claimsTable.querySelectorAll('tbody tr, tr');
                    
                    // AngularJS Scope Extraction: ECW uses ng-binding/ng-repeat 
                    // Data is in Angular's scope, not in rendered DOM text
                    let usedAngular = false;
                    try {
                        if (window.angular) {
                            // First, try to re-enable debug info if it was disabled in production
                            try {
                                const injector = angular.element(document.body).injector();
                                if (injector) {
                                    const $rootScope = injector.get('$rootScope');
                                    const $compile = injector.get('$compileProvider');
                                }
                            } catch(e) {}
                            
                            // Try getting scope from the table rows
                            const tbodyRows = claimsTable.querySelectorAll('tbody tr');
                            let sampleScope = null;
                            for (const row of tbodyRows) {
                                try {
                                    const s = angular.element(row).scope();
                                    if (s && s.obj) {
                                        sampleScope = s;
                                        break;
                                    }
                                    // Also check isolateScope and parent scope
                                    const is = angular.element(row).isolateScope && angular.element(row).isolateScope();
                                    if (is && is.obj) {
                                        sampleScope = is;
                                        break;
                                    }
                                } catch(e) {}
                            }
                            
                            if (sampleScope) {
                                // We found the scope! Extract from all rows
                                for (const row of tbodyRows) {
                                    try {
                                        const scope = angular.element(row).scope();
                                        if (!scope || !scope.obj) continue;
                                        const obj = scope.obj;
                                        
                                        const claim = {
                                            claim_number: obj.claimNo || obj.claimNumber || obj.claim_no || obj.columnClaimNumber || '',
                                            patient_name: obj.patName || obj.patientName || obj.patient || obj.columnPatient || '',
                                            payer: obj.payerName || obj.payer || obj.insuranceName || obj.columnPayer || '',
                                            service_date: obj.serviceDate || obj.srvDate || obj.dos || obj.columnServiceDate || '',
                                            provider: obj.providerName || obj.rendering || obj.provider || obj.columnRendering || '',
                                            pos: obj.pos || obj.placeOfService || obj.columnPos || '',
                                            cpt: obj.cpt || obj.cptCode || obj.procCode || obj.columnCpt || '',
                                            status: obj.claimStatus || obj.status || obj.columnStatus || '',
                                            charges: obj.amount || obj.charges || obj.totalCharge || '',
                                            payments: obj.payment || obj.payments || '',
                                            adjustment: obj.NetAdj || obj.adjustment || '',
                                            withheld: obj.NetWithHeld || obj.withheld || '',
                                            balance: obj.balance || '',
                                            assigned_to: obj.assignedTo || obj.assigned || '',
                                            encounter_id: obj.encounterId || obj.encId || '',
                                            _scope_keys: Object.keys(obj).slice(0, 40),
                                        };
                                        
                                        if (claim.patient_name || claim.service_date || claim.claim_number) {
                                            claims.push(claim);
                                            usedAngular = true;
                                        }
                                    } catch(e) {}
                                }
                            } else {
                                // Scope not accessible — try getting the data from controller
                                // Look for ng-controller on a parent element
                                let ctrl = claimsTable.closest('[ng-controller]');
                                if (!ctrl) ctrl = document.querySelector('[ng-controller]');
                                if (ctrl) {
                                    try {
                                        const ctrlScope = angular.element(ctrl).scope();
                                        if (ctrlScope) {
                                            // Dump scope keys for debugging
                                            const keys = Object.keys(ctrlScope).filter(k => !k.startsWith('$') && !k.startsWith('_'));
                                            claims.push({ _ctrlScopeKeys: keys.slice(0, 50) });
                                            
                                            // Look for arrays that might contain claims data
                                            for (const key of keys) {
                                                const val = ctrlScope[key];
                                                if (Array.isArray(val) && val.length > 0 && typeof val[0] === 'object') {
                                                    // Check if this looks like claims data
                                                    const first = val[0];
                                                    if (first.patName || first.patientName || first.claimNo || first.serviceDate || first.columnPatient) {
                                                        for (const obj of val) {
                                                            const claim = {
                                                                claim_number: obj.claimNo || obj.claimNumber || obj.claim_no || obj.columnClaimNumber || '',
                                                                patient_name: obj.patName || obj.patientName || obj.patient || obj.columnPatient || '',
                                                                payer: obj.payerName || obj.payer || obj.insuranceName || obj.columnPayer || '',
                                                                service_date: obj.serviceDate || obj.srvDate || obj.dos || obj.columnServiceDate || '',
                                                                provider: obj.providerName || obj.rendering || obj.provider || obj.columnRendering || '',
                                                                pos: obj.pos || '',
                                                                cpt: obj.cpt || obj.cptCode || '',
                                                                status: obj.claimStatus || obj.status || obj.columnStatus || '',
                                                                charges: obj.amount || obj.charges || '',
                                                                payments: obj.payment || obj.payments || '',
                                                                adjustment: obj.NetAdj || '',
                                                                withheld: obj.NetWithHeld || '',
                                                                balance: obj.balance || '',
                                                                encounter_id: obj.encounterId || obj.encId || '',
                                                                _scope_keys: Object.keys(obj).slice(0, 40),
                                                            };
                                                            if (claim.patient_name || claim.service_date || claim.claim_number) {
                                                                claims.push(claim);
                                                                usedAngular = true;
                                                            }
                                                        }
                                                        break;
                                                    }
                                                }
                                            }
                                        }
                                    } catch(e) {}
                                }
                            }
                        }
                    } catch(e) {
                        // Angular not available
                    }
                    
                    // Fallback: DOM-based extraction if Angular didn't work
                    if (!usedAngular) {
                        for (const row of rows) {
                            const cells = row.querySelectorAll('td');
                            if (cells.length >= 6) {
                                const cellTexts = Array.from(cells).map(c => {
                                    let txt = (c.innerText || '').trim();
                                    if (!txt) {
                                        const anchor = c.querySelector('a');
                                        if (anchor) txt = (anchor.innerText || anchor.textContent || '').trim();
                                    }
                                    return txt;
                                });
                                if (cellTexts.join('').includes('CLAIM #')) continue;
                                if (cellTexts.filter(t => t.length > 0).length < 3) continue;
                                claims.push({ cell_texts: cellTexts.slice(0, 20) });
                            }
                        }
                    }
                    
                    return { 
                        claims: claims.slice(0, 100), 
                        total: claims.length,
                        tableInfo: tableInfo,
                        extractionMethod: usedAngular ? 'angular_scope' : 'dom_fallback'
                    };
                }''')

                # Handle the response
                if isinstance(claims_data, dict) and 'error' in claims_data:
                    logger.error(f"❌ {claims_data['error']} (tables on page: {claims_data.get('tableCount', '?')})")
                    claims_list = []
                elif isinstance(claims_data, dict):
                    raw_claims = claims_data.get('claims', [])
                    table_info = claims_data.get('tableInfo', {})
                    extraction_method = claims_data.get('extractionMethod', 'unknown')
                    
                    # Separate debug entries from actual claims
                    claims_list = []
                    for c in raw_claims:
                        if '_ctrlScopeKeys' in c:
                            logger.info(f"🔍 Controller scope keys: {c['_ctrlScopeKeys']}")
                        else:
                            claims_list.append(c)
                    
                    logger.info(f"📋 Found {len(claims_list)} claim rows (method: {extraction_method})")
                    logger.info(f"   Table: class='{table_info.get('className', '')}' id='{table_info.get('id', '')}' rows={table_info.get('rowCount', '?')}")
                    # Log first 5 rows
                    for i, c in enumerate(claims_list[:5]):
                        if 'claim_number' in c:
                            logger.info(f"  ROW[{i}]: claim={c.get('claim_number')} patient={c.get('patient_name')} date={c.get('service_date')} payer={c.get('payer')} status={c.get('status')}")
                            if i == 0 and '_scope_keys' in c:
                                logger.info(f"  SCOPE KEYS: {c['_scope_keys']}")
                        else:
                            texts = c.get('cell_texts', [])
                            logger.info(f"  ROW[{i}]: {texts}")
                else:
                    claims_list = claims_data if isinstance(claims_data, list) else []
                    logger.info(f"📋 Found {len(claims_list)} ECW claims")

                # Check XHR-captured data as alternative/preferred source
                if captured_claims_data:
                    logger.info(f"📡 XHR captured {len(captured_claims_data)} responses")
                    for capture in captured_claims_data:
                        items = capture['data']
                        if len(items) > 0:
                            sample = items[0]
                            sample_keys = list(sample.keys())[:20] if isinstance(sample, dict) else []
                            logger.info(f"  XHR {capture['type']}: {len(items)} items, keys={sample_keys}")
                            
                            # Check if this looks like claims data
                            if isinstance(sample, dict) and any(k in str(sample_keys).lower() for k in ['claim', 'patient', 'pat', 'service', 'payer']):
                                logger.info("✅ Using XHR-captured data as primary source!")
                                claims_list = []
                                for obj in items:
                                    claim = {
                                        'claim_number': str(obj.get('claimNo', obj.get('claimNumber', obj.get('claim_no', obj.get('columnClaimNumber', ''))))),
                                        'patient_name': str(obj.get('patName', obj.get('patientName', obj.get('patient', obj.get('columnPatient', ''))))),
                                        'payer': str(obj.get('payerName', obj.get('payer', obj.get('insuranceName', obj.get('columnPayer', ''))))),
                                        'service_date': str(obj.get('serviceDate', obj.get('srvDate', obj.get('dos', obj.get('columnServiceDate', ''))))),
                                        'provider': str(obj.get('providerName', obj.get('rendering', obj.get('provider', obj.get('columnRendering', ''))))),
                                        'pos': str(obj.get('pos', obj.get('placeOfService', ''))),
                                        'cpt': str(obj.get('cpt', obj.get('cptCode', obj.get('procCode', '')))),
                                        'status': str(obj.get('claimStatus', obj.get('status', obj.get('columnStatus', '')))),
                                        'charges': str(obj.get('amount', obj.get('charges', obj.get('totalCharge', '')))),
                                        'payments': str(obj.get('payment', obj.get('payments', ''))),
                                        'adjustment': str(obj.get('NetAdj', obj.get('adjustment', ''))),
                                        'withheld': str(obj.get('NetWithHeld', obj.get('withheld', ''))),
                                        'balance': str(obj.get('balance', '')),
                                        'encounter_id': str(obj.get('encounterId', obj.get('encId', ''))),
                                        '_scope_keys': list(obj.keys())[:40],
                                    }
                                    if claim['patient_name'] or claim['service_date'] or claim['claim_number']:
                                        claims_list.append(claim)
                                
                                logger.info(f"📡 XHR extraction: {len(claims_list)} claims")
                                for i, c in enumerate(claims_list[:3]):
                                    logger.info(f"  XHR_ROW[{i}]: claim={c.get('claim_number')} patient={c.get('patient_name')} date={c.get('service_date')}")
                                    if i == 0:
                                        logger.info(f"  XHR KEYS: {c.get('_scope_keys')}")
                                break  # Use the first matching XHR response

                # Save to JSON for dashboard
                ecw_output = {
                    'claims': claims_list,
                    'total_found': len(claims_list),
                    'filter_status': 'Ready to Submit to Symplisend',
                    'filter_date_from': '06/01/2025',
                    'scraped_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())
                }
                with open('/opt/helixona-agent/ecw_claims.json', 'w') as ecw_f:
                    json.dump(ecw_output, ecw_f, indent=2)
                logger.info("✅ ECW claims saved to ecw_claims.json")

                # Save to DynamoDB
                try:
                    claims_table = aws_client.dynamodb.Table('helixona-claims')
                    saved_count = 0
                    for idx, claim in enumerate(claims_list):
                        # Handle both Angular scope extraction (named fields) 
                        # and DOM fallback (cell_texts array)
                        if 'claim_number' in claim:
                            # Angular/XHR extraction — fields are already named
                            claim_number = claim.get('claim_number', '').strip()
                            patient = claim.get('patient_name', '').strip()
                            date = claim.get('service_date', '').strip()
                            
                            # Generate composite key if no claim number
                            if not claim_number:
                                if patient and date:
                                    claim_number = f"ENC-{date.replace('/', '')}-{patient.replace(' ', '_').replace(',', '')}"
                                elif claim.get('encounter_id', ''):
                                    claim_number = str(claim['encounter_id'])
                                else:
                                    claim_number = f"ENC-{idx}"
                            
                            save_item = {
                                'claim_id': str(claim_number),
                                'source': 'ECW',
                                'patient_name': patient,
                                'payer': str(claim.get('payer', '')),
                                'service_date': date,
                                'provider': str(claim.get('provider', '')),
                                'pos': str(claim.get('pos', '')),
                                'cpt': str(claim.get('cpt', '')),
                                'charges': str(claim.get('charges', '')),
                                'status': str(claim.get('status', '')),
                                'payments_adjs': str(claim.get('payments', '')),
                                'adjustment': str(claim.get('adjustment', '')),
                                'withheld': str(claim.get('withheld', '')),
                                'balance': str(claim.get('balance', '')),
                                'assigned_to': str(claim.get('assigned_to', '')),
                                'encounter_id': str(claim.get('encounter_id', '')),
                                'state': 1,
                                'scraped_at': ecw_output['scraped_at'],
                            }
                        else:
                            # DOM fallback — cell_texts array 
                            # Observed column mapping from debug:
                            # [0]=checkbox, [1-4]=icon/indicator cols (empty), [5]=claim#(empty for encounters),
                            # [6]=Service Date, [7]=POS/Rendering, [8]=Patient Name
                            texts = claim.get('cell_texts', [])
                            date = texts[6] if len(texts) > 6 else ''
                            rendering = texts[7] if len(texts) > 7 else ''
                            patient = texts[8] if len(texts) > 8 else ''
                            claim_number = texts[5] if len(texts) > 5 else ''
                            
                            # Generate composite key if no claim number 
                            if not claim_number.strip():
                                if patient and date:
                                    claim_number = f"ENC-{date.replace('/', '')}-{patient.replace(' ', '_').replace(',', '')}"
                                else:
                                    claim_number = f"ENC-ROW-{idx}"
                            
                            if not patient and not date:
                                continue  # Skip truly empty rows
                            
                            save_item = {
                                'claim_id': claim_number.strip(),
                                'source': 'ECW',
                                'patient_name': patient,
                                'payer': '',
                                'service_date': date,
                                'provider': rendering,
                                'charges': '',
                                'status': 'Ready to Submit to Symplisend',
                                'state': 1,
                                'scraped_at': ecw_output['scraped_at'],
                            }
                        
                        claims_table.put_item(Item=save_item)
                        saved_count += 1
                    logger.info(f"✅ {saved_count} claims saved to DynamoDB")

                    # ── Sync: Remove stale claims no longer in ECW ──
                    # Get the set of claim IDs we just extracted
                    extracted_ids = set()
                    for idx2, claim2 in enumerate(claims_list):
                        if 'claim_number' in claim2:
                            cid = claim2.get('claim_number', '').strip()
                            if not cid:
                                patient2 = claim2.get('patient_name', '').strip()
                                date2 = claim2.get('service_date', '').strip()
                                if patient2 and date2:
                                    cid = f"ENC-{date2.replace('/', '')}-{patient2.replace(' ', '_').replace(',', '')}"
                                elif claim2.get('encounter_id', ''):
                                    cid = str(claim2['encounter_id'])
                                else:
                                    cid = f"ENC-{idx2}"
                            extracted_ids.add(str(cid))
                        else:
                            texts = claim2.get('cell_texts', [])
                            cid = texts[5] if len(texts) > 5 else ''
                            if not cid.strip():
                                patient2 = texts[8] if len(texts) > 8 else ''
                                date2 = texts[6] if len(texts) > 6 else ''
                                if patient2 and date2:
                                    cid = f"ENC-{date2.replace('/', '')}-{patient2.replace(' ', '_').replace(',', '')}"
                                else:
                                    cid = f"ENC-ROW-{idx2}"
                            extracted_ids.add(cid.strip())

                    # Reconcile with ECW WITHOUT deleting — claims no longer visible in
                    # ECW are archived (ecw_visible=False), never removed, so submitted/
                    # historical claims are preserved.
                    try:
                        _sync_ecw_claim_visibility(claims_table, extracted_ids)
                    except Exception as sync_e:
                        logger.warning(f"DynamoDB visibility sync failed: {sync_e}")

                except Exception as e:
                    logger.error(f"DynamoDB save failed: {e}")

            except Exception as e:
                logger.error(f"Claims extraction failed: {e}")

            # Screenshot for verification
            try:
                page.screenshot(path='/tmp/ecw_claims_result.png', full_page=True)
            except Exception:
                pass

            logger.info("STEP 0 complete — ECW Claims extracted.")

            # Auto-chain: trigger HCFA generation (skip if called from bs_missing_docs pipeline)
            if not body.get('skip_auto_chain'):
                try:
                    claims_table = aws_client.dynamodb.Table('helixona-claims')
                    scan_result = claims_table.scan()
                    state1_count = sum(1 for c in scan_result.get('Items', []) if int(c.get('state', 0)) == 1)
                    if state1_count > 0:
                        import json as _json
                        aws_client.sqs.send_message(
                            QueueUrl='https://sqs.us-west-2.amazonaws.com/148274106093/helixona-agent-tasks',
                            MessageBody=_json.dumps({'task_type': 'generate_hcfa'})
                        )
                        logger.info(f"🔗 Auto-queued generate_hcfa for {state1_count} claim(s) in state 1")
                    else:
                        logger.info("No claims in state 1 — skipping HCFA auto-trigger")
                except Exception as e:
                    logger.warning(f"Auto-chain failed: {e}")
            else:
                logger.info("Skipping auto-chain (called from bs_missing_docs pipeline)")

        except Exception as e:
            logger.error(f"STEP 0 failed: {e}")
            time.sleep(3)
        finally:
            manager.stop()

    # ──────────────────────────────────────────────────────
    # STEP 0C: HCFA Generation — Open claim popup & click
    #          the red icon next to "Print HCFA (02-12)"
    # ──────────────────────────────────────────────────────
    elif task_type == 'generate_hcfa':
        logger.info("STEP 0C: HCFA Form Generation from ECW...")

        # Testing mode: process only 1 claim for fast iteration
        testing_mode = body.get('testing_mode', False)
        test_claim_id = body.get('test_claim_id', '').strip()
        if test_claim_id:
            testing_mode = True
            logger.info(f"🧪 TESTING MODE: Will process only claim {test_claim_id}")
        elif testing_mode:
            logger.info("🧪 TESTING MODE: Will process only 1 claim")

        # Claims will be discovered directly from the ECW Claims page
        claims_to_process = []

        # ── 1. Launch browser & login to ECW ──
        manager = BrowserManager().start(proxy_config=None)
        try:
            page = manager.new_page()
            creds = aws_client.get_secret("ecw_credentials")
            ecw_url = creds.get('url', 'https://zjdvbjliaam2udi33vapp.ecwcloud.com/mobiledoc/jsp/webemr/login/newLogin.jsp')

            logger.info(f"Navigating to ECW: {ecw_url}")
            page.goto(ecw_url, wait_until="domcontentloaded", timeout=30000)
            try:
                page.wait_for_load_state('networkidle', timeout=15000)
            except Exception:
                pass
            time.sleep(1)

            # Dismiss "Restore pages?" dialog if present
            try:
                dismiss_btn = page.query_selector('button:has-text("Dismiss"), button:has-text("Don\'t restore")')
                if dismiss_btn:
                    dismiss_btn.click()
                    logger.info("Dismissed 'Restore pages' dialog")
                    time.sleep(1)
                else:
                    page.keyboard.press('Escape')
                    time.sleep(0.5)
            except Exception:
                pass

            current_url = page.url
            needs_login = 'login' in current_url.lower() or 'Login' in (page.title() or '')

            if needs_login:
                logger.info("Login required — ECW two-step flow...")

                # STEP 1: Enter username
                username_input = page.query_selector('#doctorID')
                if username_input:
                    username_input.click()
                    time.sleep(0.3)
                    username_input.fill('')
                    username_input.type(creds['username'], delay=random.uniform(30, 60))
                    logger.info(f"✅ Entered username")

                time.sleep(0.5)

                # Click "Next Step"
                next_btn = page.query_selector('#nextStep')
                if next_btn:
                    next_btn.click()
                    logger.info("✅ Clicked Next Step")
                else:
                    sub = page.query_selector('input[type="submit"]')
                    if sub:
                        sub.click()

                try:
                    page.wait_for_load_state('networkidle', timeout=10000)
                except Exception:
                    pass
                time.sleep(1)

                # Re-discover password frame
                login_frame = page
                for f in page.frames:
                    try:
                        pw = f.query_selector('#passwordField, input[type="password"]')
                        if pw:
                            login_frame = f
                            break
                    except Exception:
                        continue

                # STEP 2: Enter password
                password_input = (
                    login_frame.query_selector('#passwordField') or
                    login_frame.query_selector('input[type="password"]:visible') or
                    login_frame.query_selector('input[type="password"]')
                )
                if password_input:
                    password_input.click()
                    password_input.fill(creds['password'])
                    logger.info("✅ Entered password")

                    # Populate hidden fields
                    try:
                        login_frame.evaluate(f'''
                            const usernameHidden = document.querySelector('#doctorIDVal');
                            const passwordHidden = document.querySelector('#password');
                            if (usernameHidden) usernameHidden.value = "{creds['username']}";
                            if (passwordHidden) passwordHidden.value = document.querySelector('#passwordField')?.value || "";
                        ''')
                    except Exception:
                        pass

                # Scroll to show Turnstile
                try:
                    login_frame.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                    time.sleep(1)
                except Exception:
                    pass

                # STEP 3: Solve Turnstile
                logger.info("Solving Turnstile...")
                token_found = False
                turnstile_sitekey = "0x4AAAAAAAEasHHO0y-ORWXs"

                try:
                    token = login_frame.evaluate('document.querySelector("[name=\'cf-turnstile-response\']")?.value || ""')
                    if token:
                        logger.info(f"✅ Turnstile auto-solved! ({len(token)} chars)")
                        token_found = True
                except Exception:
                    pass

                if not token_found:
                    captcha_key = None
                    try:
                        captcha_creds = aws_client.get_secret("captcha_api_key")
                        captcha_key = captcha_creds.get('api_key', '')
                    except Exception:
                        pass

                    if captcha_key:
                        import requests as req
                        try:
                            turnstile_page_url = login_frame.url or page.url
                            resp = req.post("https://2captcha.com/in.php", data={
                                'key': captcha_key,
                                'method': 'turnstile',
                                'sitekey': turnstile_sitekey,
                                'pageurl': turnstile_page_url,
                                'json': 1,
                            }, timeout=15)
                            task_data = resp.json()
                            if task_data.get('status') == 1:
                                task_id = task_data['request']
                                for poll in range(40):
                                    time.sleep(3)
                                    res = req.get(
                                        f"https://2captcha.com/res.php?key={captcha_key}&action=get&id={task_id}&json=1",
                                        timeout=15
                                    )
                                    result = res.json()
                                    if result.get('status') == 1:
                                        solved_token = result['request']
                                        login_frame.evaluate('''
                                            (token) => {
                                                const cfResp = document.querySelector("[name='cf-turnstile-response']");
                                                const tsResp = document.querySelector("#turnstileResponse");
                                                if (cfResp) cfResp.value = token;
                                                if (tsResp) tsResp.value = token;
                                                if (window.turnstile) {
                                                    try {
                                                        const containers = document.querySelectorAll('[id*="cf-turnstile"], .cf-turnstile');
                                                        containers.forEach(c => {
                                                            const cbName = c.getAttribute('data-callback');
                                                            if (cbName && window[cbName]) window[cbName](token);
                                                        });
                                                    } catch(e) {}
                                                }
                                            }
                                        ''', solved_token)
                                        token_found = True
                                        logger.info("✅ Turnstile solved via 2Captcha")
                                        break
                                    elif result.get('request') != 'CAPCHA_NOT_READY':
                                        break
                        except Exception as e:
                            logger.error(f"2Captcha failed: {e}")

                # STEP 4: Click Log In
                login_btn = login_frame.query_selector(
                    '#Login, input[value="Log In"], input[type="submit"], '
                    'button:has-text("Log In")'
                )
                if login_btn:
                    login_btn.click()
                    logger.info("✅ Clicked Log In")
                else:
                    page.keyboard.press('Enter')

                time.sleep(random.uniform(2, 3))

                # Handle error=24 re-auth
                post_url = page.url
                if 'error=24' in post_url or 'getPwdPage' in post_url:
                    logger.info("🔐 Re-authentication detected — handling...")
                    time.sleep(1)
                    reauth_frame = page
                    for f in page.frames:
                        pwd_f = f.query_selector('#passwordField, input[type="password"]')
                        if pwd_f:
                            reauth_frame = f
                            break
                    pwd_field = reauth_frame.query_selector('#passwordField, input[type="password"]')
                    if pwd_field:
                        pwd_field.fill(creds['password'])
                        try:
                            reauth_frame.evaluate(f'''(pwd) => {{
                                const hidden = document.querySelector('#password');
                                if (hidden) hidden.value = pwd;
                            }}''', creds['password'])
                        except Exception:
                            pass
                        time.sleep(0.5)
                        submit_btn = reauth_frame.query_selector(
                            '#Login, input[type="submit"], button:has-text("Log In")'
                        )
                        if submit_btn:
                            submit_btn.click()
                        else:
                            reauth_frame.evaluate('document.querySelector("form")?.submit()')
                        time.sleep(random.uniform(3, 5))

                # Handle V12 Plugin popup
                try:
                    for f in page.frames:
                        ignore_btn = f.query_selector('text="Ignore and continue"')
                        if ignore_btn:
                            ignore_btn.click()
                            logger.info("✅ Dismissed V12 Plugin popup")
                            time.sleep(1)
                            break
                except Exception:
                    pass

                post_url = page.url
                if 'login' not in post_url.lower():
                    logger.info("🎉 ECW LOGIN SUCCESS!")
                else:
                    logger.warning("Login may have failed — still on login page")
            else:
                logger.info("🎉 Already logged in via persistent session!")

            # ── 3. Wait for ECW SPA to load ──
            logger.info("Waiting for ECW app to finish loading...")
            for wait_i in range(30):
                try:
                    body_text = page.inner_text('body')
                    if 'Building your user experience' not in body_text:
                        logger.info(f"✅ ECW app loaded after {wait_i}s")
                        break
                except Exception:
                    pass
                time.sleep(1)
            time.sleep(2)

            # ── 4. Navigate to Claims page ──
            logger.info("Navigating to Claims page...")
            try:
                page.evaluate('''() => {
                    window.location.hash = '/mobiledoc/jsp/webemr/webpm/claimLookup.jsp';
                }''')
                time.sleep(4)
                logger.info("✅ Hash navigation to Claims triggered")
            except Exception as e:
                logger.warning(f"Hash navigation failed: {e}")
                claims_base = ecw_url.split('/mobiledoc/')[0]
                claims_page_url = claims_base + '/mobiledoc/jsp/webemr/index.jsp#/mobiledoc/jsp/webemr/webpm/claimLookup.jsp'
                page.goto(claims_page_url, wait_until='domcontentloaded', timeout=60000)
                time.sleep(5)

            # Wait for claims UI
            for wait_i in range(30):
                try:
                    service_dt = page.query_selector('text=Service Dt')
                    claim_status = page.query_selector('text=Claim Status')
                    if service_dt or claim_status:
                        logger.info(f"✅ Claims form visible after {wait_i}s")
                        break
                except Exception:
                    pass
                time.sleep(1)
            time.sleep(3)

            # ── 5. Set filters & Lookup (same as extraction) ──
            logger.info("Setting date filter: from 06/01/2025")
            date_set = False
            from_selectors = [
                'input[ng-model*="fromDate"]',
                'input[ng-model*="serviceDateFrom"]',
                'input[ng-model*="fromServiceDate"]',
                'input[ng-model*="svcFromDate"]',
                'input[ng-model*="fromDt"]',
            ]
            for sel in from_selectors:
                try:
                    inp = page.query_selector(sel)
                    if inp and inp.is_visible():
                        inp.click(click_count=3)
                        time.sleep(0.3)
                        inp.fill('06/01/2025')
                        inp.dispatch_event('change')
                        inp.dispatch_event('blur')
                        logger.info(f"✅ FROM date set via: {sel}")
                        date_set = True
                        break
                except Exception:
                    continue

            if not date_set:
                all_visible_inputs = page.locator('input[type="text"]:visible').all()
                for i, inp in enumerate(all_visible_inputs[:15]):
                    try:
                        val = inp.input_value()
                        if val and '/' in val and len(val) == 10:
                            inp.click(click_count=3)
                            time.sleep(0.3)
                            inp.fill('06/01/2025')
                            inp.dispatch_event('change')
                            inp.dispatch_event('blur')
                            logger.info(f"✅ FROM date set via visible input[{i}]")
                            date_set = True
                            break
                    except Exception:
                        continue

            # Dismiss datepicker
            try:
                page.keyboard.press('Escape')
                time.sleep(0.5)
                page.click('body', position={'x': 10, 'y': 10}, force=True)
                time.sleep(0.3)
                page.evaluate('''() => {
                    const dp = document.querySelector('#ui-datepicker-div, .ui-datepicker');
                    if (dp) { dp.style.display = 'none'; dp.style.visibility = 'hidden'; }
                    if (typeof jQuery !== 'undefined' && jQuery.datepicker) jQuery.datepicker._hideDatepicker();
                }''')
            except Exception:
                pass
            time.sleep(0.5)

            # Set claim status filter
            logger.info("Setting claim status filter...")
            try:
                selects = page.locator('select:visible').all()
                for sel in selects:
                    try:
                        ng_model = sel.get_attribute('ng-model') or ''
                        options_text = sel.evaluate('''el => {
                            return Array.from(el.options).map(o => o.text.trim()).join('|');
                        }''')
                        if 'Ready to Submit' in options_text or 'Symplisend' in options_text or 'claimStatus' in ng_model.lower():
                            target_value = sel.evaluate('''el => {
                                for (const opt of el.options) {
                                    if (opt.text.includes('Ready to Submit to Symplisend')) return opt.value;
                                }
                                return null;
                            }''')
                            if target_value:
                                sel.select_option(value=target_value)
                                time.sleep(0.3)
                                sel.dispatch_event('change')
                                logger.info(f"✅ Status set to 'Ready to Submit to Symplisend'")
                                break
                    except Exception:
                        continue
            except Exception as e:
                logger.warning(f"Status filter failed: {e}")

            time.sleep(1)

            # Click Lookup
            logger.info("Clicking Lookup...")
            lookup_clicked = False
            try:
                lookup_buttons = page.locator('button:has-text("Lookup"), input[type="button"][value*="Lookup"]').all()
                for btn in lookup_buttons:
                    btn_text = btn.inner_text() if btn.evaluate('el => el.tagName') == 'BUTTON' else btn.get_attribute('value') or ''
                    if 'Patient' in btn_text:
                        continue
                    if btn.is_visible():
                        btn.click()
                        lookup_clicked = True
                        logger.info(f"✅ Clicked Lookup: '{btn_text.strip()}'")
                        break
            except Exception:
                pass

            if not lookup_clicked:
                try:
                    page.evaluate('''() => {
                        const btns = document.querySelectorAll('button, input[type="button"]');
                        for (const btn of btns) {
                            const text = (btn.innerText || btn.value || '').trim();
                            if (text.includes('Patient')) continue;
                            if (text === 'Lookup' || text === 'Look up') {
                                if (window.angular) {
                                    try { angular.element(btn).triggerHandler('click'); } catch(e) { btn.click(); }
                                } else { btn.click(); }
                                return;
                            }
                        }
                    }''')
                    lookup_clicked = True
                    logger.info("✅ Lookup clicked (JS fallback)")
                except Exception:
                    pass

            # Wait for results
            time.sleep(2)
            try:
                page.wait_for_load_state('networkidle', timeout=15000)
            except Exception:
                pass
            time.sleep(3)

            try:
                page.screenshot(path='/tmp/hcfa_claims_list.png')
                logger.info("📸 HCFA: Claims list screenshot saved")
            except Exception:
                pass

            # ── 6. Discover claims from ALL pages on the ECW claims list ──
            # Wait for Angular claim rows to render
            for wait_row in range(15):
                row_count = page.evaluate('() => document.querySelectorAll("tr[ng-repeat*=lstClaimReport]").length')
                if row_count and row_count > 0:
                    logger.info(f"Found {row_count} claim rows after {wait_row}s")
                    break
                time.sleep(1)
            
            # Detect pagination: "of N" tells us total pages
            total_pages = page.evaluate('''() => {
                const span = document.querySelector('span[ng-bind*="lastNavPage"]');
                if (span) {
                    const m = (span.innerText || '').match(/of\s+(\d+)/);
                    if (m) return parseInt(m[1]);
                }
                return 1;
            }''') or 1
            logger.info(f"📄 ECW claims pagination: {total_pages} page(s)")
            
            # Scrape function — extracts claims from the current page
            SCRAPE_JS = '''() => {
                const rows = document.querySelectorAll('tr[ng-repeat*="lstClaimReport"]');
                const claims = [];
                for (const row of rows) {
                    const cells = row.querySelectorAll('td');
                    if (cells.length < 6) continue;
                    
                    // Find claim ID: the td with class highlight-bgcolor-blue
                    let claimId = null;
                    for (const cell of cells) {
                        if (cell.className.includes('highlight-bgcolor-blue')) {
                            const span = cell.querySelector('span.ng-binding');
                            if (span) {
                                const text = (span.innerText || '').trim();
                                if (/^\d+$/.test(text)) {
                                    claimId = text;
                                    break;
                                }
                            }
                        }
                    }
                    if (!claimId) continue;
                    
                    // Extract all span texts from the row
                    const spans = row.querySelectorAll('span.ng-binding');
                    const spanTexts = Array.from(spans).map(s => (s.innerText || '').trim());
                    
                    let serviceDate = '';
                    let patient = '';
                    let payer = '';
                    let charges = '';
                    let status = '';
                    for (const t of spanTexts) {
                        if (/^\d{2}\/\d{2}\/\d{4}$/.test(t) && !serviceDate) serviceDate = t;
                        else if (t.includes(',') && !patient && t.length > 3 && !/\d/.test(t)) patient = t;
                        else if (t.includes('Blue Shield') || t.includes('Shield')) payer = t;
                        else if (/^\s*[\d,]+\.\d{2}\s*$/.test(t) && !charges) charges = t.trim().replace(',','');
                        else if (t.includes('Ready to Sub') || t.includes('Symplisend')) status = t;
                    }
                    claims.push({ claimId, serviceDate, patient, payer, charges, status, pageNum: window._currentPage || 1 });
                }
                return claims;
            }'''
            
            # Scrape page 1
            page.evaluate('window._currentPage = 1')
            page_claims = page.evaluate(SCRAPE_JS) or []
            logger.info(f"📋 Page 1: found {len(page_claims)} claims")
            
            # Scrape remaining pages
            for pg in range(2, total_pages + 1):
                try:
                    # Click Next button
                    next_clicked = page.evaluate('''() => {
                        const btn = document.querySelector('#nextBtn');
                        if (btn && !btn.classList.contains('disabled') && !btn.disabled) {
                            btn.click();
                            return true;
                        }
                        return false;
                    }''')
                    if not next_clicked:
                        logger.info(f"📄 No more pages (Next button disabled at page {pg})")
                        break
                    
                    # Wait for rows to refresh
                    time.sleep(2)
                    for wait_row in range(10):
                        new_count = page.evaluate('() => document.querySelectorAll("tr[ng-repeat*=lstClaimReport]").length')
                        if new_count and new_count > 0:
                            break
                        time.sleep(1)
                    
                    # Scrape this page
                    page.evaluate(f'window._currentPage = {pg}')
                    more_claims = page.evaluate(SCRAPE_JS) or []
                    logger.info(f"📋 Page {pg}: found {len(more_claims)} claims")
                    page_claims.extend(more_claims)
                except Exception as e:
                    logger.warning(f"Pagination error on page {pg}: {e}")
                    break
            
            # Navigate back to page 1 for processing
            if total_pages > 1:
                page.evaluate('''() => {
                    const btn = document.querySelector('#firstBtn');
                    if (btn) btn.click();
                }''')
                time.sleep(2)
            
            # Deduplicate by claim ID (in case of overlap)
            seen = set()
            unique_claims = []
            for pc in page_claims:
                if pc['claimId'] not in seen:
                    seen.add(pc['claimId'])
                    unique_claims.append(pc)
            page_claims = unique_claims
            logger.info(f"📋 Total unique claims across all pages: {len(page_claims)}")
            
            if page_claims and len(page_claims) > 0:
                logger.info(f"📋 Discovered {len(page_claims)} claims on ECW page")
                claims_table = aws_client.dynamodb.Table('helixona-claims')
                claims_to_process = []  # Reset — use page-discovered claims
                
                # ── SHORTCUT: If test_claim_id is set, skip full sync ──
                if test_claim_id:
                    # Find the claim in the scraped list
                    target_pc = None
                    for pc in page_claims:
                        if pc['claimId'] == str(test_claim_id):
                            target_pc = pc
                            break
                    
                    if target_pc:
                        cid = target_pc['claimId']
                        logger.info(f"🧪 Fast path: found claim {cid} in scraped list, skipping full sync")
                        # Quick upsert just this one claim
                        claims_table.update_item(
                            Key={'claim_id': cid},
                            UpdateExpression='SET #s = if_not_exists(#s, :state), patient_name = :name, service_date = :dos, dos = :dos, payer = :payer, charges = :charges, #src = :src',
                            ExpressionAttributeNames={'#s': 'state', '#src': 'source'},
                            ExpressionAttributeValues={
                                ':state': 1, ':name': target_pc.get('patient', ''),
                                ':dos': target_pc.get('serviceDate', ''), ':payer': target_pc.get('payer', ''),
                                ':charges': target_pc.get('charges', ''), ':src': 'ECW'
                            }
                        )
                        item = claims_table.get_item(Key={'claim_id': cid}).get('Item', {})
                        claims_to_process.append({
                            'claim_id': cid,
                            'patient_name': target_pc.get('patient', ''),
                            'service_date': target_pc.get('serviceDate', ''),
                            'needs_hcfa': not bool(item.get('hcfa_s3_path')),
                            'needs_prog_notes': (not bool(item.get('prog_notes_s3_path'))
                                                 or bool(item.get('iv_note_patient_mismatch'))
                                                 or (not bool(item.get('office_visit')) and not bool(item.get('iv_note_rx_start_date')))),
                            'needs_subscriber_id': not bool(item.get('subscriber_id')),
                            'needs_encounter_date': not bool(item.get('encounter_date')),
                            'needs_encounter_file': not bool(item.get('encounter_file_s3_path')),
                            'page_num': int(target_pc.get('pageNum', 1)),
                        })
                        logger.info(f"🧪 Claim {cid} needs: {[k for k,v in claims_to_process[0].items() if k.startswith('needs_') and v]}")
                    else:
                        logger.warning(f"🧪 Claim {test_claim_id} not found in {len(page_claims)} scraped claims!")
                        # Fallback: load from DynamoDB and use direct claim lookup (filter returned no rows)
                        try:
                            db_item = claims_table.get_item(Key={'claim_id': str(test_claim_id)}).get('Item')
                            if db_item:
                                cid = str(test_claim_id)
                                logger.info(f"🧪 DDB fallback: loaded claim {cid} from DynamoDB, using direct claim lookup")
                                claims_to_process.append({
                                    'claim_id': cid,
                                    'patient_name': db_item.get('patient_name', ''),
                                    'service_date': db_item.get('service_date', '') or db_item.get('dos', ''),
                                    'needs_hcfa': not bool(db_item.get('hcfa_s3_path')),
                                    'needs_prog_notes': (not bool(db_item.get('prog_notes_s3_path'))
                                                     or bool(db_item.get('iv_note_patient_mismatch'))
                                                     or (not bool(db_item.get('office_visit')) and not bool(db_item.get('iv_note_rx_start_date')))),
                                    'needs_subscriber_id': not bool(db_item.get('subscriber_id')),
                                    'needs_encounter_date': not bool(db_item.get('encounter_date')),
                                    'needs_encounter_file': not bool(db_item.get('encounter_file_s3_path')),
                                    'page_num': 1,
                                })
                                logger.info(f"🧪 Claim {cid} needs: {[k for k,v in claims_to_process[0].items() if k.startswith('needs_') and v]}")
                            else:
                                logger.warning(f"🧪 Claim {test_claim_id} not found in DynamoDB either")
                        except Exception as _fb_err:
                            logger.warning(f"🧪 DDB fallback failed: {_fb_err}")
                    # Skip normal processing below
                else:
                    # ── Normal full sync ──
                    for pc in page_claims:
                        cid = pc['claimId']
                        logger.info(f"  → Claim {cid}: {pc.get('patient','')} | {pc.get('serviceDate','')}")
                        # Upsert into DynamoDB — set state=1 only if new
                        claims_table.update_item(
                            Key={'claim_id': cid},
                            UpdateExpression='SET #s = if_not_exists(#s, :state), patient_name = :name, service_date = :dos, dos = :dos, payer = :payer, charges = :charges, claim_status = :status, #src = :src',
                            ExpressionAttributeNames={'#s': 'state', '#src': 'source'},
                            ExpressionAttributeValues={
                                ':state': 1,
                                ':name': pc.get('patient', 'Unknown'),
                                ':dos': pc.get('serviceDate', ''),
                                ':payer': pc.get('payer', ''),
                                ':charges': pc.get('charges', ''),
                                ':status': pc.get('status', ''),
                                ':src': 'ECW'
                            }
                        )
                        # Check if this claim needs processing
                        item = claims_table.get_item(Key={'claim_id': cid}).get('Item', {})
                        claim_state = int(item.get('state', 0))
                        has_prog_notes = bool(item.get('prog_notes_s3_path'))
                        has_hcfa = bool(item.get('hcfa_s3_path'))
                        hcfa_size = int(item.get('hcfa_pdf_size', 0))
                        
                        if claim_state == 1:
                            # Needs full processing (HCFA + Prog Notes)
                            claims_to_process.append({
                                'claim_id': cid,
                                'patient_name': pc.get('patient', ''),
                                'service_date': pc.get('serviceDate', ''),
                                'needs_hcfa': True,
                                'needs_prog_notes': True,
                                'page_num': int(pc.get('pageNum', 1)),
                            })
                        elif claim_state >= 2:
                            # Validate existing data quality
                            needs_hcfa_redo = False
                            # Force re-capture of IV Note when ANY signal indicates the
                            # stored note is bad: missing, wrong patient (mismatch flagged),
                            # or stale from the pre-Rx-Start-Date code (non-office claim
                            # without an iv_note_rx_start_date stored). Fase 5 + mismatch
                            # guard will skip the save if the content is still wrong.
                            needs_pn = (not has_prog_notes
                                        or bool(item.get('iv_note_patient_mismatch'))
                                        or (not bool(item.get('office_visit')) and not bool(item.get('iv_note_rx_start_date'))))
                            reasons = []
                            
                            # Check HCFA PDF quality
                            if not has_hcfa:
                                needs_hcfa_redo = True
                                reasons.append('no HCFA PDF')
                            elif hcfa_size > 0 and hcfa_size < 500000:
                                # Real HCFA PDFs with CMS-1500 background are >1MB
                                # Small ones (<500KB) are likely damaged HTML renders
                                needs_hcfa_redo = True
                                reasons.append(f'HCFA too small ({hcfa_size//1024}KB, likely damaged)')
                            
                            if needs_pn:
                                reasons.append('missing Prog Notes')
                            
                            # Check subscriber ID for first-time submissions
                            needs_subscriber = False
                            sub_type = item.get('submission_type', '')
                            if sub_type == 'First Time Submission' and not item.get('subscriber_id'):
                                needs_subscriber = True
                                reasons.append('missing Subscriber ID')
                            
                            # Check encounter date
                            needs_enc_date = not item.get('encounter_date')
                            if needs_enc_date:
                                reasons.append('missing Encounter Date')
                            
                            # Check encounter file — also force re-extraction when:
                            #  • marked Needs Review (encounter_revision_needed)
                            #  • IV-therapy (non-office) claim whose stored encounter_date
                            #    doesn't match the IV Note's Rx Start Date (stale from the
                            #    pre-Rx-Start-Date code, or a wrong-patient capture). With
                            #    the EXACT-only matching strategy, those two MUST be equal.
                            _rx_d = item.get('iv_note_rx_start_date')
                            _enc_d = item.get('encounter_date')
                            _is_office = bool(item.get('office_visit'))
                            needs_enc_file = (not item.get('encounter_file_s3_path')
                                              or bool(item.get('encounter_revision_needed'))
                                              or (not _is_office and bool(_rx_d) and _enc_d != _rx_d))
                            if needs_enc_file:
                                reasons.append('missing/stale Encounter File')
                            
                            if needs_hcfa_redo or needs_pn or needs_subscriber or needs_enc_date or needs_enc_file:
                                claims_to_process.append({
                                    'claim_id': cid,
                                    'patient_name': pc.get('patient', ''),
                                    'service_date': pc.get('serviceDate', ''),
                                    'needs_hcfa': needs_hcfa_redo,
                                    'needs_prog_notes': needs_pn,
                                    'needs_subscriber_id': needs_subscriber,
                                    'needs_encounter_date': needs_enc_date,
                                    'needs_encounter_file': needs_enc_file,
                                    'page_num': int(pc.get('pageNum', 1)),
                                })
                                logger.info(f"  ℹ️ Claim {cid}: state {claim_state}, needs redo: {', '.join(reasons)}")
                            else:
                                logger.info(f"  ✅ Claim {cid}: fully complete (HCFA {hcfa_size//1024}KB + Prog Notes + Enc File)")
                    logger.info(f"📋 {len(claims_to_process)} claims to process")
                    
                    # ── 6b. Sync with ECW WITHOUT deleting — archive (don't remove)
                    #        claims no longer on ECW so history is preserved. ──
                    ecw_claim_ids = set(pc['claimId'] for pc in page_claims)
                    try:
                        _sync_ecw_claim_visibility(claims_table, ecw_claim_ids)
                    except Exception as e:
                        logger.warning(f"Claim sync failed: {e}")
            else:
                logger.warning("No claims found on the ECW page after Lookup")
                # Fallback: if test_claim_id is set, load from DDB and use direct claim lookup
                if test_claim_id:
                    try:
                        claims_table = aws_client.dynamodb.Table('helixona-claims')
                        db_item = claims_table.get_item(Key={'claim_id': str(test_claim_id)}).get('Item')
                        if db_item:
                            cid = str(test_claim_id)
                            logger.info(f"🧪 DDB fallback (empty page): loaded claim {cid} from DynamoDB, using direct claim lookup")
                            claims_to_process.append({
                                'claim_id': cid,
                                'patient_name': db_item.get('patient_name', ''),
                                'service_date': db_item.get('service_date', '') or db_item.get('dos', ''),
                                'needs_hcfa': not bool(db_item.get('hcfa_s3_path')),
                                'needs_prog_notes': (not bool(db_item.get('prog_notes_s3_path'))
                                                     or bool(db_item.get('iv_note_patient_mismatch'))
                                                     or (not bool(db_item.get('office_visit')) and not bool(db_item.get('iv_note_rx_start_date')))),
                                'needs_subscriber_id': not bool(db_item.get('subscriber_id')),
                                'needs_encounter_date': not bool(db_item.get('encounter_date')),
                                'needs_encounter_file': not bool(db_item.get('encounter_file_s3_path')),
                                'page_num': 1,
                            })
                            logger.info(f"🧪 Claim {cid} needs: {[k for k,v in claims_to_process[0].items() if k.startswith('needs_') and v]}")
                        else:
                            logger.warning(f"🧪 Claim {test_claim_id} not found in DynamoDB either")
                    except Exception as _fb_err:
                        logger.warning(f"🧪 DDB fallback failed: {_fb_err}")

            # ── 7. Process each claim: open detail popup → HCFA + Prog Notes ──
            hcfa_success_count = 0
            hcfa_fail_count = 0

            # If a specific claim ID was requested, filter the list
            if test_claim_id:
                original_count = len(claims_to_process)
                claims_to_process = [c for c in claims_to_process if str(c.get('claim_id', '')) == str(test_claim_id)]
                if claims_to_process:
                    logger.info(f"🧪 Filtered to claim {test_claim_id} (from {original_count} total)")
                else:
                    logger.warning(f"🧪 Claim {test_claim_id} not found in {original_count} scraped claims!")

            for claim_record in claims_to_process:
                # Testing mode: process only 1 claim (when no specific claim ID given)
                if testing_mode and not test_claim_id and hcfa_success_count + hcfa_fail_count >= 1:
                    logger.info("🧪 Testing Mode: stopping after 1 claim")
                    break
                claim_id = claim_record['claim_id']
                patient_name = claim_record.get('patient_name', '')
                logger.info(f"═══ Processing claim {claim_id} ({patient_name}) ═══")

                # ── Clean slate: nuke every leftover modal/iframe from the previous
                # claim BEFORE opening this one. Otherwise stale ProgNoteViwerFrame
                # iframes / encounter-preview modals from claim N-1 hang around in
                # the DOM and the bot can read the wrong patient's content. Audit
                # showed 66/88 IV notes and 30/43 encounters were captured wrong
                # for this exact reason. Fixing it at the source means each claim
                # starts with a DOM that has only its own documentation in it. ──
                try:
                    for _frm in [page] + list(page.frames):
                        try:
                            _frm.evaluate('''() => {
                                // 1) Remove every ProgNoteViwerFrame iframe (stale viewer from prior claim).
                                document.querySelectorAll('iframe[name="ProgNoteViwerFrame"], iframe[id*="ProgNote"]').forEach(el => el.remove());
                                // 2) Remove encounter preview modals (have closeEncouterPreviewDialog or encounterPreviewContent).
                                document.querySelectorAll('button[ng-click*="closeEncouterPreviewDialog"]').forEach(btn => {
                                    const modal = btn.closest('.modal, [role="dialog"]'); if (modal) modal.remove();
                                });
                                document.querySelectorAll('#encounterPreviewContent').forEach(el => {
                                    const modal = el.closest('.modal, [role="dialog"]'); if (modal) modal.remove() ; else el.remove();
                                });
                                // 3) Remove any Prog. Notes modal that lingered (has the Print button for ProgNoteViwerFrame).
                                document.querySelectorAll('button[ng-click*="printProgressNote"]').forEach(btn => {
                                    const modal = btn.closest('.modal, [role="dialog"]'); if (modal) modal.remove();
                                });
                                // 4) Remove orphan backdrops left behind when their modal was already removed.
                                document.querySelectorAll('.modal-backdrop').forEach(el => el.remove());
                            }''')
                        except Exception:
                            continue
                    # Belt-and-suspenders: press Escape a couple times to dismiss any
                    # native dialog the JS pass missed.
                    try:
                        page.keyboard.press('Escape')
                        time.sleep(0.1)
                        page.keyboard.press('Escape')
                    except Exception:
                        pass
                except Exception as _clean_err:
                    logger.warning(f"Leftover-modal cleanup failed: {_clean_err}")

                # Write current step to DDB so the dashboard can highlight the row
                # in real time (polls /api/claims). The "processing_at" timestamp is
                # what the UI uses to decide which row is live (within last ~90s).
                def _set_step(step):
                    try:
                        aws_client.update_claim_status(claim_id, {
                            'current_step': step,
                            'processing_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                        })
                    except Exception:
                        pass
                _set_step('opening claim popup')

                # Ensure DDB has `cpt` populated as soon as the claim enters the loop,
                # BEFORE any needs_only_metadata short-circuit, so the dashboard's
                # office-visit detection works for every claim — not only those that
                # reach Fase 5. Backfill from the stored HCFA PDF in S3 if missing.
                if not claim_record.get('cpt'):
                    try:
                        _ddb = aws_client.dynamodb.Table('helixona-claims').get_item(
                            Key={'claim_id': claim_id}).get('Item', {})
                        if _ddb.get('cpt'):
                            claim_record['cpt'] = _ddb.get('cpt')
                        elif _ddb.get('hcfa_s3_path'):
                            import re as _re_cb
                            _m_cb = _re_cb.match(r's3://([^/]+)/(.+)', _ddb.get('hcfa_s3_path'))
                            if _m_cb:
                                _lp_cb = f'/tmp/{claim_id}_cpt_backfill.pdf'
                                try:
                                    aws_client.s3.download_file(_m_cb.group(1), _m_cb.group(2), _lp_cb)
                                    _cpt_bf = _cpt_from_hcfa_pdf(_lp_cb)
                                    if _cpt_bf:
                                        claim_record['cpt'] = _cpt_bf
                                        aws_client.update_claim_status(claim_id, {'cpt': _cpt_bf})
                                        logger.info(f"📋 CPT backfilled from S3 HCFA for claim {claim_id}: {_cpt_bf}")
                                    try:
                                        os.remove(_lp_cb)
                                    except Exception:
                                        pass
                                except Exception as _bf_e:
                                    logger.warning(f"CPT backfill failed for {claim_id}: {_bf_e}")
                    except Exception:
                        pass

                # ── Pre-flight: verify any STORED IV note / encounter PDF still
                # belongs to this patient. Without this check, an old wrong-patient
                # capture (from before patient verification was added) tricks the
                # bot into the metadata-only path; the bot then trusts the bad IV
                # note's Rx Start Date and searches encounters for the wrong date.
                # If anything fails verification, clear it and force re-capture.
                try:
                    _surname_pf = (claim_record.get('patient_name', '') or '').split(',')[0].strip().upper()
                    if _surname_pf:
                        _ddb_pf = aws_client.dynamodb.Table('helixona-claims').get_item(
                            Key={'claim_id': claim_id}).get('Item', {})

                        def _pdf_patient_ok(s3_uri):
                            """Download a stored PDF, return True if the expected surname appears
                            in the document's HEADER (first ~500 chars). False = definitely wrong
                            patient. None = couldn't verify (don't punish).

                            Why head-only: substrings anywhere in the text falsely match other
                            people named in the note (e.g. claim 5305 had "Note verified by:
                            Juan Castaneda, Paramedic" in Felicia's note → CASTANEDA passed
                            even though the actual patient on the note was REVERS).
                            """
                            if not s3_uri:
                                return None
                            try:
                                import re as _re_pf, pdfplumber as _pp_pf
                                _m = _re_pf.match(r's3://([^/]+)/(.+)', s3_uri)
                                if not _m:
                                    return None
                                _lp = f'/tmp/{claim_id}_preflight_{_m.group(2).split("/")[-1]}'
                                aws_client.s3.download_file(_m.group(1), _m.group(2), _lp)
                                with _pp_pf.open(_lp) as _doc:
                                    # Only read the first page's text — patient header lives there.
                                    _first_page_text = (_doc.pages[0].extract_text() or '') if _doc.pages else ''
                                try: os.remove(_lp)
                                except Exception: pass
                                _head = _first_page_text[:500].upper()
                                if not _head.strip():
                                    return None  # image-only PDF — can't verify, leave alone
                                return _surname_pf in _head
                            except Exception:
                                return None

                        # IV Note
                        _iv_stored = _ddb_pf.get('prog_notes_s3_path')
                        if _iv_stored:
                            _ok = _pdf_patient_ok(_iv_stored)
                            if _ok is False:
                                logger.warning(f"🧹 Pre-flight: stored IV Note for claim {claim_id} is for a DIFFERENT patient. Clearing so it re-captures.")
                                try:
                                    aws_client.dynamodb.Table('helixona-claims').update_item(
                                        Key={'claim_id': claim_id},
                                        UpdateExpression='SET prog_notes_s3_path = :empty, iv_note_patient_mismatch = :true, '
                                                         'prog_notes_capture_failed = :why, prog_notes_capture_failed_at = :now '
                                                         'REMOVE iv_note_rx_start_date',
                                        ExpressionAttributeValues={
                                            ':empty': '', ':true': True,
                                            ':why': 'preflight_patient_mismatch',
                                            ':now': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                        })
                                    claim_record['needs_prog_notes'] = True
                                    claim_record.pop('iv_note_rx_start_date', None)
                                except Exception as _pf_err:
                                    logger.warning(f"Pre-flight IV-note clear failed: {_pf_err}")

                        # Encounter file
                        _enc_stored = _ddb_pf.get('encounter_file_s3_path')
                        if _enc_stored:
                            _ok = _pdf_patient_ok(_enc_stored)
                            if _ok is False:
                                logger.warning(f"🧹 Pre-flight: stored Encounter PDF for claim {claim_id} is for a DIFFERENT patient. Clearing so it re-captures.")
                                try:
                                    aws_client.dynamodb.Table('helixona-claims').update_item(
                                        Key={'claim_id': claim_id},
                                        UpdateExpression='SET encounter_file_s3_path = :empty, encounter_files_s3_paths = :emptylist, '
                                                         'encounter_files_count = :zero, encounter_files_visit_types = :emptylist, '
                                                         'encounter_revision_needed = :true, encounter_capture_failed = :why, '
                                                         'encounter_capture_failed_at = :now '
                                                         'REMOVE encounter_file_captured_at, encounter_date, encounter_pick_strategy',
                                        ExpressionAttributeValues={
                                            ':empty': '', ':emptylist': [], ':zero': 0, ':true': True,
                                            ':why': 'preflight_patient_mismatch',
                                            ':now': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                        })
                                    claim_record['needs_encounter_file'] = True
                                except Exception as _pf_err:
                                    logger.warning(f"Pre-flight encounter clear failed: {_pf_err}")
                except Exception as _pf_top:
                    logger.warning(f"Pre-flight patient verification failed: {_pf_top}")

                # Fase 0 — neutraliza window.print en page + todos los frames
                # para que printProgressNote(), printenc() y cualquier print()
                # accidental jamás abra un dialog que bloquee el flujo.
                try:
                    page.evaluate('''() => {
                        window.print = () => {};
                        try {
                            for (let i = 0; i < window.frames.length; i++) {
                                try { window.frames[i].print = () => {}; } catch(e) {}
                            }
                        } catch(e) {}
                    }''')
                    for _frm in page.frames:
                        try:
                            _frm.evaluate('''() => {
                                window.print = () => {};
                                try {
                                    for (let i = 0; i < window.frames.length; i++) {
                                        try { window.frames[i].print = () => {}; } catch(e) {}
                                    }
                                } catch(e) {}
                            }''')
                        except Exception:
                            continue
                except Exception:
                    pass

                # ── Use Claim Lookup to open the claim directly ──
                lookup_success = False
                for ctx in [page] + list(page.frames):
                    try:
                        # Find the claim lookup input and type the claim number
                        result = ctx.evaluate("""(claimId) => {
                            // Find the claim lookup input (id starts with claimLookupIpt)
                            const input = document.querySelector('input[id^="claimLookupIpt"]') 
                                       || document.querySelector('input[ng-model="_InvId"]');
                            if (!input) return { ok: false, reason: 'lookup input not found' };
                            
                            // Clear and set value
                            input.value = '';
                            input.focus();
                            
                            // Use Angular's scope to set the model value
                            const ngEl = window.angular && window.angular.element(input);
                            if (ngEl && ngEl.controller) {
                                const scope = ngEl.scope();
                                if (scope) {
                                    scope._InvId = claimId;
                                    scope.$apply();
                                }
                            }
                            
                            // Also set via native input event
                            input.value = claimId;
                            input.dispatchEvent(new Event('input', { bubbles: true }));
                            input.dispatchEvent(new Event('change', { bubbles: true }));
                            
                            return { ok: true, inputId: input.id };
                        }""", str(claim_id))
                        
                        if result and result.get('ok'):
                            logger.info(f"✅ Set claim lookup to {claim_id} (input: {result.get('inputId')})")
                            time.sleep(0.3)
                            
                            # Click the Lookup button or press Enter
                            clicked_lookup = ctx.evaluate("""() => {
                                // Find Lookup button
                                const btns = document.querySelectorAll('button, input[type="button"]');
                                for (const btn of btns) {
                                    const text = (btn.textContent || btn.value || '').trim();
                                    if (text === 'Lookup' || text === 'Look Up') {
                                        btn.click();
                                        return 'button';
                                    }
                                }
                                return null;
                            }""")
                            
                            if clicked_lookup:
                                logger.info(f"✅ Clicked Lookup button")
                            else:
                                # Fallback: press Enter in the input field
                                try:
                                    input_el = ctx.query_selector('input[id^="claimLookupIpt"]') or ctx.query_selector('input[ng-model="_InvId"]')
                                    if input_el:
                                        input_el.press('Enter')
                                        logger.info(f"✅ Pressed Enter on lookup input")
                                except Exception:
                                    pass
                            
                            lookup_success = True
                            break
                    except Exception:
                        continue

                if not lookup_success:
                    logger.error(f"❌ Could not find claim lookup input — skipping {claim_id}")
                    hcfa_fail_count += 1
                    continue

                # Wait for the claim detail popup to open
                time.sleep(2)

                # ── 6b. Verify the popup ACTUALLY opened ──
                # Must find Cancel/OK/Prog.Notes buttons — NOT just listing page text
                popup_found = False
                for f_ctx in page.frames:
                    try:
                        has_popup = f_ctx.evaluate("""() => {
                            const btns = document.querySelectorAll('button');
                            let hasCancel = false, hasProgNotes = false, hasOK = false;
                            for (const btn of btns) {
                                const t = (btn.value || btn.textContent || '').trim();
                                if (t === 'Cancel') hasCancel = true;
                                if (t.includes('Prog') && t.includes('Note')) hasProgNotes = true;
                                if (t === 'OK') hasOK = true;
                            }
                            return (hasCancel || hasOK) && hasProgNotes;
                        }""")
                        if has_popup:
                            popup_found = True
                            logger.info("✅ Claim popup confirmed (Cancel/OK + Prog Notes found in iframe)")
                            break
                    except Exception:
                        continue
                
                if not popup_found:
                    try:
                        popup_found = page.evaluate("""() => {
                            const bodyText = document.body.innerText || '';
                            return bodyText.includes('Prog. Notes') && bodyText.includes('Cancel');
                        }""")
                        if popup_found:
                            logger.info("✅ Claim popup confirmed (Prog Notes + Cancel in main page)")
                    except Exception:
                        pass


                if not popup_found:
                    # Last resort: double-click the row
                    logger.info("Popup not detected — trying double-click...")
                    try:
                        page.evaluate('''(claimId) => {
                            const rows = document.querySelectorAll('tbody tr');
                            for (const row of rows) {
                                if ((row.innerText || '').includes(claimId)) {
                                    const evt = new MouseEvent('dblclick', { bubbles: true, cancelable: true });
                                    row.dispatchEvent(evt);
                                    return true;
                                }
                            }
                            return false;
                        }''', str(claim_id))
                        time.sleep(1.5)
                        # Re-check for popup with proper detection
                        for f_ctx in page.frames:
                            try:
                                has_popup = f_ctx.evaluate("""() => {
                                    const btns = document.querySelectorAll('button');
                                    for (const btn of btns) {
                                        const t = (btn.value || btn.textContent || '').trim();
                                        if (t === 'Cancel' || (t.includes('Prog') && t.includes('Note'))) return true;
                                    }
                                    return false;
                                }""")
                                if has_popup:
                                    popup_found = True
                                    break
                            except Exception:
                                continue
                    except Exception:
                        pass

                if not popup_found:
                    logger.error(f"❌ Claim popup did not open for {claim_id} — skipping")
                    hcfa_fail_count += 1
                    continue

                logger.info(f"✅ Claim detail popup is open for {claim_id}")

                try:
                    page.screenshot(path=f'/tmp/hcfa_popup_{claim_id}.png')
                    logger.info(f"📸 Popup screenshot saved for {claim_id}")
                except Exception:
                    pass

                # ── READ-ONLY PROBE: dump the insurance grid DOM/scope, then skip ──
                # Diagnostic only (body flag dump_insurance_grid). Reuses the login,
                # navigation and popup-open machinery above; reads the grid and stops
                # without touching HCFA/submission. Used to resolve the unknowns needed
                # to fix the Subscriber No extraction (primary-row marker, ng-repeat,
                # stale-popup check). See _dump_insurance_grid.
                if body.get('dump_insurance_grid'):
                    _dump_insurance_grid(page, claim_id, aws_client)
                    continue

                # ── 6c. Set up download handler BEFORE clicking the red icon ──
                download_path = f'/tmp/hcfa_{claim_id}.pdf'

                # ── 6d. Find and click the RED ICON next to "Print HCFA" ──
                hcfa_clicked = False
                hcfa_context = page
                # Search iframes for the ACTUAL claim detail popup (not the claims listing).
                # The real popup has visible buttons like Cancel, Prog. Notes, etc.
                best_frame = None
                best_score = 0
                for ctx in page.frames:
                    try:
                        frame_info = ctx.evaluate('''() => {
                            const text = document.body?.innerText || '';
                            const hasHCFA = text.includes('Print HCFA') || text.includes('HCFA');
                            const hasClaimNo = text.includes('Claim No');
                            const hasProgNotes = text.includes('Prog. Notes') || text.includes('Prog Notes');
                            const hasCancel = !!document.querySelector('button:has(span), input[value="Cancel"]');
                            // Check for visible Prog. Notes or Cancel buttons
                            let hasVisibleBtn = false;
                            const btns = document.querySelectorAll('button');
                            for (const btn of btns) {
                                const t = (btn.value || btn.textContent || '').trim();
                                if ((t === 'Cancel' || t.includes('Prog') || t.includes('viewProgNote')) && btn.offsetWidth > 0) {
                                    hasVisibleBtn = true;
                                    break;
                                }
                            }
                            return { hasHCFA, hasClaimNo, hasProgNotes, hasCancel, hasVisibleBtn, 
                                     btnCount: btns.length };
                        }''')
                        score = 0
                        if frame_info.get('hasHCFA'): score += 1
                        if frame_info.get('hasClaimNo'): score += 1
                        if frame_info.get('hasProgNotes'): score += 2  # Strong signal
                        if frame_info.get('hasVisibleBtn'): score += 3  # Strongest signal
                        if frame_info.get('btnCount', 0) > 5: score += 1
                        if score > best_score:
                            best_score = score
                            best_frame = ctx
                    except Exception:
                        continue
                
                if best_frame and best_score >= 2:
                    hcfa_context = best_frame
                    logger.info(f"Found claim popup in iframe (score={best_score}): {best_frame.url[:80] if hasattr(best_frame, 'url') else 'frame'}")
                else:
                    # Fallback to main page only if no iframe found
                    try:
                        main_text = page.evaluate('document.body?.innerText || ""')
                        if 'Print HCFA' in main_text:
                            logger.info("Print HCFA found in main page (no iframe match)")
                    except Exception:
                        pass

                # ── Extract Subscriber ID from the claim popup ──
                subscriber_id = None
                # True only once the value has been confirmed from the HCFA box 1a
                # (the reliable source). A DOM-scrape-only value stays unverified and
                # is gated out of auto-submission below.
                subscriber_id_verified = False
                try:
                    subscriber_id = hcfa_context.evaluate('''() => {
                        // Strategy 1: Angular binding ng-bind="insurance.SubscriberNo"
                        const subSpan = document.querySelector('span[ng-bind="insurance.SubscriberNo"]');
                        if (subSpan && subSpan.textContent.trim()) return subSpan.textContent.trim();
                        
                        // Strategy 2: Look for "Subscriber No" header in insurance table
                        const ths = document.querySelectorAll('th, td');
                        for (const th of ths) {
                            const t = (th.textContent || '').trim();
                            if (t === 'Subscriber No' || t === 'Subscriber ID') {
                                const idx = th.cellIndex;
                                const row = th.closest('tr');
                                if (row && row.nextElementSibling) {
                                    const cell = row.nextElementSibling.cells[idx];
                                    if (cell && cell.textContent.trim()) return cell.textContent.trim();
                                }
                            }
                        }
                        return null;
                    }''')
                    if subscriber_id:
                        logger.info(f"📋 Subscriber ID: {subscriber_id}")
                except Exception as e:
                    logger.warning(f"Could not extract Subscriber ID: {e}")

                # ── Extract Encounter Date ──
                # Old flow used PHM Hub → Care Plan → IV Therapy Edit → Start Date.
                # New flow: read "Rx Start Date" directly from the IV Note (Prog Notes) HTML.
                # The PHM Hub path is kept as a DISABLED fallback (set needs_encounter=False
                # by default so we skip the slow Hub flow entirely).
                encounter_date = None
                needs_encounter = False  # DISABLED — Rx Start Date comes from IV Note instead
                if needs_encounter:
                    try:
                        logger.info(f"🔍 Extracting encounter date for claim {claim_id}...")
                        
                        # Step 1: Click "Hub" button in the claim popup
                        hub_clicked = hcfa_context.evaluate('''() => {
                            const btns = document.querySelectorAll('button');
                            for (const btn of btns) {
                                const text = (btn.textContent || '').trim();
                                if (text === 'Hub' || text.includes('Hub')) {
                                    const id = btn.id || '';
                                    if (id.includes('claimPatientHub') || btn.getAttribute('ng-click')?.includes('loadPatientHubnInfo')) {
                                        btn.click();
                                        return { ok: true, id: id };
                                    }
                                }
                            }
                            return { ok: false };
                        }''')
                        
                        if hub_clicked and hub_clicked.get('ok'):
                            logger.info(f"✅ Clicked Hub button: {hub_clicked.get('id')}")
                            time.sleep(2)  # Wait for Hub popup to open
                            
                            # Step 2: Click PHM Hub dropdown, then PHM Hub link
                            phm_clicked = False
                            for attempt in range(5):
                                for f_ctx in page.frames:
                                    try:
                                        result = f_ctx.evaluate('''() => {
                                            // Find PHM Hub dropdown button
                                            const btns = document.querySelectorAll('button');
                                            for (const btn of btns) {
                                                const text = (btn.textContent || '').trim();
                                                if (text.includes('PHM Hub') && btn.getAttribute('data-toggle') === 'dropdown') {
                                                    btn.click();
                                                    return { ok: true, step: 'dropdown' };
                                                }
                                            }
                                            return { ok: false };
                                        }''')
                                        if result and result.get('ok'):
                                            logger.info("✅ Opened PHM Hub dropdown")
                                            time.sleep(0.5)
                                            
                                            # Click PHM Hub link in the dropdown
                                            f_ctx.evaluate('''() => {
                                                const links = document.querySelectorAll('a');
                                                for (const link of links) {
                                                    const ngClick = link.getAttribute('ng-click') || '';
                                                    if (ngClick.includes("CCMRHub")) {
                                                        link.click();
                                                        return true;
                                                    }
                                                }
                                                // Fallback: find link with text PHM Hub
                                                for (const link of links) {
                                                    if ((link.textContent || '').trim().includes('PHM Hub')) {
                                                        link.click();
                                                        return true;
                                                    }
                                                }
                                                return false;
                                            }''')
                                            phm_clicked = True
                                            break
                                    except Exception:
                                        continue
                                if phm_clicked:
                                    break
                                time.sleep(1)
                            
                            if phm_clicked:
                                logger.info("✅ Clicked PHM Hub link")
                                time.sleep(3)  # Wait for PHM Hub popup
                                
                                # Step 3: Click "Care Plan" tab
                                care_plan_clicked = False
                                for f_ctx in page.frames:
                                    try:
                                        result = f_ctx.evaluate('''() => {
                                            const tabs = document.querySelectorAll('a');
                                            for (const tab of tabs) {
                                                const text = (tab.textContent || '').trim();
                                                if (text === 'Care Plan') {
                                                    tab.click();
                                                    return true;
                                                }
                                            }
                                            return false;
                                        }''')
                                        if result:
                                            care_plan_clicked = True
                                            break
                                    except Exception:
                                        continue
                                
                                if care_plan_clicked:
                                    logger.info("✅ Clicked Care Plan tab")
                                    time.sleep(2)
                                    
                                    # Step 4: Click Expand All
                                    for f_ctx in page.frames:
                                        try:
                                            f_ctx.evaluate('''() => {
                                                const expanders = document.querySelectorAll('[ng-click*="expandAll"], i.icon.add_blue');
                                                for (const el of expanders) {
                                                    el.click();
                                                    return true;
                                                }
                                                // Fallback: look for "Expand All" text
                                                const spans = document.querySelectorAll('span');
                                                for (const s of spans) {
                                                    if ((s.textContent || '').trim() === 'Expand') {
                                                        s.parentElement.click();
                                                        return true;
                                                    }
                                                }
                                                return false;
                                            }''')
                                        except Exception:
                                            continue
                                    
                                    time.sleep(2)
                                    logger.info("✅ Clicked Expand All")
                                    
                                    # Step 5: Find "IV Therapy" INTERVENTION and click its 3-dots → Edit
                                    # Strategy: Find the text node "IV Therapy", walk up to its 
                                    # ng-repeat container, then find the more-0 button WITHIN that 
                                    # specific container (not any other container).
                                    intervention_edit_opened = False
                                    for f_ctx in page.frames:
                                        try:
                                            # Step 5: Click the correct more-0 button (same Y as IV Therapy)
                                            # then click Edit → openEditInterventionPopover
                                            click_result = f_ctx.evaluate('''() => {
                                                // Check if IV Therapy exists in this frame
                                                const bodyText = (document.body.textContent || '');
                                                if (!bodyText.includes('IV Therapy')) {
                                                    return { ok: false, reason: 'no-iv-therapy-text' };
                                                }
                                                
                                                // Find IV Therapy text Y-position
                                                const walker = document.createTreeWalker(
                                                    document.body, NodeFilter.SHOW_TEXT
                                                );
                                                let ivY = 0;
                                                while (walker.nextNode()) {
                                                    const t = (walker.currentNode.textContent || '').trim();
                                                    if (t.includes('IV Therapy')) {
                                                        const r = walker.currentNode.parentElement.getBoundingClientRect();
                                                        ivY = r.top;
                                                        break;
                                                    }
                                                }
                                                if (ivY === 0) return { ok: false, reason: 'iv-therapy-no-position' };
                                                
                                                // Find the more-0 button closest to IV Therapy Y
                                                const allBtns = document.querySelectorAll('button[id^="more-"]');
                                                let closestBtn = null;
                                                let closestDist = Infinity;
                                                for (const btn of allBtns) {
                                                    if (btn.offsetWidth === 0) continue;
                                                    const rect = btn.getBoundingClientRect();
                                                    const dist = Math.abs(rect.top - ivY);
                                                    if (dist < closestDist) {
                                                        closestDist = dist;
                                                        closestBtn = btn;
                                                    }
                                                }
                                                
                                                if (!closestBtn || closestDist > 30) {
                                                    return { ok: false, reason: 'no-more-btn-near-iv', closestDist: closestDist };
                                                }
                                                
                                                // Click the 3-dots button to open dropdown
                                                closestBtn.scrollIntoView({ block: 'center' });
                                                closestBtn.click();
                                                
                                                return { 
                                                    ok: true, 
                                                    step: 'dots-clicked',
                                                    btnId: closestBtn.id,
                                                    btnY: Math.round(closestBtn.getBoundingClientRect().top),
                                                    ivY: Math.round(ivY),
                                                    dist: Math.round(closestDist)
                                                };
                                            }''')
                                            
                                            if not click_result or not click_result.get('ok'):
                                                reason = click_result.get('reason', 'unknown') if click_result else 'null'
                                                if reason == 'no-iv-therapy-text':
                                                    logger.info("📋 IV Therapy NOT found in this frame")
                                                else:
                                                    logger.warning(f"📋 IV Therapy issue: {reason}")
                                                continue
                                            
                                            # 3-dots clicked — now click the Edit menu item
                                            # with ng-click='openEditInterventionPopover'
                                            logger.info(f"✅ Clicked 3-dots near IV Therapy (btnY={click_result.get('btnY')}, ivY={click_result.get('ivY')}, dist={click_result.get('dist')})")
                                            time.sleep(0.5)
                                            
                                            # Click the specific Edit with openEditInterventionPopover
                                            edit_clicked = f_ctx.evaluate('''(ivY) => {
                                                // Find Edit Intervention elements (any tag type)
                                                const links = document.querySelectorAll('[ng-click*="openEditIntervention"]');
                                                if (links.length === 0) {
                                                    // Debug: count all ng-click elements
                                                    const allNgClick = document.querySelectorAll('[ng-click]');
                                                    let editCount = 0;
                                                    for (const el of allNgClick) {
                                                        const nk = el.getAttribute('ng-click') || '';
                                                        if (nk.includes('Intervention') || nk.includes('intervention')) editCount++;
                                                    }
                                                    return { ok: false, reason: 'no-links-found', total: 0, allNgClick: allNgClick.length, interventionNgClick: editCount };
                                                }
                                                
                                                // Pick the one closest to IV Therapy Y position
                                                let best = null;
                                                let bestDist = Infinity;
                                                for (const link of links) {
                                                    const rect = link.getBoundingClientRect();
                                                    const dist = Math.abs(rect.top - ivY);
                                                    if (dist < bestDist) {
                                                        bestDist = dist;
                                                        best = link;
                                                    }
                                                }
                                                
                                                if (best) {
                                                    best.click();
                                                    return { 
                                                        ok: true, 
                                                        total: links.length,
                                                        tag: best.tagName,
                                                        dist: Math.round(bestDist),
                                                        ngClick: (best.getAttribute('ng-click') || '').substring(0, 60)
                                                    };
                                                }
                                                return { ok: false, reason: 'no-best-link' };
                                            }''', click_result.get('ivY', 0))
                                            
                                            if not edit_clicked or not edit_clicked.get('ok'):
                                                logger.warning(f"📋 Could not find Edit Intervention link: {edit_clicked}")
                                                continue
                                            
                                            logger.info(f"✅ Clicked Edit Intervention menu item")
                                            time.sleep(1)
                                            
                                            # Dismiss "unsaved data" dialog if present
                                            f_ctx.evaluate('''() => {
                                                const btns = document.querySelectorAll('button');
                                                for (const btn of btns) {
                                                    const t = (btn.textContent || '').trim();
                                                    if (t === "Don't Save" && btn.offsetWidth > 0) {
                                                        btn.click(); return true;
                                                    }
                                                }
                                                return false;
                                            }''')
                                            time.sleep(2)
                                            
                                            # The Edit Intervention opens within the Edit Goal modal
                                            # The intervention data (including start date) is inside it
                                            intervention_edit_opened = True
                                            edit_result = {'ok': True, 'step': 'intervention-edit-opened'}
                                            
                                            if edit_result and edit_result.get('ok'):
                                                
                                                # Take screenshot for debugging
                                                try:
                                                    page.screenshot(path=f'/tmp/enc_edit_{claim_id}.png')
                                                    logger.info(f"📸 Edit modal screenshot: /tmp/enc_edit_{claim_id}.png")
                                                except Exception:
                                                    pass
                                                
                                                # Step 6: Extract Start Date from Edit Intervention modal
                                                # The date is rendered by Angular's datepicker — the visual 
                                                # text "22.10.2025" is NOT in input.value. We must read it
                                                # via Angular scope ($viewValue or $modelValue).
                                                date_info = f_ctx.evaluate('''() => {
                                                    const results = [];
                                                    const inputs = document.querySelectorAll('input');
                                                    for (const inp of inputs) {
                                                        const visible = inp.offsetWidth > 0 && inp.offsetHeight > 0;
                                                        if (!visible) continue;
                                                        
                                                        // Try standard value first
                                                        let val = inp.value || '';
                                                        
                                                        // Try Angular scope for the real value
                                                        let angularVal = '';
                                                        try {
                                                            if (window.angular) {
                                                                const ngEl = window.angular.element(inp);
                                                                const ctrl = ngEl.controller ? ngEl.controller('ngModel') : null;
                                                                if (ctrl) {
                                                                    // Try $viewValue (formatted display string)
                                                                    if (ctrl.$viewValue) angularVal = String(ctrl.$viewValue);
                                                                    // Try $modelValue (could be Date object)
                                                                    if (!angularVal && ctrl.$modelValue) {
                                                                        const mv = ctrl.$modelValue;
                                                                        if (mv instanceof Date) {
                                                                            const dd = String(mv.getDate()).padStart(2, '0');
                                                                            const mm = String(mv.getMonth() + 1).padStart(2, '0');
                                                                            const yyyy = mv.getFullYear();
                                                                            angularVal = dd + '.' + mm + '.' + yyyy;
                                                                        } else {
                                                                            angularVal = String(mv);
                                                                        }
                                                                    }
                                                                }
                                                                // Also try scope binding
                                                                if (!angularVal) {
                                                                    const scope = ngEl.scope ? ngEl.scope() : null;
                                                                    const ngModel = inp.getAttribute('ng-model');
                                                                    if (scope && ngModel) {
                                                                        const parts = ngModel.split('.');
                                                                        let v = scope;
                                                                        for (const p of parts) {
                                                                            v = v ? v[p] : undefined;
                                                                        }
                                                                        if (v instanceof Date) {
                                                                            const dd = String(v.getDate()).padStart(2, '0');
                                                                            const mm = String(v.getMonth() + 1).padStart(2, '0');
                                                                            const yyyy = v.getFullYear();
                                                                            angularVal = dd + '.' + mm + '.' + yyyy;
                                                                        } else if (v) {
                                                                            angularVal = String(v);
                                                                        }
                                                                    }
                                                                }
                                                            }
                                                        } catch (e) {}
                                                        
                                                        const effectiveVal = angularVal || val;
                                                        if (effectiveVal) {
                                                            results.push({
                                                                value: val,
                                                                angularValue: angularVal,
                                                                effectiveValue: effectiveVal,
                                                                ngModel: inp.getAttribute('ng-model') || '',
                                                                placeholder: inp.placeholder || '',
                                                                type: inp.type || ''
                                                            });
                                                        }
                                                    }
                                                    
                                                    // Also check modal title
                                                    const headers = document.querySelectorAll('h4, h3, .modal-title, [class*="header"]');
                                                    let modalTitle = '';
                                                    for (const h of headers) {
                                                        const t = (h.textContent || '').trim();
                                                        if (t.includes('Edit')) {
                                                            modalTitle = t;
                                                            break;
                                                        }
                                                    }
                                                    
                                                    return { inputs: results, modalTitle: modalTitle };
                                                }''')
                                                
                                                logger.info(f"📋 Edit modal: title='{date_info.get('modalTitle', '')}', inputs found={len(date_info.get('inputs', []))}")
                                                for inp in (date_info or {}).get('inputs', []):
                                                    logger.info(f"  📋 input: value='{inp.get('value')}', angular='{inp.get('angularValue')}', ngModel='{inp.get('ngModel')}', placeholder='{inp.get('placeholder')}'")
                                                
                                                # Find the Start Date value
                                                # Priority: dateObj.singleDatepicker (intervention dates)
                                                # The FIRST dateObj.singleDatepicker = Start Date
                                                import re
                                                date_val = None
                                                
                                                # First pass: look for dateObj.singleDatepicker inputs
                                                for inp in (date_info or {}).get('inputs', []):
                                                    ng_model = inp.get('ngModel', '')
                                                    if 'singleDatepicker' in ng_model:
                                                        val = inp.get('effectiveValue', '') or inp.get('value', '')
                                                        # Match YYYY-MM-DD
                                                        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', val)
                                                        if m:
                                                            date_val = val
                                                            logger.info(f"📋 Found START date from singleDatepicker: {val}")
                                                            break
                                                        # Match DD.MM.YYYY
                                                        m2 = re.match(r'^(\d{2})\.(\d{2})\.(\d{4})$', val)
                                                        if m2:
                                                            day, month, year = m2.groups()
                                                            date_val = f"{year}-{month}-{day}"
                                                            logger.info(f"📋 Found START date (DD.MM.YYYY): {val} → {date_val}")
                                                            break
                                                
                                                # Fallback: any date-like input
                                                if not date_val:
                                                    for inp in (date_info or {}).get('inputs', []):
                                                        val = inp.get('effectiveValue', '') or inp.get('value', '')
                                                        m = re.match(r'^(\d{2})\.(\d{2})\.(\d{4})$', val)
                                                        if m:
                                                            day, month, year = m.groups()
                                                            date_val = f"{year}-{month}-{day}"
                                                            logger.info(f"📋 Found date DD.MM.YYYY (fallback): {val} → {date_val}")
                                                            break
                                                        m3 = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', val)
                                                        if m3:
                                                            date_val = val
                                                            logger.info(f"📋 Found date YYYY-MM-DD (fallback): {val}")
                                                            break
                                                
                                                if date_val:
                                                    encounter_date = date_val
                                                    logger.info(f"📋 Encounter Date: {encounter_date}")
                                                else:
                                                    logger.warning(f"No date found in Edit modal inputs")
                                                
                                                # Close the Edit Intervention modal → click Cancel
                                                try:
                                                    f_ctx.evaluate('''() => {
                                                        const btns = document.querySelectorAll('button');
                                                        for (const btn of btns) {
                                                            const text = (btn.textContent || '').trim();
                                                            if (text === 'Cancel' && btn.offsetWidth > 0) {
                                                                btn.click();
                                                                return true;
                                                            }
                                                        }
                                                        return false;
                                                    }''')
                                                except Exception:
                                                    pass
                                                time.sleep(1)
                                                
                                                # Dismiss any "Don't Save" dialog that appears after Cancel
                                                try:
                                                    f_ctx.evaluate('''() => {
                                                        const btns = document.querySelectorAll('button');
                                                        for (const btn of btns) {
                                                            const t = (btn.textContent || '').trim();
                                                            if (t === "Don't Save" && btn.offsetWidth > 0) {
                                                                btn.click(); return true;
                                                            }
                                                        }
                                                        return false;
                                                    }''')
                                                except Exception:
                                                    pass
                                                time.sleep(0.5)
                                                
                                                break
                                        except Exception as e:
                                            logger.debug(f"Frame search error: {e}")
                                            continue
                                else:
                                    logger.warning("Could not click Care Plan tab")
                            else:
                                logger.warning("Could not click PHM Hub")
                        else:
                            logger.warning("Could not find Hub button")
                        
                        # ── Close all nested popups properly ──
                        # Order: Edit modal Cancel → PHM Hub X → Patient Hub X → any alert OK
                        _close_all_hub_popups(page)

                    except Exception as e:
                        logger.warning(f"Encounter date extraction failed: {e}")
                        _close_all_hub_popups(page)

                # ── Re-open the claim popup after PHM Hub closure ──
                # PHM Hub / Patient Hub closure can also dismiss the underlying claim
                # detail popup. Re-opening via Claim Lookup is the simplest robust path
                # to ensure HCFA / Prog Notes / Encounter File extraction still works.
                if needs_encounter:
                    logger.info(f"🔄 Re-opening claim {claim_id} popup after PHM Hub flow…")
                    reopened, refreshed_frame = _open_claim_popup_via_lookup(page, claim_id)
                    if reopened:
                        if refreshed_frame:
                            hcfa_context = refreshed_frame
                            logger.info(f"✅ Claim popup re-opened; refreshed hcfa_context to best frame")
                        else:
                            logger.info(f"✅ Claim popup re-opened (using main page as context)")
                    else:
                        logger.warning(f"⚠️ Could not re-open claim {claim_id} popup — downstream extraction may fail")

                # ══ ENCOUNTER FILE EXTRACTION ══
                # From the claim detail popup, use the Smart Panel → Encounters tab
                # If encounter_date wasn't extracted fresh, load from DynamoDB
                # Determine the CLAIM's DOS — this is what we filter encounters by.
                # claim_record has 'dos' / 'service_date' in MM/DD/YYYY format.
                claim_dos = claim_record.get('dos') or claim_record.get('service_date') or ''
                if not claim_dos:
                    # Fallback: load from DDB
                    try:
                        db_item = aws_client.dynamodb.Table('helixona-claims').get_item(Key={'claim_id': claim_id}).get('Item', {})
                        claim_dos = db_item.get('dos') or db_item.get('service_date') or ''
                    except Exception:
                        pass

                if not encounter_date:
                    try:
                        db_item = aws_client.dynamodb.Table('helixona-claims').get_item(Key={'claim_id': claim_id}).get('Item', {})
                        encounter_date = db_item.get('encounter_date', '')
                        if encounter_date:
                            logger.info(f"📂 Stored encounter_date in DynamoDB: {encounter_date} (kept for reference)")
                    except Exception:
                        pass


                # ── Encounter / Progress Note extraction (driven by IV Note Rx Start Date) ──
                # Rx Start Date is stored YYYY-MM-DD; the encounter search wants MM/DD/YYYY.
                def _rx_to_mmddyyyy(rx):
                    rx = (rx or '').strip()
                    if not rx:
                        return ''
                    if '-' in rx:
                        try:
                            y, m, d = rx.split('-')
                            return f"{m}/{d}/{y}"
                        except Exception:
                            return ''
                    return rx
                # Defined here; invoked AFTER the IV Note is captured (Rx Start Date ready)
                # and from the metadata-only path below. Single code path — no duplication.
                def _extract_encounter_file(enc_date_mmddyyyy):
                    # Office visit (E/M CPT 99201-99205 / 99211-99215): only HCFA +
                    # IV Note are required — NO Progress Note. Flag it and bail (this
                    # is NOT a "Needs Review" case — the Progress Note simply doesn't apply).
                    _cpt = claim_record.get('cpt')
                    if not _cpt:
                        try:
                            _cpt = aws_client.dynamodb.Table('helixona-claims').get_item(
                                Key={'claim_id': claim_id}).get('Item', {}).get('cpt')
                        except Exception:
                            _cpt = None
                    if is_office_visit(_cpt):
                        logger.info(f"🏥 Claim {claim_id} is an office visit (CPT {_cpt}) — Progress Note not required")
                        try:
                            aws_client.update_claim_status(claim_id, {
                                'office_visit': True,
                                'progress_note_required': False,
                                'encounter_revision_needed': False,
                            })
                        except Exception:
                            pass
                        return
                    if not enc_date_mmddyyyy:
                        logger.warning(f"📂 No Rx Start Date for claim {claim_id} — encounter marked Needs Review")
                        try:
                            aws_client.update_claim_status(claim_id, {
                                'encounter_revision_needed': True,
                                'encounter_capture_failed': 'no_rx_start_date',
                                'encounter_capture_failed_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                            })
                        except Exception:
                            pass
                        return
                    if not claim_record.get('encounter_file_s3_path'):
                        try:
                            logger.info(f"📂 Starting Encounter File extraction (claim DOS: {claim_dos or 'unknown'})")

                            # Target date = the IV Note's Rx Start Date (MM/DD/YYYY), passed in as
                            # enc_date_mmddyyyy. (The empty case is handled by the guard above.)
                            logger.info(f"📂 Looking for encounter by Rx Start Date: {enc_date_mmddyyyy}")
                            logger.info(f"📂 Looking for date: {enc_date_mmddyyyy or '(none — will pick most recent matching type)'}")
                        
                            import subprocess
                            import base64
                        
                            # Step 1: Click "Encounters" in the Smart Panel (right side of claim detail popup)
                            # The claim detail popup has a smart panel with tabs like Encounters, Documents, etc.
                            enc_tab_clicked = False
                            for frm in page.frames:
                                try:
                                    result = frm.evaluate('''() => {
                                        // Strategy 1: <a> with title="Encounters"
                                        const links = document.querySelectorAll('a[title="Encounters"]');
                                        for (const a of links) {
                                            if (a.offsetWidth > 0) { a.click(); return 'title-attr'; }
                                        }
                                        // Strategy 2: <label> text "Encounters" → click parent <a>
                                        const labels = document.querySelectorAll('label');
                                        for (const lbl of labels) {
                                            if ((lbl.textContent || '').trim() === 'Encounters') {
                                                const parent = lbl.closest('a');
                                                if (parent) { parent.click(); return 'label-parent'; }
                                                lbl.click();
                                                return 'label-direct';
                                            }
                                        }
                                        // Strategy 3: Any <a> with ng-click containing loadSmartPanel and Encounters text
                                        const all = document.querySelectorAll('a.cursor, a[ng-click*="loadSmartPanel"]');
                                        for (const a of all) {
                                            if ((a.textContent || '').trim().includes('Encounters')) {
                                                a.click();
                                                return 'smart-panel-text';
                                            }
                                        }
                                        return null;
                                    }''')
                                    if result:
                                        enc_tab_clicked = True
                                        logger.info(f"📂 Clicked Encounters tab via: {result}")
                                        break
                                except Exception:
                                    continue
                        
                            if not enc_tab_clicked:
                                logger.warning("📂 Could not find Encounters tab in Smart Panel")
                                try:
                                    aws_client.update_claim_status(claim_id, {
                                        'encounter_capture_failed': 'tab_not_found',
                                        'encounter_capture_failed_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                        'encounter_revision_needed': True,
                                    })
                                except Exception as _flag_err:
                                    logger.warning(f"Failed to write encounter_capture_failed flag: {_flag_err}")
                            else:
                                # Poll directly for the SAME selector the scan will read
                                # (td[ng-click*="viewEncClick"]) instead of guessing at the
                                # row wrapper — that way we know rows are actually queryable
                                # before we snapshot. Poll across all frames (Encounters
                                # may render inside an iframe). Give up after ~15s; if still
                                # empty by then, dump frame diagnostics so we can see why.
                                _cells_count = 0
                                for _i in range(30):  # 30 × 500ms = 15s max
                                    _max_cells = 0
                                    for _frm_w in [page] + list(page.frames):
                                        try:
                                            _c = _frm_w.evaluate(
                                                '() => document.querySelectorAll(\'td[ng-click*="viewEncClick"]\').length')
                                            if _c > _max_cells:
                                                _max_cells = _c
                                        except Exception:
                                            continue
                                    if _max_cells > 0 and _max_cells == _cells_count:
                                        # Stable: rows aren't growing anymore.
                                        break
                                    _cells_count = _max_cells
                                    time.sleep(0.5)
                                logger.info(f"📂 Encounters list ready: {_cells_count} clickable cells across all frames")

                                # If we still have 0 rows after 15s, log frame-level
                                # diagnostics so we can tell whether the tab click
                                # missed, ECW responded slowly, or the patient just
                                # has no encounters.
                                if _cells_count == 0:
                                    try:
                                        _diag = []
                                        for _frm_d in [page] + list(page.frames):
                                            try:
                                                _info = _frm_d.evaluate('''() => ({
                                                    rows: document.querySelectorAll('tr[ng-repeat]').length,
                                                    encTabs: document.querySelectorAll('a[title="Encounters"]').length,
                                                    tables: document.querySelectorAll('table').length,
                                                    spinners: document.querySelectorAll('.spinner, [class*="loading"]').length,
                                                    bodyChars: (document.body ? document.body.innerText.length : 0)
                                                })''')
                                                _diag.append(_info)
                                            except Exception:
                                                continue
                                        logger.warning(f"📂 0 encounter cells — frame diagnostics: {_diag}")
                                    except Exception:
                                        pass

                                # Step 2: Find encounter rows matching the CLAIM'S DOS first.
                                # Allowed visit types for billing documentation:
                                ALLOWED_VISIT_TYPES = [
                                    'NP', 'NP F/U', 'F/U', 'WEB', 'TEL',
                                    'ESTP', 'LVL 2 NP', 'LVL 2 F/U', 'LVL 3 NP', 'LVL 3 F/U',
                                    'PRO – PRO', 'PRO – TP', 'Telemed', 'TOCPT',
                                ]
                                # Also accept ECW's IV-infusion encounter visit types via regex —
                                # rendered as "Flow N" (e.g. "Flow 20", "Flow 60") where N is the
                                # flow rate in mL/hr. These are the actual progress notes for IV
                                # sessions billed under CPT 96365/96367/etc., and were missing
                                # from the static allowlist (claim 5459 had a Flow 20 on the
                                # right date that was getting filtered out).
                                ALLOWED_VISIT_PATTERNS = [
                                    r'^Flow \d+$',     # Flow 20, Flow 30, Flow 60, etc.
                                ]
                                # Picking strategy: ONLY exact-date match on the IV Note's Rx
                                # Start Date with an allowed visit type. Fallbacks were dropped
                                # — closest-before / most-recent matched unrelated Web Encounters
                                # / eMessages. No exact match → Needs Review (no submission).
                                matches = []
                                sample_rows = []
                                selected_date = None
                                selected_strategy = None
                                failure_diag = {}
                                for frm in page.frames:
                                    try:
                                        search_result = frm.evaluate('''(args) => {
                                            const allowed = args.allowedTypes.map(t => t.toLowerCase());
                                            const patterns = (args.allowedPatterns || []).map(p => new RegExp(p, 'i'));
                                            const targetDate = args.targetDate;  // claim DOS as MM/DD/YYYY (may be empty)
                                            function parse(d) {
                                                const m = (d || '').match(/^(\d{2})\/(\d{2})\/(\d{4})$/);
                                                if (!m) return null;
                                                return new Date(parseInt(m[3]), parseInt(m[1]) - 1, parseInt(m[2])).getTime();
                                            }
                                            const targetTs = parse(targetDate);
                                            const isAllowed = (vt) => {
                                                if (!vt) return false;
                                                if (allowed.includes(vt.toLowerCase())) return true;
                                                return patterns.some(re => re.test(vt));
                                            };

                                            // The Encounters tab is already scoped to the current
                                            // claim's patient (popup header shows the patient name;
                                            // each row's columns are date/time/type/status/provider
                                            // — no patient column). So we DON'T patient-filter rows.
                                            const clickCells = document.querySelectorAll('td[ng-click*="viewEncClick"]');
                                            const allRows = [];
                                            const typeMatched = [];
                                            clickCells.forEach((cell, idx) => {
                                                const row = cell.closest('tr');
                                                if (!row) return;
                                                const dateTd = row.querySelector('td[ng-bind="enc.date"]');
                                                const dateText = dateTd ? dateTd.textContent.trim() : '';
                                                const visitSpan = cell.querySelector('span[ng-bind="enc.visitType"]');
                                                const visitType = visitSpan ? visitSpan.textContent.trim() : cell.textContent.trim();
                                                const rowSnippet = (row.textContent || '').replace(/\s+/g, ' ').trim().substring(0, 120);
                                                allRows.push({ idx, date: dateText, visitType, rowText: rowSnippet });
                                                if (isAllowed(visitType)) {
                                                    typeMatched.push({ idx, date: dateText, visitType });
                                                }
                                            });

                                            // EXACT match only: the Progress Note must be on the
                                            // IV Note's Rx Start Date. Fallback strategies were
                                            // dropped because closest-before / most-recent-overall
                                            // ended up matching unrelated Web Encounters / eMessages
                                            // (e.g. claim 4777 picked a 12/22/2025 WEB eMessage when
                                            // the Rx Start Date was 01/09/2026). If no exact match,
                                            // we return no matches → Python marks Needs Review.
                                            if (targetDate) {
                                                const exact = typeMatched.filter(r => r.date === targetDate);
                                                if (exact.length > 0) {
                                                    return {
                                                        matches: exact, selectedDate: targetDate,
                                                        strategy: 'exact-rx-start-date',
                                                        typeMatchedCount: typeMatched.length,
                                                        allRows: allRows.slice(0, 15)
                                                    };
                                                }
                                            }
                                            // Diagnostic detail for the failure case:
                                            // - dateMatched: rows on the target date (any type)
                                            // - typeMatchedSample: rows with allowed type (any date)
                                            // - distinctDates: all dates present in the list
                                            const dateMatched = targetDate
                                                ? allRows.filter(r => r.date === targetDate)
                                                : [];
                                            const distinctDates = Array.from(new Set(
                                                allRows.map(r => r.date).filter(d => d)
                                            )).sort();
                                            return {
                                                matches: [],
                                                selectedDate: null,
                                                strategy: 'no-exact-rx-date-match',
                                                typeMatchedCount: typeMatched.length,
                                                typeMatchedSample: typeMatched.slice(0, 10),
                                                dateMatched: dateMatched,
                                                distinctDates: distinctDates,
                                                allRowsCount: allRows.length,
                                                allRows: allRows.slice(0, 15)
                                            };
                                        }''', {'allowedTypes': ALLOWED_VISIT_TYPES,
                                              'allowedPatterns': ALLOWED_VISIT_PATTERNS,
                                              'targetDate': enc_date_mmddyyyy or ''})

                                        if search_result and search_result.get('matches'):
                                            matches = search_result['matches']
                                            selected_date = search_result.get('selectedDate')
                                            selected_strategy = search_result.get('strategy')
                                            sample_rows = search_result.get('allRows', [])
                                            logger.info(
                                                f"📂 Encounter pick: strategy={selected_strategy}, "
                                                f"{search_result.get('typeMatchedCount', 0)} typed matches, "
                                                f"picked {len(matches)} from {selected_date}: "
                                                f"{[m['visitType'] for m in matches]}"
                                            )
                                            break
                                        elif search_result is not None:
                                            # Don't blindly overwrite — keep whichever frame
                                            # has the MOST rows, so the final sample_rows is
                                            # actually informative for diagnosis (instead of
                                            # the last frame iterated, which is often empty).
                                            _these_rows = search_result.get('allRows', [])
                                            if len(_these_rows) > len(sample_rows):
                                                sample_rows = _these_rows
                                                failure_diag = search_result
                                    except Exception:
                                        continue

                                if not matches and sample_rows:
                                    # No NP/F-U visit found at all — log what types DO exist for triage.
                                    types_present = sorted(set(r['visitType'] for r in sample_rows if r.get('visitType')))
                                    logger.info(f"📂 No NP/F-U encounter found; types present: {types_present}")

                                # Use the selected_date (actual NP/F-U visit date) for matching the
                                # click logic below, instead of the IV-Therapy intervention date.
                                if selected_date:
                                    enc_date_mmddyyyy = selected_date

                                if not matches:
                                    # Surface WHY we couldn't match: how many allowed-type rows
                                    # exist (any date), the dates available, rows ON the target
                                    # date (so we know whether the date is even in the list).
                                    logger.warning(
                                        f"❌ No matching encounter for claim {claim_id} (target={enc_date_mmddyyyy}). "
                                        f"allowed-type rows={failure_diag.get('typeMatchedCount', 0)}/{failure_diag.get('allRowsCount', len(sample_rows))}, "
                                        f"rows ON target date (any type)={len(failure_diag.get('dateMatched', []))}, "
                                        f"distinct dates seen={failure_diag.get('distinctDates', [])[:20]}"
                                    )
                                    if failure_diag.get('dateMatched'):
                                        logger.warning(
                                            f"   ↳ rows on {enc_date_mmddyyyy} (visit types not in allowed list): "
                                            f"{failure_diag.get('dateMatched')}"
                                        )
                                    if failure_diag.get('typeMatchedSample'):
                                        logger.warning(
                                            f"   ↳ sample of allowed-type rows on other dates: "
                                            f"{failure_diag.get('typeMatchedSample')}"
                                        )
                                    logger.warning(f"   ↳ first 5 sample rows: {sample_rows[:5]}")
                                    aws_client.update_claim_status(claim_id, {
                                        'encounter_revision_needed': True,
                                        # Not a date mismatch anymore — it's a type-not-present case.
                                        'encounter_date_mismatch': False,
                                        'encounter_capture_failed': 'no_np_or_fu_visit_type',
                                        'encounter_capture_failed_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                    })
                                else:
                                    # Only ONE Progress Note per claim — submission needs a
                                    # single doc, and ECW's Encounters list often shows the
                                    # same encounter under multiple visit-type columns (e.g.
                                    # claim 5425 had ['TEL','TEL','Flow 25','Flow 25'] — same
                                    # session listed 4× → 4 duplicate PDFs uploaded to
                                    # Symplisend (6 files total). Trim to the first match.
                                    if len(matches) > 1:
                                        logger.info(f"📂 Encounters tab returned {len(matches)} matching rows {[m.get('visitType') for m in matches]}; capturing ONLY the first")
                                        matches = matches[:1]

                                    # Step 3: For each matching encounter, click → capture PDF → close modal → next
                                    captured_paths = []
                                    captured_visit_types = []

                                    # Register popup listener once for Strategy B (printID popup capture)
                                    popup_pages = []
                                    def _on_popup(popup_page):
                                        popup_pages.append(popup_page)
                                        try:
                                            popup_page.evaluate('''() => { window.print = function() {}; }''')
                                        except Exception:
                                            pass
                                    page.context.on('page', _on_popup)

                                    for enc_idx, match in enumerate(matches, start=1):
                                        visit_type = match.get('visitType', '?')
                                        logger.info(f"📂 Capture attempt {enc_idx}/{len(matches)} (visitType={visit_type})")

                                        # Re-click the Encounters tab between iterations — the list re-renders
                                        # after each preview modal closes, so DOM indices are not stable
                                        if enc_idx > 1:
                                            re_tab_ok = False
                                            for frm in page.frames:
                                                try:
                                                    re_tab = frm.evaluate('''() => {
                                                        const links = document.querySelectorAll('a[title="Encounters"]');
                                                        for (const a of links) {
                                                            if (a.offsetWidth > 0) { a.click(); return true; }
                                                        }
                                                        const labels = document.querySelectorAll('label');
                                                        for (const lbl of labels) {
                                                            if ((lbl.textContent || '').trim() === 'Encounters') {
                                                                const parent = lbl.closest('a');
                                                                if (parent) { parent.click(); return true; }
                                                            }
                                                        }
                                                        return false;
                                                    }''')
                                                    if re_tab:
                                                        re_tab_ok = True
                                                        break
                                                except Exception:
                                                    continue
                                            if not re_tab_ok:
                                                logger.warning(f"  ❌ Could not re-click Encounters tab for iteration {enc_idx}")
                                                continue
                                            time.sleep(3)

                                        # Click the (enc_idx-1)-th encounter row matching date AND allowed type.
                                        # The Encounters tab is already scoped to the current claim's
                                        # patient at the popup level — no per-row patient filter needed.
                                        #
                                        # Two-step click: (1) tag the target cell with a unique
                                        # data attribute via evaluate(); (2) use Playwright's
                                        # mouse.click at the cell's bounding-box center for a
                                        # TRUSTED user gesture. Plain `cell.click()` inside
                                        # evaluate() is a JS click and doesn't always trigger
                                        # Angular's `viewEncClick` for certain row types — claim
                                        # 5074 (Brown) saw the JS click no-op while the same code
                                        # worked for claim 5305 (Castaneda). A real mouse click
                                        # behaves identically to a user click.
                                        nth_to_click = enc_idx - 1
                                        click_ok = False
                                        click_box = None  # {x, y, frame_url} of the cell
                                        for frm in page.frames:
                                            try:
                                                _tag_result = frm.evaluate('''(args) => {
                                                    const targetDate = args.targetDate;
                                                    const allowed = args.allowedTypes.map(t => t.toLowerCase());
                                                    const patterns = (args.allowedPatterns || []).map(p => new RegExp(p, 'i'));
                                                    const nth = args.nth;
                                                    const isAllowed = (vt) => {
                                                        if (!vt) return false;
                                                        if (allowed.includes(vt.toLowerCase())) return true;
                                                        return patterns.some(re => re.test(vt));
                                                    };
                                                    // Clear any prior tagging
                                                    document.querySelectorAll('[data-helixona-enc-target="1"]').forEach(e => e.removeAttribute('data-helixona-enc-target'));
                                                    const cells = document.querySelectorAll('td[ng-click*="viewEncClick"]');
                                                    let count = 0;
                                                    for (const cell of cells) {
                                                        const row = cell.closest('tr');
                                                        if (!row) continue;
                                                        const dateTd = row.querySelector('td[ng-bind="enc.date"]');
                                                        if (!dateTd || dateTd.textContent.trim() !== targetDate) continue;
                                                        const visitSpan = cell.querySelector('span[ng-bind="enc.visitType"]');
                                                        const visitType = (visitSpan ? visitSpan.textContent.trim() : cell.textContent.trim());
                                                        if (!isAllowed(visitType)) continue;
                                                        if (count === nth) {
                                                            cell.setAttribute('data-helixona-enc-target', '1');
                                                            cell.scrollIntoView({block: 'center', inline: 'center'});
                                                            const r = cell.getBoundingClientRect();
                                                            return { ok: true, x: r.left + r.width/2, y: r.top + r.height/2 };
                                                        }
                                                        count += 1;
                                                    }
                                                    return { ok: false };
                                                }''', {'targetDate': enc_date_mmddyyyy,
                                                      'allowedTypes': ALLOWED_VISIT_TYPES,
                                                      'allowedPatterns': ALLOWED_VISIT_PATTERNS,
                                                      'nth': nth_to_click})
                                                if _tag_result and _tag_result.get('ok'):
                                                    click_box = _tag_result
                                                    click_box['frame'] = frm
                                                    break
                                            except Exception:
                                                continue
                                        if click_box:
                                            try:
                                                # Trusted Playwright click at the cell coords.
                                                page.mouse.click(click_box['x'], click_box['y'])
                                                click_ok = True
                                                logger.info(f"  🖱️ Trusted click on encounter row at ({click_box['x']:.0f},{click_box['y']:.0f}) (enc_idx={enc_idx})")
                                            except Exception as _mc_err:
                                                logger.warning(f"  mouse.click failed: {_mc_err}")
                                                # Fall back to dispatch_event on the tagged cell.
                                                try:
                                                    click_box['frame'].evaluate('''() => {
                                                        const cell = document.querySelector('[data-helixona-enc-target="1"]');
                                                        if (cell) {
                                                            cell.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
                                                            return true;
                                                        }
                                                        return false;
                                                    }''')
                                                    click_ok = True
                                                    logger.info(f"  🔄 dispatch_event fallback click on encounter row (enc_idx={enc_idx})")
                                                except Exception:
                                                    pass
                                        if not click_ok:
                                            logger.warning(f"  ❌ Could not click {enc_idx}-th match (visitType={visit_type})")
                                            continue

                                        # Wait for the preview modal or web encounter view to
                                        # render. Try element-driven first; fall back to a short
                                        # settle if no known marker appears.
                                        try:
                                            page.wait_for_selector(
                                                'button:has-text("Print"), input[value="Print"], '
                                                'iframe[name*="enc"], #divEncounterPreview, '
                                                '.preview-container, [ng-click*="print"]',
                                                timeout=5000)
                                            time.sleep(0.5)
                                        except Exception:
                                            time.sleep(1.5)

                                        enc_file_path = f'/tmp/encounter_file_{claim_id}_{enc_idx}.pdf'
                                        this_captured = False
                                        is_web_or_tel = visit_type.upper() in ('WEB', 'TEL')

                                        # WEB/TEL: Click "Print Report" → ECW builds the print-
                                        # formatted document. Chromium's print preview renders it
                                        # from a frame or popup that we can target directly with
                                        # CDP — no need to actually open or interact with the
                                        # print-preview UI. Strategy:
                                        #   1. Suppress window.print + snapshot existing frames/popups
                                        #   2. Click "Print Report"
                                        #   3. Wait ~3s for ECW to render the print-formatted DOM
                                        #   4. Scan for a NEW frame or popup whose body contains the
                                        #      current claim's surname (proves it's the right doc)
                                        #   5. CDP `Page.printToPDF` directly on that frame/popup
                                        #      with print-media emulation — yields the same PDF
                                        #      Chrome's print preview "Save" button would produce
                                        #   6. If no matching frame/popup found, log diagnostics and
                                        #      fall through to Strategy A (HTML extract) / B (printID)
                                        # WEB/TEL encounter capture — SAME approach as IV Note:
                                        # 1. Suppress window.print so Chrome's print preview never
                                        #    opens (avoids modal blocking of evaluate() and
                                        #    eliminates the xdotool/file-chooser path entirely).
                                        # 2. Click "Print Report" so ECW builds the print-formatted
                                        #    DOM into a frame (sets up the same DOM Chrome would
                                        #    render in its preview).
                                        # 3. Scan all frames for the one containing this claim's
                                        #    surname in the header — that's the print DOM frame.
                                        # 4. Extract its full HTML and render via Playwright's
                                        #    page.pdf() in a fresh page. Same call IV Note uses
                                        #    successfully ([src/main.py:7099-7105]).
                                        # WEB/TEL encounter capture — capture the HTML that ECW
                                        # passes to window.print(), then render to PDF via
                                        # Playwright (same as IV Note flow). Approach:
                                        #   1. Inject a `print()` override that stashes the
                                        #      caller frame's outerHTML on `window.__helixonaPrintHTML`
                                        #      INSTEAD of opening Chrome's print preview. We use
                                        #      `add_init_script` on the context so the override
                                        #      applies to every iframe ECW creates dynamically —
                                        #      simple evaluate() at click time misses iframes
                                        #      created AFTER the inject (which is how ECW's
                                        #      "Print Report" path actually works).
                                        #   2. Click "Print Report" — ECW builds the print DOM in
                                        #      a new iframe and calls iframe.contentWindow.print(),
                                        #      which is now our override → HTML captured, no
                                        #      preview opens.
                                        #   3. Scan all frames for the captured HTML and render
                                        #      via Playwright's page.pdf() — same call IV Note
                                        #      uses ([src/main.py:7099-7105]).
                                        if is_web_or_tel:
                                            try:
                                                _expected_sn = (claim_record.get('patient_name', '') or '').split(',')[0].strip().upper()

                                                # The override + capture script. Stashes the calling
                                                # frame's outerHTML so we can read it after click.
                                                _capture_js = '''
                                                () => {
                                                    try {
                                                        window.__helixonaPrintHTML = null;
                                                        window.__helixonaPrintAt = 0;
                                                        const _orig = window.print;
                                                        window.print = function() {
                                                            try {
                                                                window.__helixonaPrintHTML = document.documentElement.outerHTML;
                                                                window.__helixonaPrintAt = Date.now();
                                                            } catch(e) {}
                                                        };
                                                    } catch(e) {}
                                                }
                                                '''
                                                # Apply to ALL existing frames now.
                                                try: page.evaluate(_capture_js)
                                                except Exception: pass
                                                for frm in page.frames:
                                                    try: frm.evaluate(_capture_js)
                                                    except Exception: continue
                                                # And to every NEW document loaded into the context
                                                # from this point on — Playwright runs init scripts
                                                # on every frame in the context as it loads.
                                                _init_script_id = None
                                                try:
                                                    page.context.add_init_script('''
                                                    (() => {
                                                        try {
                                                            window.__helixonaPrintHTML = null;
                                                            window.__helixonaPrintAt = 0;
                                                            window.print = function() {
                                                                try {
                                                                    window.__helixonaPrintHTML = document.documentElement.outerHTML;
                                                                    window.__helixonaPrintAt = Date.now();
                                                                } catch(e) {}
                                                            };
                                                        } catch(e) {}
                                                    })();
                                                    ''')
                                                    _init_script_id = True
                                                except Exception as _is_err:
                                                    logger.warning(f"  add_init_script failed: {_is_err}")

                                                # Click Print Report — ECW will build the print
                                                # DOM and call window.print() (which is now our
                                                # capturing override → no preview opens).
                                                report_clicked = None
                                                for frm in page.frames:
                                                    try:
                                                        clicked = frm.evaluate('''() => {
                                                            const btns = document.querySelectorAll('button, input[type="button"], input[type="submit"]');
                                                            for (const btn of btns) {
                                                                const txt = (btn.value || btn.textContent || '').trim().toLowerCase();
                                                                const ngClick = (btn.getAttribute('ng-click') || '').toLowerCase();
                                                                if (txt === 'print report' || txt.includes('print report')
                                                                    || ngClick.includes('printreport') || ngClick.includes("'print report'")) {
                                                                    if (btn.offsetWidth > 0 && btn.offsetHeight > 0) {
                                                                        btn.click();
                                                                        return { ok: true, txt: txt };
                                                                    }
                                                                }
                                                            }
                                                            return { ok: false };
                                                        }''')
                                                        if clicked and clicked.get('ok'):
                                                            report_clicked = clicked
                                                            logger.info(f"  🖨️ Clicked Print Report for {visit_type}: {clicked}")
                                                            break
                                                    except Exception:
                                                        continue

                                                if report_clicked:
                                                    # Two ways to find the encounter HTML:
                                                    #   A. `window.__helixonaPrintHTML` set by our
                                                    #      window.print override (works if ECW
                                                    #      actually calls print() on a frame).
                                                    #   B. Body innerText of any frame contains the
                                                    #      patient surname + encounter markers
                                                    #      (works regardless of how ECW renders the
                                                    #      print content — it could just inject
                                                    #      HTML into a frame without ever calling
                                                    #      print()). Since the preview doesn't open
                                                    #      anymore (override prevented it), frames
                                                    #      are not modally blocked → evaluate()
                                                    #      returns quickly.
                                                    target_frame_html = None
                                                    target_frame_url = None
                                                    target_method = None
                                                    _deadline_cap = time.monotonic() + 8.0
                                                    while time.monotonic() < _deadline_cap and not target_frame_html:
                                                        time.sleep(0.3)
                                                        for _f in page.frames:
                                                            # Method A: captured via print() override
                                                            try:
                                                                _captured = _f.evaluate(
                                                                    '() => window.__helixonaPrintHTML || null'
                                                                )
                                                            except Exception:
                                                                _captured = None
                                                            if _captured and len(_captured) > 500:
                                                                if _expected_sn and _expected_sn in _captured.upper()[:2000]:
                                                                    target_frame_html = _captured
                                                                    target_frame_url = _f.url or '<no-url>'
                                                                    target_method = 'window.print override'
                                                                    break
                                                            # Method B: frame body contains the
                                                            # encounter document directly.
                                                            try:
                                                                _head = (_f.evaluate(
                                                                    '() => document.body ? document.body.innerText.substring(0, 600) : ""'
                                                                ) or '').upper()
                                                            except Exception:
                                                                continue
                                                            if not _expected_sn or _expected_sn not in _head:
                                                                continue
                                                            if not (('PROVIDER' in _head and 'DOB' in _head)
                                                                    or 'TELEPHONE ENCOUNTER' in _head
                                                                    or ('DOS' in _head and 'ACC NO' in _head)):
                                                                continue
                                                            try:
                                                                _html = _f.evaluate(
                                                                    '() => document.documentElement ? document.documentElement.outerHTML : ""'
                                                                )
                                                            except Exception:
                                                                continue
                                                            if _html and len(_html) > 500:
                                                                target_frame_html = _html
                                                                target_frame_url = _f.url or '<no-url>'
                                                                target_method = 'frame body scan'
                                                                break
                                                        if target_frame_html:
                                                            logger.info(f"  🎯 Captured encounter HTML via {target_method} from {target_frame_url[:80]} ({len(target_frame_html)} chars)")

                                                    # Remove the init script override so subsequent
                                                    # claims don't get it permanently.
                                                    try:
                                                        page.context.add_init_script(
                                                            '(() => { try { delete window.print; } catch(e){} })();'
                                                        )
                                                    except Exception:
                                                        pass

                                                    if target_frame_html:
                                                        # Same render call IV Note uses — open a
                                                        # fresh Playwright page, set_content, pdf().
                                                        try:
                                                            pdf_page = page.context.new_page()
                                                            pdf_page.set_content(target_frame_html, wait_until='domcontentloaded')
                                                            time.sleep(0.5)
                                                            pdf_page.pdf(
                                                                path=enc_file_path,
                                                                format='Letter',
                                                                print_background=True,
                                                                margin={'top': '0.4in', 'bottom': '0.4in',
                                                                        'left': '0.4in', 'right': '0.4in'}
                                                            )
                                                            pdf_page.close()
                                                            _sz = os.path.getsize(enc_file_path) if os.path.exists(enc_file_path) else 0
                                                            if _sz > 5000:
                                                                this_captured = True
                                                                logger.info(f"  ✅ Strategy WEB/TEL (window.print override → Playwright.pdf) captured ({_sz} bytes)")
                                                            else:
                                                                logger.warning(f"  ⚠️ HTML-extract PDF too small ({_sz} bytes)")
                                                        except Exception as _rend_err:
                                                            logger.warning(f"  HTML-extract render failed: {_rend_err}")
                                                    else:
                                                        # Detailed diagnostic — for each frame log:
                                                        # url, whether the override script ran
                                                        # (__helixonaPrintHTML defined even if null),
                                                        # body head text, and any iframes within.
                                                        # This tells us if ECW uses a print path
                                                        # we don't intercept yet.
                                                        _diag = []
                                                        for _f in page.frames:
                                                            try:
                                                                _info = _f.evaluate('''() => {
                                                                    const out = {
                                                                        scriptRan: typeof window.__helixonaPrintHTML !== "undefined",
                                                                        printCalled: !!window.__helixonaPrintHTML,
                                                                        head: document.body ? document.body.innerText.substring(0, 200) : "",
                                                                        iframes: Array.from(document.querySelectorAll("iframe"))
                                                                                      .map(i => ({src: (i.src||"").substring(0,80), name: i.name || ""}))
                                                                                      .slice(0, 5),
                                                                    };
                                                                    return out;
                                                                }''')
                                                            except Exception:
                                                                _info = {'scriptRan': False, 'head': '<eval-failed>'}
                                                            _diag.append({
                                                                'url': (_f.url or '')[:80],
                                                                'scriptRan': _info.get('scriptRan'),
                                                                'printCalled': _info.get('printCalled'),
                                                                'head': (_info.get('head') or '').replace('\n', ' ')[:120],
                                                                'iframes': _info.get('iframes') or [],
                                                            })
                                                        try:
                                                            subprocess.run(['xdotool', 'key', 'Escape'],
                                                                           env={**os.environ, 'DISPLAY': ':99'},
                                                                           timeout=3)
                                                        except Exception:
                                                            pass
                                                        logger.warning(
                                                            f"  ⚠️ No frame captured encounter HTML for {visit_type} "
                                                            f"(expected surname={_expected_sn!r}). frames_count={len(page.frames)}"
                                                        )
                                                        for _d in _diag:
                                                            logger.warning(f"    frame: {_d}")
                                                else:
                                                    logger.warning(f"  ⚠️ Print Report button not found for {visit_type}")
                                                    # Diagnostic: dump frame bodies to see WHERE
                                                    # the modal opened (or if it opened at all).
                                                    for _f in page.frames:
                                                        try:
                                                            _info = _f.evaluate('''(sn) => {
                                                                const body = document.body;
                                                                const txt = body ? body.innerText : '';
                                                                const head = txt.substring(0, 250);
                                                                const btns = Array.from(document.querySelectorAll('button, input[type="button"]'))
                                                                    .filter(b => b.offsetWidth > 0 && b.offsetHeight > 0)
                                                                    .map(b => ((b.value || b.textContent || '').trim()).substring(0, 30))
                                                                    .filter(t => t)
                                                                    .slice(0, 30);
                                                                return {
                                                                    head: head.replace(/\\n+/g, ' ').substring(0, 200),
                                                                    hasSurname: sn ? txt.toUpperCase().includes(sn) : false,
                                                                    hasTelEnc: txt.toLowerCase().includes('telephone encounter'),
                                                                    hasPrintReport: txt.toLowerCase().includes('print report'),
                                                                    btnCount: btns.length,
                                                                    btns: btns,
                                                                };
                                                            }''', _expected_sn)
                                                        except Exception as _e:
                                                            _info = {'error': str(_e)[:80]}
                                                        logger.warning(f"    frame {(_f.url or '<no-url>')[:60]}: {_info}")
                                            except Exception as _tel_err:
                                                logger.warning(f"  WEB/TEL print capture failed: {_tel_err}")

                                        # Legacy xdotool fallback removed — it relied on driving
                                        # Chrome's print preview UI via keyboard/mouse on Xvfb,
                                        # which was unreliable (GTK focus issues, no _NET_ACTIVE_WINDOW,
                                        # Chromium evaluates blocked by modal print preview). The
                                        # primary HTML-extract strategy above replicates the IV-Note
                                        # capture path and works without driving any OS dialogs.
                                        if False and is_web_or_tel and not this_captured:
                                            try:
                                                _xenv = {**os.environ, 'DISPLAY': ':99'}
                                                _dl_dir = os.path.expanduser('~/Downloads')
                                                try: os.makedirs(_dl_dir, exist_ok=True)
                                                except Exception: pass
                                                # Restore window.print across all frames — the CDP
                                                # path above suppressed it to prevent a blocking OS
                                                # dialog, but the xdotool fallback NEEDS that dialog
                                                # to open so we can drive it with keystrokes. Without
                                                # this, ECW's Print Report click is a no-op and
                                                # Chrome never shows the print preview.
                                                try:
                                                    page.evaluate('() => { try { delete window.print; } catch(e) { window.print = window.__proto__.print; } }')
                                                except Exception:
                                                    pass
                                                for _rfrm in page.frames:
                                                    try:
                                                        _rfrm.evaluate('() => { try { delete window.print; } catch(e) { window.print = window.__proto__.print; } }')
                                                    except Exception:
                                                        continue
                                                # Snapshot PDFs currently in Downloads so we can spot
                                                # the new one Chrome creates after we hit Save.
                                                _before = set()
                                                try:
                                                    _before = {f for f in os.listdir(_dl_dir) if f.lower().endswith('.pdf')}
                                                except Exception:
                                                    pass

                                                # Pre-attempt cleanup — if a previous iteration left
                                                # a stuck print/file dialog, kill it BEFORE we click
                                                # Print Report. (Doing this AFTER the click would
                                                # close the freshly-opened print preview!)
                                                try:
                                                    for _e in range(4):
                                                        subprocess.run(['xdotool', 'key', '--clearmodifiers', 'Escape'],
                                                                       env=_xenv, timeout=2)
                                                        time.sleep(0.2)
                                                except Exception:
                                                    pass

                                                # Click Print Report — let Chrome open the OS dialog.
                                                report_clicked2 = None
                                                for frm in page.frames:
                                                    try:
                                                        clicked2 = frm.evaluate('''() => {
                                                            const btns = document.querySelectorAll('button, input[type="button"], input[type="submit"]');
                                                            for (const btn of btns) {
                                                                const txt = (btn.value || btn.textContent || '').trim().toLowerCase();
                                                                const ngClick = (btn.getAttribute('ng-click') || '').toLowerCase();
                                                                if (txt === 'print report' || txt.includes('print report')
                                                                    || ngClick.includes('printreport')) {
                                                                    if (btn.offsetWidth > 0 && btn.offsetHeight > 0) {
                                                                        btn.click();
                                                                        return { ok: true };
                                                                    }
                                                                }
                                                            }
                                                            return { ok: false };
                                                        }''')
                                                        if clicked2 and clicked2.get('ok'):
                                                            report_clicked2 = True
                                                            logger.info(f"  🖨️ Clicked Print Report for {visit_type} (xdotool path)")
                                                            break
                                                    except Exception:
                                                        continue

                                                if report_clicked2:
                                                    # Helper: list all visible X11 windows with titles.
                                                    def _list_windows():
                                                        try:
                                                            _w = subprocess.run(['xdotool', 'search', '--name', '.'],
                                                                                env=_xenv, capture_output=True, text=True, timeout=5)
                                                            ids = [w for w in (_w.stdout or '').strip().split('\n') if w]
                                                            out = []
                                                            for _wid in ids:
                                                                try:
                                                                    _t = subprocess.run(['xdotool', 'getwindowname', _wid],
                                                                                        env=_xenv, capture_output=True, text=True, timeout=2)
                                                                    out.append((_wid, (_t.stdout or '').strip()))
                                                                except Exception:
                                                                    pass
                                                            return out
                                                        except Exception:
                                                            return []

                                                    # Helper: focus a window THEN send Enter to its
                                                    # focused widget. Under Xvfb without a real WM,
                                                    # `--window <wid>` alone delivers the X event but
                                                    # GTK file-chooser ignores keystrokes when its
                                                    # widget tree never got input focus — so the
                                                    # Save button never fires. windowfocus uses
                                                    # XSetInputFocus directly (no _NET_ACTIVE_WINDOW
                                                    # required) and the followup --window keypress
                                                    # then lands on the focused button/entry.
                                                    def _send_enter(window_id):
                                                        try:
                                                            subprocess.run(['xdotool', 'windowfocus', '--sync', window_id],
                                                                           env=_xenv, timeout=3)
                                                        except Exception:
                                                            pass
                                                        time.sleep(0.4)  # let GTK settle on focus
                                                        subprocess.run(['xdotool', 'key', '--clearmodifiers',
                                                                        '--window', window_id, 'Return'],
                                                                       env=_xenv, timeout=3)

                                                    # Wait for the print dialog to render in Chrome.
                                                    time.sleep(2.5)
                                                    wins_before = _list_windows()
                                                    logger.info(f"  🪟 Windows before Save#1: {[f'{w}:{t[:50]}' for w,t in wins_before]}")

                                                    # First Save: directed at Chrome's main eCW window
                                                    # (print preview is rendered inside that window).
                                                    chrome_win = next((wid for wid, t in wins_before
                                                                       if 'eCW' in t or 'Chrome' in t), None)
                                                    if chrome_win:
                                                        _send_enter(chrome_win)
                                                        logger.info(f"  ⏎ Sent Enter to print preview (win {chrome_win})")
                                                    else:
                                                        subprocess.run(['xdotool', 'key', '--clearmodifiers', 'Return'],
                                                                       env=_xenv, timeout=3)

                                                    # Wait for the GTK file-chooser dialog to appear as
                                                    # a NEW window. Poll until we see one that wasn't
                                                    # there before.
                                                    file_dialog_win = None
                                                    seen_ids = {w for w, _ in wins_before}
                                                    for _w in range(15):  # up to ~7.5s
                                                        time.sleep(0.5)
                                                        wins_after = _list_windows()
                                                        for wid, title in wins_after:
                                                            if wid in seen_ids:
                                                                continue
                                                            # GTK file chooser titles often contain
                                                            # "Save" or the page title.
                                                            if 'Save' in title or 'save' in title or '.pdf' in title.lower():
                                                                file_dialog_win = wid
                                                                logger.info(f"  📁 Detected file dialog: {wid}:{title[:60]}")
                                                                break
                                                            # Sometimes the dialog has no title — just
                                                            # take the first new window.
                                                            if not file_dialog_win:
                                                                file_dialog_win = wid
                                                        if file_dialog_win:
                                                            break

                                                    if file_dialog_win:
                                                        # Give GTK a beat to finish rendering the
                                                        # dialog widgets before we send keys to it.
                                                        # Otherwise focus may not be on the filename
                                                        # entry yet (where Enter triggers Save).
                                                        time.sleep(0.6)
                                                        _send_enter(file_dialog_win)
                                                        logger.info(f"  ⏎ Sent Enter to file dialog (win {file_dialog_win})")
                                                    else:
                                                        logger.warning(f"  ⚠️ No file-chooser dialog appeared after Save#1 — sending blind Enter")
                                                        subprocess.run(['xdotool', 'key', '--clearmodifiers', 'Return'],
                                                                       env=_xenv, timeout=3)

                                                    # Poll ~/Downloads for the NEW PDF to appear.
                                                    new_path = None
                                                    for _w in range(40):  # up to ~20s
                                                        try:
                                                            now = {f for f in os.listdir(_dl_dir) if f.lower().endswith('.pdf')}
                                                            new_files = now - _before
                                                            if new_files:
                                                                # Pick the most recently modified.
                                                                candidates = [(os.path.getmtime(os.path.join(_dl_dir, f)), f)
                                                                              for f in new_files]
                                                                candidates.sort(reverse=True)
                                                                cand = os.path.join(_dl_dir, candidates[0][1])
                                                                # Ensure the file finished writing (size stable).
                                                                _s1 = os.path.getsize(cand)
                                                                time.sleep(0.5)
                                                                _s2 = os.path.getsize(cand)
                                                                if _s1 == _s2 and _s1 > 5000:
                                                                    new_path = cand
                                                                    break
                                                        except Exception:
                                                            pass
                                                        time.sleep(0.5)

                                                    if new_path:
                                                        # Move/copy the downloaded PDF to the slot the
                                                        # rest of the pipeline expects.
                                                        try:
                                                            import shutil as _sh
                                                            _sh.move(new_path, enc_file_path)
                                                        except Exception:
                                                            try:
                                                                with open(new_path, 'rb') as _src, open(enc_file_path, 'wb') as _dst:
                                                                    _dst.write(_src.read())
                                                                try: os.remove(new_path)
                                                                except Exception: pass
                                                            except Exception as _mv_err:
                                                                logger.warning(f"  Failed to move downloaded PDF: {_mv_err}")
                                                        _sz2 = os.path.getsize(enc_file_path) if os.path.exists(enc_file_path) else 0
                                                        if _sz2 > 5000:
                                                            this_captured = True
                                                            logger.info(f"  ✅ Strategy WEB/TEL (xdotool → ~/Downloads) captured ({_sz2} bytes)")
                                                    else:
                                                        logger.warning(f"  ⚠️ No new PDF appeared in {_dl_dir} after xdotool Save")
                                                        # Dismiss any stuck dialogs.
                                                        for _ in range(3):
                                                            subprocess.run(['xdotool', 'key', 'Escape'], env=_xenv, timeout=3)
                                                            time.sleep(0.3)
                                            except Exception as _xd_err:
                                                logger.warning(f"  xdotool save strategy failed: {_xd_err}")

                                        # Strategy A: extract #encounterPreviewContent HTML and render as PDF
                                        if not this_captured:
                                         for frm in page.frames:
                                            if this_captured:
                                                break
                                            try:
                                                preview_html = frm.evaluate('''() => {
                                                    const preview = document.getElementById('encounterPreviewContent');
                                                    if (preview && preview.innerHTML.length > 500) {
                                                        return preview.outerHTML;
                                                    }
                                                    return null;
                                                }''')
                                                if preview_html and len(preview_html) > 500:
                                                    full_html = (
                                                        '<!DOCTYPE html><html><head><meta charset="UTF-8"><style>'
                                                        'body { font-family: Verdana, Arial, Helvetica, sans-serif; font-size: 9pt; margin: 20px; }'
                                                        'table { border-collapse: collapse; width: 99%; }'
                                                        '.font1 { font-family: Verdana, Arial, Helvetica; font-size: 10.5pt; }'
                                                        '.table1 { font-family: Verdana, Arial, Helvetica; font-size: 9pt; width: 99%; }'
                                                        '.PNSig { font-family: Verdana; font-size: 9pt; font-weight: bold; }'
                                                        '.fixedTableLayout { table-layout: fixed; }'
                                                        '.TableFooter { font-family: Arial; font-size: 10pt; font-weight: bold; }'
                                                        'td { word-break: break-word; }'
                                                        '</style></head><body>' + preview_html + '</body></html>'
                                                    )
                                                    pdf_page = page.context.new_page()
                                                    pdf_page.set_content(full_html, wait_until='domcontentloaded')
                                                    time.sleep(2)
                                                    pdf_page.pdf(
                                                        path=enc_file_path,
                                                        format='Letter',
                                                        print_background=True,
                                                        margin={'top': '0.4in', 'bottom': '0.4in', 'left': '0.4in', 'right': '0.4in'}
                                                    )
                                                    pdf_page.close()
                                                    if os.path.exists(enc_file_path) and os.path.getsize(enc_file_path) > 10000:
                                                        this_captured = True
                                                        logger.info(f"  ✅ Strategy A captured ({os.path.getsize(enc_file_path)} bytes)")
                                            except Exception:
                                                pass

                                        # Strategy B: #printID button popup
                                        if not this_captured:
                                            # Reset popup_pages for this iteration
                                            del popup_pages[:]
                                            for frm in page.frames:
                                                try:
                                                    frm.evaluate('''() => { window.print = function() {}; }''')
                                                except Exception:
                                                    pass

                                            print_clicked = False
                                            for frm in page.frames:
                                                try:
                                                    result = frm.evaluate('''() => {
                                                        const btn = document.getElementById('printID');
                                                        if (btn && btn.offsetWidth > 0) { btn.click(); return 'printID'; }
                                                        const ngBtn = document.querySelector('button[ng-click*="printDialog"]');
                                                        if (ngBtn && ngBtn.offsetWidth > 0) { ngBtn.click(); return 'printDialog'; }
                                                        return null;
                                                    }''')
                                                    if result:
                                                        print_clicked = True
                                                        break
                                                except Exception:
                                                    continue

                                            if print_clicked:
                                                # Poll for the popup at 250ms intervals (was a flat
                                                # 5s + up to 6s of 2s ticks ≈ 11s worst case).
                                                # Most popups appear in <1s.
                                                for _wait in range(20):
                                                    if popup_pages:
                                                        break
                                                    time.sleep(0.25)
                                                if popup_pages:
                                                    for pp in popup_pages:
                                                        try:
                                                            pp.wait_for_load_state('domcontentloaded', timeout=10000)
                                                            # Short settle for any in-flight render
                                                            # (was 3s flat). domcontentloaded above
                                                            # already guarantees the DOM is parsed.
                                                            time.sleep(0.5)
                                                            cdp = page.context.new_cdp_session(pp)
                                                            pdf_result = cdp.send('Page.printToPDF', {
                                                                'landscape': False,
                                                                'displayHeaderFooter': False,
                                                                'printBackground': True,
                                                                'paperWidth': 8.5,
                                                                'paperHeight': 11,
                                                                'marginTop': 0.4,
                                                                'marginBottom': 0.4,
                                                                'marginLeft': 0.4,
                                                                'marginRight': 0.4,
                                                            })
                                                            pdf_data = base64.b64decode(pdf_result['data'])
                                                            with open(enc_file_path, 'wb') as f:
                                                                f.write(pdf_data)
                                                            cdp.detach()
                                                            if os.path.getsize(enc_file_path) > 10000:
                                                                this_captured = True
                                                                logger.info(f"  ✅ Strategy B captured ({os.path.getsize(enc_file_path)} bytes)")
                                                                break
                                                        except Exception as _pp_err:
                                                            logger.warning(f"  Strategy B popup PDF failed: {_pp_err}")
                                                for pp in popup_pages:
                                                    try:
                                                        pp.close()
                                                    except Exception:
                                                        pass
                                            try:
                                                subprocess.run(['xdotool', 'key', 'Escape'],
                                                              env={**os.environ, 'DISPLAY': ':99'}, timeout=5)
                                            except Exception:
                                                pass
                                            time.sleep(1)

                                        # Patient-mismatch defense for the encounter PDF:
                                        # a stale claim popup / wrong Encounters tab can yield a
                                        # PDF for the WRONG patient (e.g., Brewer's progress note
                                        # ending up on Revers' claim). Verify the captured PDF text
                                        # contains the expected surname before uploading to S3.
                                        # Fail-SAFE: if we can't verify for ANY reason (missing
                                        # patient_name, pdfplumber crash, image-only PDF) we DROP
                                        # the file. Better to leave the claim as "Needs Review"
                                        # than to silently attach the wrong patient's document.
                                        if this_captured and os.path.exists(enc_file_path):
                                            _verified = False
                                            _drop_reason = None
                                            _expected_surname = (claim_record.get('patient_name', '') or '').split(',')[0].strip().upper()
                                            if not _expected_surname:
                                                _drop_reason = 'no expected patient_name on record'
                                            else:
                                                try:
                                                    import pdfplumber as _pp_lib
                                                    with _pp_lib.open(enc_file_path) as _pdf_doc:
                                                        # Header-only: patient name lives at the
                                                        # top of the first page. Whole-document
                                                        # substring matches other people named in
                                                        # the note (paramedic, provider, etc.).
                                                        _enc_head = ((_pdf_doc.pages[0].extract_text() or '')
                                                                     if _pdf_doc.pages else '')[:500].upper()
                                                    if not _enc_head.strip():
                                                        _drop_reason = 'PDF text extraction returned empty (image-only or scanned PDF)'
                                                    elif _expected_surname not in _enc_head:
                                                        _drop_reason = f'PDF is for a DIFFERENT patient (expected {_expected_surname}, header={_enc_head[:120].strip()!r})'
                                                    else:
                                                        _verified = True
                                                        logger.info(f"  ✓ Encounter PDF #{enc_idx} patient verified: {_expected_surname}")
                                                except Exception as _vp:
                                                    _drop_reason = f'pdfplumber verification crashed: {_vp}'

                                            if not _verified:
                                                logger.warning(f"  ⛔ Encounter PDF #{enc_idx} for claim {claim_id} DROPPED — {_drop_reason}. NOT uploading.")
                                                this_captured = False
                                                try:
                                                    os.remove(enc_file_path)
                                                except Exception:
                                                    pass

                                        # Upload this PDF to S3 if captured (and patient verified)
                                        if this_captured and os.path.exists(enc_file_path):
                                            try:
                                                s3_key = f"encounter_files/{claim_id}_encounter_{enc_idx}.pdf"
                                                enc_s3_path = aws_client.upload_to_s3(enc_file_path, s3_key)
                                                captured_paths.append(enc_s3_path)
                                                captured_visit_types.append(visit_type)
                                                logger.info(f"  ✅ Uploaded to S3: {enc_s3_path}")
                                            except Exception as _up_err:
                                                logger.error(f"  ❌ S3 upload failed: {_up_err}")
                                        else:
                                            logger.warning(f"  ⚠️ Skipping {visit_type} (capture failed or patient mismatch)")

                                        # Close ONLY the encounter preview modal — never the claim popup.
                                        # Skip the fallback .close click that previously could dismiss
                                        # the underlying claim popup and break HCFA/Prog Notes capture.
                                        for frm in page.frames:
                                            try:
                                                frm.evaluate('''() => {
                                                    const CLAIM_MARKERS = ['Print HCFA', 'Prog. Notes', 'Prog Notes', 'view_hcfa', 'Adjustments'];
                                                    const closeBtn = document.querySelector('button[ng-click*="closeEncouterPreviewDialog"]');
                                                    if (closeBtn) { closeBtn.click(); return 'specific'; }
                                                    // Fallback: only click .close inside a modal that has the encounter preview content
                                                    const previewModals = document.querySelectorAll('.modal, [role="dialog"]');
                                                    for (const m of previewModals) {
                                                        const text = m.textContent || '';
                                                        // Identify the preview modal (has encounterPreviewContent or "Encounter" but NOT claim markers)
                                                        if (CLAIM_MARKERS.some(mk => text.includes(mk))) continue;
                                                        if (m.querySelector('#encounterPreviewContent') || text.includes('Vitals') || text.includes('Reason for Visit')) {
                                                            const cb = m.querySelector('.close, .modal-header .close');
                                                            if (cb && cb.offsetWidth > 0) { cb.click(); return 'preview-fallback'; }
                                                        }
                                                    }
                                                    return null;
                                                }''')
                                            except Exception:
                                                continue
                                        time.sleep(2)

                                    # Detach popup listener after the loop
                                    try:
                                        page.context.remove_listener('page', _on_popup)
                                    except Exception:
                                        pass

                                    # After all attempts: persist results to DDB.
                                    # encounter_date is now the actual NP/F-U visit date (the date of
                                    # the captured documentation), not the IV-Therapy intervention date.
                                    if captured_paths:
                                        logger.info(f"✅ Captured {len(captured_paths)}/{len(matches)} encounter PDFs ({captured_visit_types})")
                                        # encounter_date = the captured encounter's actual date
                                        # (matches the Progress Note PDF). Use the strategy that picked it.
                                        actual_enc_date = None
                                        if selected_date:
                                            try:
                                                mm, dd, yyyy = selected_date.split('/')
                                                actual_enc_date = f"{yyyy}-{mm}-{dd}"
                                            except Exception:
                                                pass
                                        update_data = {
                                            'encounter_file_s3_path': captured_paths[0],
                                            'encounter_files_s3_paths': captured_paths,
                                            'encounter_files_visit_types': captured_visit_types,
                                            'encounter_files_count': len(captured_paths),
                                            'encounter_file_captured_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                            'encounter_pick_strategy': selected_strategy or 'unknown',
                                            # Clear any stale "needs review" flags from prior runs
                                            'encounter_revision_needed': False,
                                            'encounter_date_mismatch': False,
                                            'encounter_date_source': 'captured_encounter',
                                        }
                                        if actual_enc_date:
                                            update_data['encounter_date'] = actual_enc_date
                                        aws_client.update_claim_status(claim_id, update_data)
                                    else:
                                        logger.warning(f"❌ Captured 0/{len(matches)} encounter PDFs")
                                        try:
                                            # Clear any stale wrong-patient encounter paths from
                                            # prior runs so submission readiness blocks this claim
                                            # until a valid encounter is captured.
                                            aws_client.update_claim_status(claim_id, {
                                                'encounter_capture_failed': 'pdf_capture_failed_after_click',
                                                'encounter_capture_failed_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                                'encounter_capture_attempted_count': len(matches),
                                                'encounter_revision_needed': True,
                                                'encounter_file_s3_path': '',
                                                'encounter_files_s3_paths': [],
                                            })
                                        except Exception as _flag_err:
                                            logger.warning(f"Failed to write encounter_capture_failed flag: {_flag_err}")
                        
                            # No `_close_all_hub_popups` here — no hubs are open after encounter
                            # file extraction; calling it risks closing the claim popup.

                        except Exception as enc_err:
                            logger.warning(f"Encounter file extraction failed: {enc_err}")
                            try:
                                aws_client.update_claim_status(claim_id, {
                                    'encounter_capture_failed': f'exception: {str(enc_err)[:200]}',
                                    'encounter_capture_failed_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                    'encounter_revision_needed': True,
                                })
                            except Exception as _flag_err:
                                logger.warning(f"Failed to write encounter_capture_failed flag: {_flag_err}")



                # On the metadata-only path the fresh-HCFA parse block below does not
                # run, so source the subscriber id from the already-stored HCFA box 1a
                # (reliable) instead of the flaky DOM scrape.
                _meta_only = not claim_record.get('needs_hcfa', False) and not claim_record.get('needs_prog_notes', False)
                if _meta_only:
                    try:
                        _hcfa_s3 = claim_record.get('hcfa_s3_path') or ''
                        if not _hcfa_s3:
                            _md = aws_client.dynamodb.Table('helixona-claims').get_item(Key={'claim_id': claim_id}).get('Item', {})
                            _hcfa_s3 = _md.get('hcfa_s3_path') or ''
                        if _hcfa_s3.startswith('s3://'):
                            _b, _k = _hcfa_s3[5:].split('/', 1)
                            _lp = f'/tmp/hcfa_meta_{claim_id}.pdf'
                            aws_client.s3.download_file(_b, _k, _lp)
                            _pdf_sub = _subscriber_id_from_hcfa_pdf(_lp)
                            if _pdf_sub:
                                if subscriber_id and subscriber_id.upper() != _pdf_sub:
                                    logger.warning(f"📋 Subscriber id corrected (metadata path) for {claim_id}: {subscriber_id!r} → {_pdf_sub!r}")
                                else:
                                    logger.info(f"📋 Subscriber id from stored HCFA box 1a for {claim_id}: {_pdf_sub}")
                                subscriber_id = _pdf_sub
                                subscriber_id_verified = True
                    except Exception as _msub_err:
                        logger.warning(f"Subscriber id from stored HCFA failed for {claim_id}: {_msub_err}")

                # If we only needed metadata (subscriber_id / encounter_date), save and skip
                needs_only_metadata = not claim_record.get('needs_hcfa', False) and not claim_record.get('needs_prog_notes', False)
                if needs_only_metadata and (subscriber_id or encounter_date):
                    try:
                        update_data = {}
                        if subscriber_id:
                            update_data['subscriber_id'] = subscriber_id
                            # gate flag: only a box-1a-confirmed value is trusted for auto-submit
                            update_data['subscriber_id_unverified'] = not subscriber_id_verified
                        if encounter_date:
                            update_data['encounter_date'] = encounter_date
                        aws_client.update_claim_status(claim_id, update_data)
                        logger.info(f"✅ Saved metadata for claim {claim_id}: {list(update_data.keys())}")
                        hcfa_success_count += 1
                    except Exception as e:
                        logger.error(f"Failed to save metadata: {e}")
                    # Encounter / Progress Note still needed — capture via stored Rx Start Date
                    try:
                        _rx_a = claim_record.get('iv_note_rx_start_date')
                        if not _rx_a:
                            try:
                                _dba = aws_client.dynamodb.Table('helixona-claims').get_item(Key={'claim_id': claim_id}).get('Item', {})
                                _rx_a = _dba.get('iv_note_rx_start_date')
                            except Exception:
                                _rx_a = None
                        _extract_encounter_file(_rx_to_mmddyyyy(_rx_a))
                    except Exception as _enc_call_err:
                        logger.warning(f"Encounter extraction (metadata-only) failed: {_enc_call_err}")
                    # Close popup and move to next claim. Use a short click timeout
                    # so we don't burn Playwright's 30s default waiting for a hidden/
                    # covered Cancel button — Escape fallback is fast and reliable.
                    try:
                        cancel_btn = hcfa_context.query_selector('button:has-text("Cancel"), input[value="Cancel"]')
                        if cancel_btn:
                            try:
                                cancel_btn.click(timeout=2000)
                            except Exception:
                                page.keyboard.press('Escape')
                        else:
                            page.keyboard.press('Escape')
                    except Exception:
                        try:
                            page.keyboard.press('Escape')
                        except Exception:
                            pass
                    time.sleep(0.3)
                    # Clear current_step so the dashboard's "live processing" badge
                    # doesn't linger on the last-touched claim after we move on.
                    try:
                        aws_client.update_claim_status(claim_id, {'current_step': ''})
                    except Exception:
                        pass
                    continue

                # Safety net: if the encounter-file flow inadvertently closed the claim popup,
                # re-open it via Claim Lookup so HCFA + Prog Notes capture has a real popup.
                try:
                    popup_visible = False
                    for f_ctx in [hcfa_context] + list(page.frames):
                        try:
                            visible = f_ctx.evaluate('''() => {
                                const btns = document.querySelectorAll('button');
                                let visProg = false, visCancel = false;
                                for (const btn of btns) {
                                    if (!(btn.offsetWidth > 0 && btn.offsetHeight > 0)) continue;
                                    const t = (btn.value || btn.textContent || '').trim();
                                    if ((t.includes('Prog') && t.includes('Note')) || t === 'Prog. Notes') visProg = true;
                                    if (t === 'Cancel') visCancel = true;
                                }
                                return visProg && visCancel;
                            }''')
                            if visible:
                                popup_visible = True
                                break
                        except Exception:
                            continue
                    if not popup_visible:
                        logger.info(f"🔄 Claim popup not visible — re-opening claim {claim_id} for HCFA + Prog Notes…")
                        reopened, refreshed_frame = _open_claim_popup_via_lookup(page, claim_id)
                        if reopened and refreshed_frame:
                            hcfa_context = refreshed_frame
                            logger.info(f"✅ Claim popup re-opened for HCFA section")
                        elif reopened:
                            logger.info(f"✅ Claim popup re-opened (main page context)")
                        else:
                            logger.warning(f"⚠️ Re-open before HCFA failed — capture likely to fail")
                except Exception as _re_err:
                    logger.warning(f"Popup visibility check failed: {_re_err}")

                # First: Find the exact "Print HCFA" button (NOT a container TD)
                # and dump its surrounding DOM for debugging
                try:
                    dom_info = hcfa_context.evaluate('''() => {
                        // ONLY match the actual button/input — NOT parent containers
                        // ECW typically uses <input type="button" value="Print HCFA (02-12)">
                        const inputs = document.querySelectorAll('input[type="button"], button');
                        let hcfaBtn = null;
                        
                        for (const el of inputs) {
                            const val = (el.value || '').trim();
                            const txt = (el.textContent || '').trim();
                            if (val.includes('HCFA') || txt.includes('HCFA')) {
                                hcfaBtn = el;
                                break;
                            }
                        }
                        
                        if (!hcfaBtn) {
                            // Fallback: look for any element whose own textContent is JUST "Print HCFA"
                            const spans = document.querySelectorAll('span, a, div');
                            for (const el of spans) {
                                if (el.children.length === 0 && (el.textContent || '').trim().includes('HCFA')) {
                                    hcfaBtn = el;
                                    break;
                                }
                            }
                        }
                        
                        if (!hcfaBtn) return { found: false, reason: 'No HCFA button/input found' };
                        
                        // Get the parent and dump its HTML for debugging
                        const parent = hcfaBtn.parentElement;
                        const parentParent = parent ? parent.parentElement : null;
                        
                        // Get elements near the HCFA button (siblings, etc.)
                        const siblings = parent ? Array.from(parent.children) : [];
                        const hcfaIdx = siblings.indexOf(hcfaBtn);
                        
                        const nearbyInfo = siblings.map((el, i) => ({
                            idx: i,
                            tag: el.tagName,
                            id: el.id || '',
                            cls: (el.className || '').toString().substring(0, 80),
                            val: (el.value || el.textContent || '').substring(0, 50),
                            onclick: el.getAttribute('onclick') || el.getAttribute('ng-click') || '',
                            src: el.getAttribute('src') || '',
                            w: el.offsetWidth,
                            h: el.offsetHeight,
                            type: el.type || ''
                        }));
                        
                        return {
                            found: true,
                            hcfa_tag: hcfaBtn.tagName,
                            hcfa_value: hcfaBtn.value || hcfaBtn.textContent || '',
                            hcfa_id: hcfaBtn.id || '',
                            hcfa_onclick: hcfaBtn.getAttribute('onclick') || hcfaBtn.getAttribute('ng-click') || '',
                            hcfa_idx: hcfaIdx,
                            parent_tag: parent?.tagName || '',
                            parent_id: parent?.id || '',
                            parent_html: (parent?.innerHTML || '').substring(0, 800),
                            pp_tag: parentParent?.tagName || '',
                            siblings: nearbyInfo
                        };
                    }''')
                    
                    if dom_info.get('found'):
                        logger.info(f"HCFA button found: <{dom_info.get('hcfa_tag')}> value='{dom_info.get('hcfa_value', '')[:60]}'")
                        logger.info(f"  id='{dom_info.get('hcfa_id', '')}' onclick='{dom_info.get('hcfa_onclick', '')[:120]}'")
                        logger.info(f"  Parent: <{dom_info.get('parent_tag', '')}> id='{dom_info.get('parent_id', '')}'")
                        logger.info(f"  HCFA index among siblings: {dom_info.get('hcfa_idx')}")
                        for sib in dom_info.get('siblings', []):
                            logger.info(f"  Sibling[{sib['idx']}]: <{sib['tag']}> cls='{sib.get('cls', '')[:40]}' val='{sib.get('val', '')[:30]}' onclick='{sib.get('onclick', '')[:60]}' src='{sib.get('src', '')[:50]}' w={sib['w']}")
                        logger.info(f"  Parent HTML: {dom_info.get('parent_html', '')[:300]}")
                    else:
                        logger.warning(f"HCFA button not found: {dom_info}")
                except Exception as e:
                    logger.warning(f"DOM inspection failed: {e}")

                # Strategy: Click the VISIBLE "Print HCFA (02-12)" button in the 
                # claim popup's bottom toolbar. NOT the hidden #claimLookupBtn22 which
                # opens the wrong "Batch Claims" dialog.
                # From user screenshot: bottom toolbar has:
                # Header | Data | Option ▲ | [doc icon] | [RED ICON] | Print HCFA (02-12) | ▲ | Adjustments | ...
                
                # First: dump the bottom toolbar to find all visible buttons
                try:
                    toolbar_info = hcfa_context.evaluate('''() => {
                        // Find ALL visible buttons/inputs in the page
                        const allBtns = document.querySelectorAll('input[type="button"], button, a');
                        const visibleBtns = [];
                        for (const btn of allBtns) {
                            if (btn.offsetWidth > 0 && btn.offsetHeight > 0) {
                                const val = (btn.value || btn.textContent || '').trim().substring(0, 40);
                                visibleBtns.push({
                                    tag: btn.tagName,
                                    id: btn.id || '',
                                    val: val,
                                    cls: (btn.className || '').toString().substring(0, 50),
                                    onclick: (btn.getAttribute('ng-click') || btn.getAttribute('onclick') || '').substring(0, 80),
                                    w: btn.offsetWidth,
                                    h: btn.offsetHeight,
                                    y: btn.getBoundingClientRect().top
                                });
                            }
                        }
                        // Sort by Y position (bottom ones first — the toolbar is at bottom)
                        visibleBtns.sort((a, b) => b.y - a.y);
                        return visibleBtns.slice(0, 20);
                    }''')
                    logger.info(f"Bottom toolbar buttons ({len(toolbar_info)} found):")
                    for btn in toolbar_info:
                        if 'hcfa' in btn.get('val', '').lower() or 'print' in btn.get('val', '').lower() or 'hcfa' in btn.get('onclick', '').lower():
                            logger.info(f"  🎯 <{btn['tag']}> id='{btn['id']}' val='{btn['val']}' onclick='{btn['onclick']}' w={btn['w']} y={btn['y']:.0f}")
                        elif btn['y'] > 500:  # Only log buttons near the bottom
                            logger.info(f"  <{btn['tag']}> val='{btn['val'][:25]}' onclick='{btn['onclick'][:40]}' y={btn['y']:.0f}")
                except Exception as e:
                    logger.warning(f"Toolbar inspection failed: {e}")

                # Click the RED ICON (billingClaimBtn33) which opens the HCFA PDF viewer
                # billingClaimBtn33 → view_hcfa('view',0,$event,'02-12') = PDF VIEWER popup
                # billingClaimBtn34 → view_hcfa('print',1,$event,'02-12') = PRINT dialog (wrong!)
                _set_step('capturing HCFA')
                needs_hcfa = claim_record.get('needs_hcfa', True)
                needs_prog_notes = claim_record.get('needs_prog_notes', True)
                hcfa_clicked = False
                hcfa_popup_url_from_pw = None
                hcfa_popup_page_pw = None
                if not needs_hcfa:
                    logger.info(f"⏭️ Skipping HCFA for claim {claim_id} (already captured) — {'Prog Notes only' if needs_prog_notes else 'validation only'}")
                    hcfa_clicked = True  # Skip all HCFA code paths

                # ── Priority 0: Real Playwright click on billingClaimBtn33 ──
                # view_hcfa builds a dynamic <form target="_blank"> and submits it. That submit
                # only works inside a TRUSTED user gesture — JS .click() via page.evaluate() is
                # NOT trusted, so the popup is silently dropped. Playwright's locator.click()
                # IS trusted, so the form actually submits and the popup opens.
                if needs_hcfa:
                    new_pages_during_click = []
                    def _on_hcfa_popup(p):
                        new_pages_during_click.append(p)
                        try:
                            p.evaluate('() => { window.print = () => {}; }')
                        except Exception:
                            pass
                    page.context.on('page', _on_hcfa_popup)
                    try:
                        btn33 = None
                        btn33_frame = None
                        for ctx in [hcfa_context, page] + list(page.frames):
                            try:
                                el = ctx.query_selector('#billingClaimBtn33')
                                if el:
                                    btn33 = el
                                    btn33_frame = ctx
                                    break
                            except Exception:
                                continue
                        if btn33 and btn33_frame:
                            logger.info("🎯 Found #billingClaimBtn33 — unhiding + clicking via Playwright (trusted user gesture)")
                            # Force-unhide the button: ng-hide class, parent display, etc.
                            # Then scrollIntoView so Playwright sees it as actionable.
                            try:
                                btn33_frame.evaluate('''() => {
                                    const btn = document.querySelector('#billingClaimBtn33');
                                    if (!btn) return false;
                                    // Climb 5 levels and unhide everything
                                    let n = btn;
                                    for (let i = 0; i < 6 && n; i++) {
                                        n.classList.remove('ng-hide');
                                        n.classList.remove('hidden');
                                        n.style.removeProperty('display');
                                        n.style.visibility = 'visible';
                                        n.style.opacity = '1';
                                        n.style.pointerEvents = 'auto';
                                        n = n.parentElement;
                                    }
                                    btn.scrollIntoView({block: 'center', inline: 'center'});
                                    return true;
                                }''')
                            except Exception:
                                pass
                            time.sleep(0.4)
                            # Re-query (DOM may have been touched by ng-hide removal)
                            try:
                                btn33 = btn33_frame.query_selector('#billingClaimBtn33') or btn33
                            except Exception:
                                pass
                            clicked = False
                            # Attempt 1: standard Playwright click (now visible)
                            try:
                                btn33.click(timeout=5000)
                                clicked = True
                                hcfa_clicked = True
                                logger.info("✅ Playwright click on btn33 succeeded after unhide")
                            except Exception as _click_err:
                                logger.warning(f"Playwright click after unhide failed: {_click_err}")
                            # Attempt 2: mouse.click at bounding box center (real input event)
                            if not clicked:
                                try:
                                    box = btn33.bounding_box()
                                    if box:
                                        cx = box['x'] + box['width'] / 2
                                        cy = box['y'] + box['height'] / 2
                                        page.mouse.click(cx, cy)
                                        clicked = True
                                        hcfa_clicked = True
                                        logger.info(f"✅ page.mouse.click on btn33 at ({cx:.0f},{cy:.0f}) succeeded")
                                except Exception as _mouse_err:
                                    logger.warning(f"page.mouse.click on btn33 failed: {_mouse_err}")
                            # Attempt 3: dispatch_event — last resort, no actionability checks
                            if not clicked:
                                try:
                                    btn33.dispatch_event('click')
                                    clicked = True
                                    hcfa_clicked = True
                                    logger.info("✅ dispatch_event('click') on btn33 succeeded")
                                except Exception as _disp_err:
                                    logger.warning(f"dispatch_event on btn33 failed: {_disp_err} — falling back to JS")
                            # Wait briefly for the new tab to spawn
                            for _w in range(8):
                                if new_pages_during_click:
                                    break
                                time.sleep(0.4)
                            for p in new_pages_during_click:
                                try:
                                    p.wait_for_load_state('domcontentloaded', timeout=10000)
                                except Exception:
                                    pass
                                p_url = p.url
                                if p_url and ('hcfa' in p_url.lower() or 'HCFAClaim' in p_url or '.pdf' in p_url.lower()):
                                    hcfa_popup_url_from_pw = p_url
                                    hcfa_popup_page_pw = p
                                    logger.info(f"✅ HCFA popup via Playwright click: {p_url[:160]}")
                                    break
                            if hcfa_clicked and not hcfa_popup_url_from_pw and new_pages_during_click:
                                # Take whatever popup we got — let downstream validate it
                                p = new_pages_during_click[0]
                                hcfa_popup_url_from_pw = p.url
                                hcfa_popup_page_pw = p
                                logger.info(f"ℹ️ Playwright captured popup (URL unclear): {p.url[:160]}")
                        else:
                            logger.info("ℹ️ #billingClaimBtn33 not in DOM for this claim — falling through to JS strategies")
                    finally:
                        try:
                            page.context.remove_listener('page', _on_hcfa_popup)
                        except Exception:
                            pass

                click_result = None
                try:
                  if needs_hcfa and not hcfa_popup_url_from_pw:
                    logger.info("Falling back to JS-based HCFA trigger...")
                    click_result = hcfa_context.evaluate('''() => {
                        // Priority 1: Intercept window.open + call scope.view_hcfa('view',…).
                        // The Angular function builds a PDF URL and calls window.open(url, '_blank').
                        // Programmatic calls aren't a user gesture so popup blockers can swallow
                        // them silently — but we just need the URL. Patch window.open to capture
                        // it, then return it for XHR download.
                        try {
                            if (window.angular) {
                                const candidates = [
                                    document.querySelector('#billingClaimBtn33'),
                                    document.querySelector('[ng-click*="view_hcfa"]'),
                                    document.querySelector('[ng-controller]'),
                                    document.body
                                ].filter(Boolean);
                                for (const el of candidates) {
                                    let scope = null;
                                    try { scope = angular.element(el).scope(); } catch(e) {}
                                    if (scope && typeof scope.view_hcfa === 'function') {
                                        const capturedUrls = [];
                                        const origOpen = window.open;
                                        window.open = function(url, name) {
                                            if (url) capturedUrls.push(url);
                                            return null;
                                        };
                                        try {
                                            scope.view_hcfa('view', 0, null, '02-12');
                                            try { scope.$apply(); } catch(e) {}
                                        } catch(e) {}
                                        window.open = origOpen;
                                        if (capturedUrls.length > 0) {
                                            return { ok: true, method: 'angular_scope_view_hcfa', pdf_url: capturedUrls[0] };
                                        }
                                        return { ok: true, method: 'angular_scope_view_hcfa', pdf_url: null };
                                    }
                                }
                            }
                        } catch(e) {}

                        // Priority 2: Angular scope openPrintHCFADialog(0) — legacy claims.
                        try {
                            if (window.angular) {
                                const el = document.querySelector('#claimLookupBtn22')
                                        || document.querySelector('[ng-click*="openPrintHCFADialog"]');
                                if (el) {
                                    let scope = null;
                                    try { scope = angular.element(el).scope(); } catch(e) {}
                                    if (scope && typeof scope.openPrintHCFADialog === 'function') {
                                        scope.openPrintHCFADialog(0);
                                        try { scope.$apply(); } catch(e) {}
                                        return { ok: true, method: 'angular_scope_openPrintHCFADialog' };
                                    }
                                }
                            }
                        } catch(e) {}

                        // Priority 3 (UI): #billingClaimBtn33 if visible
                        const viewBtn = document.querySelector('#billingClaimBtn33');
                        if (viewBtn && viewBtn.offsetWidth > 0 && viewBtn.offsetHeight > 0) {
                            viewBtn.click();
                            return { ok: true, id: 'billingClaimBtn33', method: 'ui_billingClaimBtn33' };
                        }

                        // Priority 4 (UI): any visible button with view_hcfa('view',...) onclick
                        const allBtns = document.querySelectorAll('button');
                        for (const btn of allBtns) {
                            if (btn.offsetWidth === 0) continue;
                            const ngClick = btn.getAttribute('ng-click') || '';
                            if (ngClick.includes("view_hcfa") && ngClick.includes("'view'")) {
                                btn.click();
                                return { ok: true, method: 'ui_ng-click-view', onclick: ngClick.substring(0, 100) };
                            }
                        }

                        // Priority 5 (UI): #claimLookupBtn22 if visible
                        const printHcfaBtn = document.querySelector('#claimLookupBtn22')
                            || document.querySelector('[ng-click*="openPrintHCFADialog"]');
                        if (printHcfaBtn && printHcfaBtn.offsetWidth > 0 && printHcfaBtn.offsetHeight > 0) {
                            printHcfaBtn.click();
                            return { ok: true, id: printHcfaBtn.id || 'openPrintHCFADialog', method: 'ui_claimLookupBtn22' };
                        }

                        // Last resort: any visible button containing "HCFA" in text
                        for (const btn of allBtns) {
                            if (btn.offsetWidth === 0) continue;
                            const text = (btn.value || btn.textContent || '').toLowerCase();
                            if (text.includes('hcfa')) {
                                btn.click();
                                return { ok: true, method: 'ui_text_match', text: text.substring(0, 50) };
                            }
                        }

                        return { ok: false, reason: 'No HCFA trigger found' };
                    }''')
                    
                    if isinstance(click_result, dict) and click_result.get('ok'):
                        logger.info(f"✅ Clicked HCFA trigger: {click_result}")
                        hcfa_clicked = True
                        
                        # If openPrintHCFADialog was used, handle the form version dialog
                        if click_result.get('method', '').startswith('angular_scope_openPrintHCFADialog') \
                           or click_result.get('method') == 'openPrintHCFADialog':
                            time.sleep(1.5)
                            try:
                                dialog_result = hcfa_context.evaluate('''() => {
                                    // Pick the 02-12 radio if available (matches view_hcfa('view',0,...,'02-12'))
                                    const radios = document.querySelectorAll('input[type="radio"]');
                                    let picked = false;
                                    for (const r of radios) {
                                        const label = (r.parentElement && r.parentElement.textContent || '') + ' ' + (r.value || '');
                                        if (label.includes('02-12') || label.includes('02/12')) {
                                            r.click();
                                            picked = true;
                                            break;
                                        }
                                    }
                                    if (!picked && radios.length > 0) radios[0].click();

                                    // Prefer the VIEW button (PDF viewer) over Print (print dialog).
                                    const dlgBtns = document.querySelectorAll('button, input[type="button"], input[type="submit"]');
                                    const preference = ['view', 'preview', 'ok', 'submit', 'print'];
                                    for (const pref of preference) {
                                        for (const btn of dlgBtns) {
                                            const text = (btn.value || btn.textContent || '').toLowerCase().trim();
                                            if (text === pref && (btn.offsetWidth > 0 || btn.offsetHeight > 0)) {
                                                btn.click();
                                                return { ok: true, text: text };
                                            }
                                        }
                                    }
                                    return { ok: false };
                                }''')
                                if dialog_result and dialog_result.get('ok'):
                                    logger.info(f"✅ HCFA dialog confirmed: {dialog_result}")
                                else:
                                    logger.info("ℹ️ No HCFA dialog found (may have opened directly)")
                            except Exception as e:
                                logger.warning(f"HCFA dialog handling: {e}")
                    else:
                        logger.warning(f"Red icon click result: {click_result}")
                except Exception as e:
                    logger.error(f"Red icon click failed: {e}")

                # Fallback: Use Playwright's text selector
                if not hcfa_clicked:
                    try:
                        logger.info("Fallback: Playwright text selector for Print HCFA...")
                        for btn_text in ['Print HCFA (02-12)', 'Print HCFA']:
                            hcfa_btn = page.locator(f'button:visible:has-text("{btn_text}")').first
                            if hcfa_btn.count() > 0:
                                hcfa_btn.click(timeout=5000)
                                hcfa_clicked = True
                                logger.info(f"✅ Clicked via Playwright locator: '{btn_text}'")
                                break
                    except Exception as e:
                        logger.warning(f"Playwright locator failed: {e}")


                if not hcfa_clicked:
                    logger.error(f"❌ Could not click Print HCFA for {claim_id}")
                    hcfa_fail_count += 1
                    try:
                        cancel_btn = hcfa_context.query_selector('button:has-text("Cancel"), input[value="Cancel"]')
                        if cancel_btn:
                            cancel_btn.click()
                            time.sleep(0.3)
                    except Exception:
                        pass
                    continue

                if not needs_hcfa:
                    pass  # Skip PDF loading — go directly to prog notes below
                elif hcfa_clicked:
                    # Now wait for the HCFA PDF to load in the in-page modal
                    # The "HCFA(02-12)" viewer is an overlay/modal within the same page,
                    # NOT a separate browser window. It embeds the PDF via <embed>, <iframe>,
                    # or <object> tag.
                    logger.info("Waiting for HCFA PDF viewer modal...")
                    time.sleep(1.5)  # give the PDF time to load
                
                hcfa_popup = None
                hcfa_pdf_captured = False
                pdf_url = None
                original_ref_no = None
                resubmission_code = None
                hcfa_cpt = None  # CPT/HCPCS parsed from the HCFA PDF (office-visit detection)
                # Start UNKNOWN — never lie with "First Time Submission" before we have evidence.
                # Promoted to "First Time Submission" or "Resubmission" by Box 22 parsing
                # (from PDF or DOM). If neither succeeds, stay "Unknown" so downstream / dashboard
                # can flag it for manual review.
                submission_type = 'Unknown'

                # Prefer Playwright-captured popup URL (Priority 0 — real user gesture)
                if needs_hcfa and hcfa_popup_url_from_pw:
                    pdf_url = hcfa_popup_url_from_pw
                    if hcfa_popup_page_pw is not None:
                        hcfa_popup = hcfa_popup_page_pw
                    logger.info(f"📥 HCFA URL from Playwright popup: {pdf_url[:150]}")

                # Otherwise, if the JS-fallback Angular scope window.open intercept captured one,
                # use that instead of scanning embeds/iframes.
                if not pdf_url and needs_hcfa and isinstance(click_result, dict) and click_result.get('pdf_url'):
                    pdf_url = click_result['pdf_url']
                    logger.info(f"📥 HCFA URL captured via window.open intercept: {pdf_url[:150]}")

                if not needs_hcfa:
                    pass  # Skip PDF capture entirely
                elif pdf_url:
                    # Already captured via window.open intercept — skip the embed/iframe scan
                    pass
                elif hcfa_clicked:
                  # Strategy 1: Search ALL frames for embedded PDF elements
                  for frame in [page] + page.frames:
                    try:
                        result = frame.evaluate('''() => {
                            // Check <embed> elements
                            const embeds = document.querySelectorAll('embed');
                            for (const e of embeds) {
                                const src = e.src || e.getAttribute('data') || '';
                                if (src && (src.includes('.pdf') || src.includes('hcfa') || src.includes('HCFA') || src.includes('HCFAClaim') || e.type === 'application/pdf')) {
                                    return { type: 'embed', url: src };
                                }
                            }
                            
                            // Check <iframe> elements for PDF  
                            const iframes = document.querySelectorAll('iframe');
                            for (const f of iframes) {
                                const src = f.src || '';
                                if (src && (src.includes('.pdf') || src.includes('hcfa') || src.includes('HCFA') || src.includes('HCFAClaim'))) {
                                    return { type: 'iframe', url: src };
                                }
                            }
                            
                            // Check <object> elements
                            const objects = document.querySelectorAll('object');
                            for (const o of objects) {
                                const data = o.data || o.getAttribute('data') || '';
                                if (data && (data.includes('.pdf') || data.includes('hcfa') || data.includes('HCFA'))) {
                                    return { type: 'object', url: data };
                                }
                            }
                            
                            // Check for ANY embed/iframe/object (might just be a PDF viewer)
                            if (embeds.length > 0) {
                                return { type: 'embed_any', url: embeds[0].src || '', count: embeds.length };
                            }
                            
                            return null;
                        }''')
                        if result and result.get('url'):
                            pdf_url = result['url']
                            logger.info(f"✅ Found PDF in {result['type']}: {pdf_url[:150]}")
                            hcfa_popup = frame  # use this frame as context for download
                            break
                        elif result:
                            logger.info(f"Found {result['type']} without URL: {result}")
                    except Exception:
                        continue
                
                # Strategy 2: Check for new browser windows (fallback)
                if not pdf_url and needs_hcfa:
                    try:
                        all_pages = page.context.pages
                        for p in all_pages:
                            if p == page:
                                continue
                            p_url = p.url
                            if 'hcfa' in p_url.lower() or 'HCFAClaim' in p_url or 'pdf' in p_url.lower():
                                hcfa_popup = p
                                pdf_url = p_url
                                logger.info(f"✅ HCFA popup window: {p_url[:120]}")
                                break
                            try:
                                title = p.title()
                                if 'HCFA' in title:
                                    hcfa_popup = p
                                    pdf_url = p_url
                                    logger.info(f"✅ HCFA popup by title: '{title}'")
                                    break
                            except Exception:
                                pass
                    except Exception:
                        pass
                
                # Strategy 3: Inspect all new frames for PDF-like URLs
                if not pdf_url and needs_hcfa:
                    for frame in page.frames:
                        try:
                            f_url = frame.url
                            if 'hcfa' in f_url.lower() or 'HCFAClaim' in f_url or f_url.endswith('.pdf'):
                                pdf_url = f_url
                                hcfa_popup = frame
                                logger.info(f"✅ PDF found in frame URL: {f_url[:120]}")
                                break
                        except Exception:
                            continue
                
                if pdf_url and needs_hcfa:
                    try:
                        # Validate the PDF URL matches the current claim
                        import re as _re
                        claimno_match = _re.search(r'claimno=(\d+)', pdf_url)
                        if claimno_match and claimno_match.group(1) != str(claim_id):
                            old_claimno = claimno_match.group(1)
                            pdf_url = pdf_url.replace(f'claimno={old_claimno}', f'claimno={claim_id}')
                            logger.warning(f"⚠️ Corrected stale iframe URL: claimno {old_claimno} → {claim_id}")
                        
                        logger.info(f"PDF URL for claim {claim_id}: {pdf_url[:150]}")
                        
                        # Screenshot the current state
                        try:
                            page.screenshot(path=f'/tmp/hcfa_viewer_{claim_id}.png')
                            logger.info(f"📸 HCFA viewer screenshot saved for {claim_id}")
                        except Exception:
                            pass

                        # ── Download the actual HCFA PDF ──
                        # Method A: Use XMLHttpRequest from the claim detail iframe
                        # (shares session cookies + CSRF context — avoids 412 errors)
                        try:
                            logger.info(f"Downloading PDF via iframe XHR: {pdf_url[:100]}")
                            pdf_b64_result = hcfa_context.evaluate('''(url) => {
                                try {
                                    const xhr = new XMLHttpRequest();
                                    xhr.open('GET', url, false);
                                    xhr.responseType = 'arraybuffer';
                                    xhr.send();
                                    if (xhr.status === 200) {
                                        const bytes = new Uint8Array(xhr.response);
                                        let binary = '';
                                        const chunk = 8192;
                                        for (let i = 0; i < bytes.length; i += chunk) {
                                            binary += String.fromCharCode.apply(null, bytes.subarray(i, Math.min(i + chunk, bytes.length)));
                                        }
                                        return { ok: true, data: btoa(binary), size: bytes.length, type: xhr.getResponseHeader('Content-Type') };
                                    }
                                    return { ok: false, status: xhr.status, text: xhr.statusText };
                                } catch(e) {
                                    return { ok: false, error: e.message };
                                }
                            }''', pdf_url)
                            
                            if pdf_b64_result and pdf_b64_result.get('ok'):
                                import base64
                                pdf_bytes = base64.b64decode(pdf_b64_result['data'])
                                logger.info(f"XHR response: size={len(pdf_bytes)}, type={pdf_b64_result.get('type','?')}, starts_with={pdf_bytes[:10]}")
                                if pdf_bytes[:5] == b'%PDF-':
                                    with open(download_path, 'wb') as f:
                                        f.write(pdf_bytes)
                                    hcfa_pdf_captured = True
                                    logger.info(f"✅ HCFA PDF via XHR: {download_path} ({len(pdf_bytes)} bytes)")
                                else:
                                    logger.warning(f"XHR response is not PDF (first bytes: {pdf_bytes[:20]})")
                            else:
                                logger.warning(f"XHR result: {pdf_b64_result}")
                        except Exception as e:
                            logger.warning(f"XHR download failed: {e}")
                        
                        # Method B: Use requests with browser cookies + Referer
                        if not hcfa_pdf_captured:
                            try:
                                import requests as _requests
                                cookies = page.context.cookies()
                                cookie_dict = {c['name']: c['value'] for c in cookies}
                                headers = {
                                    'Referer': page.url,
                                    'Accept': 'application/pdf,*/*',
                                    'User-Agent': page.evaluate('() => navigator.userAgent'),
                                }
                                logger.info(f"Trying requests with {len(cookies)} cookies + Referer...")
                                resp = _requests.get(pdf_url, cookies=cookie_dict, headers=headers, timeout=15, verify=False)
                                logger.info(f"Requests: status={resp.status_code}, type={resp.headers.get('content-type','')}, size={len(resp.content)}")
                                if resp.content[:5] == b'%PDF-':
                                    with open(download_path, 'wb') as f:
                                        f.write(resp.content)
                                    hcfa_pdf_captured = True
                                    logger.info(f"✅ HCFA PDF via requests: {download_path} ({len(resp.content)} bytes)")
                            except Exception as e:
                                logger.warning(f"Requests download failed: {e}")
                        
                        # Method C: DISABLED — page.pdf() renders HTML without the CMS-1500 form 
                        # background image, producing damaged-looking PDFs. Better to leave as
                        # "pending" and retry later than to upload a broken HCFA.
                        if not hcfa_pdf_captured:
                            logger.warning(f"⚠️ Could not download real HCFA PDF for {claim_id} (XHR + requests both failed)")

                        # ── Extract Box 22 (RESUBMISSION CODE + ORIGINAL REF. NO.) from the PDF ──
                        # Box 22 on CMS-1500: left side = resubmission code (e.g. "7")
                        #                     right side = original ref. no. (e.g. "202512302615")
                        # If Box 22 is EMPTY → First Time Submission
                        # If Box 22 has values → Resubmission (ref is the BlueShield claim ID)
                        # NOTE: don't reset submission_type here — it's already 'Unknown' by default,
                        # and any DOM-based detection above this block has already promoted it.

                        if hcfa_pdf_captured and os.path.exists(download_path):
                            try:
                                import pdfplumber
                                logger.info(f"Parsing Box 22 from {download_path}...")
                                with pdfplumber.open(download_path) as pdf_doc:
                                    for pg_num, pg in enumerate(pdf_doc.pages):
                                        words = pg.extract_words()
                                        
                                        # Calibrated from actual HCFA PDF analysis:
                                        # Box 22 RESUBMISSION CODE: x0≈370-410, top≈460-475
                                        # Box 22 ORIGINAL REF. NO.: x0≈440-550, top≈460-475
                                        # Provider NPI (Box 24J):   x0≈500-575, top≈530-660 (NOT box 22!)
                                        
                                        for w in words:
                                            x0 = float(w.get('x0', 0))
                                            top = float(w.get('top', 0))
                                            val = w['text'].strip()
                                            
                                            # Box 22 row: top ≈ 460-480
                                            if 458 <= top <= 480:
                                                # Resubmission code (left): x0 ≈ 370-420
                                                if 365 <= x0 <= 425 and val.isdigit() and len(val) <= 2:
                                                    resubmission_code = val
                                                    logger.info(f"Box 22 RESUBMISSION CODE: {val} (x0={x0:.0f}, top={top:.0f})")
                                                
                                                # Original Ref. No. (right): x0 ≈ 440-555
                                                if 435 <= x0 <= 555 and len(val) >= 8 and val.isdigit():
                                                    original_ref_no = val
                                                    logger.info(f"Box 22 ORIGINAL REF. NO.: {val} (x0={x0:.0f}, top={top:.0f})")
                                        
                                        # Strategy 2: Text-based fallback
                                        # Look for the ICD line with resubmission code + ref
                                        # e.g. "G939 E8840 K9089 G629 7 260673998400"
                                        if not original_ref_no:
                                            import re as _re
                                            text = pg.extract_text() or ''
                                            for line in text.split('\n'):
                                                # Pattern: ICD codes followed by single digit + long number
                                                m = _re.search(r'[A-Z]\d{3,4}\b.*?\b(\d{1,2})\s+(\d{10,15})$', line)
                                                if m:
                                                    code_val = m.group(1)
                                                    ref_val = m.group(2)
                                                    # Exclude provider NPI (appears on service lines)
                                                    # Service lines have format: dates + CPT + amounts + NPI
                                                    # Box 22 line has ICD codes + code + ref
                                                    if ref_val not in ('1407241524',) and not line.strip().startswith(('01 ', '02 ', '03 ', '04 ', '05 ', '06 ', '07 ', '08 ', '09 ', '10 ', '11 ', '12 ')):
                                                        resubmission_code = code_val
                                                        original_ref_no = ref_val
                                                        logger.info(f"Box 22 via text: code={code_val}, ref={ref_val}")
                                                        break
                                        
                                        if original_ref_no:
                                            break  # found it

                                # CPT/HCPCS from the HCFA (Box 24D service lines) — used to
                                # detect office-visit claims (E/M 99201-99205 / 99211-99215),
                                # which need only HCFA + IV Note (no Progress Note).
                                try:
                                    hcfa_cpt = _cpt_from_hcfa_pdf(download_path)
                                    logger.info(f"📋 CPT/HCPCS from HCFA for claim {claim_id}: {hcfa_cpt or '(none found)'}")
                                except Exception as _cpt_err:
                                    logger.warning(f"CPT parse from HCFA failed for {claim_id}: {_cpt_err}")

                                # ── Subscriber id from HCFA box 1a (RELIABLE source) ──
                                # The on-screen grid scrape (subscriber_id, set above from
                                # the popup) reads stale popups / the wrong insurance row and
                                # is non-deterministic. ECW's own box 1a is the source of
                                # truth, so prefer it whenever the PDF yields a value.
                                try:
                                    _pdf_sub = _subscriber_id_from_hcfa_pdf(download_path)
                                    if _pdf_sub:
                                        if subscriber_id and subscriber_id.upper() != _pdf_sub:
                                            logger.warning(f"📋 Subscriber id corrected for {claim_id}: DOM scrape={subscriber_id!r} → HCFA box 1a={_pdf_sub!r}")
                                        else:
                                            logger.info(f"📋 Subscriber id from HCFA box 1a for {claim_id}: {_pdf_sub}")
                                        subscriber_id = _pdf_sub
                                        subscriber_id_verified = True
                                    elif subscriber_id:
                                        logger.warning(f"📋 HCFA box 1a yielded no subscriber id for {claim_id}; keeping DOM scrape {subscriber_id!r} (UNVERIFIED — gated from auto-submit)")
                                    else:
                                        logger.warning(f"📋 No subscriber id from HCFA box 1a or DOM for {claim_id}")
                                except Exception as _sub_err:
                                    logger.warning(f"Subscriber id parse from HCFA failed for {claim_id}: {_sub_err}")

                                if original_ref_no:
                                    submission_type = "Resubmission"
                                    logger.info(f"📋 Box 22: code={resubmission_code}, ref={original_ref_no} → Resubmission")
                                elif resubmission_code and resubmission_code not in ('0', '1'):
                                    submission_type = "Resubmission"
                                    logger.info(f"📋 Box 22: code={resubmission_code} (no ref) → Resubmission")
                                else:
                                    # Box 22 is empty in a successfully captured PDF → genuinely first-time
                                    submission_type = "First Time Submission"
                                    logger.info(f"📋 Box 22 empty in PDF → First Time Submission")
                                    
                            except ImportError:
                                logger.warning("pdfplumber not installed — skipping Box 22 extraction")
                            except Exception as e:
                                logger.warning(f"Box 22 extraction failed: {e}")
                            
                    except Exception as e:
                        logger.error(f"HCFA PDF capture failed: {e}")
                elif needs_hcfa:
                    logger.warning(f"⚠️ No HCFA PDF URL found for {claim_id}")
                    try:
                        page.screenshot(path=f'/tmp/hcfa_after_click_{claim_id}.png')
                        logger.info(f"📸 Post-click screenshot saved for {claim_id}")
                    except Exception:
                        pass
                
                # ── DOM-based Box 22 fallback when PDF wasn't captured ──
                if needs_hcfa and not original_ref_no:
                    diagnostics = []
                    try:
                        for f_ctx in [hcfa_context] + list(page.frames):
                            try:
                                box22_data = f_ctx.evaluate('''() => {
                                    const KEYWORDS_CODE = ['resubmission', 'frequencycode', 'freqcode', 'frequency', 'box22', 'box-22', 'resubcode', 'resubmissioncode', 'claimfreq'];
                                    const KEYWORDS_REF  = ['originalref', 'original_ref', 'origref', 'originalrefno', 'originalpayerid', 'payerclaimid', 'originalclaimid', 'priorpayerid', 'box22ref'];
                                    let resubCode = null, origRef = null;
                                    const inspected = [];

                                    function keyMatches(s, list) {
                                        const lower = (s || '').toLowerCase().replace(/[\s_-]/g, '');
                                        return list.some(k => lower.includes(k.replace(/[\s_-]/g, '')));
                                    }

                                    // Strategy 1: scan every input/textarea/select for id/name/ng-model keyword hits
                                    const fields = document.querySelectorAll('input, textarea, select');
                                    for (const el of fields) {
                                        const val = (el.value || '').trim();
                                        const id = el.id || '';
                                        const name = el.name || '';
                                        const ngModel = el.getAttribute('ng-model') || '';
                                        const placeholder = el.placeholder || '';
                                        const sig = id + '|' + name + '|' + ngModel + '|' + placeholder;
                                        if (val && (keyMatches(sig, KEYWORDS_CODE) || keyMatches(sig, KEYWORDS_REF))) {
                                            inspected.push({ sig: sig.substring(0, 80), val: val.substring(0, 30) });
                                        }
                                        if (val && keyMatches(sig, KEYWORDS_CODE) && val !== '0') resubCode = resubCode || val;
                                        if (val && keyMatches(sig, KEYWORDS_REF) && val.length >= 6) origRef = origRef || val;
                                    }

                                    // Strategy 2: walk every <label>/<th>/<td> looking for "Resubmission Code" /
                                    // "Original Ref" labels, then read the adjacent input/cell value.
                                    function readAdjacent(labelEl) {
                                        // Sibling input
                                        let n = labelEl.nextElementSibling;
                                        while (n) {
                                            if (n.tagName === 'INPUT' || n.tagName === 'TEXTAREA' || n.tagName === 'SELECT') {
                                                if (n.value) return n.value.trim();
                                            }
                                            const inp = n.querySelector && n.querySelector('input, textarea, select');
                                            if (inp && inp.value) return inp.value.trim();
                                            if (n.textContent && n.textContent.trim() && n !== labelEl) {
                                                return n.textContent.trim();
                                            }
                                            n = n.nextElementSibling;
                                        }
                                        // Parent/next-row pattern
                                        const parent = labelEl.parentElement;
                                        if (parent) {
                                            const inp = parent.querySelector('input, textarea, select');
                                            if (inp && inp !== labelEl && inp.value) return inp.value.trim();
                                        }
                                        return null;
                                    }

                                    const labels = document.querySelectorAll('label, th, td, div, span, b');
                                    for (const lbl of labels) {
                                        const t = (lbl.textContent || '').trim().toLowerCase();
                                        if (!t || t.length > 80) continue;
                                        if ((t.includes('original') && t.includes('ref')) ||
                                            t === 'original ref. no.' || t === 'original ref no' ||
                                            t.includes('payer claim id') || t.includes('original claim')) {
                                            const v = readAdjacent(lbl);
                                            if (v && /^\d{6,}$/.test(v)) {
                                                origRef = origRef || v;
                                                inspected.push({ sig: 'label:' + t.substring(0, 40), val: v });
                                            }
                                        }
                                        if (t.includes('resubmission') ||
                                            (t.includes('frequency') && t.includes('code'))) {
                                            const v = readAdjacent(lbl);
                                            if (v && v !== '0' && v !== '1' && /^\d{1,2}$/.test(v)) {
                                                resubCode = resubCode || v;
                                                inspected.push({ sig: 'label:' + t.substring(0, 40), val: v });
                                            }
                                        }
                                    }

                                    // Strategy 3: Angular scope FormData / claimData
                                    if (!origRef || !resubCode) {
                                        try {
                                            const ctrls = document.querySelectorAll('[ng-controller]');
                                            for (const c of ctrls) {
                                                let scope = null;
                                                try { scope = angular.element(c).scope(); } catch(e) {}
                                                if (!scope) continue;
                                                const fd = scope.FormData || scope.formData || scope.claimData ||
                                                           scope.Claim || scope.claim || {};
                                                const refCands = [fd.OriginalRefNo, fd.originalRefNo, fd.original_ref_no,
                                                                  fd.PayerClaimId, fd.payerClaimId, fd.OriginalPayerID,
                                                                  fd.OriginalClaimId, fd.originalClaimId];
                                                const codeCands = [fd.FrequencyCode, fd.frequencyCode, fd.ResubmissionCode,
                                                                   fd.resubmissionCode, fd.ClaimFrequencyCode];
                                                for (const r of refCands) {
                                                    if (r && r.toString().length >= 6) { origRef = origRef || r.toString(); break; }
                                                }
                                                for (const c2 of codeCands) {
                                                    if (c2 && c2 !== '0' && c2 !== 0 && c2 !== '1' && c2 !== 1) {
                                                        resubCode = resubCode || c2.toString(); break;
                                                    }
                                                }
                                                if (origRef && resubCode) break;
                                            }
                                        } catch(e) {}
                                    }

                                    return { resubCode, origRef, inspected: inspected.slice(0, 20) };
                                }''')

                                if box22_data:
                                    if box22_data.get('inspected'):
                                        diagnostics.extend(box22_data['inspected'])
                                    if box22_data.get('origRef'):
                                        original_ref_no = box22_data['origRef']
                                        resubmission_code = box22_data.get('resubCode') or resubmission_code
                                        submission_type = "Resubmission"
                                        logger.info(f"📋 Box 22 from DOM: code={resubmission_code}, ref={original_ref_no} → Resubmission")
                                        break
                                    elif box22_data.get('resubCode') and box22_data['resubCode'] not in ('0', '1'):
                                        resubmission_code = box22_data['resubCode']
                                        submission_type = "Resubmission"
                                        logger.info(f"📋 Box 22 from DOM: resubmission code={resubmission_code} (no ref) → Resubmission")
                                        break
                            except Exception:
                                continue
                        if submission_type == 'Unknown' and not original_ref_no:
                            if diagnostics:
                                logger.warning(f"📋 Box 22 DOM scan found candidates but none matched: {diagnostics[:5]}")
                            else:
                                logger.warning(f"📋 Box 22 DOM scan found nothing — submission_type left as Unknown")
                    except Exception as e:
                        logger.warning(f"DOM Box 22 extraction failed: {e}")
                
                # ── 6f. Upload HCFA PDF to S3 and update DynamoDB ──
                if hcfa_pdf_captured and os.path.exists(download_path):
                    try:
                        s3_key = f"hcfa_forms/{claim_id}_hcfa.pdf"
                        s3_path = aws_client.upload_to_s3(download_path, s3_key)
                        logger.info(f"✅ HCFA PDF uploaded to S3: {s3_path}")

                        update_data = {
                            'state': 2,
                            'hcfa_s3_path': s3_path,
                            'hcfa_pdf_size': os.path.getsize(download_path),
                            'hcfa_generated_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                            'submission_type': submission_type,
                        }
                        if original_ref_no:
                            update_data['original_ref_no'] = original_ref_no
                        if resubmission_code:
                            update_data['resubmission_code'] = resubmission_code
                        if hcfa_cpt:
                            update_data['cpt'] = hcfa_cpt
                            claim_record['cpt'] = hcfa_cpt  # so _extract_encounter_file sees it this run
                        if subscriber_id:
                            update_data['subscriber_id'] = subscriber_id
                            # gate flag: only a box-1a-confirmed value is trusted for auto-submit
                            update_data['subscriber_id_unverified'] = not subscriber_id_verified
                        if encounter_date:
                            update_data['encounter_date'] = encounter_date
                        aws_client.update_claim_status(claim_id, update_data)
                        logger.info(f"✅ Claim {claim_id} updated to state 2 with HCFA PDF in S3")
                        hcfa_success_count += 1
                    except Exception as e:
                        logger.error(f"S3 upload / DynamoDB update failed for {claim_id}: {e}")
                        hcfa_fail_count += 1
                elif needs_hcfa:
                    # HCFA was triggered but PDF not captured
                    logger.warning(f"⚠️ HCFA triggered for {claim_id} but PDF not captured")
                    try:
                        aws_client.update_claim_status(claim_id, {
                            'state': 2,
                            'hcfa_triggered': True,
                            'hcfa_pdf_captured': False,
                            'hcfa_generated_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                            'submission_type': submission_type,
                        })
                        hcfa_success_count += 1
                        logger.info(f"✅ Claim {claim_id} updated to state 2 (HCFA triggered, PDF pending)")
                    except Exception as e:
                        logger.error(f"DynamoDB update failed for {claim_id}: {e}")
                        hcfa_fail_count += 1

                if needs_hcfa:
                  # ── 6h. Close ONLY the HCFA viewer overlay — keep the claim detail popup open
                  # so Prog Notes / Encounter File capture can still find their buttons.
                  try:
                    hcfa_closed = page.evaluate('''() => {
                        const CLAIM_MARKERS = ['Print HCFA', 'Prog. Notes', 'Prog Notes', 'view_hcfa', 'Adjustments'];
                        // Find a close button whose containing modal is the HCFA viewer
                        // (contains <embed>/<iframe> for the PDF, or has 'HCFA' in title).
                        const closeBtns = document.querySelectorAll('.close, [class*="close"], button[ng-click*="close"]');
                        for (const btn of closeBtns) {
                            if (!(btn.offsetWidth > 0 && btn.offsetHeight > 0)) continue;
                            const rect = btn.getBoundingClientRect();
                            if (rect.top >= 200) continue;
                            const modal = btn.closest('.modal, [role="dialog"], [class*="popup"], [class*="Popup"]')
                                          || btn.parentElement;
                            if (!modal) continue;
                            const modalText = modal.textContent || '';
                            // Skip claim detail popup
                            if (CLAIM_MARKERS.some(m => modalText.includes(m))) continue;
                            // Prefer modals that look like an HCFA viewer (PDF embed or "HCFA" in header)
                            const hasPdf = modal.querySelector('embed, iframe[src*="pdf" i], iframe[src*="HCFA" i], object[data*="pdf" i]');
                            const looksLikeHcfa = hasPdf || /HCFA|CMS\s*1500/i.test(modalText.substring(0, 200));
                            if (looksLikeHcfa) {
                                btn.click();
                                return 'closed-hcfa-modal';
                            }
                        }
                        return 'no-hcfa-modal';
                    }''')
                    if hcfa_closed == 'closed-hcfa-modal':
                        logger.info("✅ Closed HCFA modal via X button")
                    else:
                        logger.info("ℹ️ No HCFA viewer modal to close (claim popup left intact)")
                    time.sleep(0.3)
                  except Exception:
                    time.sleep(0.3)
                
                # ── Office-visit determination FIRST (from CPT, parsed off the HCFA PDF) ──
                # If this is an office visit (E/M 99201-99205 / 99211-99215) it needs only
                # HCFA + IV Note — NO Progress Note (encounter) and NO Rx Start Date.
                # Decide it here, before the IV Note / Rx logic, and persist the flag.
                # If CPT is unknown (HCFA capture skipped because already generated on a
                # prior run → hcfa_cpt never parsed), backfill it: use the DDB `cpt` if a
                # newer run stored it, else download the HCFA PDF from S3 and parse it.
                if not (claim_record.get('cpt') or hcfa_cpt):
                    try:
                        _ddb_item = aws_client.dynamodb.Table('helixona-claims').get_item(
                            Key={'claim_id': claim_id}).get('Item', {})
                        if _ddb_item.get('cpt'):
                            hcfa_cpt = _ddb_item.get('cpt')
                        else:
                            _hcfa_s3 = _ddb_item.get('hcfa_s3_path') or ''
                            import re as _re_s3
                            _m = _re_s3.match(r's3://([^/]+)/(.+)', _hcfa_s3)
                            if _m:
                                _lp = f'/tmp/{claim_id}_hcfa_cptscan.pdf'
                                aws_client.s3.download_file(_m.group(1), _m.group(2), _lp)
                                hcfa_cpt = _cpt_from_hcfa_pdf(_lp)
                                try:
                                    os.remove(_lp)
                                except Exception:
                                    pass
                                logger.info(f"📋 CPT backfilled from S3 HCFA for claim {claim_id}: {hcfa_cpt or '(none found)'}")
                                if hcfa_cpt:
                                    try:
                                        aws_client.update_claim_status(claim_id, {'cpt': hcfa_cpt})
                                    except Exception:
                                        pass
                    except Exception as _bf_err:
                        logger.warning(f"CPT backfill from S3 HCFA failed for {claim_id}: {_bf_err}")
                if hcfa_cpt and not claim_record.get('cpt'):
                    claim_record['cpt'] = hcfa_cpt
                is_office_visit_claim = is_office_visit(claim_record.get('cpt') or hcfa_cpt)
                if is_office_visit_claim:
                    logger.info(f"🏥 Claim {claim_id} is an OFFICE VISIT (CPT {claim_record.get('cpt') or hcfa_cpt}) — IV Note + HCFA only; skipping Rx Start Date / Progress Note")
                    claim_record['office_visit'] = True
                    try:
                        aws_client.update_claim_status(claim_id, {
                            'office_visit': True,
                            'progress_note_required': False,
                            'encounter_revision_needed': False,
                        })
                    except Exception:
                        pass

                _set_step('capturing IV Note')
                # ── Fase 5: Capture Progress Notes PDF (the "IV Note") ──
                prog_notes_captured = False
                prog_notes_path = f"/tmp/prog_notes_{claim_id}.pdf"
                # Defensive: delete any stale /tmp file from a previous run of
                # this same claim. Without this, the "PDF captured" check at
                # the end of the iframe-extract path could mark prior-run
                # wrong-patient content as a fresh capture and upload it.
                try:
                    if os.path.exists(prog_notes_path):
                        os.remove(prog_notes_path)
                except Exception:
                    pass
                iv_note_rx_start_date = None        # parsed from "Rx Start Date" in HTML
                iv_note_patient_mismatch = False    # set True if captured note is for the wrong patient


                if not needs_prog_notes:
                    logger.info(f"⏭️ Skipping Prog Notes for claim {claim_id} (already captured)")

                if needs_prog_notes:
                  # ALWAYS close+reopen the claim popup before Prog. Notes — even if it
                  # looks alive. ECW's Angular $scope on a reused popup can serve the
                  # PREVIOUS claim's progress note (the "REVERS, Felicia" cascade — 195
                  # claims contaminated in a single run). A fresh popup forces ECW to
                  # re-instantiate the controller with this claim's context.
                  try:
                      # Step 1: close any existing claim popup so the lookup opens a
                      # brand-new one. Try Cancel button first, Escape as fallback.
                      try:
                          for _cctx in [hcfa_context] + list(page.frames):
                              try:
                                  did_close = _cctx.evaluate('''() => {
                                      const btns = document.querySelectorAll('button, input[type="button"]');
                                      for (const btn of btns) {
                                          if (!(btn.offsetWidth > 0 && btn.offsetHeight > 0)) continue;
                                          const t = (btn.value || btn.textContent || '').trim();
                                          if (t === 'Cancel') { btn.click(); return true; }
                                      }
                                      return false;
                                  }''')
                                  if did_close:
                                      break
                              except Exception:
                                  continue
                          time.sleep(0.5)
                          try:
                              page.keyboard.press('Escape')
                          except Exception:
                              pass
                          time.sleep(0.3)
                      except Exception:
                          pass

                      # Step 2: reopen the claim popup via lookup → fresh Angular scope.
                      logger.info(f"🔄 Forcing fresh claim popup before Prog Notes for {claim_id} (anti-stale-state)…")
                      reopened, refreshed_frame = _open_claim_popup_via_lookup(page, claim_id)
                      if reopened and refreshed_frame:
                          hcfa_context = refreshed_frame
                          logger.info(f"✅ Claim popup re-opened for Prog Notes (fresh scope)")
                      else:
                          logger.warning(f"⚠️ Could not re-open claim popup for {claim_id} before Prog Notes — falling back to existing context")
                  except Exception as _pn_re_err:
                      logger.warning(f"Prog Notes popup refresh failed: {_pn_re_err}")

                if needs_prog_notes:
                  try:
                    # ── Reset ProgNoteViwerFrame BEFORE clicking Prog. Notes ──
                    # If the trusted click fails for any reason (ng-click not fired),
                    # the iframe would otherwise still hold the PREVIOUS claim's
                    # progress note → every subsequent claim's capture would be the
                    # same wrong-patient content. Clearing here means: either the
                    # new click repopulates it, or extraction fails cleanly via
                    # the mismatch guard. Logs showed 195 mismatches per run all
                    # against the same "REVERS, Felicia" — that's the cascade.
                    for _rctx in [hcfa_context] + list(page.frames):
                        try:
                            _rctx.evaluate('''() => {
                                const ifs = document.querySelectorAll('iframe[name="ProgNoteViwerFrame"], iframe[id*="ProgNote"]');
                                for (const f of ifs) {
                                    try {
                                        if (f.contentDocument) {
                                            f.contentDocument.open();
                                            f.contentDocument.write('');
                                            f.contentDocument.close();
                                        }
                                        try { f.src = 'about:blank'; } catch(e) {}
                                    } catch(e) {}
                                }
                                return ifs.length;
                            }''')
                        except Exception:
                            continue

                    prog_clicked = None
                    prog_note_frame = None  # track which frame has the button

                    # Priority 0: Real Playwright click on the visible "Prog. Notes" button.
                    # The button's id is dynamic (id="claimProgressNoteBtnXXXXXXXXXX"), but
                    # ng-click="viewProgNote()" is stable. We need a TRUSTED user gesture so
                    # the right claim's progress note loads — scope.viewProgNote() alone often
                    # serves a stale popup state (wrong patient/date). Same root cause as HCFA.
                    prog_btn_handle = None
                    prog_btn_frame = None
                    for ctx in [hcfa_context, page] + list(page.frames):
                        try:
                            el = ctx.query_selector('button[ng-click="viewProgNote()"]')
                            if not el:
                                el = ctx.query_selector('button[id^="claimProgressNoteBtn"]')
                            if el:
                                prog_btn_handle = el
                                prog_btn_frame = ctx
                                break
                        except Exception:
                            continue

                    if prog_btn_handle and prog_btn_frame:
                        logger.info("🎯 Found Prog. Notes button — unhiding + clicking via Playwright")
                        # Force-unhide and un-disable the button so Playwright can click it.
                        try:
                            prog_btn_frame.evaluate('''() => {
                                const btn = document.querySelector('button[ng-click="viewProgNote()"]')
                                          || document.querySelector('button[id^="claimProgressNoteBtn"]');
                                if (!btn) return false;
                                let n = btn;
                                for (let i = 0; i < 6 && n; i++) {
                                    n.classList.remove('ng-hide');
                                    n.classList.remove('hidden');
                                    n.style.removeProperty('display');
                                    n.style.visibility = 'visible';
                                    n.style.opacity = '1';
                                    n.style.pointerEvents = 'auto';
                                    n = n.parentElement;
                                }
                                btn.disabled = false;
                                btn.removeAttribute('disabled');
                                btn.scrollIntoView({block: 'center', inline: 'center'});
                                return true;
                            }''')
                        except Exception:
                            pass
                        time.sleep(0.3)
                        try:
                            prog_btn_handle = prog_btn_frame.query_selector(
                                'button[ng-click="viewProgNote()"]'
                            ) or prog_btn_frame.query_selector(
                                'button[id^="claimProgressNoteBtn"]'
                            ) or prog_btn_handle
                        except Exception:
                            pass

                        clicked = False
                        # Attempt 1: standard Playwright click (trusted user gesture)
                        # Tight timeout — when this fails (logs show ~197/run), the
                        # dispatch_event fallback below is what actually opens the
                        # modal. 5s here was burning ~16 min/run on dead clicks.
                        try:
                            prog_btn_handle.click(timeout=1500)
                            clicked = True
                            prog_clicked = {'ok': True, 'method': 'playwright_click_visible'}
                            prog_note_frame = prog_btn_frame
                            logger.info("✅ Prog. Notes clicked via Playwright (trusted gesture)")
                        except Exception as _pc_err:
                            logger.warning(f"Playwright click on Prog. Notes failed: {_pc_err}")
                        # Attempt 2: page.mouse.click at bounding box center (real input event)
                        if not clicked:
                            try:
                                box = prog_btn_handle.bounding_box()
                                if box:
                                    cx = box['x'] + box['width'] / 2
                                    cy = box['y'] + box['height'] / 2
                                    page.mouse.click(cx, cy)
                                    clicked = True
                                    prog_clicked = {'ok': True, 'method': 'mouse_click_at_box'}
                                    prog_note_frame = prog_btn_frame
                                    logger.info(f"✅ Prog. Notes mouse.click at ({cx:.0f},{cy:.0f})")
                            except Exception as _m_err:
                                logger.warning(f"mouse.click failed: {_m_err}")
                        # Attempt 3: dispatch_event — last resort
                        if not clicked:
                            try:
                                prog_btn_handle.dispatch_event('click')
                                clicked = True
                                prog_clicked = {'ok': True, 'method': 'dispatch_event_click'}
                                prog_note_frame = prog_btn_frame
                                logger.info("✅ Prog. Notes dispatch_event('click') fired")
                            except Exception as _d_err:
                                logger.warning(f"dispatch_event failed: {_d_err}")

                    # Strategy 0 (FALLBACK only): Angular scope viewProgNote().
                    # Used only when no button is in DOM (extremely rare).
                    if not prog_clicked:
                     for f_ctx in [hcfa_context] + list(page.frames):
                        try:
                            scope_clicked = f_ctx.evaluate('''() => {
                                if (!window.angular) return { ok: false };
                                const candidates = [
                                    document.querySelector('button[ng-click="viewProgNote()"]'),
                                    document.querySelector('[ng-click*="viewProgNote"]'),
                                    document.querySelector('[ng-controller]'),
                                    document.body
                                ].filter(Boolean);
                                for (const el of candidates) {
                                    let scope = null;
                                    try { scope = angular.element(el).scope(); } catch(e) {}
                                    if (scope && typeof scope.viewProgNote === 'function') {
                                        scope.viewProgNote();
                                        try { scope.$apply(); } catch(e) {}
                                        return { ok: true, method: 'angular_scope_viewProgNote' };
                                    }
                                }
                                return { ok: false };
                            }''')
                            if scope_clicked and scope_clicked.get('ok'):
                                prog_clicked = scope_clicked
                                prog_note_frame = f_ctx
                                logger.info(f"✅ Prog Notes via Angular scope: {scope_clicked}")
                                break
                        except Exception:
                            continue

                    # Fallback: visible button click (only if scope path didn't work)
                    if not prog_clicked:
                     for attempt in range(15):  # Wait up to 15 seconds
                        # Search all frames for the Prog. Notes button
                        for f_ctx in [hcfa_context] + list(page.frames):
                            try:
                                result = f_ctx.evaluate('''() => {
                                    // Try ng-click selector first
                                    const btn = document.querySelector('button[ng-click="viewProgNote()"]');
                                    if (btn) {
                                        if (btn.disabled) return { found: true, disabled: true };
                                        if (btn.offsetWidth > 0) {
                                            btn.click();
                                            return { ok: true, id: btn.id, text: btn.textContent.trim() };
                                        }
                                    }
                                    // Fallback: text matching
                                    const all = document.querySelectorAll('button');
                                    for (const b of all) {
                                        const text = b.textContent.trim();
                                        if ((text.includes('Prog') && text.includes('Note')) || text === 'Prog. Notes') {
                                            if (b.disabled) return { found: true, disabled: true };
                                            if (b.offsetWidth > 0) {
                                                b.click();
                                                return { ok: true, id: b.id, text: text };
                                            }
                                        }
                                    }
                                    return { found: false };
                                }''')
                                if result and result.get('ok'):
                                    prog_clicked = result
                                    prog_note_frame = f_ctx
                                    break
                                elif result and result.get('found') and result.get('disabled'):
                                    logger.info(f"Prog. Notes button found but disabled, waiting... ({attempt+1}s)")
                                    prog_note_frame = f_ctx  # remember which frame
                                    break
                            except Exception:
                                continue
                        
                        if prog_clicked:
                            break
                        time.sleep(0.5)
                    
                    if prog_clicked and prog_clicked.get('ok'):
                        logger.info(f"✅ Clicked Prog. Notes: {prog_clicked}")
                        time.sleep(2)  # Wait for Progress Note modal to load
                        
                        # Dismiss the "Keyboard Shortcuts" help overlay if it appears
                        try:
                            dismissed = page.evaluate('''() => {
                                // Check if keyboard shortcuts overlay is visible
                                const allText = document.body.innerText || '';
                                if (!allText.includes('Keyboard Shortcuts')) return 'not-present';
                                
                                // Find the page-level X close button (top-right, outside modals)
                                const xBtns = document.querySelectorAll('.close, [class*="close"]');
                                for (const btn of xBtns) {
                                    const rect = btn.getBoundingClientRect();
                                    if (rect.right > 980 && rect.top < 50 && btn.offsetWidth > 0) {
                                        btn.click();
                                        return 'closed-x';
                                    }
                                }
                                
                                // Fallback: hide the overlay div directly
                                const divs = document.querySelectorAll('div');
                                for (const div of divs) {
                                    if (div.innerText && div.innerText.includes('Keyboard Shortcuts:')
                                        && div.offsetWidth > 300 && !div.closest('.modal')) {
                                        div.style.display = 'none';
                                        return 'hidden';
                                    }
                                }
                                return 'not-dismissed';
                            }''')
                            if dismissed and dismissed != 'not-present':
                                logger.info(f"Dismissed keyboard shortcuts overlay: {dismissed}")
                                time.sleep(0.5)
                        except Exception:
                            pass
                        
                        # Screenshot of the modal
                        try:
                            page.screenshot(path=f'/tmp/prog_notes_popup_{claim_id}.png')
                            logger.info("📸 Prog Notes modal screenshot saved")
                        except Exception:
                            pass

                        # ── Office-visit pre-modal handling ──
                        # For office visits the Prog. Notes modal FIRST shows a
                        # visualization-only view (no Print button, just Close).
                        # Closing that view reveals the actual document modal.
                        # IV-therapy claims go straight to the document, so this
                        # loop is a no-op for them (Print is already present).
                        for _adv in range(3):
                            _has_print = False
                            for _vctx in [hcfa_context] + list(page.frames):
                                try:
                                    if _vctx.query_selector('button[ng-click*="printProgressNote"]'):
                                        _has_print = True
                                        break
                                except Exception:
                                    continue
                            if _has_print:
                                break  # already on the document view
                            _adv_closed = False
                            for _vctx in [hcfa_context] + list(page.frames):
                                try:
                                    _vctx.evaluate('() => { window.print = () => {}; }')
                                except Exception:
                                    pass
                                try:
                                    _ok = _vctx.evaluate('''() => {
                                        const btns = document.querySelectorAll('button');
                                        for (const b of btns) {
                                            const t = (b.textContent || b.value || '').trim();
                                            if (t === 'Close' && b.offsetWidth > 0) {
                                                const r = b.getBoundingClientRect();
                                                if (r.top > 300 && r.left > 200) { b.click(); return true; }
                                            }
                                        }
                                        return false;
                                    }''')
                                    if _ok:
                                        _adv_closed = True
                                        logger.info(f"🔁 Closed visualization view (office visit pre-modal), attempt {_adv+1}")
                                        break
                                except Exception:
                                    continue
                            if not _adv_closed:
                                break
                            time.sleep(2)

                        # Use printProgressNote() but intercept window.print to avoid blocking.
                        # Key insight: We must override print() AFTER the iframe document is
                        # written but before focus()/print() — but printProgressNote() does
                        # write+print atomically. So we override print as a no-op FIRST,
                        # then call printProgressNote() which writes + tries to print (no-op).
                        # For the 2nd claim, the iframe already has stale content, so we
                        # call printProgressNote() via Angular scope directly — it uses
                        # document.write() which implicitly calls document.open() first,
                        # replacing ALL prior content.
                        
                        # Override window.print() to avoid blocking print dialogs
                        pn_ctx = prog_note_frame if prog_note_frame else hcfa_context
                        page.evaluate('''() => {
                            window.print = () => {};
                            try {
                                for (let i = 0; i < window.frames.length; i++) {
                                    try { window.frames[i].print = () => {}; } catch(e) {}
                                }
                            } catch(e) {}
                        }''')
                        try:
                            pn_ctx.evaluate('''() => {
                                window.print = () => {};
                                try {
                                    for (let i = 0; i < window.frames.length; i++) {
                                        try { window.frames[i].print = () => {}; } catch(e) {}
                                    }
                                } catch(e) {}
                            }''')
                        except Exception:
                            pass
                        
                        # FIRST: trusted Playwright click on the real "Print" button
                        # <button ng-click="printProgressNote('ProgNoteViwerFrame','ProgNoteViwerFrame')">Print</button>
                        # ECW's ng-click needs a TRUSTED gesture; the JS .click()/scope
                        # call below often no-ops ("Could not call printProgressNote") and
                        # the modal just sits there. The button lives in the modal's frame,
                        # so search every frame, not just pn_ctx.
                        # NOTE: do NOT gate on is_visible() — ECW modal buttons frequently
                        # report not-visible even when present (that's why the prior attempt
                        # logged "Could not call printProgressNote"). dispatch_event('click')
                        # is the method that proved reliable for the Prog. Notes button in
                        # this exact modal, so try it after a forced Playwright click.
                        print_result = None
                        for _pctx in [pn_ctx, hcfa_context] + list(page.frames):
                            try:
                                _pbtn = _pctx.query_selector('button[ng-click*="printProgressNote"]')
                                if not _pbtn:
                                    continue
                                try:
                                    _pctx.evaluate('() => { window.print = () => {}; }')
                                except Exception:
                                    pass
                                _clicked = False
                                try:
                                    _pbtn.click(timeout=2500, force=True)
                                    _clicked = True
                                    logger.info("✅ Print button: Playwright force-click")
                                except Exception:
                                    pass
                                if not _clicked:
                                    try:
                                        _pbtn.dispatch_event('click')
                                        _clicked = True
                                        logger.info("✅ Print button: dispatch_event('click')")
                                    except Exception:
                                        pass
                                if not _clicked:
                                    try:
                                        _ok = _pctx.evaluate('''() => {
                                            const b = document.querySelector('button[ng-click*="printProgressNote"]');
                                            if (!b) return false;
                                            try {
                                                const s = angular.element(b).scope();
                                                if (s && s.printProgressNote) {
                                                    s.printProgressNote('ProgNoteViwerFrame', 'ProgNoteViwerFrame');
                                                    return true;
                                                }
                                            } catch(e) {}
                                            b.click();
                                            return true;
                                        }''')
                                        _clicked = bool(_ok)
                                        if _clicked:
                                            logger.info("✅ Print button: Angular scope / JS click")
                                    except Exception:
                                        pass
                                if _clicked:
                                    print_result = {'ok': True, 'method': 'print-button'}
                                    break
                            except Exception:
                                continue

                        # Fallback: call printProgressNote via Angular scope in the correct frame
                        if not (print_result and print_result.get('ok')):
                          print_result = pn_ctx.evaluate('''() => {
                            // Try calling the Angular scope function directly
                            try {
                                const scope = angular.element(document.querySelector('[ng-click*="printProgressNote"]')).scope();
                                if (scope && scope.printProgressNote) {
                                    // Override print on the iframe before calling
                                    const iframe = document.querySelector('iframe[name="ProgNoteViwerFrame"]');
                                    if (iframe && iframe.contentWindow) {
                                        try { iframe.contentWindow.print = () => {}; } catch(e) {}
                                    }
                                    scope.printProgressNote('ProgNoteViwerFrame', 'ProgNoteViwerFrame');
                                    return { ok: true, method: 'angular-scope' };
                                }
                            } catch(e) {}
                            
                            // Fallback: click the Print button
                            const btns = document.querySelectorAll('button[ng-click*="printProgressNote"]');
                            for (const btn of btns) {
                                if (btn.offsetWidth > 0) {
                                    btn.click();
                                    return { ok: true, method: 'click' };
                                }
                            }
                            
                            // Fallback 2: try viewProgNote function directly
                            try {
                                if (typeof viewProgNote === 'function') {
                                    viewProgNote();
                                    return { ok: true, method: 'viewProgNote' };
                                }
                            } catch(e) {}
                            
                            return { ok: false };
                        }''')
                        
                        # ProgNoteViwerFrame is the modal's VIEWER iframe — it is populated
                        # when the Prog. Notes modal opens, INDEPENDENT of whether the
                        # Print / printProgressNote invocation "succeeded" (pressing Print
                        # just opens Chrome's native dialog). So always extract the frame.
                        if True:
                            logger.info(f"📄 Extracting Progress Note (print_result={print_result})")
                            time.sleep(1)  # Wait for content write
                            
                            # Re-override print on iframe (document.write creates a new doc)
                            try:
                                pn_ctx.evaluate('''() => {
                                    const iframe = document.querySelector('iframe[name="ProgNoteViwerFrame"]');
                                    if (iframe && iframe.contentWindow) {
                                        try { iframe.contentWindow.print = () => {}; } catch(e) {}
                                    }
                                }''')
                            except Exception:
                                pass
                            
                            # Find the note frame.
                            #
                            # ECW leaves a STALE ProgNoteViwerFrame iframe from every
                            # previous claim's popup in the DOM — so picking the first
                            # frame whose name matches yields the WRONG patient's note
                            # almost every time (audit showed 66/88 captures were stale,
                            # nearly all containing "REVERS, FELICIA" — that was the
                            # last popup that successfully populated). Fix: scan EVERY
                            # candidate frame and pick the one whose body text actually
                            # contains THIS claim's surname. Retry — the modal loads
                            # async, so the correct iframe may not be populated yet.
                            _expected_surname = (patient_name or '').split(',')[0].strip().upper()
                            prog_frame = None
                            _fallback_frame = None  # any named ProgNote frame, used only if no surname match after retries
                            for _try in range(10):
                                for frame in page.frames:
                                    try:
                                        is_named = 'ProgNote' in (frame.name or '')
                                        _ft = frame.evaluate('() => (document.body ? document.body.innerText : "").substring(0, 1500)') or ''
                                        _u = _ft.upper()
                                        looks_like_note = ('SUBJECTIVE' in _u or 'CHIEF COMPLAINT' in _u or 'DOS:' in _u)
                                        # Best: named ProgNote frame whose content shows our patient
                                        if is_named and _expected_surname and _expected_surname in _u:
                                            prog_frame = frame
                                            logger.info(f"Found ProgNoteViwerFrame by name+content (surname={_expected_surname}): {frame.name}")
                                            break
                                        # Second-best: ANY frame whose content shows our patient + note markers
                                        if looks_like_note and _expected_surname and _expected_surname in _u:
                                            prog_frame = frame
                                            logger.info(f"Found Prog Note frame by content (surname={_expected_surname})")
                                            break
                                        # Last-resort fallback: a named frame, even if surname not yet visible.
                                        # Only used if the retry loop ends with no surname match.
                                        if is_named and len(_ft.strip()) > 100 and _fallback_frame is None:
                                            _fallback_frame = frame
                                    except Exception:
                                        continue
                                if prog_frame:
                                    break
                                time.sleep(1)
                            if not prog_frame and _fallback_frame is not None:
                                prog_frame = _fallback_frame
                                logger.warning(f"⚠️ No surname-match found after retries — falling back to first non-empty ProgNote frame; downstream verifier will drop if wrong patient")

                            if prog_frame:
                                try:
                                    frame_html = prog_frame.evaluate('() => document.documentElement ? document.documentElement.outerHTML : ""')
                                    
                                    if frame_html and len(frame_html) > 500:
                                        logger.info(f"ProgNoteViwerFrame content: {len(frame_html)} chars")

                                        # Extract FULL iframe text once for BOTH verification
                                        # and parsing. No substring limit — earlier 3000/15000
                                        # cutoffs missed "Rx Start Date" lines that sit past
                                        # the medications section. Surname verification below
                                        # still slices to first 500 chars in Python so the
                                        # head-only check isn't affected.
                                        try:
                                            frame_text = prog_frame.evaluate('() => document.body ? document.body.innerText : ""')
                                        except Exception:
                                            frame_text = ''
                                        logger.info(f"Prog Note preview: {(frame_text or '')[:200]}")

                                        # VERIFY PATIENT FIRST — only parse Rx Start Date / render
                                        # PDF if the iframe text matches this claim's patient.
                                        # ECW's ProgNoteViwerFrame frequently serves stale data
                                        # from a previous patient (the "REVERS, Felicia" cascade).
                                        # If we parsed BEFORE verifying, we'd save the wrong
                                        # patient's Rx Start Date into DDB (claim 5197 had
                                        # 01/28/2026 from Felicia instead of the real 01/09/2026
                                        # from Gray's note).
                                        _iv_verified = False
                                        _iv_drop_reason = None
                                        _iv_expected_surname = (patient_name or '').split(',')[0].strip().upper()
                                        if not _iv_expected_surname:
                                            _iv_drop_reason = 'claim has no patient_name on record'
                                        else:
                                            # Head-only: patient header lives in the first ~500
                                            # chars. Wider windows match other people named later
                                            # in the note (paramedic/provider lines), masking
                                            # cross-patient bleed (claim 5305 Felicia leak).
                                            _iv_text_head = (frame_text or '')[:500].upper()
                                            if not _iv_text_head.strip():
                                                _iv_drop_reason = 'IV Note iframe text empty (failed to extract)'
                                            elif _iv_expected_surname not in _iv_text_head:
                                                _iv_drop_reason = (f'IV Note iframe is for a DIFFERENT patient '
                                                                   f'(expected {_iv_expected_surname}, iframe begins with '
                                                                   f'{_iv_text_head[:80].strip()!r})')
                                            else:
                                                _iv_verified = True
                                                logger.info(f"  ✓ IV Note iframe patient verified: {_iv_expected_surname}")

                                        # Only parse DOS / Rx Start Date if we verified the
                                        # iframe is this patient's note. Stale (wrong-patient)
                                        # content gets dropped without contaminating DDB.
                                        if _iv_verified:
                                            import re
                                            try:
                                                dos_match = re.search(r'DOS:\s*(\d{2}/\d{2}/\d{4})', frame_text)
                                                if dos_match:
                                                    logger.info(f"📋 DOS: {dos_match.group(1)} for claim {claim_id}")
                                                if is_office_visit_claim:
                                                    logger.info(f"🏥 Office visit — skipping Rx Start Date parse for claim {claim_id}")
                                                else:
                                                    rx_match = re.search(r'Rx\s*Start\s*Date\s*:?\s*(\d{2}/\d{2}/\d{4})', frame_text, re.IGNORECASE)
                                                    if rx_match:
                                                        mm, dd, yyyy = rx_match.group(1).split('/')
                                                        iv_note_rx_start_date = f"{yyyy}-{mm}-{dd}"
                                                        logger.info(f"📋 Rx Start Date from IV Note: {iv_note_rx_start_date} (claim {claim_id})")
                                                    else:
                                                        logger.info(f"📋 No 'Rx Start Date' found in IV Note for claim {claim_id}")
                                            except Exception as _rx_err:
                                                logger.warning(f"Rx Start Date parse failed: {_rx_err}")

                                        if not _iv_verified:
                                            logger.warning(f"⛔ IV Note for claim {claim_id} DROPPED — {_iv_drop_reason}. PDF NOT rendered or uploaded.")
                                            # Sync the local var so the final update at
                                            # the bottom of this block doesn't clobber the
                                            # mismatch=True we're about to write to DDB.
                                            iv_note_patient_mismatch = True
                                            # Delete any stale /tmp file from a prior run — the
                                            # captured-check below this if/else uses os.path.exists
                                            # and would otherwise mark this claim as "captured"
                                            # using last-run's (wrong-patient) PDF, which then
                                            # gets uploaded to S3 overwriting the cleared path.
                                            try:
                                                if os.path.exists(prog_notes_path):
                                                    os.remove(prog_notes_path)
                                            except Exception:
                                                pass
                                            try:
                                                aws_client.update_claim_status(claim_id, {
                                                    'iv_note_patient_mismatch': True,
                                                    'prog_notes_capture_failed': (_iv_drop_reason or 'patient_mismatch')[:200],
                                                    'prog_notes_capture_failed_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                                })
                                            except Exception:
                                                pass
                                        else:
                                            # Render the verified iframe content as PDF.
                                            pdf_page = page.context.new_page()
                                            pdf_page.set_content(frame_html, wait_until='domcontentloaded')
                                            time.sleep(0.3)
                                            pdf_page.pdf(
                                                path=prog_notes_path,
                                                format='Letter',
                                                print_background=True,
                                                margin={'top': '0.4in', 'bottom': '0.4in', 'left': '0.4in', 'right': '0.4in'}
                                            )
                                            pdf_page.close()

                                        if os.path.exists(prog_notes_path) and os.path.getsize(prog_notes_path) > 500:
                                            prog_notes_captured = True
                                            logger.info(f"✅ Progress Notes PDF captured: {prog_notes_path} ({os.path.getsize(prog_notes_path)} bytes)")
                                    else:
                                        logger.warning(f"ProgNoteViwerFrame content too small ({len(frame_html) if frame_html else 0} chars)")
                                except Exception as e:
                                    logger.warning(f"Failed to extract iframe content: {e}")
                        if not prog_notes_captured:
                            logger.warning(f"⚠️ Progress Note frame capture did not succeed for claim {claim_id} — trying Print-button fallback")
                        
                        # ── Fallback: press the modal's "Print" button and capture the
                        # resulting popup PDF. Office-visit progress notes don't populate
                        # the ProgNoteViwerFrame iframe, so the approach above yields
                        # nothing and the modal would just sit there ("se queda trabado").
                        # This presses Print like a user would. Bounded — never hangs. ──
                        if not prog_notes_captured:
                            try:
                                import base64 as _b64
                                _pn_pop = []
                                def _on_pn_popup(p):
                                    _pn_pop.append(p)
                                page.context.on('page', _on_pn_popup)
                                try:
                                    _printed = page.evaluate('''() => {
                                        const els = document.querySelectorAll('button, input[type="button"], input[type="submit"], a');
                                        for (const b of els) {
                                            const t = (b.value || b.textContent || '').trim().toLowerCase();
                                            if ((t === 'print' || t === 'print note' || t === 'print progress note')
                                                && b.offsetWidth > 0 && b.offsetHeight > 0) {
                                                const r = b.getBoundingClientRect();
                                                if (r.top > 300) { b.click(); return true; }
                                            }
                                        }
                                        return false;
                                    }''')
                                    logger.info(f"🖨️ Prog Notes 'Print' button clicked: {_printed} (claim {claim_id})")
                                except Exception as _pce:
                                    _printed = False
                                    logger.warning(f"Print button click failed: {_pce}")
                                if _printed:
                                    for _w in range(12):
                                        if _pn_pop:
                                            break
                                        time.sleep(1)
                                    # Expected surname for patient verification — same rule
                                    # the iframe path uses ([src/main.py:7068-7094]). The
                                    # Print popup serves whatever ECW's modal scope holds,
                                    # so it can return a stale (wrong-patient) note. We
                                    # MUST verify before writing the PDF.
                                    _pp_expected_surname = (patient_name or '').split(',')[0].strip().upper()
                                    for _pg in _pn_pop:
                                        try:
                                            try:
                                                _pg.wait_for_load_state('domcontentloaded', timeout=10000)
                                            except Exception:
                                                pass
                                            time.sleep(2)
                                            # Extract FULL popup text — no substring limit so
                                            # the Rx Start Date is never cut off. Surname check
                                            # below uses Python slice [:500] for head-only.
                                            try:
                                                _txt = _pg.evaluate('() => document.body ? document.body.innerText : ""') or ''
                                            except Exception:
                                                _txt = ''
                                            if not is_office_visit_claim and not iv_note_rx_start_date:
                                                try:
                                                    import re as _re_rx
                                                    _m = _re_rx.search(r'Rx\s*Start\s*Date\s*:?\s*(\d{2}/\d{2}/\d{4})', _txt, _re_rx.IGNORECASE)
                                                    if _m:
                                                        _mm, _dd, _yy = _m.group(1).split('/')
                                                        iv_note_rx_start_date = f"{_yy}-{_mm}-{_dd}"
                                                        logger.info(f"📋 Rx Start Date from printed IV Note: {iv_note_rx_start_date} (claim {claim_id})")
                                                except Exception:
                                                    pass
                                            # Patient verification — drop if popup is for a
                                            # different patient than this claim.
                                            _pp_verified = False
                                            _pp_drop_reason = None
                                            if not _pp_expected_surname:
                                                _pp_drop_reason = 'claim has no patient_name on record'
                                            else:
                                                # Head-only (first 500 chars) — wider windows can
                                                # match a paramedic/provider whose name happens to
                                                # match this claim's surname (see claim 5305).
                                                _pp_head = _txt[:500].upper()
                                                if not _pp_head.strip():
                                                    _pp_drop_reason = 'Print popup text empty (failed to extract)'
                                                elif _pp_expected_surname not in _pp_head:
                                                    _pp_drop_reason = (f'Print popup is for a DIFFERENT patient '
                                                                       f'(expected {_pp_expected_surname}, popup begins with '
                                                                       f'{_pp_head[:80].strip()!r})')
                                                else:
                                                    _pp_verified = True
                                                    logger.info(f"  ✓ Print-popup IV Note patient verified: {_pp_expected_surname}")
                                            if not _pp_verified:
                                                logger.warning(f"⛔ Print-popup IV Note for claim {claim_id} DROPPED — {_pp_drop_reason}. PDF NOT written.")
                                                iv_note_patient_mismatch = True
                                                try:
                                                    aws_client.update_claim_status(claim_id, {
                                                        'iv_note_patient_mismatch': True,
                                                        'prog_notes_capture_failed': (_pp_drop_reason or 'patient_mismatch')[:200],
                                                        'prog_notes_capture_failed_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                                    })
                                                except Exception:
                                                    pass
                                                continue
                                            _cdp = page.context.new_cdp_session(_pg)
                                            _pdf = _cdp.send('Page.printToPDF', {
                                                'landscape': False, 'displayHeaderFooter': False,
                                                'printBackground': True, 'paperWidth': 8.5, 'paperHeight': 11,
                                                'marginTop': 0.4, 'marginBottom': 0.4, 'marginLeft': 0.4, 'marginRight': 0.4,
                                            })
                                            with open(prog_notes_path, 'wb') as _f:
                                                _f.write(_b64.b64decode(_pdf['data']))
                                            _cdp.detach()
                                            if os.path.exists(prog_notes_path) and os.path.getsize(prog_notes_path) > 500:
                                                prog_notes_captured = True
                                                logger.info(f"✅ Progress Notes PDF captured via Print button ({os.path.getsize(prog_notes_path)} bytes)")
                                                break
                                        except Exception as _ppe:
                                            logger.warning(f"Print-popup capture failed: {_ppe}")
                                    for _pg in _pn_pop:
                                        try:
                                            _pg.close()
                                        except Exception:
                                            pass
                                    if not _pn_pop:
                                        logger.warning(f"⚠️ Print clicked but no popup spawned for claim {claim_id}")
                                try:
                                    page.context.remove_listener('page', _on_pn_popup)
                                except Exception:
                                    pass
                            except Exception as _fb_err:
                                logger.warning(f"Prog Notes Print fallback failed for {claim_id}: {_fb_err}")

                        # DEEP RETRY — when iframe + popup both gave wrong-patient
                        # content, ECW's session has stale state we can't clear
                        # by just closing the modal. Hard-reload the page so the
                        # SPA re-initializes and ECW re-fetches fresh data, then
                        # re-open the claim popup and re-attempt the Prog Notes
                        # capture once. This is what surfaces the right patient's
                        # note when the first attempt got someone else's.
                        if (not prog_notes_captured) and iv_note_patient_mismatch:
                            logger.info(f"🔁 Deep retry for claim {claim_id} — page reload to clear stale ECW state")
                            try:
                                # Defensive: remove any stale local file before retry.
                                try:
                                    if os.path.exists(prog_notes_path):
                                        os.remove(prog_notes_path)
                                except Exception:
                                    pass
                                # Reset the mismatch flag for the retry attempt;
                                # if the retry also fails it gets set again below.
                                iv_note_patient_mismatch = False

                                # Hard reload the page → kills client-side
                                # Angular state + iframe caches.
                                try:
                                    page.reload(wait_until='domcontentloaded', timeout=30000)
                                    time.sleep(5)
                                except Exception as _rl_err:
                                    logger.warning(f"  page reload failed: {_rl_err}")

                                # Re-open the claim popup via lookup → fresh
                                # Angular scope for the claim controller.
                                reopened, refreshed_frame = _open_claim_popup_via_lookup(page, claim_id)
                                if not reopened:
                                    logger.warning(f"  ⚠️ Deep retry: could not re-open claim popup after reload")
                                else:
                                    if refreshed_frame:
                                        hcfa_context = refreshed_frame

                                    # Click Prog. Notes button (same logic as
                                    # the primary path, condensed). Trust the
                                    # ng-click; fall back to dispatch_event.
                                    _dr_btn = None
                                    _dr_frame = None
                                    for _dr_ctx in [hcfa_context, page] + list(page.frames):
                                        try:
                                            _el = _dr_ctx.query_selector('button[ng-click="viewProgNote()"]') \
                                                  or _dr_ctx.query_selector('button[id^="claimProgressNoteBtn"]')
                                            if _el:
                                                _dr_btn = _el
                                                _dr_frame = _dr_ctx
                                                break
                                        except Exception:
                                            continue
                                    if _dr_btn:
                                        try:
                                            _dr_btn.click(timeout=2000)
                                        except Exception:
                                            try:
                                                _dr_btn.dispatch_event('click')
                                            except Exception:
                                                pass
                                        time.sleep(3)

                                        # Find ProgNoteViwerFrame whose body
                                        # contains the expected surname.
                                        _dr_expected = (patient_name or '').split(',')[0].strip().upper()
                                        _dr_prog_frame = None
                                        for _retry_wait in range(8):  # up to 8s
                                            for _fr in page.frames:
                                                try:
                                                    is_named = 'ProgNote' in (_fr.name or '')
                                                    _txt = (_fr.evaluate('() => document.body ? document.body.innerText.substring(0,1500) : ""') or '').upper()
                                                except Exception:
                                                    continue
                                                if _dr_expected and _dr_expected in _txt[:500] and (
                                                    'RX START DATE' in _txt or 'DOS:' in _txt or 'PATIENT:' in _txt
                                                ):
                                                    _dr_prog_frame = _fr
                                                    break
                                            if _dr_prog_frame:
                                                break
                                            time.sleep(1)

                                        if _dr_prog_frame:
                                            try:
                                                _dr_html = _dr_prog_frame.evaluate(
                                                    '() => document.documentElement ? document.documentElement.outerHTML : ""'
                                                )
                                                _dr_text = _dr_prog_frame.evaluate(
                                                    '() => document.body ? document.body.innerText : ""'
                                                ) or ''
                                                if _dr_html and len(_dr_html) > 500 and _dr_expected in _dr_text[:500].upper():
                                                    logger.info(f"  ✅ Deep retry: found correct patient's IV Note ({_dr_expected})")
                                                    # Parse DOS + Rx Start Date NOW that we know it's the right patient.
                                                    import re as _re_dr
                                                    _dos_m = _re_dr.search(r'DOS:\s*(\d{2}/\d{2}/\d{4})', _dr_text)
                                                    if _dos_m:
                                                        logger.info(f"📋 DOS: {_dos_m.group(1)} (deep retry, claim {claim_id})")
                                                    if not is_office_visit_claim:
                                                        _rx_m = _re_dr.search(r'Rx\s*Start\s*Date\s*:?\s*(\d{2}/\d{2}/\d{4})',
                                                                              _dr_text, _re_dr.IGNORECASE)
                                                        if _rx_m:
                                                            _mm, _dd, _yy = _rx_m.group(1).split('/')
                                                            iv_note_rx_start_date = f"{_yy}-{_mm}-{_dd}"
                                                            logger.info(f"📋 Rx Start Date (deep retry): {iv_note_rx_start_date} (claim {claim_id})")
                                                    # Render PDF via Playwright.
                                                    _pdf_page = page.context.new_page()
                                                    _pdf_page.set_content(_dr_html, wait_until='domcontentloaded')
                                                    time.sleep(0.3)
                                                    _pdf_page.pdf(
                                                        path=prog_notes_path, format='Letter',
                                                        print_background=True,
                                                        margin={'top': '0.4in', 'bottom': '0.4in',
                                                                'left': '0.4in', 'right': '0.4in'}
                                                    )
                                                    _pdf_page.close()
                                                    if os.path.exists(prog_notes_path) and os.path.getsize(prog_notes_path) > 500:
                                                        prog_notes_captured = True
                                                        logger.info(f"✅ Progress Notes PDF captured via DEEP RETRY ({os.path.getsize(prog_notes_path)} bytes)")
                                                else:
                                                    logger.warning(f"  ⚠️ Deep retry: frame found but patient mismatch persists")
                                                    iv_note_patient_mismatch = True
                                            except Exception as _dr_rend_err:
                                                logger.warning(f"  Deep retry render failed: {_dr_rend_err}")
                                                iv_note_patient_mismatch = True
                                        else:
                                            logger.warning(f"  ⚠️ Deep retry: no frame matched surname {_dr_expected!r}")
                                            iv_note_patient_mismatch = True
                                    else:
                                        logger.warning(f"  ⚠️ Deep retry: Prog. Notes button not found")
                                        iv_note_patient_mismatch = True
                            except Exception as _dr_err:
                                logger.warning(f"Deep retry failed for {claim_id}: {_dr_err}")
                                iv_note_patient_mismatch = True

                        # Close the Progress Notes modal
                        try:
                            page.evaluate('''() => {
                                const btns = document.querySelectorAll('button');
                                for (const btn of btns) {
                                    if (btn.textContent.trim() === 'Close' && btn.offsetWidth > 0) {
                                        const rect = btn.getBoundingClientRect();
                                        if (rect.top > 400) { btn.click(); return true; }
                                    }
                                }
                                // Fallback: X close button
                                const xBtns = document.querySelectorAll('.close, [class*="close"]');
                                for (const btn of xBtns) {
                                    if (btn.offsetWidth > 0) { btn.click(); return true; }
                                }
                                return false;
                            }''')
                            logger.info("✅ Closed Progress Notes modal")
                        except Exception:
                            page.keyboard.press('Escape')
                        time.sleep(0.3)
                    else:
                        logger.warning(f"Could not find Prog. Notes button for claim {claim_id}")
                    
                    # Upload Progress Notes PDF to S3 if captured
                    if prog_notes_captured and os.path.exists(prog_notes_path):
                        try:
                            s3_key = f"prog_notes/{claim_id}_prog_notes.pdf"
                            prog_s3_path = aws_client.upload_to_s3(prog_notes_path, s3_key)
                            logger.info(f"✅ Progress Notes uploaded to S3: {prog_s3_path}")
                            iv_note_update = {
                                'prog_notes_s3_path': prog_s3_path,
                                'prog_notes_size': os.path.getsize(prog_notes_path),
                                'prog_notes_captured_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                # Persist patient-mismatch flag so the dashboard can show "Needs Review"
                                # when ECW's popup state served the wrong patient's note.
                                'iv_note_patient_mismatch': bool(iv_note_patient_mismatch),
                            }
                            # Store the Rx Start Date in its own field so it doesn't clobber the
                            # encounter_date (which holds the Progress Note / captured encounter date).
                            if iv_note_rx_start_date:
                                iv_note_update['iv_note_rx_start_date'] = iv_note_rx_start_date
                            aws_client.update_claim_status(claim_id, iv_note_update)
                            logger.info(
                                f"✅ Claim {claim_id} IV Note saved "
                                f"(mismatch={iv_note_patient_mismatch}, rx={iv_note_rx_start_date or 'n/a'})"
                            )
                        except Exception as e:
                            logger.error(f"Prog Notes S3 upload failed for {claim_id}: {e}")
                  except Exception as e:
                    logger.warning(f"Progress Notes capture failed for {claim_id}: {e}")
                
                _set_step('looking up Progress Note')
                # ── Encounter / Progress Note: capture using the IV Note Rx Start Date ──
                try:
                    _rx_b = iv_note_rx_start_date or claim_record.get('iv_note_rx_start_date')
                    _extract_encounter_file(_rx_to_mmddyyyy(_rx_b))
                except Exception as _enc_call_err:
                    logger.warning(f"Encounter extraction (post-IV-note) failed: {_enc_call_err}")
                # Close the claim detail popup properly. Short click timeout so
                # we don't burn Playwright's 30s default if the button is hidden/
                # covered — Escape fallback is fast and reliable.
                try:
                    close_btn = hcfa_context.query_selector(
                        'button:has-text("Cancel"), input[value="Cancel"], '
                        'button:has-text("OK"), input[value="OK"]'
                    )
                    if close_btn:
                        try:
                            close_btn.click(timeout=2000)
                            logger.info("✅ Closed claim popup")
                        except Exception:
                            page.keyboard.press('Escape')
                    else:
                        page.keyboard.press('Escape')
                except Exception:
                    try:
                        page.keyboard.press('Escape')
                    except Exception:
                        pass
                time.sleep(0.3)
                # Clear current_step so the dashboard's "live processing" badge
                # doesn't linger on the last-touched claim after we move on.
                try:
                    aws_client.update_claim_status(claim_id, {'current_step': ''})
                except Exception:
                    pass
                
                # Dismiss any lingering modals with Escape
                try:
                    page.keyboard.press('Escape')
                except Exception:
                    pass
                time.sleep(0.3)

            logger.info(f"═══ HCFA Generation complete: {hcfa_success_count} success, {hcfa_fail_count} failed ═══")

        except Exception as e:
            logger.error(f"HCFA generation failed: {e}")
            import traceback
            logger.error(traceback.format_exc())
        finally:
            manager.stop()

    # ──────────────────────────────────────
    # STEP 0B: TEST — Blue Shield Portal Login
    # ──────────────────────────────────────
    elif task_type == 'test_blueshield_login':
        logger.info("STEP 0B: Blue Shield Provider login...")
        manager = BrowserManager().start()
        try:
            page = manager.new_page()
            creds = aws_client.get_secret("blueshield_credentials")
            gmail_creds = aws_client.get_secret("gmail_credentials")

            # ─── Navigate to BS portal ───
            target_url = "https://www.blueshieldca.com/providerwebapp/claims/claimStatus"
            logger.info(f"Navigating to: {target_url}")
            page.goto(target_url, wait_until="domcontentloaded", timeout=60000)

            # Wait for page to fully settle (redirects may happen)
            try:
                page.wait_for_load_state('networkidle', timeout=20000)
            except Exception:
                pass
            time.sleep(1)

            # Dismiss cookie consent banner if present
            try:
                cookie_btn = page.query_selector('button:has-text("Continue"), button:has-text("Accept"), button:has-text("OK")')
                if cookie_btn and cookie_btn.is_visible():
                    cookie_btn.click()
                    logger.info("Dismissed cookie consent banner")
                    time.sleep(1)
            except Exception:
                pass

            # ─── Check: already logged in? ───
            current_url = page.url
            logger.info(f"Landed on: {page.title()} | {current_url[:120]}")

            # Check the URL PATH (not query params) to avoid false positives
            from urllib.parse import urlparse
            parsed = urlparse(current_url)
            is_on_portal = 'providerwebapp' in parsed.path

            if is_on_portal:
                logger.info("🎉 Already logged in via persistent session!")
                needs_login = False
            else:
                needs_login = True
                logger.info(f"Login required. Page: {page.title()}")

                # ─── STEP 1: Login form ───
                # Wait for any username input (PingFederate or Provider Connection)
                try:
                    page.wait_for_selector(
                        'input[name="pf.username"], input#username, '
                        'input[placeholder="Username"], input[placeholder="Password"]',
                        state='visible', timeout=30000
                    )
                except Exception:
                    pass
                time.sleep(random.uniform(1.0, 2.0))

                # Dismiss cookie banner again if it reappeared
                try:
                    cookie_btn = page.query_selector('button:has-text("Continue"), button:has-text("Accept")')
                    if cookie_btn and cookie_btn.is_visible():
                        cookie_btn.click()
                        time.sleep(1)
                except Exception:
                    pass

                # Find the username field (try multiple selectors)
                username_field = (
                    page.query_selector('input[name="pf.username"]') or
                    page.query_selector('input#username') or
                    page.query_selector('input[placeholder="Username"]') or
                    page.query_selector('input[type="text"][placeholder*="sername"]')
                )
                # Find password field (type=password is universal)
                password_field = (
                    page.query_selector('input[type="password"]') or
                    page.query_selector('input[placeholder="Password"]') or
                    page.query_selector('input[name="pf.pass"]') or
                    page.query_selector('input#password')
                )

                if username_field:
                    # Clear any auto-filled value first
                    username_field.click()
                    time.sleep(0.2)
                    page.keyboard.press('Control+a')
                    page.keyboard.press('Backspace')
                    time.sleep(random.uniform(0.3, 0.8))
                    for char in creds['username']:
                        page.keyboard.type(char, delay=random.uniform(40, 120))
                    logger.info("✅ Username entered")
                else:
                    logger.error("❌ Username field not found")

                time.sleep(random.uniform(0.5, 1.2))

                if password_field:
                    # Clear any auto-filled value first
                    password_field.click()
                    time.sleep(0.2)
                    page.keyboard.press('Control+a')
                    page.keyboard.press('Backspace')
                    time.sleep(random.uniform(0.3, 0.7))
                    for char in creds['password']:
                        page.keyboard.type(char, delay=random.uniform(40, 100))
                    logger.info("✅ Password entered")
                else:
                    logger.error("❌ Password field not found")

                time.sleep(random.uniform(0.8, 2.0))

                # Submit login — click Log in / Sign in button
                try:
                    page.click('button:has-text("Log in")', timeout=3000)
                    logger.info("✅ Clicked 'Log in'")
                except Exception:
                    try:
                        page.click('button:has-text("Sign in"), input[type="submit"]', timeout=3000)
                        logger.info("✅ Clicked Sign In")
                    except Exception:
                        page.evaluate('''() => {
                            const form = document.querySelector('form');
                            if (form) {
                                const okField = document.querySelector('input[name="pf.ok"]');
                                if (okField) okField.value = "clicked";
                                form.submit();
                            }
                        }''')
                        logger.info("✅ Form submitted (fallback)")
                time.sleep(random.uniform(2, 4))
                logger.info(f"After login — Title: {page.title()}")

                # ─── STEP 2: Handle SSO redirects until MFA or portal ───
                for sso_attempt in range(8):
                    current_title = page.title()
                    current_url = page.url
                    page_text = page.inner_text('body').lower()

                    # Check if we reached the portal
                    if 'providerwebapp' in urlparse(current_url).path:
                        logger.info("🎉 Reached portal after SSO!")
                        break

                    # Check if we're on the MFA page
                    if 'select authentication' in page_text or '2-step' in page_text or 'email at' in page_text:
                        logger.info("🔐 MFA page detected")
                        break

                    # Check if there's a visible login form (real login page, NOT SSO pass-through)
                    has_login_form = page.query_selector(
                        'input[name="pf.username"], input#username, input[placeholder="Username"]'
                    )
                    if has_login_form and has_login_form.is_visible():
                        logger.info(f"On login form page — credentials may not have submitted correctly")
                        # Try clicking the Sign In button
                        try:
                            page.click('button:has-text("Sign in")', timeout=3000)
                            logger.info("✅ Clicked Sign In")
                        except Exception:
                            try:
                                page.click('input[type="submit"]', timeout=3000)
                                logger.info("✅ Clicked submit input")
                            except Exception:
                                page.keyboard.press('Enter')
                                logger.info("✅ Pressed Enter to submit")
                        time.sleep(random.uniform(2, 4))
                        continue

                    # SSO pass-through page (no visible login form, just hidden form)
                    if current_title == 'Sign On' or (current_title == '' and 'authorization.ping' in current_url):
                        logger.info(f"SSO pass-through #{sso_attempt + 1}")
                        page.evaluate('''() => {
                            const form = document.querySelector('form');
                            if (form) {
                                const okField = document.querySelector('input[name="pf.ok"]');
                                if (okField) okField.value = "clicked";
                                form.submit();
                            }
                        }''')
                        time.sleep(random.uniform(2, 4))
                        continue

                    # Unknown page — stop
                    logger.info(f"Stopped at: {current_title[:50]} | {current_url[:80]}")
                    break

                # Re-check page state after SSO redirects
                page_text = page.inner_text('body').lower()

                # ─── STEP 3: MFA if needed ───
                if '2-step' in page_text or 'authentication method' in page_text or 'email at' in page_text:
                    logger.info("🔐 MFA required")
                    time.sleep(random.uniform(1.0, 2.0))

                    # Select Email
                    try:
                        page.click('text=Email at', timeout=5000)
                        logger.info("✅ Selected Email method")
                    except Exception:
                        radio = page.query_selector('input[type="radio"]')
                        if radio:
                            radio.click()
                            logger.info("✅ Clicked radio button")

                    time.sleep(random.uniform(0.5, 1.5))

                    # Clear old emails BEFORE triggering send
                    from src.utils.gmail_mfa import clear_old_mfa_emails, fetch_mfa_code
                    clear_old_mfa_emails(gmail_creds['email'], gmail_creds['app_password'])

                    # Click Send code
                    try:
                        page.click('text=Send code', timeout=5000)
                        logger.info("✅ Clicked 'Send code'")
                    except Exception:
                        page.evaluate('''() => {
                            const form = document.querySelector('form');
                            if (form) { form.submit(); }
                        }''')
                        logger.info("✅ Form submitted")

                    time.sleep(3)

                    # Fetch fresh code from Gmail
                    mfa_code = fetch_mfa_code(
                        gmail_user=gmail_creds['email'],
                        gmail_app_password=gmail_creds['app_password'],
                        max_wait_seconds=90,
                        poll_interval=3
                    )

                    if mfa_code:
                        logger.info(f"✅ MFA code: {mfa_code}")
                        code_field = page.query_selector(
                            'input[name="pf.challengeResponse"], '
                            'input[type="text"]:not([name="pf.username"])'
                        )
                        if code_field:
                            code_field.fill(mfa_code)
                            logger.info("✅ Code entered")
                            time.sleep(random.uniform(0.5, 1.0))

                            # Debug: log all buttons on the page
                            buttons = page.query_selector_all('button, input[type="submit"], a.button')
                            for b in buttons:
                                b_text = b.inner_text().strip() if b.inner_text() else ''
                                b_type = b.get_attribute('type') or ''
                                logger.info(f"  BTN: '{b_text}' type={b_type}")

                            # Submit verification — try many button labels
                            submitted = False
                            for btn_text in ['Confirm', 'Verify', 'Submit', 'Continue', 'Sign on', 'Next']:
                                try:
                                    page.click(f'text={btn_text}', timeout=2000)
                                    logger.info(f"✅ Clicked '{btn_text}'")
                                    submitted = True
                                    break
                                except Exception:
                                    continue

                            if not submitted:
                                # Try any submit button
                                try:
                                    page.click('input[type="submit"], button[type="submit"]', timeout=3000)
                                    logger.info("✅ Clicked submit button")
                                    submitted = True
                                except Exception:
                                    pass

                            if not submitted:
                                page.evaluate('''() => {
                                    const form = document.querySelector('form');
                                    if (form) {
                                        const okField = document.querySelector('input[name="pf.ok"]');
                                        if (okField) okField.value = "clicked";
                                        form.submit();
                                    }
                                }''')
                                logger.info("✅ Form submitted (fallback)")

                            time.sleep(4)
                            logger.info(f"After verify — URL: {page.url}")

                            # Handle SSO redirects after MFA
                            for i in range(5):
                                if 'providerwebapp' in page.url:
                                    break
                                if page.title() == 'Sign On':
                                    page.evaluate('''() => {
                                        const form = document.querySelector('form');
                                        if (form) { form.submit(); }
                                    }''')
                                    time.sleep(3)
                                    logger.info(f"  SSO #{i+1} → {page.title()}")
                                else:
                                    break
                    else:
                        logger.error("❌ MFA code not received")

            # ─── Final status ───
            final_parsed = urlparse(page.url)
            if 'providerwebapp' in final_parsed.path:
                logger.info("🎉 LOGIN SUCCESS — Inside Blue Shield Provider Portal!")
            else:
                logger.info(f"Final: {page.title()} | {page.url[:120]}")

            # ─── STEP 4: Navigate to Claims & Apply Filters ───
            claims_url = "https://www.blueshieldca.com/providerwebapp/claims/claimStatus"

            # Handle any intermediate page (e.g. "Remember Me")
            for _ in range(5):
                cur = urlparse(page.url)
                if 'providerwebapp' in cur.path:
                    break
                if 'ping-ext' in page.url:
                    logger.info(f"Intermediate page: {page.title()} — auto-submitting")
                    page.evaluate('''() => {
                        const form = document.querySelector('form');
                        if (form) { form.submit(); }
                    }''')
                    time.sleep(3)

            if 'claimStatus' not in page.url:
                logger.info("Navigating to claims page...")
                page.goto(claims_url, wait_until="domcontentloaded", timeout=60000)
                try:
                    page.wait_for_load_state('networkidle', timeout=20000)
                except Exception:
                    pass
                time.sleep(2)

            # Dismiss cookie banner if present
            try:
                cookie_btn = page.query_selector('button:has-text("Continue"), button:has-text("Accept")')
                if cookie_btn and cookie_btn.is_visible():
                    cookie_btn.click()
                    time.sleep(1)
            except Exception:
                pass

            logger.info("Applying claim filters...")
            time.sleep(random.uniform(1.0, 2.0))

            # 4A: Set Claim Status = "In process" via Angular dropdown
            try:
                # Use JS to find and log all form elements
                form_info = page.evaluate('''() => {
                    const inputs = document.querySelectorAll('input, select, mat-select, textarea');
                    return Array.from(inputs).map((el, i) => ({
                        idx: i,
                        tag: el.tagName,
                        type: el.type || '',
                        name: el.name || '',
                        id: el.id || '',
                        placeholder: el.placeholder || '',
                        ariaLabel: el.getAttribute('aria-label') || '',
                        formcontrolname: el.getAttribute('formcontrolname') || '',
                        value: el.value || ''
                    }));
                }''')
                for fi in form_info:
                    logger.info(f"  FORM[{fi['idx']}]: <{fi['tag']}> type={fi['type']} name={fi['name']} "
                                f"fc={fi['formcontrolname']} placeholder={fi['placeholder']} "
                                f"aria={fi['ariaLabel']} val={fi['value'][:30]}")
            except Exception as e:
                logger.warning(f"Could not enumerate form: {e}")

            # Try to set Claim status dropdown (Angular Material mat-select)
            try:
                # Find mat-select elements via JS
                mat_selects = page.evaluate('''() => {
                    const sels = document.querySelectorAll('mat-select, [role="combobox"]');
                    return Array.from(sels).map((el, i) => ({
                        idx: i,
                        fc: el.getAttribute('formcontrolname') || '',
                        ariaLabel: el.getAttribute('aria-label') || '',
                        text: el.textContent.trim().substring(0, 50)
                    }));
                }''')
                for ms in mat_selects:
                    logger.info(f"  MAT-SEL[{ms['idx']}]: fc={ms['fc']} aria={ms['ariaLabel']} text='{ms['text']}'")

                # Click the Claim status mat-select (it's the one with text "Claim status")
                claim_dropdown = page.evaluate('''() => {
                    const sels = document.querySelectorAll('mat-select, [role="combobox"]');
                    for (let i = 0; i < sels.length; i++) {
                        if (sels[i].textContent.trim().includes('Claim status')) {
                            return i;
                        }
                    }
                    return -1;
                }''')

                if claim_dropdown >= 0:
                    dropdowns = page.query_selector_all('mat-select, [role="combobox"]')
                    dropdowns[claim_dropdown].scroll_into_view_if_needed()
                    dropdowns[claim_dropdown].click()
                    time.sleep(2)

                    # Debug: log all visible mat-option elements
                    options = page.evaluate('''() => {
                        const opts = document.querySelectorAll('mat-option, [role="option"]');
                        return Array.from(opts).map(o => o.textContent.trim());
                    }''')
                    logger.info(f"  Dropdown options: {options}")

                    # Try selecting "In process"
                    selected = False
                    for selector in [
                        'mat-option:has-text("In process")',
                        '[role="option"]:has-text("In process")',
                        'span:has-text("In process")',
                        'text=In process'
                    ]:
                        try:
                            page.click(selector, timeout=2000)
                            logger.info(f"✅ Claim status = 'In process' (via {selector})")
                            selected = True
                            break
                        except Exception:
                            continue

                    if not selected:
                        # Try JS click
                        page.evaluate('''() => {
                            const opts = document.querySelectorAll('mat-option, [role="option"]');
                            for (const opt of opts) {
                                if (opt.textContent.trim().includes('In process')) {
                                    opt.click();
                                    return true;
                                }
                            }
                            return false;
                        }''')
                        logger.info("✅ Claim status set via JS click")

                    time.sleep(0.5)
                    page.keyboard.press('Escape')
                    time.sleep(0.3)
                else:
                    logger.warning("mat-select for claimStatus not found")
            except Exception as e:
                logger.warning(f"Claim status dropdown: {e}")

            time.sleep(random.uniform(1.0, 2.0))

            # 4B: Set Dates of Service = last 48 hours
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(hours=48)
            start_str = start_date.strftime("%m/%d/%Y")
            end_str = end_date.strftime("%m/%d/%Y")
            logger.info(f"Setting date range: {start_str} to {end_str}")

            try:
                # Find ALL input elements and filter by matInput or date-like attributes
                all_inputs = page.query_selector_all('input')
                date_inputs = []
                for inp in all_inputs:
                    placeholder = inp.get_attribute('placeholder') or ''
                    aria = inp.get_attribute('aria-label') or ''
                    fc = inp.get_attribute('formcontrolname') or ''
                    mat = inp.get_attribute('matinput')
                    if any(kw in placeholder.lower() for kw in ['start', 'end', 'date']) or \
                       any(kw in aria.lower() for kw in ['start', 'end', 'date']) or \
                       any(kw in fc.lower() for kw in ['start', 'end', 'date', 'from', 'to']):
                        date_inputs.append(inp)
                        logger.info(f"  DATE input: placeholder='{placeholder}' fc='{fc}' aria='{aria}'")

                if len(date_inputs) >= 4:
                    # Second pair = Status/payment date (index 2,3)
                    si, ei = date_inputs[2], date_inputs[3]
                    logger.info("Using Status/payment date fields (2nd pair)")
                elif len(date_inputs) >= 2:
                    si, ei = date_inputs[0], date_inputs[1]
                    logger.info("Using first date pair (only 2 found)")
                else:
                    si, ei = None, None
                    logger.warning(f"Found {len(date_inputs)} date inputs (need 2+)")

                if si and ei:

                    si.click()
                    time.sleep(0.3)
                    page.keyboard.press('Control+a')
                    page.keyboard.press('Backspace')
                    page.keyboard.type(start_str, delay=random.uniform(30, 60))
                    page.keyboard.press('Tab')
                    logger.info(f"✅ Start date = {start_str}")
                    time.sleep(0.5)

                    ei.click()
                    time.sleep(0.3)
                    page.keyboard.press('Control+a')
                    page.keyboard.press('Backspace')
                    page.keyboard.type(end_str, delay=random.uniform(30, 60))
                    page.keyboard.press('Tab')
                    logger.info(f"✅ End date = {end_str}")
                    time.sleep(0.5)
            except Exception as e:
                logger.warning(f"Could not set date range: {e}")

            time.sleep(random.uniform(1.0, 2.0))

            # 4C: Click Search button
            try:
                page.click('button:has-text("Search")', timeout=5000)
                logger.info("✅ Clicked Search")
            except Exception:
                try:
                    page.click('button[type="submit"]', timeout=5000)
                    logger.info("✅ Clicked submit")
                except Exception as e:
                    logger.warning(f"Search button not found: {e}")

            # Wait for results to load
            time.sleep(random.uniform(5, 8))
            logger.info(f"After search — Title: {page.title()}")

            # 4D: Click "Show more claims" to load all results
            show_more_count = 0
            while True:
                try:
                    show_more = page.query_selector('button:has-text("Show more claims")')
                    if show_more and show_more.is_visible():
                        time.sleep(random.uniform(1.5, 3.0))
                        show_more.click()
                        show_more_count += 1
                        logger.info(f"✅ Show more claims #{show_more_count}")
                        time.sleep(random.uniform(3, 5))
                    else:
                        break
                except Exception:
                    break

            if show_more_count > 0:
                logger.info(f"Loaded {show_more_count} additional pages of claims")

            # 4E: Scrape claim results from the page
            try:
                # First: take a snapshot of what the page shows
                page_text = page.inner_text('body')
                logger.info(f"Page text length: {len(page_text)}")

                # Find the "Showing 1–200 of 200 claims" count
                import re
                # Try "of X claims" first (e.g. "Showing 1–200 of 200 claims")
                show_match = re.search(r'of\s+(\d+)\s+claims?', page_text)
                if not show_match:
                    show_match = re.search(r'Showing\s+(\d+)\s+claims?', page_text)
                total_shown = int(show_match.group(1)) if show_match else 0
                logger.info(f"BS Portal shows: {total_shown} claims")

                # Check for no results
                no_results = "\u2019t find any claims" in page_text or "couldn't find" in page_text

                # Find the claims section — between "Showing X claims:" and footer
                claims_text = ''
                show_idx = page_text.find('Showing')
                if show_idx > -1:
                    claims_text = page_text[show_idx:]
                elif not no_results:
                    hide_idx = page_text.find('Hide search')
                    if hide_idx > -1:
                        claims_text = page_text[hide_idx:]

                logger.info(f"Claims text length: {len(claims_text)}")
                # Trim footer
                for footer_marker in ['PROVIDER TOOLS', 'QUICK LINKS', '© California']:
                    footer_idx = claims_text.find(footer_marker)
                    if footer_idx > 0:
                        claims_text = claims_text[:footer_idx]
                        break
                if claims_text:
                    logger.info(f"Claims text (first 800): {claims_text[:800]}")

                # Also discover custom Angular tags
                dom_info = page.evaluate('''() => {
                    const tags = {};
                    document.querySelectorAll('*').forEach(el => {
                        const t = el.tagName.toLowerCase();
                        if (t.startsWith('app-') || t.startsWith('mat-') || t.startsWith('cdk-'))
                            tags[t] = (tags[t] || 0) + 1;
                    });
                    const btns = Array.from(document.querySelectorAll('button'))
                        .map(b => b.textContent.trim())
                        .filter(t => t.length > 0 && t.length < 50);
                    return { custom_tags: tags, buttons: btns.slice(0, 20) };
                }''')
                logger.info(f"Buttons: {dom_info.get('buttons', [])}")
                logger.info(f"Custom tags: {dom_info.get('custom_tags', {})}")

                # Parse individual claims from text lines
                # Structure: every claim = 3 lines: STATUS, DATE, DATA(tab-separated)
                # Headers: Claim#, Type, DOS, Check/EFT, Patient, MemberID, Provider, Billed, Paid, PatientResp
                claim_lines = []
                if claims_text:
                    lines = claims_text.split('\n')
                    raw_lines = [l.strip() for l in lines if l.strip() and len(l.strip()) > 3]

                    # Skip header lines (first ~14 lines until we hit "IN PROCESS")
                    start_idx = 0
                    for i, line in enumerate(raw_lines):
                        if line.upper() == 'IN PROCESS' or line.upper() == 'FINALIZED':
                            start_idx = i
                            break

                    # Parse claims in groups
                    i = start_idx
                    while i < len(raw_lines) - 1:
                        status_line = raw_lines[i].strip()
                        if status_line.upper() not in ('IN PROCESS', 'FINALIZED', 'FINALIZED-DENIED'):
                            i += 1
                            continue

                        date_line = raw_lines[i + 1].strip() if i + 1 < len(raw_lines) else ''

                        # Find the data line (may be i+2 or i+3 if there's an adjustment note)
                        data_line = ''
                        for j in range(i + 2, min(i + 5, len(raw_lines))):
                            if '\t' in raw_lines[j]:
                                data_line = raw_lines[j]
                                i = j + 1
                                break
                        else:
                            i += 2
                            continue

                        # Parse tab-separated data
                        parts = data_line.split('\t')
                        claim = {
                            'status': status_line,
                            'last_modified': date_line,
                            'claim_number': parts[0] if len(parts) > 0 else '',
                            'type': parts[1] if len(parts) > 1 else '',
                            'dos': parts[2] if len(parts) > 2 else '',
                            'check_eft': parts[3] if len(parts) > 3 else '',
                            'patient_name': parts[4] if len(parts) > 4 else '',
                            'member_id': parts[5] if len(parts) > 5 else '',
                            'provider': parts[6] if len(parts) > 6 else '',
                            'amount_billed': parts[7] if len(parts) > 7 else '',
                            'amount_paid': parts[8] if len(parts) > 8 else '',
                            'patient_resp': parts[9] if len(parts) > 9 else ''
                        }
                        claim_lines.append(claim)

                logger.info(f"Parsed {len(claim_lines)} structured claims")

                # Save to JSON for dashboard
                import json as json_mod
                bs_data = {
                    'scraped_at': datetime.now().isoformat(),
                    'filter': {
                        'status': 'In process',
                        'date_from': start_str,
                        'date_to': end_str
                    },
                    'total_shown': total_shown,
                    'no_results': no_results,
                    'claims': claim_lines[:300],
                    'result_text_sample': claims_text[:3000]
                }
                with open('/opt/helixona-agent/bs_claims.json', 'w') as f:
                    json_mod.dump(bs_data, f, indent=2, default=str)
                logger.info(f"✅ Saved {len(claim_lines)} claim lines to bs_claims.json")

            except Exception as e:
                logger.warning(f"Could not scrape claims: {e}")

            logger.info("✅ Claims filters applied and results loaded")

            logger.info("Browser open on noVNC for 60s...")
            time.sleep(60)
            logger.info("STEP 0B complete.")
        except Exception as e:
            logger.error(f"STEP 0B failed: {e}")
            time.sleep(15)
        finally:
            manager.stop()

    # ──────────────────────────────────────
    # STAGE 1: Nightly Bulk ECW Extraction
    # ──────────────────────────────────────
    elif task_type == 'nightly_bulk_extract':
        manager = BrowserManager().start()
        try:
            page = manager.new_page()
            credentials = {"username": "test", "password": "password"}
            if perform_ecw_login(page, credentials):
                processed = process_nightly_bulk_claims(page, aws_client)
                logger.info(f"Stage 1 complete: {len(processed)} claims processed.")
        finally:
            manager.stop()

    # ──────────────────────────────────────
    # STAGE 3: Blue Shield Claim # Capture
    # ──────────────────────────────────────
    elif task_type == 'capture_blueshield_claim':
        manager = BrowserManager().start()
        try:
            page = manager.new_page()
            portal = BlueShieldPortal(page)
            credentials = {"username": "bs_user", "password": "password"}

            if portal.login(credentials):
                claim_id = body.get('claim_id')
                bs_claim_number = portal.capture_claim_number(
                    body.get('subscriber_id', ''),
                    body.get('dos', ''),
                    body.get('dos', ''),
                    body.get('patient_name', '')
                )
                if bs_claim_number:
                    aws_client.update_claim_status(claim_id, {
                        'blueshield_claim_number': bs_claim_number,
                        'current_state': 6
                    })
                    logger.info(f"Stage 3 complete: BS Claim # {bs_claim_number} captured.")
                else:
                    logger.warning(f"Stage 3: Could not capture BS claim # for {claim_id}.")
        finally:
            manager.stop()

    # ──────────────────────────────────────
    # STAGE 4: AI Medical Record Verification
    # ──────────────────────────────────────
    elif task_type == 'verify_medical_record':
        claim_id = body.get('claim_id')
        medical_record_text = body.get('medical_record_text', '')
        dos = body.get('dos', '')

        verifier = IVVerificationService(aws_client)
        verified = verifier.verify_iv_prescription(medical_record_text, dos)

        aws_client.update_claim_status(claim_id, {
            'medical_record_verified': verified,
            'current_state': 8 if verified else 7
        })

        # Rule 6 check
        result = rules.rule_6_ai_verification_gate({'medical_record_verified': verified})
        if result['action'] != 'passed':
            # Create exception task
            task_id = str(uuid.uuid4())
            table = aws_client.dynamodb.Table('helixona-tasks')
            table.put_item(Item={
                'task_id': task_id,
                'claim_id': claim_id,
                'task_type': 'Exception Review',
                'status': 'Open',
                'owner': 'billing_team',
                'notes': 'AI could not verify IV prescription in medical record'
            })
            logger.warning(f"Stage 4: Verification failed. Exception task {task_id} created.")
        else:
            logger.info(f"Stage 4 complete: Medical record verified for {claim_id}.")

    # ──────────────────────────────────────
    # STAGE 5: Cover Letter Generation
    # ──────────────────────────────────────
    elif task_type == 'generate_cover_letter':
        claim_id = body.get('claim_id')
        claim_data = body.get('claim_data', {})
        claim_data['claim_id'] = claim_id

        # Pre-submission rules check (Rules 1-6)
        rule_results = rules.validate_claim_pre_submission(claim_data)
        blocked = [r for r in rule_results if r['action'] not in ('passed', 'auto_corrected')]

        if blocked:
            logger.error(f"Stage 5: Blocked by rules {[r['rule'] for r in blocked]}. Cannot generate cover letter.")
        else:
            pdf_path = generate_cover_letter_pdf(claim_data)
            s3_key = f"cover_letters/{claim_id}_{claim_data.get('blueshield_claim_number', 'pending')}.pdf"
            s3_path = aws_client.upload_to_s3(pdf_path, s3_key)

            aws_client.update_claim_status(claim_id, {
                'cover_letter_s3_path': s3_path,
                'current_state': 9
            })
            logger.info(f"Stage 5 complete: Cover letter uploaded to {s3_path}.")

    # ──────────────────────────────────────
    # STAGE 6: SympliSend Full Submission
    # ──────────────────────────────────────
    elif task_type == 'symplisend_submission':
        claim_id = body.get('claim_id')
        submission_data = body.get('submission_data', {})
        file_paths = body.get('file_paths', [])

        # Rule 7 pre-check for Itemization
        if submission_data.get('submission_type') == 'Itemization':
            r7 = rules.rule_7_itemized_bill(submission_data)
            if r7['action'] != 'passed':
                logger.error(f"Stage 6: Rule 7 blocked — {r7['reason']}")
                return

        manager = BrowserManager().start()
        try:
            page = manager.new_page()
            symplisend = SympliSendSubmission(page)
            credentials = {"username": "ss_user", "password": "password"}

            if symplisend.login(credentials):
                aws_client.update_claim_status(claim_id, {'current_state': 10})

                result = symplisend.execute_full_submission(submission_data, file_paths)
                fln_number = result.get('fln_number')

                # Store submission in DynamoDB
                submission_id = str(uuid.uuid4())
                table = aws_client.dynamodb.Table('helixona-submissions')
                table.put_item(Item={
                    'submission_id': submission_id,
                    'claim_id': claim_id,
                    'submission_type': result.get('submission_type', ''),
                    'subscriber_id': submission_data.get('subscriber_id', ''),
                    'status': result.get('status', 'Unknown'),
                    'fln_number': fln_number or 'MISSING',
                    'submitted_at': str(time.time()),
                })

                # Rule 8: FLN# capture check
                r8 = rules.rule_8_fln_capture({'status': 'Completed', 'fln_number': fln_number})
                if r8['action'] != 'passed':
                    logger.error("Stage 6: FLN# NOT captured! Creating alert task.")
                else:
                    aws_client.update_claim_status(claim_id, {
                        'current_state': 11,
                        'fln_number': fln_number
                    })
                    logger.info(f"Stage 6 complete: FLN# {fln_number} captured for {claim_id}.")
        finally:
            manager.stop()

    # ──────────────────────────────────────
    # STAGE 8: Adjudication Processing
    # ──────────────────────────────────────
    elif task_type == 'process_adjudication':
        claim_id = body.get('claim_id')
        eob_data = body.get('eob_data', {})
        result = process_adjudication(claim_id, eob_data, aws_client)
        logger.info(f"Stage 8 complete: {result.get('outcome')} for {claim_id}.")

    elif task_type == 'bs_missing_docs':
        # BlueShield Claims First case (Missing documentation)
        # Single session: login → discover claims → HCFA + Prog Notes for all
        logger.info("═══ BlueShield Missing Documentation — Full Pipeline ═══")
        
        # generate_hcfa now discovers claims from the ECW page directly,
        # upserts them into DynamoDB, and processes all state=1 claims
        # — all in a single browser session with one login.
        step_body = {'task_type': 'generate_hcfa'}
        # Forward testing_mode and test_claim_id flags if present
        if body.get('testing_mode'):
            step_body['testing_mode'] = True
        if body.get('test_claim_id'):
            step_body['test_claim_id'] = body['test_claim_id']
            step_body['testing_mode'] = True
            logger.info(f"🧪 Testing Mode ON — will process only claim {body['test_claim_id']}")
        elif body.get('testing_mode'):
            logger.info("🧪 Testing Mode ON — will process only 1 claim")
        process_message({'Body': json.dumps(step_body)}, aws_client)
        
        logger.info("═══ BlueShield Missing Documentation pipeline complete ═══")

    # ──────────────────────────────────────
    # BLUESHIELD SUBMISSIONS — Upload docs to SympliSend
    # ──────────────────────────────────────
    elif task_type == 'blueshield_submissions':
        logger.info("═══ BlueShield Submissions — Upload Documentation to SympliSend ═══")
        
        test_claim_id = body.get('test_claim_id')
        testing_mode = body.get('testing_mode', False)
        
        # Step 1: Find claims ready for submission
        claims_table = aws_client.dynamodb.Table('helixona-claims')
        try:
            scan_result = claims_table.scan()
            all_claims = scan_result.get('Items', [])
        except Exception as e:
            logger.error(f"Failed to scan claims: {e}")
            return
        
        ready_claims = []
        for item in all_claims:
            cid = item.get('claim_id', '')
            has_hcfa = bool(item.get('hcfa_s3_path'))
            has_prog_notes = bool(item.get('prog_notes_s3_path'))
            has_enc_file = bool(item.get('encounter_file_s3_path'))
            # Require a subscriber id that was CONFIRMED from the HCFA box 1a.
            # A DOM-scrape-only value (subscriber_id_unverified=True) is held back
            # so we never auto-submit an unconfirmed member number to the payer.
            has_subscriber = bool(item.get('subscriber_id')) and not bool(item.get('subscriber_id_unverified'))
            already_submitted = bool(item.get('symplisend_submitted'))
            # Office visits need only HCFA + IV Note — Progress Note not required.
            office = is_office_visit(item.get('cpt'))
            # Manual override: reviewer flagged this claim as not needing a
            # Progress Note (set via dashboard for diagnostic-only / mixed CPT
            # claims that ECW has no encounter doc for). Treated like an office
            # visit for the purposes of the submission filter.
            no_prog_note_manual = bool(item.get('progress_note_not_required'))
            # Reject claims with known document quality issues — even when paths
            # are set, the content may be wrong-patient or stale.
            iv_mismatch = bool(item.get('iv_note_patient_mismatch'))
            enc_needs_review = bool(item.get('encounter_revision_needed'))

            # Submission rule:
            #   1. HCFA always required
            #   2. IV Note (prog_notes_s3_path) required + verified (no mismatch)
            #   3. Progress Note (encounter_file_s3_path) required for IV Therapy,
            #      and the encounter must not be flagged for revision.
            #      Office visits AND manually-flagged claims skip the Progress
            #      Note requirement entirely.
            #   4. Subscriber ID present
            #   5. Not already submitted
            if not (has_hcfa and has_prog_notes and has_subscriber and not already_submitted):
                continue
            if iv_mismatch:
                continue  # IV Note content known-bad
            if office or no_prog_note_manual:
                # Office visit or manual override: HCFA + IV Note is enough.
                ready_claims.append(item)
            else:
                # IV Therapy: must have a verified Progress Note.
                if has_enc_file and not enc_needs_review:
                    ready_claims.append(item)
        
        logger.info(f"📋 {len(ready_claims)} claims ready for submission (out of {len(all_claims)} total)")
        
        # Filter to test claim if specified
        if test_claim_id:
            # For test mode: find the claim regardless of submitted status
            test_item = claims_table.get_item(Key={'claim_id': str(test_claim_id)}).get('Item', {})
            if test_item:
                has_hcfa = bool(test_item.get('hcfa_s3_path'))
                has_prog_notes = bool(test_item.get('prog_notes_s3_path'))
                has_enc_file = bool(test_item.get('encounter_file_s3_path'))
                has_subscriber = bool(test_item.get('subscriber_id')) and not bool(test_item.get('subscriber_id_unverified'))
                office = is_office_visit(test_item.get('cpt'))
                no_prog_note_manual = bool(test_item.get('progress_note_not_required'))
                iv_mismatch = bool(test_item.get('iv_note_patient_mismatch'))
                enc_needs_review = bool(test_item.get('encounter_revision_needed'))

                # Same submission rule as the main loop above.
                ok = has_hcfa and has_prog_notes and has_subscriber and not iv_mismatch
                if office or no_prog_note_manual:
                    ok = ok  # office / manual override: HCFA + IV Note enough
                else:
                    ok = ok and has_enc_file and not enc_needs_review

                if ok:
                    if test_item.get('symplisend_submitted'):
                        logger.info(f"🧪 Claim {test_claim_id} already submitted to SympliSend, skipping")
                        return
                    ready_claims = [test_item]
                    logger.info(f"🧪 Forced test claim {test_claim_id}")
                else:
                    missing = []
                    if not has_hcfa: missing.append('HCFA')
                    if not has_prog_notes: missing.append('IV Note')
                    if iv_mismatch: missing.append('IV Note has patient mismatch')
                    if not office and not has_enc_file: missing.append('Progress Note')
                    if not office and enc_needs_review: missing.append('Progress Note needs review')
                    if not has_subscriber:
                        if test_item.get('subscriber_id') and test_item.get('subscriber_id_unverified'):
                            missing.append('Subscriber ID UNVERIFIED (HCFA box 1a not read)')
                        else:
                            missing.append('Subscriber ID')
                    logger.warning(f"🧪 Claim {test_claim_id} not submittable: {', '.join(missing)}")
                    return
            else:
                logger.warning(f"🧪 Claim {test_claim_id} not found in DynamoDB")
                return
        
        if not ready_claims:
            logger.info("No claims ready for submission")
            return
        
        # Step 2: Login to Blue Shield and navigate to SympliSend
        manager = BrowserManager().start()
        try:
            page = manager.new_page()
            creds = aws_client.get_secret("blueshield_credentials")
            gmail_creds = aws_client.get_secret("gmail_credentials")
            
            # ─── Login to Blue Shield Portal ───
            target_url = "https://www.blueshieldca.com/providerwebapp/claims/claimStatus"
            logger.info(f"Navigating to: {target_url}")
            page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
            
            try:
                page.wait_for_load_state('networkidle', timeout=20000)
            except Exception:
                pass
            time.sleep(1)
            
            # Dismiss cookie consent
            try:
                cookie_btn = page.query_selector('button:has-text("Continue"), button:has-text("Accept"), button:has-text("OK")')
                if cookie_btn and cookie_btn.is_visible():
                    cookie_btn.click()
                    time.sleep(1)
            except Exception:
                pass
            
            from urllib.parse import urlparse
            current_url = page.url
            parsed = urlparse(current_url)
            is_on_portal = 'providerwebapp' in parsed.path
            
            if is_on_portal:
                logger.info("🎉 Already logged in via persistent session!")
            else:
                logger.info(f"Login required. Page: {page.title()}")
                
                # Wait for login form
                try:
                    page.wait_for_selector(
                        'input[name="pf.username"], input#username, '
                        'input[placeholder="Username"], input[placeholder="Password"]',
                        state='visible', timeout=30000
                    )
                except Exception:
                    pass
                time.sleep(random.uniform(1.0, 2.0))
                
                # Dismiss cookie banner again
                try:
                    cookie_btn = page.query_selector('button:has-text("Continue"), button:has-text("Accept")')
                    if cookie_btn and cookie_btn.is_visible():
                        cookie_btn.click()
                        time.sleep(1)
                except Exception:
                    pass
                
                # Enter username
                username_field = (
                    page.query_selector('input[name="pf.username"]') or
                    page.query_selector('input#username') or
                    page.query_selector('input[placeholder="Username"]')
                )
                if username_field:
                    username_field.click()
                    time.sleep(0.2)
                    page.keyboard.press('Control+a')
                    page.keyboard.press('Backspace')
                    time.sleep(random.uniform(0.3, 0.8))
                    for char in creds['username']:
                        page.keyboard.type(char, delay=random.uniform(40, 120))
                    logger.info("✅ Username entered")
                else:
                    logger.error("❌ Username field not found")
                
                time.sleep(random.uniform(0.5, 1.2))
                
                # Enter password
                password_field = (
                    page.query_selector('input[type="password"]') or
                    page.query_selector('input[name="pf.pass"]')
                )
                if password_field:
                    password_field.click()
                    time.sleep(0.2)
                    page.keyboard.press('Control+a')
                    page.keyboard.press('Backspace')
                    time.sleep(random.uniform(0.3, 0.7))
                    for char in creds['password']:
                        page.keyboard.type(char, delay=random.uniform(40, 100))
                    logger.info("✅ Password entered")
                else:
                    logger.error("❌ Password field not found")
                
                time.sleep(random.uniform(0.8, 2.0))
                
                # Submit login
                try:
                    page.click('button:has-text("Log in")', timeout=3000)
                    logger.info("✅ Clicked 'Log in'")
                except Exception:
                    try:
                        page.click('button:has-text("Sign in"), input[type="submit"]', timeout=3000)
                        logger.info("✅ Clicked Sign In")
                    except Exception:
                        page.evaluate('''() => {
                            const form = document.querySelector('form');
                            if (form) {
                                const okField = document.querySelector('input[name="pf.ok"]');
                                if (okField) okField.value = "clicked";
                                form.submit();
                            }
                        }''')
                        logger.info("✅ Form submitted (fallback)")
                time.sleep(random.uniform(2, 4))
                
                # Handle SSO redirects
                for sso_attempt in range(8):
                    current_url = page.url
                    if 'providerwebapp' in urlparse(current_url).path:
                        logger.info("🎉 Reached portal after SSO!")
                        break
                    page_text = page.inner_text('body').lower()
                    if '2-step' in page_text or 'authentication method' in page_text or 'email at' in page_text:
                        logger.info("🔐 MFA page detected")
                        break
                    
                    has_login_form = page.query_selector('input[name="pf.username"], input#username')
                    if has_login_form and has_login_form.is_visible():
                        try:
                            page.click('button:has-text("Sign in")', timeout=3000)
                        except Exception:
                            page.keyboard.press('Enter')
                        time.sleep(random.uniform(2, 4))
                        continue
                    
                    if page.title() == 'Sign On' or 'authorization.ping' in current_url:
                        page.evaluate('''() => {
                            const form = document.querySelector('form');
                            if (form) {
                                const okField = document.querySelector('input[name="pf.ok"]');
                                if (okField) okField.value = "clicked";
                                form.submit();
                            }
                        }''')
                        time.sleep(random.uniform(2, 4))
                        continue
                    break
                
                # Handle MFA if needed
                page_text = page.inner_text('body').lower()
                if '2-step' in page_text or 'authentication method' in page_text or 'email at' in page_text:
                    logger.info("🔐 MFA required")
                    time.sleep(random.uniform(1.0, 2.0))
                    
                    try:
                        page.click('text=Email at', timeout=5000)
                        logger.info("✅ Selected Email method")
                    except Exception:
                        radio = page.query_selector('input[type="radio"]')
                        if radio:
                            radio.click()
                    
                    time.sleep(random.uniform(0.5, 1.5))
                    
                    from src.utils.gmail_mfa import clear_old_mfa_emails, fetch_mfa_code
                    clear_old_mfa_emails(gmail_creds['email'], gmail_creds['app_password'])
                    
                    try:
                        page.click('text=Send code', timeout=5000)
                        logger.info("✅ Clicked 'Send code'")
                    except Exception:
                        page.evaluate('() => { const f = document.querySelector("form"); if (f) f.submit(); }')
                    
                    time.sleep(3)
                    
                    mfa_code = fetch_mfa_code(
                        gmail_user=gmail_creds['email'],
                        gmail_app_password=gmail_creds['app_password'],
                        max_wait_seconds=90,
                        poll_interval=3
                    )
                    
                    if mfa_code:
                        logger.info(f"✅ MFA code: {mfa_code}")
                        code_field = page.query_selector(
                            'input[name="pf.challengeResponse"], '
                            'input[type="text"]:not([name="pf.username"])'
                        )
                        if code_field:
                            code_field.fill(mfa_code)
                            time.sleep(random.uniform(0.5, 1.0))
                            
                            for btn_text in ['Confirm', 'Verify', 'Submit', 'Continue', 'Sign on', 'Next']:
                                try:
                                    page.click(f'text={btn_text}', timeout=2000)
                                    logger.info(f"✅ Clicked '{btn_text}'")
                                    break
                                except Exception:
                                    continue
                            
                            time.sleep(4)
                            
                            # Handle SSO redirects after MFA
                            for i in range(5):
                                if 'providerwebapp' in page.url:
                                    break
                                if page.title() == 'Sign On':
                                    page.evaluate('() => { const f = document.querySelector("form"); if (f) f.submit(); }')
                                    time.sleep(3)
                                else:
                                    break
                    else:
                        logger.error("❌ MFA code not received")
                        return
            # Handle intermediate pages (e.g. "Remember Me", "Trust this device")
            for _attempt in range(8):
                cur_url = page.url
                cur_parsed = urlparse(cur_url)
                
                if 'providerwebapp' in cur_parsed.path:
                    logger.info("🎉 Reached portal!")
                    break
                
                # Handle ping-ext / SSO pass-through pages
                if 'ping-ext' in cur_url or page.title() == 'Sign On':
                    logger.info(f"Intermediate page: {page.title()} — auto-submitting")
                    page.evaluate('''() => {
                        const form = document.querySelector('form');
                        if (form) {
                            const okField = document.querySelector('input[name="pf.ok"]');
                            if (okField) okField.value = "clicked";
                            form.submit();
                        }
                    }''')
                    time.sleep(3)
                    continue
                
                # Handle "Remember me" / "Trust device" / "Stay signed in" pages
                page_text = page.inner_text('body').lower()
                if any(kw in page_text for kw in ['remember', 'trust', 'stay signed', '30 days', 'keep me']):
                    logger.info(f"🔐 Trust/Remember page detected: {page.title()}")
                    # Try clicking Yes/Accept/Continue/Submit
                    clicked = False
                    for btn_text in ['Yes', 'Accept', 'Continue', 'OK', 'Submit', "Don't ask again"]:
                        try:
                            page.click(f'text={btn_text}', timeout=2000)
                            logger.info(f"✅ Clicked '{btn_text}' on trust page")
                            clicked = True
                            break
                        except Exception:
                            continue
                    if not clicked:
                        # Fallback: submit any form
                        page.evaluate('''() => {
                            const form = document.querySelector('form');
                            if (form) form.submit();
                        }''')
                        logger.info("✅ Auto-submitted trust page form")
                    time.sleep(3)
                    continue
                
                # Unknown page — log and try submitting form
                logger.info(f"Unknown intermediate page: {page.title()[:50]} | {cur_url[:80]}")
                try:
                    page.evaluate('() => { const f = document.querySelector("form"); if (f) f.submit(); }')
                except Exception:
                    pass
                time.sleep(3)
            
            # Verify we're logged in
            final_parsed = urlparse(page.url)
            if 'providerwebapp' not in final_parsed.path:
                # Try navigating directly
                page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
                time.sleep(5)
                if 'providerwebapp' not in urlparse(page.url).path:
                    logger.error("❌ Could not login to Blue Shield portal")
                    return
            
            logger.info("🎉 LOGIN SUCCESS — Blue Shield Portal")
            
            # ─── Navigate to SympliSend via SSO ───
            logger.info("📂 Navigating to SympliSend...")
            page.goto("https://www.blueshieldca.com/en/provider/claims/how-to-submit", 
                      wait_until="domcontentloaded", timeout=60000)
            time.sleep(3)
            
            # Click the "SympliSend" link
            symplisend_clicked = False
            try:
                # Listen for new tab/popup (the link has target="_blank")
                with page.context.expect_page(timeout=30000) as new_page_info:
                    page.click('a[title="SympliSend tool"], a:has-text("SympliSend")', timeout=10000)
                    logger.info("✅ Clicked SympliSend link")
                
                symplisend_page = new_page_info.value
                symplisend_page.wait_for_load_state('domcontentloaded', timeout=60000)
                symplisend_clicked = True
            except Exception as e:
                logger.warning(f"SympliSend link click failed: {e}")
                # Fallback: try direct SSO URL
                try:
                    page.goto("https://www.blueshieldca.com/providerwebapp/externalSSO?partnerId=FirstSource&parentPage=howtoSubmit&tab=same",
                             wait_until="domcontentloaded", timeout=60000)
                    symplisend_page = page
                    symplisend_clicked = True
                except Exception as e2:
                    logger.error(f"Could not reach SympliSend: {e2}")
                    return
            
            # Wait for SympliSend dashboard to load (SSO redirect takes time)
            logger.info(f"📂 SympliSend page: {symplisend_page.url[:100]}")
            time.sleep(5)
            
            # Wait for the dashboard to appear
            for _wait in range(10):
                if 'symplisendbscprovider' in symplisend_page.url or 'dashboard' in symplisend_page.url:
                    break
                time.sleep(3)
                logger.info(f"  Waiting for SympliSend redirect... URL: {symplisend_page.url[:80]}")
            
            logger.info(f"✅ SympliSend Dashboard: {symplisend_page.url[:100]}")
            
            try:
                symplisend_page.screenshot(path='/tmp/symplisend_dashboard.png')
                logger.info("📸 SympliSend dashboard screenshot saved")
            except Exception:
                pass
            
            # ─── Step 3: Process each claim ───
            submission_count = 0
            for claim_item in ready_claims:
                claim_id = claim_item.get('claim_id', '')
                submission_type = claim_item.get('submission_type', 'First Time Submission')
                subscriber_id = claim_item.get('subscriber_id', '')
                
                logger.info(f"═══ Submitting claim {claim_id}: {submission_type} | Subscriber: {subscriber_id} ═══")
                
                logger.info(f"  Submission type: {submission_type}")

                # Audit state, set up before the try so the failure handler can
                # always report what had been gathered when things went wrong.
                # (label, s3_path, local_path) for every document that will
                # actually be handed to the uploader — this is what the audit
                # log itemises.
                audit_docs = []
                # What we ended up selecting in the SympliSend form. Recorded
                # as-observed so the log can expose a mismatch with the claim's
                # own submission_type.
                form_type_selected = ''

                try:
                    # Download the PDFs from S3 to /tmp/
                    # Encounter files may be multiple (one per encounter visit on the same date)
                    import re as _re
                    s3_bucket = None
                    files_to_upload = []

                    # Build the list of (doc_name, s3_path) tuples to download.
                    # Mirror the dashboard exactly: HCFA + IV Note + ONE Progress
                    # Note (the primary `encounter_file_s3_path`). The dashboard's
                    # link uses this single path, so the submission should too —
                    # uploading every entry in `encounter_files_s3_paths` causes
                    # duplicates when ECW lists the same encounter under multiple
                    # visit-type columns (claim 5425 had 4 dup PDFs → 6 files).
                    docs_to_download = [
                        ('hcfa', claim_item.get('hcfa_s3_path', '')),
                        ('prog_notes', claim_item.get('prog_notes_s3_path', '')),
                    ]
                    single_enc = claim_item.get('encounter_file_s3_path', '')
                    if single_enc:
                        docs_to_download.append(('encounter', single_enc))

                    for doc_name, s3_path in docs_to_download:
                        if not s3_path:
                            logger.warning(f"  Missing {doc_name} for claim {claim_id}")
                            continue
                        match = _re.match(r's3://([^/]+)/(.+)', s3_path)
                        if match:
                            bucket = match.group(1)
                            key = match.group(2)
                            local_path = f'/tmp/{claim_id}_{doc_name}.pdf'
                            try:
                                aws_client.s3.download_file(bucket, key, local_path)
                                files_to_upload.append(local_path)
                                audit_docs.append((doc_name, s3_path, local_path))
                                logger.info(f"  ✅ Downloaded {doc_name}: {os.path.getsize(local_path)} bytes")
                            except Exception as e:
                                logger.error(f"  ❌ Failed to download {doc_name}: {e}")
                        else:
                            logger.warning(f"  Invalid S3 path: {s3_path}")

                    # 2 files (HCFA + IV Note) is enough for office visits AND
                    # for claims the reviewer manually flagged as not needing a
                    # Progress Note. Otherwise the encounter PDF is required.
                    _no_pn_required = is_office_visit(claim_item.get('cpt')) \
                                      or bool(claim_item.get('progress_note_not_required'))
                    _min_files = 2 if _no_pn_required else 3
                    if len(files_to_upload) < _min_files:
                        logger.warning(f"  Only {len(files_to_upload)}/{_min_files} files downloaded, skipping submission")
                        record_submission(
                            aws_client, claim_item,
                            outcome=OUTCOME_BLOCKED,
                            documents=describe_documents(audit_docs),
                            error=f'only {len(files_to_upload)}/{_min_files} documents available',
                            notes='nothing was uploaded to the payer',
                        )
                        continue
                    
                    # Click "New Submission"
                    try:
                        symplisend_page.click('button[name="addNewSubmission"], button:has-text("New Submission")', timeout=10000)
                        logger.info("✅ Clicked 'New Submission'")
                    except Exception as e:
                        logger.error(f"❌ Could not click New Submission: {e}")
                        try:
                            symplisend_page.screenshot(path=f'/tmp/symplisend_no_btn_{claim_id}.png')
                        except Exception:
                            pass
                        record_submission(
                            aws_client, claim_item,
                            outcome=OUTCOME_FAILED,
                            documents=describe_documents(audit_docs),
                            error=f'could not open New Submission form: {e}',
                            notes='nothing was uploaded to the payer',
                        )
                        continue
                    
                    # Wait for Angular to finish re-rendering the form after New Submission
                    time.sleep(3)

                    # ─── Set Submission Type dropdown to "Provider First Submission Claim" ───
                    try:
                        sel_result = symplisend_page.evaluate('''() => {
                            const target = 'Provider First Submission Claim';
                            const selects = document.querySelectorAll('select');
                            for (const sel of selects) {
                                let matchOpt = null;
                                for (const opt of sel.options) {
                                    if ((opt.text || '').trim().toLowerCase() === target.toLowerCase()) {
                                        matchOpt = opt;
                                        break;
                                    }
                                }
                                if (matchOpt) {
                                    sel.value = matchOpt.value;
                                    sel.dispatchEvent(new Event('change', { bubbles: true }));
                                    sel.dispatchEvent(new Event('input', { bubbles: true }));
                                    // Angular ng-model sync
                                    if (window.angular) {
                                        try {
                                            const $scope = window.angular.element(sel).scope();
                                            const ngModel = sel.getAttribute('ng-model');
                                            if ($scope && ngModel) {
                                                $scope.$apply(() => { $scope[ngModel.split('.').pop()] = matchOpt.value; });
                                            }
                                        } catch (e) {}
                                    }
                                    return { ok: true, value: matchOpt.value, text: matchOpt.text };
                                }
                            }
                            return { ok: false };
                        }''')
                        if sel_result and sel_result.get('ok'):
                            form_type_selected = (sel_result.get('text') or 'Provider First Submission Claim').strip()
                            logger.info(f"✅ Submission Type set to 'Provider First Submission Claim' (value={sel_result.get('value')})")
                        else:
                            # Fallback: Playwright's select_option by visible label
                            try:
                                symplisend_page.select_option('select', label='Provider First Submission Claim', timeout=5000)
                                form_type_selected = 'Provider First Submission Claim'
                                logger.info("✅ Submission Type set via select_option(label=...)")
                            except Exception as e_sel:
                                logger.warning(f"⚠️ Could not set Submission Type dropdown: {e_sel}")
                                try:
                                    symplisend_page.screenshot(path=f'/tmp/symplisend_no_subtype_{claim_id}.png')
                                except Exception:
                                    pass
                        time.sleep(1)
                    except Exception as e:
                        logger.warning(f"Submission Type dropdown set failed: {e}")

                    # Enter Subscriber ID using page-level methods (avoids stale element references)
                    try:
                        # Wait for the input to appear and be stable
                        symplisend_page.wait_for_selector('input.form-control', state='visible', timeout=10000)
                        time.sleep(0.3)
                        
                        # Use page.click + page.fill (re-queries DOM each call, no detachment)
                        symplisend_page.click('input.form-control')
                        time.sleep(0.3)
                        symplisend_page.fill('input.form-control', subscriber_id)
                        time.sleep(0.3)
                        symplisend_page.keyboard.press('Tab')
                        time.sleep(0.5)
                        logger.info(f"✅ Entered Subscriber ID: {subscriber_id}")
                    except Exception as e:
                        logger.warning(f"Subscriber ID fill failed ({e}), trying type()...")
                        try:
                            symplisend_page.click('input.form-control', timeout=5000)
                            time.sleep(0.3)
                            symplisend_page.keyboard.press('Control+a')
                            time.sleep(0.2)
                            symplisend_page.keyboard.type(subscriber_id, delay=30)
                            symplisend_page.keyboard.press('Tab')
                            time.sleep(1)
                            logger.info(f"✅ Entered Subscriber ID via type(): {subscriber_id}")
                        except Exception as e2:
                            logger.error(f"❌ Subscriber ID entry failed: {e2}")
                            symplisend_page.screenshot(path=f'/tmp/symplisend_no_subid_{claim_id}.png')
                    
                    time.sleep(1)
                    
                    # Answer "No" to radio button questions
                    # Question 1 is always visible; Question 2 appears conditionally after Q1 is answered
                    total_answered = 0
                    try:
                        for q_round in range(3):  # Up to 3 rounds of questions
                            radio_count = symplisend_page.evaluate('''() => {
                                const radios = document.querySelectorAll('input[type="radio"][value="False"]');
                                let unanswered = 0;
                                for (const r of radios) {
                                    if (!r.checked) unanswered++;
                                }
                                return unanswered;
                            }''')
                            
                            if radio_count == 0:
                                break
                            
                            logger.info(f"📂 Round {q_round+1}: Found {radio_count} unanswered 'No' radio buttons")
                            
                            # Click each unanswered "No" radio
                            symplisend_page.evaluate('''() => {
                                const radios = document.querySelectorAll('input[type="radio"][value="False"]');
                                for (const r of radios) {
                                    if (!r.checked) {
                                        r.click();
                                        r.dispatchEvent(new Event('change', { bubbles: true }));
                                        r.dispatchEvent(new Event('input', { bubbles: true }));
                                    }
                                }
                            }''')
                            total_answered += radio_count
                            
                            # Wait for conditional questions to appear
                            time.sleep(1.5)
                        
                        logger.info(f"✅ Answered 'No' to {total_answered} questions total")
                    except Exception as e:
                        logger.warning(f"Radio button selection failed: {e}")
                    
                    time.sleep(1)
                    
                    # Upload files
                    try:
                        file_input = symplisend_page.query_selector('input#dropFileRef, input[type="file"]')
                        if file_input:
                            file_input.set_input_files(files_to_upload)
                            logger.info(f"✅ Uploaded {len(files_to_upload)} files")
                        else:
                            logger.error("❌ File input not found")
                    except Exception as e:
                        logger.error(f"File upload failed: {e}")
                    
                    time.sleep(1)
                    
                    # Take screenshot before submit
                    try:
                        symplisend_page.screenshot(path=f'/tmp/symplisend_before_submit_{claim_id}.png')
                        logger.info(f"📸 Pre-submit screenshot saved")
                    except Exception:
                        pass
                    
                    # Step 1: Click "Done" button — button[name="donebtn"]
                    try:
                        symplisend_page.wait_for_selector('button[name="donebtn"]', state='visible', timeout=15000)
                        time.sleep(0.5)
                        symplisend_page.click('button[name="donebtn"]', timeout=5000)
                        logger.info("✅ Clicked 'Done' (donebtn)")
                    except Exception as e:
                        logger.warning(f"Done button click failed: {e}, trying JS...")
                        try:
                            symplisend_page.evaluate('''() => {
                                const btn = document.querySelector('button[name="donebtn"]');
                                if (btn) { btn.disabled = false; btn.click(); }
                            }''')
                            logger.info("✅ Clicked 'Done' (JS)")
                        except Exception:
                            pass
                    
                    time.sleep(3)
                    
                    # Step 2: Click "Submit" button — button[name="newsubmit"]
                    done_clicked = False
                    try:
                        symplisend_page.wait_for_selector('button[name="newsubmit"]', state='visible', timeout=15000)
                        time.sleep(0.5)
                        symplisend_page.click('button[name="newsubmit"]', timeout=5000)
                        done_clicked = True
                        logger.info("✅ Clicked 'Submit' (newsubmit)")
                    except Exception as e1:
                        logger.warning(f"Submit button failed: {e1}, trying JS...")
                        try:
                            done_clicked = symplisend_page.evaluate('''() => {
                                const btn = document.querySelector('button[name="newsubmit"]');
                                if (btn) { btn.click(); return true; }
                                return false;
                            }''')
                            if done_clicked:
                                logger.info("✅ Clicked 'Submit' (JS)")
                        except Exception:
                            pass
                    
                    # Wait for page to navigate/load after Submit (confirmation page)
                    try:
                        pre_submit_url = symplisend_page.url
                        logger.info("⏳ Waiting for confirmation page after Submit...")
                        for _wait_conf in range(30):
                            time.sleep(1)
                            current_url = symplisend_page.url
                            if current_url != pre_submit_url:
                                logger.info(f"✅ Page navigated to: {current_url}")
                                time.sleep(2)
                                break
                            # Check if page content changed (submission confirmed)
                            page_text = symplisend_page.evaluate('() => document.body.innerText.substring(0, 500)')
                            if any(kw in (page_text or '').lower() for kw in ['submitted', 'success', 'confirmation', 'fln', 'received', 'reference']):
                                logger.info("✅ Confirmation content detected on page")
                                time.sleep(1)
                                break
                        else:
                            logger.warning("⚠️ No page change detected after 30s, continuing anyway")
                    except Exception as e:
                        logger.warning(f"Wait for confirmation failed: {e}")
                        time.sleep(3)
                    
                    # Capture confirmation screenshot
                    try:
                        symplisend_page.screenshot(path=f'/tmp/symplisend_after_submit_{claim_id}.png')
                        logger.info(f"📸 Post-submit screenshot saved")
                    except Exception:
                        pass
                    
                    # Try to capture FLN or confirmation number
                    fln_number = None
                    try:
                        fln_result = symplisend_page.evaluate('''() => {
                            const body = document.body.innerText || '';
                            // Look for FLN or confirmation number patterns
                            const flnMatch = body.match(/FLN[#:\\s]*(\\S+)/i);
                            if (flnMatch) return { fln: flnMatch[1] };
                            const confMatch = body.match(/(?:confirmation|reference|tracking)[#:\\s]*(\\S+)/i);
                            if (confMatch) return { fln: confMatch[1] };
                            // Look for any long number that appears after submission
                            const numMatch = body.match(/(\\d{8,15})/);
                            if (numMatch) return { fln: numMatch[1], method: 'number-scan' };
                            return null;
                        }''')
                        if fln_result:
                            fln_number = fln_result.get('fln')
                            logger.info(f"✅ Captured FLN/Confirmation: {fln_number}")
                    except Exception:
                        pass
                    
                    # Guard: only proceed if Done was actually clicked
                    if not done_clicked:
                        logger.error(f"❌ Could not submit claim {claim_id} — Done button not found")
                        try:
                            symplisend_page.screenshot(path=f'/tmp/symplisend_failed_{claim_id}.png')
                        except Exception:
                            pass
                        record_submission(
                            aws_client, claim_item,
                            outcome=OUTCOME_FAILED,
                            documents=describe_documents(audit_docs),
                            submission_form_type=form_type_selected,
                            error='Submit button never fired — packet not sent',
                        )
                        continue
                    
                    # Check for error indicators on the page
                    submission_ok = True
                    try:
                        page_text = symplisend_page.inner_text('body').lower()
                        if any(kw in page_text for kw in ['error', 'failed', 'invalid', 'required field']):
                            if not any(kw in page_text for kw in ['success', 'submitted', 'confirmation', 'fln', 'received']):
                                logger.error(f"❌ Page shows error after submit for claim {claim_id}")
                                submission_ok = False
                    except Exception:
                        pass
                    
                    if not submission_ok:
                        logger.error(f"❌ Submission failed for claim {claim_id} — NOT marking as submitted")
                        record_submission(
                            aws_client, claim_item,
                            outcome=OUTCOME_FAILED,
                            documents=describe_documents(audit_docs),
                            submission_form_type=form_type_selected,
                            error='payer form reported an error after submit',
                        )
                        continue

                    # Audit row FIRST: if the claim update below fails, we still
                    # hold dated proof of what went to the payer. The reverse
                    # order would lose the evidence on the more likely failure.
                    record_submission(
                        aws_client, claim_item,
                        outcome=OUTCOME_SUBMITTED,
                        documents=describe_documents(audit_docs),
                        method=METHOD_SYMPLISEND,
                        submission_form_type=form_type_selected,
                        fln=fln_number or '',
                        notes='' if fln_number else 'no FLN acknowledgement captured from the payer',
                    )

                    # Only mark as submitted after confirmed success
                    aws_client.update_claim_status(claim_id, {
                        'symplisend_submitted': True,
                        'symplisend_submitted_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                        'symplisend_fln': fln_number or 'PENDING',
                        'symplisend_submission_type': submission_type,
                        'state': 9,
                    })
                    logger.info(f"✅ Claim {claim_id} submitted to SympliSend")
                    submission_count += 1
                    
                    # Clean up temp files
                    for fp in files_to_upload:
                        try:
                            os.remove(fp)
                        except Exception:
                            pass
                    
                    time.sleep(2)
                    
                    # Testing mode: stop after 1
                    if testing_mode:
                        logger.info("🧪 Testing Mode: stopping after 1 submission")
                        break
                
                except Exception as claim_err:
                    logger.error(f"❌ Submission failed for claim {claim_id}: {claim_err}")
                    record_submission(
                        aws_client, claim_item,
                        outcome=OUTCOME_FAILED,
                        documents=describe_documents(audit_docs),
                        submission_form_type=form_type_selected,
                        error=f'unhandled error during submission: {claim_err}',
                    )
                    continue

            logger.info(f"═══ BlueShield Submissions complete: {submission_count} submitted ═══")
        
        finally:
            manager.stop()
        
        # ─── Step 4: Update claim status in ECW to "Claim sent via Symplisend" ───
        # Get list of claims that were just submitted (symplisend_submitted=True but ecw_status_updated!=True)
        if submission_count > 0:
            logger.info("═══ Updating claim statuses in ECW... ═══")
            
            # Re-scan for claims that need ECW status update
            try:
                scan_result = claims_table.scan()
                submitted_claims = [
                    item for item in scan_result.get('Items', [])
                    if item.get('symplisend_submitted') and not item.get('ecw_status_updated')
                ]
                if test_claim_id:
                    submitted_claims = [c for c in submitted_claims if str(c.get('claim_id', '')) == str(test_claim_id)]
            except Exception as e:
                logger.error(f"Failed to scan for submitted claims: {e}")
                submitted_claims = []
            
            if submitted_claims:
                ecw_manager = BrowserManager().start()
                try:
                    ecw_page = ecw_manager.new_page()
                    ecw_creds = aws_client.get_secret("ecw_credentials")
                    
                    # Login to ECW via the proven Turnstile-aware login flow.
                    # (A bare username/password fill never gets past the Cloudflare
                    # Turnstile captcha, leaving us stranded on the login page.)
                    if not _perform_ecw_login(ecw_page, ecw_creds, aws_client):
                        logger.error("ECW login failed — skipping ECW status updates this run")
                        raise RuntimeError("ECW login failed")

                    logger.info("Navigating to Billing → Claims (hash navigation)...")
                    try:
                        ecw_page.evaluate("window.location.hash = '/mobiledoc/jsp/webemr/webpm/claimLookup.jsp'")
                    except Exception as e:
                        logger.warning(f"Hash navigation failed: {e}")
                    time.sleep(4)
                    for wait_i in range(30):
                        try:
                            if ecw_page.query_selector('text=Claim Status') or ecw_page.query_selector('text=Service Dt'):
                                logger.info(f"✅ Claims form visible after {wait_i}s")
                                break
                        except Exception:
                            pass
                        time.sleep(1)
                    time.sleep(2)

                    for claim_item in submitted_claims:
                        cid = claim_item.get('claim_id', '')
                        logger.info(f"📋 Updating ECW status for claim {cid}...")
                        try:
                            if _set_claim_status_in_ecw(ecw_page, cid):
                                aws_client.update_claim_status(cid, {
                                    'ecw_status_updated': True,
                                    'ecw_status_code': 'Claim sent via Symplisend',
                                    'ecw_status_updated_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                                })
                                logger.info(f"  ✅ Claim {cid} ECW status updated + verified + marked in DynamoDB")
                        except Exception as ecw_err:
                            logger.error(f"  ❌ ECW status update failed for {cid}: {ecw_err}")
                            continue

                    logger.info("═══ ECW status updates complete ═══")

                except RuntimeError as login_err:
                    # Login failed — submissions already succeeded, so don't fail the
                    # whole task (would risk an SQS retry). Claims keep ecw_status_updated
                    # unset and get picked up by the standalone ecw_status_update task.
                    logger.warning(f"Skipping ECW status updates: {login_err}")
                finally:
                    ecw_manager.stop()

    elif task_type == 'ecw_status_update':
        # ─── Standalone ECW Status Update Task ───
        # Updates claim status in ECW from "Ready to Submit to Symplisend" to "Claim sent via Symplisend"
        logger.info("═══ ECW Status Update — Updating claim statuses in ECW ═══")
        
        test_claim_id = body.get('test_claim_id')
        claims_table = aws_client.dynamodb.Table('helixona-claims')
        
        try:
            scan_result = claims_table.scan()
            submitted_claims = [
                item for item in scan_result.get('Items', [])
                if item.get('symplisend_submitted') and not item.get('ecw_status_updated')
            ]
            if test_claim_id:
                submitted_claims = [c for c in submitted_claims if str(c.get('claim_id', '')) == str(test_claim_id)]
        except Exception as e:
            logger.error(f"Failed to scan for submitted claims: {e}")
            submitted_claims = []
        
        if not submitted_claims:
            logger.info("No claims need ECW status update")
            return
        
        logger.info(f"📋 {len(submitted_claims)} claims need ECW status update")
        
        ecw_manager = BrowserManager().start(proxy_config=None)
        try:
            ecw_page = ecw_manager.new_page()
            ecw_creds = aws_client.get_secret("ecw_credentials")
            
            # Login to ECW via the proven Turnstile-aware login flow.
            # (A bare username/password fill never gets past the Cloudflare Turnstile
            # captcha, leaving us stranded on the login page — "loaded after 0s".)
            if not _perform_ecw_login(ecw_page, ecw_creds, aws_client):
                logger.error("ECW login failed — aborting status update (will retry next run)")
                return

            # Navigate to Billing → Claims via direct hash routing (proven reliable;
            # the previous text-click navigation never reached the Claim Lookup screen)
            logger.info("Navigating to Billing → Claims (hash navigation)...")
            try:
                ecw_page.evaluate("window.location.hash = '/mobiledoc/jsp/webemr/webpm/claimLookup.jsp'")
            except Exception as e:
                logger.warning(f"Hash navigation failed: {e}")
            time.sleep(4)

            # Wait for the claims lookup UI to actually render
            for wait_i in range(30):
                try:
                    if ecw_page.query_selector('text=Claim Status') or ecw_page.query_selector('text=Service Dt'):
                        logger.info(f"✅ Claims form visible after {wait_i}s")
                        break
                except Exception:
                    pass
                time.sleep(1)
            time.sleep(2)

            for claim_item in submitted_claims:
                cid = claim_item.get('claim_id', '')
                logger.info(f"📋 Updating ECW status for claim {cid}...")
                try:
                    if _set_claim_status_in_ecw(ecw_page, cid):
                        aws_client.update_claim_status(cid, {
                            'ecw_status_updated': True,
                            'ecw_status_code': 'Claim sent via Symplisend',
                            'ecw_status_updated_at': time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime()),
                        })
                        logger.info(f"  ✅ Claim {cid} ECW status updated + verified + marked in DynamoDB")
                except Exception as ecw_err:
                    logger.error(f"  ❌ ECW status update failed for {cid}: {ecw_err}")
                    continue

            logger.info("═══ ECW status updates complete ═══")

        finally:
            ecw_manager.stop()

    else:
        logger.warning(f"Unknown task type: {task_type}")


def main():
    from src.config import settings as _settings
    role = _settings.bot_role or 'submissions'
    logger.info("╔══════════════════════════════════════════╗")
    logger.info("║  Helixona Billing Agent v1 — Blue Shield ║")
    logger.info("║  Running on AWS EC2 (Cloud Agent)        ║")
    logger.info(f"║  BOT_ROLE = {role:<29}║")
    logger.info("╚══════════════════════════════════════════╝")

    try:
        aws_client = AWSClient()
    except Exception as e:
        logger.error(f"Failed to initialize AWS Client: {e}")
        return

    logger.info("SQS polling loop started. Listening for tasks...")
    idle_count = 0

    while True:
        try:
            messages = aws_client.receive_sqs_messages()
            for message in messages:
                try:
                    process_message(message, aws_client)
                    aws_client.delete_sqs_message(message['ReceiptHandle'])
                except Exception as e:
                    logger.error(f"Error processing message: {e}")

            if not messages:
                idle_count += 1
                if idle_count % 6 == 1:  # Log every ~2 min instead of every 20s
                    logger.info("Agent healthy, waiting for tasks...")

        except KeyboardInterrupt:
            logger.info("Agent shutting down.")
            break
        except Exception as e:
            logger.error(f"Unexpected error in run loop: {e}")
            time.sleep(5)


if __name__ == "__main__":
    main()
