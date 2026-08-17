"""
Helixona Billing Agent — Monitoring Dashboard
Real-time web UI to see what the agent is doing.
"""
import json
import subprocess
from datetime import datetime
from flask import Flask, render_template_string, jsonify, request, send_file
import boto3
import os
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)

session = boto3.Session(
    aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
    aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    region_name=os.environ.get('AWS_REGION', 'us-west-2')
)

dynamodb = session.resource('dynamodb')
sqs = session.client('sqs')
SQS_URL = os.environ['SQS_QUEUE_URL']
SQS_URL_RESUB = os.environ.get('SQS_QUEUE_URL_RESUB', '')
SQS_URL_IV = os.environ.get('SQS_QUEUE_URL_IV', '')
EC2_IP = "54.189.175.233"
KEY_FILE = "infra/helixona-agent-key.pem"

# Per-bot routing: which SQS queue + systemd unit each tab targets.
# Each bot is its own systemd unit with its own X display, noVNC port, SQS
# queue and Chrome profile. Keep this table in step with QUEUE_BY_ROLE in
# src/aws/clients.py — the agent side of the same mapping.
BOT_ROUTING = {
    'submissions': {
        'queue_url': SQS_URL,
        'service': 'helixona-agent',
        'label': 'Blue Shield Submissions',
        'novnc_port': 6080,
    },
    'resubmissions': {
        'queue_url': SQS_URL_RESUB,
        'service': 'helixona-agent-resub',
        'label': 'Blue Shield Resubmissions',
        'novnc_port': 6081,
    },
    'iv_corrections': {
        'queue_url': SQS_URL_IV,
        'service': 'helixona-agent-iv',
        'label': 'IV Fix Coding',
        'novnc_port': 6082,
    },
}


def _bot_from_request():
    """Resolve which bot a request targets. Defaults to 'submissions'."""
    bot = (request.args.get('bot') or '').strip()
    if not bot:
        body = request.get_json(silent=True) or {}
        bot = (body.get('bot') or '').strip()
    if bot not in BOT_ROUTING:
        bot = 'submissions'
    return bot

# ── Internal 16-state mapping (used for DynamoDB) ──
STATE_LABELS = {
    1: "Claims ECW (Ready to Submit)", 2: "HCFA Generated",
    3: "Awaiting BS Claim #", 4: "BS Claim # Captured",
    5: "Medical Record Pulled", 6: "AI Verification Complete",
    7: "Cover Letter Generated", 8: "SympliSend Started",
    9: "FLN# Captured", 10: "Paid", 11: "Reduced / Denied",
    12: "Appeal in Progress", 13: "Closed",
    14: "Under Review", 15: "Resubmitted", 16: "Final"
}

# ── Simplified pipeline stages for the UI ──
# Maps internal states → display stage
PIPELINE_STAGES = {
    "documentation": {"label": "Documentation Completed", "color": "#CDB486", "states": [1, 2, 3, 4, 5, 6, 7]},
    "ready":         {"label": "Ready to Submit to SympliSend", "color": "#E8D5B0", "states": [8]},
    "submitted":     {"label": "Submitted to SympliSend", "color": "#b09968", "states": [9, 10, 13, 16]},
    "revision":      {"label": "Revision Required SympliSend", "color": "#8a7450", "states": [11, 12, 14, 15]},
}

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Helixona Billing Agent — Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;500;600;700&family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#07070a;--bg2:#0c0c12;--panel:#0f0f16;--panel2:#15151e;--card:#14141c;--card2:#1a1a24;
  --bdr:#1f1f2b;--bdr2:#2a2a38;
  --accent:#CDB486;--accent2:#b09968;--accent-glow:rgba(205,180,134,.18);
  --vio:#8b5cf6;--vio2:#a78bfa;--vio3:#c4b5fd;--viod:#6d28d9;
  --success:#10b981;--warning:#f59e0b;--bad:#ef4444;--info:#3b82f6;
  --text-primary:#f1f5f9;--text-secondary:#94a3b8;--text-muted:#64748b;--text-dim:#475569;
  --border:#1f1f2b;
}
*{margin:0;padding:0;box-sizing:border-box}
html,body{background:var(--bg);color:var(--text-primary);font-family:'Inter',sans-serif;min-height:100vh;font-size:13px;-webkit-font-smoothing:antialiased}
a{color:inherit;text-decoration:none}
input,textarea,select,button{font-family:'Inter',sans-serif}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--bdr2);border-radius:3px}

/* ========== LAYOUT ========== */
.app{display:grid;grid-template-columns:240px 1fr;min-height:100vh}
.sb{background:linear-gradient(180deg,#08080c 0%,#0a0a10 100%);border-right:1px solid var(--bdr);padding:18px 14px;display:flex;flex-direction:column;gap:14px;position:sticky;top:0;height:100vh;overflow-y:auto;z-index:50}
.shell{padding:18px 26px 40px;min-width:0}

/* ========== SIDEBAR ========== */
.brand{display:flex;align-items:center;gap:10px;padding:6px 8px 14px;border-bottom:1px solid var(--bdr)}
.brand img{height:30px;opacity:.95}
.brand-t{font-family:'Playfair Display',serif;font-size:15px;font-weight:600;color:var(--accent);line-height:1}
.brand-s{font-size:9px;color:var(--text-muted);text-transform:uppercase;letter-spacing:1.2px;margin-top:3px}

.nav{display:flex;flex-direction:column;gap:2px}
.nav a{display:flex;align-items:center;gap:10px;padding:10px 12px;border-radius:9px;font-size:12px;color:var(--text-secondary);cursor:pointer;transition:all .15s;font-weight:500}
.nav a:hover{background:var(--card);color:var(--text-primary)}
.nav a.on{background:var(--card2);color:var(--text-primary);box-shadow:inset 2px 0 0 var(--accent)}
.nav a .ico{width:16px;text-align:center;font-size:13px;opacity:.85}
.nav-cnt{margin-left:auto;background:var(--bg2);color:var(--text-muted);font-size:9px;padding:2px 7px;border-radius:8px;font-weight:600}
.nav a.on .nav-cnt{background:var(--accent-glow);color:var(--accent)}

.sb-sec{font-size:9px;text-transform:uppercase;letter-spacing:1.5px;color:var(--text-dim);font-weight:600;padding:0 10px;margin-top:8px;display:flex;align-items:center;gap:6px}
.sb-list{display:flex;flex-direction:column;gap:3px;max-height:240px;overflow-y:auto}
.sb-item{display:flex;align-items:center;gap:9px;padding:8px 10px;border-radius:8px;cursor:pointer;transition:all .15s}
.sb-item:hover{background:var(--card)}
.sb-ico{width:26px;height:26px;border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:11px;flex-shrink:0;font-weight:600}
.sb-info{min-width:0;flex:1}
.sb-l{font-size:9px;color:var(--text-muted);text-transform:uppercase;letter-spacing:.6px}
.sb-n{font-size:11px;color:var(--text-primary);font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}

.sb-foot{margin-top:auto;background:linear-gradient(135deg,rgba(239,68,68,.1),transparent);border:1px solid rgba(239,68,68,.22);border-radius:11px;padding:12px;text-align:center}
.sb-foot-t{font-size:11px;font-weight:600;color:var(--bad);margin-bottom:2px}
.sb-foot-s{font-size:9px;color:var(--text-muted)}

/* ========== TOPBAR ========== */
.tb{display:flex;align-items:center;justify-content:space-between;margin-bottom:22px;gap:14px;flex-wrap:wrap}
.tb-l{display:flex;align-items:center;gap:14px}
.acct{display:flex;align-items:center;gap:10px;background:var(--card);border:1px solid var(--bdr);padding:6px 14px 6px 6px;border-radius:30px}
.avatar{width:32px;height:32px;border-radius:50%;background:linear-gradient(135deg,var(--accent),var(--viod));display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px;color:#000}
.acct-i{line-height:1.2}
.acct-h{font-size:10px;color:var(--text-muted);display:flex;align-items:center;gap:6px}
.acct-h .pro{background:var(--accent);color:#000;font-size:8px;font-weight:700;padding:1px 5px;border-radius:4px;letter-spacing:.5px}
.acct-n{font-size:12px;font-weight:600;color:var(--text-primary)}

.tb-r{display:flex;align-items:center;gap:10px}
.icbtn{width:36px;height:36px;border-radius:9px;background:var(--card);border:1px solid var(--bdr);display:flex;align-items:center;justify-content:center;cursor:pointer;color:var(--text-secondary);position:relative;transition:all .2s;font-size:14px}
.icbtn:hover{color:var(--text-primary);border-color:var(--bdr2)}

.btn{padding:9px 16px;border-radius:9px;font-size:12px;font-weight:600;cursor:pointer;border:none;transition:all .2s;display:inline-flex;align-items:center;gap:7px;color:var(--text-primary);background:transparent}
.btn:hover{background:var(--card)}
.btn-refresh{background:transparent;border:1px solid var(--bdr);color:var(--text-secondary);padding:5px 10px;font-size:11px}
.btn-refresh:hover{border-color:var(--accent);color:var(--accent)}
.btn-primary{background:linear-gradient(135deg,var(--accent),var(--accent2));color:#000;width:100%;padding:11px;justify-content:center}
.btn-primary:hover{box-shadow:0 6px 20px rgba(205,180,134,.3)}

.stop-btn{padding:8px 16px;border-radius:9px;font-size:12px;font-weight:600;cursor:pointer;color:var(--bad);border:1px solid rgba(239,68,68,.35);background:rgba(239,68,68,.08);transition:all .2s;display:inline-flex;align-items:center;gap:6px}
.stop-btn:hover{background:rgba(239,68,68,.18);box-shadow:0 6px 20px rgba(239,68,68,.2)}
.stop-btn.stopping{opacity:.6;cursor:wait}

.novnc-link{display:flex;align-items:center;gap:6px;padding:8px 14px;border-radius:9px;font-size:11px;color:var(--text-secondary);background:var(--card);border:1px solid var(--bdr);transition:all .2s}
.novnc-link:hover{border-color:var(--accent);color:var(--accent)}

.status-badge{display:flex;align-items:center;gap:7px;background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.25);padding:6px 12px;border-radius:30px;font-size:11px;color:var(--success);font-weight:500}
.status-dot{width:6px;height:6px;background:var(--success);border-radius:50%;animation:p 2s infinite}
@keyframes p{0%,100%{box-shadow:0 0 0 0 rgba(16,185,129,.4)}50%{box-shadow:0 0 0 5px transparent}}

/* Row currently being processed by the agent */
tr.processing-row{background:rgba(59,130,246,.10) !important;animation:rowPulse 1.6s ease-in-out infinite}
@keyframes rowPulse{0%,100%{box-shadow:inset 3px 0 0 var(--info)}50%{box-shadow:inset 3px 0 0 #93c5fd}}
.proc-badge{display:inline-flex;align-items:center;gap:6px;background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.4);color:var(--info);padding:3px 9px;border-radius:30px;font-size:10px;font-weight:600;white-space:nowrap}
.proc-spin{width:8px;height:8px;border-radius:50%;background:var(--info);box-shadow:0 0 0 0 rgba(59,130,246,.5);animation:p 1.2s infinite}

/* ========== HERO ========== */
.hero{display:grid;grid-template-columns:1fr 1fr 1fr 340px;gap:14px;margin-bottom:18px}
.hero-h{display:flex;align-items:center;justify-content:space-between;grid-column:1/-1;margin-bottom:-2px}
.hero-h-l{display:flex;align-items:center;gap:12px}
.hero-h-t{font-family:'Playfair Display',serif;font-size:28px;font-weight:600;letter-spacing:-.5px}
.hero-h-l .rec{display:flex;align-items:center;gap:6px;color:var(--text-muted);font-size:11px}
.pill-ct{background:var(--card);border:1px solid var(--bdr);padding:4px 10px;border-radius:30px;font-size:10px;color:var(--text-secondary);font-weight:600}
.chip{background:var(--card);border:1px solid var(--bdr);padding:6px 12px;border-radius:8px;font-size:11px;font-weight:500;color:var(--text-secondary);cursor:pointer;display:flex;align-items:center;gap:6px;transition:all .15s}
.chip:hover{color:var(--text-primary);border-color:var(--bdr2)}
.chip .ar{font-size:8px;color:var(--text-muted)}

.bigcard{background:linear-gradient(180deg,var(--panel) 0%,var(--bg2) 100%);border:1px solid var(--bdr);border-radius:16px;padding:18px;position:relative;overflow:hidden;display:flex;flex-direction:column;min-height:200px;transition:all .25s}
.bigcard:hover{border-color:var(--bdr2);transform:translateY(-2px);box-shadow:0 10px 30px rgba(0,0,0,.4)}
.bc-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px}
.bc-id{display:flex;align-items:center;gap:9px}
.bc-icon{width:30px;height:30px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:14px}
.bc-meta{line-height:1.25}
.bc-tag{font-size:9px;color:var(--text-muted);text-transform:uppercase;letter-spacing:.7px;margin-bottom:2px}
.bc-name{font-size:13px;font-weight:600;color:var(--text-primary)}
.bc-arrow{width:28px;height:28px;border-radius:8px;background:var(--card);border:1px solid var(--bdr);display:flex;align-items:center;justify-content:center;color:var(--text-secondary);font-size:11px;cursor:pointer;transition:all .15s}
.bc-arrow:hover{color:var(--text-primary);transform:rotate(-45deg)}
.bc-lbl{font-size:10px;color:var(--text-muted);text-transform:uppercase;letter-spacing:.8px;margin-bottom:4px}
.bc-num{font-family:'Playfair Display',serif;font-size:36px;font-weight:700;line-height:1.05;letter-spacing:-1px}
.bc-num .pct{font-size:20px;color:var(--text-secondary);font-weight:500;margin-left:2px}
.bc-delta{display:inline-flex;align-items:center;gap:5px;background:rgba(16,185,129,.1);color:var(--success);font-size:11px;font-weight:600;padding:3px 9px;border-radius:30px;margin-top:8px;width:fit-content}
.bc-delta.dn{background:rgba(239,68,68,.1);color:var(--bad)}
.bc-spark{position:absolute;bottom:0;left:0;right:0;height:70px;pointer-events:none}

.promo{background:radial-gradient(120% 100% at 100% 0%,rgba(167,139,250,.28),transparent 60%),linear-gradient(160deg,#1e1531 0%,#0d0a18 100%);border:1px solid rgba(167,139,250,.25);border-radius:16px;padding:20px;position:relative;overflow:hidden;display:flex;flex-direction:column;justify-content:space-between;min-height:200px}
.promo::before{content:'';position:absolute;top:-30px;right:-30px;width:160px;height:160px;background:radial-gradient(circle,rgba(167,139,250,.18),transparent 70%);pointer-events:none}
.promo-top{display:flex;justify-content:space-between;align-items:flex-start;position:relative}
.promo-logo{display:flex;align-items:center;gap:7px;font-family:'Playfair Display',serif;font-size:13px;color:#fff;font-weight:600}
.promo-logo .d{width:14px;height:14px;border-radius:50%;background:linear-gradient(135deg,#fff,#c4b5fd)}
.promo-new{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.18);font-size:9px;color:#fff;padding:3px 8px;border-radius:30px;font-weight:600;letter-spacing:.6px}
.promo-h{font-family:'Playfair Display',serif;font-size:20px;font-weight:600;color:#fff;line-height:1.15;margin-top:10px;position:relative}
.promo-p{font-size:11px;color:rgba(255,255,255,.65);line-height:1.5;margin-top:8px;position:relative}
.promo-btns{display:flex;flex-direction:column;gap:8px;margin-top:14px;position:relative}
.promo-btn{background:rgba(255,255,255,.92);color:#1a0b3a;padding:10px;border-radius:9px;font-weight:600;font-size:12px;text-align:center;cursor:pointer;display:flex;align-items:center;justify-content:center;gap:7px;transition:all .2s;border:none}
.promo-btn:hover{background:#fff;transform:translateY(-1px)}
.promo-btn.alt{background:rgba(255,255,255,.08);color:#fff;border:1px solid rgba(255,255,255,.15)}
.promo-btn.alt:hover{background:rgba(255,255,255,.14)}

/* ========== HERO KPI (submission progress) ========== */
.hero-kpi{background:linear-gradient(135deg,var(--panel) 0%,var(--bg2) 100%);border:1px solid var(--bdr);border-radius:16px;padding:24px 26px;margin-bottom:18px;position:relative;overflow:hidden}
.hero-kpi::after{content:'';position:absolute;top:-50px;right:-50px;width:240px;height:240px;background:radial-gradient(circle,var(--accent-glow),transparent 60%);pointer-events:none}
.hero-kpi-top{display:flex;justify-content:space-between;align-items:flex-end;margin-bottom:16px;flex-wrap:wrap;gap:14px;position:relative}
.hero-kpi-headline{font-family:'Playfair Display',serif;line-height:1}
.hero-num{font-size:68px;font-weight:700;color:var(--accent);letter-spacing:-2.5px;text-shadow:0 0 28px var(--accent-glow)}
.hero-denom{font-size:18px;color:var(--text-muted);font-family:'Inter';font-weight:500;margin-left:8px}
.hero-denom #hero-total{color:var(--text-primary);font-weight:600}
.hero-kpi-pct{font-size:13px;color:var(--text-secondary);text-align:right}
.hero-kpi-pct strong{font-family:'Playfair Display',serif;font-size:22px;color:var(--success);font-weight:600;margin-right:4px}
.hero-progress{position:relative;height:8px;background:var(--card2);border-radius:4px;overflow:hidden}
.hero-progress-fill{height:100%;background:linear-gradient(90deg,var(--accent),var(--success));border-radius:4px;transition:width .5s ease;box-shadow:0 0 18px var(--accent-glow)}

/* ========== PIPELINE ========== */
.pipeline-section{margin-bottom:20px}
.section-title{font-size:13px;font-weight:600;margin:0 0 12px;display:flex;align-items:center;gap:10px;color:var(--text-primary)}
.pipeline-flow{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}
.pipeline-stage{background:linear-gradient(180deg,var(--panel) 0%,var(--bg2) 100%);border:1px solid var(--bdr);border-radius:14px;padding:18px 16px 22px;position:relative;cursor:pointer;transition:all .25s;overflow:hidden}
.pipeline-stage:hover{border-color:var(--bdr2);transform:translateY(-2px);box-shadow:0 10px 24px rgba(0,0,0,.4)}
.stage-count{font-family:'Playfair Display',serif;font-size:42px;font-weight:700;line-height:1;letter-spacing:-1.5px;margin-bottom:6px}
.stage-name{font-size:11px;color:var(--text-secondary);font-weight:500;line-height:1.35;margin-bottom:12px}
.stage-bar{position:absolute;bottom:0;left:0;right:0;height:3px;box-shadow:0 0 16px currentColor}
.stage-details{display:flex;flex-wrap:wrap;gap:4px;margin-top:8px}
.detail-chip{background:var(--card);border:1px solid var(--bdr);font-size:9.5px;padding:3px 7px;border-radius:30px;color:var(--text-secondary);font-weight:500}
.detail-chip.has-count{color:var(--text-primary)}
.stage-arrow{position:absolute;right:-12px;top:50%;transform:translateY(-50%);width:22px;height:22px;background:var(--card);border:1px solid var(--bdr);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;color:var(--text-secondary);z-index:3}

/* ========== SUB-TABS & MAIN GRID ========== */
.subtabs{display:flex;gap:18px;border-bottom:1px solid var(--bdr);padding-bottom:0;margin-bottom:16px;overflow-x:auto}
.subtab{padding:10px 0;font-size:12px;font-weight:600;color:var(--text-muted);cursor:pointer;border-bottom:2px solid transparent;transition:all .15s;white-space:nowrap;display:flex;align-items:center;gap:6px}
.subtab .s{font-size:9px;color:var(--text-dim);font-weight:500;display:block}
.subtab.on{color:var(--text-primary);border-bottom-color:var(--accent)}

.main{display:grid;grid-template-columns:1fr 360px;gap:18px;align-items:start}

/* ========== CLAIMS TABLE ========== */
.claims-section{background:var(--panel);border:1px solid var(--bdr);border-radius:14px;overflow:hidden}
.claims-section .section-title{padding:14px 18px;margin:0;border-bottom:1px solid var(--bdr)}
.claims-table-wrap{overflow-x:auto;max-height:760px;overflow-y:auto}
.claims-table-wrap.hide-payer .col-payer{display:none}
table{width:100%;border-collapse:collapse}
thead th{background:rgba(255,255,255,.015);padding:11px 14px;text-align:left;font-size:9px;font-weight:600;color:var(--text-muted);text-transform:uppercase;letter-spacing:.6px;border-bottom:1px solid var(--bdr);position:sticky;top:0;z-index:2}
tbody td{padding:11px 14px;font-size:11.5px;border-bottom:1px solid var(--bdr);color:var(--text-secondary)}
tbody tr{transition:background .15s}
tbody tr:hover{background:var(--card)}
tbody tr:last-child td{border-bottom:none}
.state-pill{display:inline-block;padding:3px 9px;border-radius:30px;font-size:10px;font-weight:600;white-space:nowrap}
.state-pill.submitted{background:rgba(16,185,129,.12);color:var(--success)}
.state-pill.documentation{background:rgba(205,180,134,.12);color:var(--accent)}
.state-pill.ready{background:rgba(59,130,246,.12);color:var(--info)}
.state-pill.revision{background:rgba(239,68,68,.12);color:var(--bad)}
.hcfa-link{cursor:pointer;color:var(--accent);font-size:11px;font-weight:500;padding:3px 7px;border-radius:6px;background:var(--accent-glow);display:inline-block;transition:all .15s}
.hcfa-link:hover{background:rgba(205,180,134,.25);box-shadow:0 0 0 1px var(--accent)}
.empty-state{text-align:center;padding:30px 20px;color:var(--text-muted);font-size:12px}

/* ========== TASK PANEL ========== */
.task-panel{display:flex;flex-direction:column;gap:14px}
.task-card{background:var(--panel);border:1px solid var(--bdr);border-radius:14px;padding:16px 18px}
.task-card h3{font-size:12px;font-weight:600;margin-bottom:12px;display:flex;align-items:center;gap:8px;color:var(--text-primary)}
.task-card label{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.7px;color:var(--text-muted);display:block;margin-bottom:6px;margin-top:8px}
.task-card label:first-of-type{margin-top:0}
.task-card select,.task-card textarea,.task-card input{width:100%;background:var(--bg2);border:1px solid var(--bdr);color:var(--text-primary);padding:9px 11px;border-radius:8px;font-size:12px;outline:none;font-family:'Inter',sans-serif;margin-bottom:6px}
.task-card select:focus,.task-card textarea:focus,.task-card input:focus{border-color:var(--accent)}
.task-card textarea{min-height:90px;font-family:'JetBrains Mono','Monaco',monospace;font-size:11px;line-height:1.5;resize:vertical}
.task-steps{background:var(--card);border:1px solid var(--bdr);border-radius:9px;padding:11px 14px;margin:10px 0 12px;font-size:11px;line-height:1.55}
.task-steps-title{font-size:11px;font-weight:600;color:var(--accent);margin-bottom:5px}
.task-steps-desc{color:var(--text-secondary);margin-bottom:6px}
.task-steps ol{margin:8px 0 0 18px;color:var(--text-secondary);font-size:10.5px;line-height:1.6}
.task-steps li{margin-bottom:4px}
.task-steps code{background:rgba(205,180,134,.1);color:var(--accent);padding:1px 5px;border-radius:4px;font-size:10px;font-family:'JetBrains Mono','Monaco',monospace}
.task-steps strong{color:var(--text-primary)}

/* logs */
.logs-panel{background:var(--panel);border:1px solid var(--bdr);border-radius:14px;display:flex;flex-direction:column;max-height:340px}
.logs-header{padding:12px 16px;border-bottom:1px solid var(--bdr);display:flex;justify-content:space-between;align-items:center}
.logs-header h3{font-size:12px;font-weight:600;display:flex;align-items:center;gap:6px}
.logs-body{flex:1;overflow-y:auto;padding:8px 14px;font-family:'JetBrains Mono','Monaco',monospace;font-size:10.5px;line-height:1.6}
.log-line{padding:2px 0;color:var(--text-secondary);word-break:break-all}
.log-line.info{color:var(--text-secondary)}
.log-line.success{color:var(--success)}
.log-line.warning{color:var(--warning)}
.log-line.error{color:var(--bad)}

/* fix coding ivs files */
.fix-ivs-run{background:var(--card);border:1px solid var(--bdr);border-radius:11px;padding:10px 12px;margin-bottom:8px;transition:all .15s}
.fix-ivs-run:hover{border-color:var(--bdr2)}
.fix-ivs-run-id{font-size:10px;color:var(--text-muted);margin-bottom:7px;display:flex;align-items:center;gap:5px}
.fix-ivs-file{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:7px 10px;margin-top:5px;border-radius:7px;color:var(--text-primary);font-size:11px;border:1px solid var(--bdr);background:var(--bg2);transition:all .15s}
.fix-ivs-file:hover{border-color:var(--accent);background:var(--card2)}
.fix-ivs-size{color:var(--text-muted);font-family:monospace;font-size:10px}

/* toast */
.toast{position:fixed;bottom:30px;right:30px;background:var(--card);border:1px solid var(--bdr);padding:13px 22px;border-radius:11px;font-size:12px;font-weight:500;color:var(--text-primary);opacity:0;transform:translateY(20px);transition:all .3s;pointer-events:none;z-index:1000;box-shadow:0 10px 30px rgba(0,0,0,.5)}
.toast.success,.toast.error{opacity:1;transform:translateY(0)}
.toast.success{border-color:var(--success);box-shadow:0 10px 30px rgba(16,185,129,.3)}
.toast.error{border-color:var(--bad);box-shadow:0 10px 30px rgba(239,68,68,.3)}

/* HCFA Lightbox */
.hcfa-lightbox{position:fixed;inset:0;background:rgba(0,0,0,.92);backdrop-filter:blur(10px);z-index:1100;display:none;flex-direction:column;padding:24px}
.hcfa-lightbox.active{display:flex}
.hcfa-lightbox-header{display:flex;align-items:center;gap:18px;padding-bottom:14px;border-bottom:1px solid var(--bdr);flex-wrap:wrap}
.hcfa-lightbox-header h3{font-size:14px;font-weight:600;color:var(--text-primary)}
.hcfa-tabs{display:flex;gap:8px;flex:1}
.hcfa-tab{background:var(--card);border:1px solid var(--bdr);color:var(--text-secondary);padding:7px 14px;border-radius:8px;font-size:11px;font-weight:600;cursor:pointer;transition:all .15s}
.hcfa-tab:hover{color:var(--text-primary)}
.hcfa-tab.active{background:var(--accent);color:#000;border-color:var(--accent)}
.hcfa-lightbox-close{background:var(--card);border:1px solid var(--bdr);color:var(--text-secondary);width:34px;height:34px;border-radius:8px;cursor:pointer;font-size:14px;transition:all .15s}
.hcfa-lightbox-close:hover{color:var(--bad);border-color:var(--bad)}
.hcfa-lightbox-body{flex:1;overflow:auto;padding:18px;display:flex;align-items:center;justify-content:center}
.hcfa-lightbox-body img{max-width:100%;max-height:100%;object-fit:contain;border-radius:8px;box-shadow:0 0 40px rgba(0,0,0,.6)}
.hcfa-placeholder{text-align:center;color:var(--text-muted);font-size:13px;padding:40px}

/* BS Drawer */
.bs-drawer-overlay{position:fixed;inset:0;background:rgba(0,0,0,.6);backdrop-filter:blur(4px);z-index:1090;opacity:0;pointer-events:none;transition:opacity .2s}
.bs-drawer-overlay.active{opacity:1;pointer-events:auto}
.bs-drawer{position:fixed;top:0;right:0;bottom:0;width:480px;max-width:96vw;background:var(--panel);border-left:1px solid var(--bdr);z-index:1095;transform:translateX(100%);transition:transform .3s ease;display:flex;flex-direction:column;box-shadow:-20px 0 40px rgba(0,0,0,.5)}
.bs-drawer.active{transform:translateX(0)}
.bs-drawer-header{padding:18px 22px;border-bottom:1px solid var(--bdr);display:flex;justify-content:space-between;align-items:center}
.bs-drawer-header h2{font-family:'Playfair Display',serif;font-size:18px;font-weight:600;color:var(--accent)}
.bs-drawer-close{background:var(--card);border:1px solid var(--bdr);color:var(--text-secondary);width:32px;height:32px;border-radius:8px;cursor:pointer;font-size:13px;transition:all .15s}
.bs-drawer-close:hover{color:var(--bad);border-color:var(--bad)}
.bs-drawer-meta{padding:12px 22px;border-bottom:1px solid var(--bdr);font-size:11px;color:var(--text-secondary)}
.bs-drawer-meta strong{color:var(--accent)}
.bs-drawer-body{flex:1;overflow-y:auto;padding:14px 22px}
.bs-claim-card{background:var(--card);border:1px solid var(--bdr);border-radius:10px;padding:12px 14px;margin-bottom:8px;transition:all .15s}
.bs-claim-card:hover{border-color:var(--accent)}

.footer{text-align:center;padding:22px 0 10px;color:var(--text-dim);font-size:10px;border-top:1px solid var(--bdr);margin-top:24px}
.footer a{color:var(--accent)}

@media(max-width:1380px){
  .hero{grid-template-columns:1fr 1fr 1fr}
  .hero .promo{grid-column:1/-1;min-height:auto;flex-direction:row;align-items:center;gap:20px}
  .hero .promo>div:nth-child(2){flex:1}
  .hero .promo-btns{flex-direction:row}
  .main{grid-template-columns:1fr}
}
@media(max-width:1080px){
  .pipeline-flow{grid-template-columns:repeat(2,1fr)}
  .hero{grid-template-columns:1fr 1fr}
}
@media(max-width:780px){
  .app{grid-template-columns:1fr}
  .sb{position:relative;height:auto}
  .hero{grid-template-columns:1fr}
  .hero .promo{flex-direction:column}
  .hero-num{font-size:46px}
}
</style>
</head>
<body>
<div class="app">

  <!-- ═══════════════ SIDEBAR ═══════════════ -->
  <aside class="sb">
    <div class="brand">
      <img src="https://helixona.com/wp-content/uploads/2025/08/Screenshot-2025-11-03-at-18.32.49-Photoroom.png" alt="Helixona">
      <div>
        <div class="brand-t">Helixona<sup style="font-size:8px;color:var(--text-muted)">®</sup></div>
        <div class="brand-s">Billing Agent · v1</div>
      </div>
    </div>

    <div class="nav">
      <a class="on" onclick="scrollToEl('.hero-kpi')"><span class="ico">▦</span>Overview</a>
      <a onclick="scrollToEl('.claims-section')"><span class="ico">📋</span>Claims<span class="nav-cnt" id="nv-cl">0</span></a>
      <a onclick="scrollToEl('#send-task-card')"><span class="ico">⚡</span>Send Task</a>
      <a onclick="scrollToEl('.logs-panel')"><span class="ico">🔴</span>Logs</a>
      <a href="/audit"><span class="ico">🧾</span>Audit Log</a>
    </div>
  </aside>

  <!-- ═══════════════ MAIN ═══════════════ -->
  <div class="shell">

    <!-- TOPBAR -->
    <div class="tb">
      <div class="tb-l">
        <button id="stopBtn" class="stop-btn" onclick="stopAgent()">⛔ Stop Agent</button>
        <a id="novnc-link" href="http://54.189.175.233:6080/vnc.html" target="_blank" class="novnc-link">🖥️ noVNC</a>
      </div>
      <div class="tb-r">
        <div class="status-badge"><div class="status-dot"></div>Agent Running · 54.189.175.233</div>
        <span style="font-size:10px;color:var(--text-muted)" id="ts">—</span>
        <div class="icbtn" onclick="loadData();loadLogs();loadFixIvsFiles()" title="Refresh">↻</div>
      </div>
    </div>

    <!-- BOT TABS -->
    <div class="subtabs" id="bot-tabs">
      <div class="subtab on" data-bot="submissions" onclick="setActiveBot('submissions')">🛡️ Blue Shield Submissions</div>
      <div class="subtab" data-bot="resubmissions" onclick="setActiveBot('resubmissions')">♻️ Blue Shield Resubmissions</div>
      <div class="subtab" data-bot="iv_corrections" onclick="setActiveBot('iv_corrections')">🩺 IV Fix Coding</div>
    </div>

    <!-- HERO KPI: Submission progress (Bot 1 — Submissions) -->
    <div class="hero-kpi" id="hero-submissions">
      <div class="hero-kpi-top">
        <div class="hero-kpi-headline">
          <span class="hero-num" id="hero-submitted">—</span>
          <span class="hero-denom"> of <span id="hero-total">—</span> claims submitted</span>
        </div>
        <div class="hero-kpi-pct"><strong id="hero-pct">—</strong> complete · <span id="hero-remaining">—</span> remaining</div>
      </div>
      <div class="hero-progress"><div class="hero-progress-fill" id="hero-progress-fill"></div></div>
    </div>

    <!-- HERO KPI: IV Fix Coding progress (Bot 2) -->
    <div class="hero-kpi" id="hero-iv" style="display:none">
      <div class="hero-kpi-top">
        <div class="hero-kpi-headline">
          <span class="hero-num" id="iv-hero-total">—</span>
          <span class="hero-denom"> IVs pending fixing</span>
          <span id="iv-not-enriched" style="display:none;margin-left:10px;font-size:11px;padding:3px 8px;border-radius:12px;background:rgba(239,68,68,.15);color:var(--bad);border:1px solid rgba(239,68,68,.3)">⚠ Latest run not enriched — showing all pending claims</span>
        </div>
        <div class="hero-kpi-pct">
          <strong id="iv-hero-pending">—</strong> Pending ·
          <strong id="iv-hero-errors">—</strong> Pending With Errors ·
          <span id="iv-hero-run" style="color:var(--text-muted)">Last run: —</span>
        </div>
      </div>
      <div class="hero-progress"><div class="hero-progress-fill" id="iv-hero-progress-fill" style="background:linear-gradient(90deg,var(--accent),var(--bad))"></div></div>
    </div>

    <!-- MAIN: claims + admin rail -->
    <div class="main">

      <!-- CLAIMS TABLE -->
      <div class="claims-section" id="claims-section-submissions">
        <div class="section-title">
          📋 Claims<span id="claims-payer-title" style="font-weight:500;color:var(--text-muted);margin-left:8px;font-size:12px;"></span>
          <span id="claims-type-counts" style="font-weight:500;margin-left:12px;font-size:11px;"></span>
          <button class="btn btn-refresh" onclick="loadData()">↻ Refresh</button>
          <span id="claims-eta" style="font-size:11px;color:var(--info);font-weight:600;margin-left:auto;display:none">⏱ </span>
          <span id="claims-meta" style="font-size:11px;color:var(--text-muted);margin-left:14px"></span>
        </div>
        <div class="claims-table-wrap">
          <table>
            <thead>
              <tr>
                <th>Claim #</th><th>Patient</th><th>DOS</th><th class="col-payer">Payer</th><th>Charges</th><th>Type</th>
                <th>HCFA</th><th>IV Note</th><th>Progress Note Date</th><th>Progress Note</th><th>Submission</th><th>Stage</th>
              </tr>
            </thead>
            <tbody id="claims-body"><tr><td colspan="11" class="empty-state">No claims yet. Send a task to begin.</td></tr></tbody>
          </table>
        </div>
      </div>

      <!-- IV CLAIMS TABLE (Bot 2 — Fix Coding IVs) -->
      <div class="claims-section" id="claims-section-iv" style="display:none">
        <div class="section-title">
          🩺 IV Claims to Fix
          <span id="iv-claims-run" style="font-weight:500;color:var(--text-muted);margin-left:8px;font-size:12px;"></span>
          <button class="btn btn-refresh" onclick="loadIvClaims()">↻ Refresh</button>
          <span id="iv-claims-meta" style="font-size:11px;color:var(--text-muted);margin-left:auto"></span>
        </div>
        <div class="claims-table-wrap">
          <table>
            <thead>
              <tr>
                <th>Claim #</th><th>Patient</th><th>DOS</th><th>Payer</th><th>Charges</th><th>Balance</th><th>Status</th><th>Provider</th>
              </tr>
            </thead>
            <tbody id="iv-claims-body"><tr><td colspan="8" class="empty-state">Loading IV claims…</td></tr></tbody>
          </table>
        </div>
      </div>

      <!-- RIGHT RAIL -->
      <div class="task-panel">

        <!-- LIVE LOGS -->
        <div class="logs-panel">
          <div class="logs-header">
            <h3>🔴 Live Agent Logs</h3>
            <button class="btn btn-refresh" onclick="loadLogs()">↻</button>
          </div>
          <div class="logs-body" id="logs-body">
            <div class="log-line info">Loading logs...</div>
          </div>
        </div>

        <!-- SEND TASK -->
        <div class="task-card" id="send-task-card">
          <h3>⚡ Send Task to Agent</h3>
          <label>Task Type</label>
          <select id="task-type" onchange="updateTaskTemplate()">
            <optgroup label="🛡️ Blue Shield Claims Bot" data-bot="submissions">
              <option value="bs_missing_docs" data-bot="submissions">📋 ECW Obtain Claims Documentation</option>
              <option value="blueshield_submissions" data-bot="submissions">📤 Blue Shield Submissions</option>
              <option value="ecw_status_update" data-bot="submissions">📝 ECW Status Update</option>
            </optgroup>
            <optgroup label="🩺 IV Fix Coding Bot" data-bot="iv_corrections">
              <option value="fix_coding_ivs__all" data-bot="iv_corrections">🔄 All Stages</option>
              <option value="fix_coding_ivs__1" data-bot="iv_corrections">1️⃣ Create Encounter Claims</option>
              <option value="fix_coding_ivs__2" data-bot="iv_corrections">2️⃣ Delete Test Claims</option>
              <option value="fix_coding_ivs__3" data-bot="iv_corrections">3️⃣ Generate Excel Documentation</option>
            </optgroup>
          </select>

          <div id="task-steps" class="task-steps"></div>

          <div style="margin:10px 0;padding:10px 12px;background:rgba(205,180,134,0.06);border:1px solid rgba(205,180,134,0.18);border-radius:8px">
            <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
              <span style="color:var(--accent);font-weight:600;font-size:11px">🧪 Test Single Claim</span>
              <span style="color:var(--text-muted);font-size:10px">(empty = all)</span>
            </div>
            <input type="text" id="test-claim-id" placeholder="e.g. 239" style="margin-bottom:0">
          </div>

          <label>Task Payload (JSON)</label>
          <textarea id="task-payload"></textarea>
          <button class="btn btn-primary" onclick="sendTask()">🚀 Send to SQS</button>
          <button class="btn" style="background:rgba(239,68,68,.1);color:var(--bad);border:1px solid rgba(239,68,68,.3);margin-top:8px;width:100%;padding:10px;border-radius:8px;font-weight:600;font-size:11.5px;cursor:pointer" onclick="deleteAllClaims()">🗑️ Delete All Claims &amp; Start Over</button>
        </div>

        <!-- REPORTS -->
        <div class="task-card" id="reports-card">
          <h3>📊 Fix Coding IVs — Saved Reports
            <button class="btn btn-refresh" style="float:right;margin-top:-2px" onclick="loadFixIvsFiles()">↻</button>
          </h3>
          <div id="fix-ivs-files">
            <div class="empty-state" style="padding:14px">Loading…</div>
          </div>
          <button class="btn" style="background:rgba(239,68,68,.1);color:var(--bad);border:1px solid rgba(239,68,68,.3);margin-top:8px;width:100%;padding:10px;border-radius:8px;font-weight:600;font-size:11.5px;cursor:pointer" onclick="deleteAllFixIvsFiles()">🗑️ Eliminar todos los saved reports</button>
        </div>

        <!-- OPEN TASKS -->
        <div class="task-card" id="tasks-card">
          <h3>🚨 Open Tasks</h3>
          <div id="tasks-list">
            <div class="empty-state" style="padding:18px">No open tasks</div>
          </div>
        </div>

      </div>
    </div>

    <div class="footer">Helixona Billing Agent · v1 · <a href="https://helixona.com" target="_blank">helixona.com</a> · Auto-refresh 10s</div>

  </div>
</div>

<div class="toast" id="toast"></div>

<!-- HCFA Lightbox Viewer -->
<div class="hcfa-lightbox" id="hcfa-lightbox">
  <div class="hcfa-lightbox-header">
    <h3 id="hcfa-lightbox-title">HCFA Form</h3>
    <div class="hcfa-tabs" id="hcfa-tabs"></div>
    <button class="hcfa-lightbox-close" onclick="closeHCFAViewer()">✕</button>
  </div>
  <div class="hcfa-lightbox-body" id="hcfa-lightbox-body">
    <div class="hcfa-placeholder">Loading...</div>
  </div>
</div>

<!-- BS Claims Drawer -->
<div class="bs-drawer-overlay" id="bs-overlay" onclick="closeBSDrawer()"></div>
<div class="bs-drawer" id="bs-drawer">
  <div class="bs-drawer-header">
    <h2 id="bs-drawer-title">Claims</h2>
    <button class="bs-drawer-close" onclick="closeBSDrawer()">✕</button>
  </div>
  <div class="bs-drawer-meta" id="bs-drawer-meta">
    <span>Loading...</span>
  </div>
  <div class="bs-drawer-body" id="bs-drawer-body">
    <div class="empty-state">No claims data yet.</div>
  </div>
</div>

<script>
// ═══════════════ Augmentations on top of the original dashboard JS ═══════════════
// Wraps loadData/loadFixIvsFiles to also render hero cards, sidebar list,
// nav counts and refreshed timestamp.
window.scrollToEl = function(sel){
  const el = document.querySelector(sel);
  if(el) el.scrollIntoView({behavior:'smooth',block:'start'});
};

(function(){
  const $ = id => document.getElementById(id);
  const getState = c => parseInt(c.state || c.current_state || 0);
  function getStages(){
    try { return PIPELINE_STAGES; } catch(e){ return {}; }
  }
  function stageKeyFor(s){
    for(const [k,v] of Object.entries(getStages()))
      if(v.states && v.states.includes(s)) return k;
    return 'documentation';
  }

  function areaSpark(values,color){
    const w=320,h=70,n=values.length||7;
    const mx=Math.max(...values,1), mn=Math.min(...values,0);
    const rng=mx-mn||1;
    const pts=values.map((v,i)=>{
      const x=(i/(n-1))*w;
      const y=h-((v-mn)/rng)*(h-16)-8;
      return [x,y];
    });
    const path=pts.map((p,i)=>(i?'L':'M')+p[0].toFixed(1)+','+p[1].toFixed(1)).join(' ');
    const area=path+` L${w},${h} L0,${h} Z`;
    const last=pts[pts.length-1];
    const gid='gs'+Math.random().toString(36).slice(2,8);
    return `<svg viewBox="0 0 ${w} ${h}" preserveAspectRatio="none" style="width:100%;height:100%">
      <defs><linearGradient id="${gid}" x1="0" x2="0" y1="0" y2="1">
        <stop offset="0%" stop-color="${color}" stop-opacity=".35"/>
        <stop offset="100%" stop-color="${color}" stop-opacity="0"/>
      </linearGradient></defs>
      <path d="${area}" fill="url(#${gid})"/>
      <path d="${path}" fill="none" stroke="${color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="filter:drop-shadow(0 0 4px ${color})"/>
      <circle cx="${last[0]}" cy="${last[1]}" r="3" fill="${color}" style="filter:drop-shadow(0 0 6px ${color})"/>
    </svg>`;
  }

  function renderHeroCards(claims){
    const stages = getStages() || {};
    const order = ['submitted','documentation','ready'];
    const iconBg = {
      submitted:'rgba(16,185,129,.15)',
      documentation:'rgba(205,180,134,.15)',
      ready:'rgba(232,213,176,.15)'
    };
    const icons = {submitted:'✅', documentation:'📄', ready:'🚀'};
    const wrap = $('hero-cards');
    if(!wrap) return;
    const t = claims.length;
    wrap.innerHTML = order.map(k=>{
      const st = stages[k];
      if(!st) return '';
      const n = claims.filter(c=>st.states.includes(getState(c))).length;
      const pct = t ? Math.round(n/t*100) : 0;
      const base = Math.max(1, Math.round(n*.6));
      const series = [base, base+1, base-1, base+2, base, base+1, n];
      const up = n >= base;
      const colors = {submitted:'#10b981', documentation:'#CDB486', ready:'#E8D5B0', revision:'#ef4444'};
      const c = colors[k] || '#CDB486';
      return `
        <div class="bigcard">
          <div class="bc-top">
            <div class="bc-id">
              <div class="bc-icon" style="background:${iconBg[k]};color:${c}">${icons[k]||'•'}</div>
              <div class="bc-meta"><div class="bc-tag">${k==='submitted'?'In Blue Shield':k==='ready'?'Docs complete':'Gathering docs'}</div><div class="bc-name">${st.label}</div></div>
            </div>
            <div class="bc-arrow">↗</div>
          </div>
          <div class="bc-lbl">Claim Count</div>
          <div class="bc-num" style="color:${c}">${n}<span class="pct">/${t||0}</span></div>
          <div class="bc-delta ${up?'':'dn'}"><span>${up?'▲':'▼'}</span>${pct}%</div>
          <div class="bc-spark">${areaSpark(series,c)}</div>
        </div>`;
    }).join('');
  }

  function renderSidebar(claims){
    const stages = getStages() || {};
    const colors = {submitted:'#10b981', documentation:'#CDB486', ready:'#E8D5B0', revision:'#ef4444'};
    const list = $('sb-list');
    if($('sb-cnt')) $('sb-cnt').textContent = claims.length;
    if($('nv-cl'))  $('nv-cl').textContent  = claims.length;
    if($('nv-pipe'))$('nv-pipe').textContent = Object.values(stages).reduce((a,st)=>a+claims.filter(c=>st.states.includes(getState(c))).length,0);
    if($('top-cnt'))$('top-cnt').textContent = (claims.length||0)+' Active';
    if(!list) return;
    if(!claims.length){
      list.innerHTML = '<div class="empty-state" style="padding:14px;font-size:10px">No claims</div>';
      return;
    }
    const sorted = [...claims].sort((a,b)=>getState(b)-getState(a)).slice(0,6);
    list.innerHTML = sorted.map(c=>{
      const sk = stageKeyFor(getState(c));
      const stg = stages[sk];
      const c0 = colors[sk] || '#CDB486';
      const initial = (c.patient_name || '?').charAt(0).toUpperCase();
      return `<div class="sb-item" title="${c.claim_id||''}">
        <div class="sb-ico" style="background:${c0}22;color:${c0}">${initial}</div>
        <div class="sb-info">
          <div class="sb-l">${(stg&&stg.label)||sk}</div>
          <div class="sb-n">${c.charges || c.claim_id || '—'}</div>
        </div>
      </div>`;
    }).join('');
  }

  function updateTs(){
    if($('ts')) $('ts').textContent = 'Updated ' + new Date().toLocaleTimeString();
  }

  // Monkey-patch loadData once it exists.
  function install(){
    if(typeof window.loadData !== 'function'){ return setTimeout(install, 60); }
    const _origLoad = window.loadData;
    window.loadData = async function(){
      const r = await _origLoad.apply(this, arguments);
      try{
        const res = await fetch('/api/claims');
        const d = await res.json();
        const claims = d.claims || [];
        renderHeroCards(claims);
        renderSidebar(claims);
        updateTs();
      }catch(e){ console.warn('augment fetch failed', e); }
      return r;
    };
    // Run once immediately so the UI does not stay blank on first load.
    window.loadData();
  }

  if(document.readyState === 'loading'){
    document.addEventListener('DOMContentLoaded', install);
  } else {
    install();
  }
})();
</script>
    <script>
        const STATE_LABELS = {{ state_labels | tojson }};
        const PIPELINE_STAGES = {{ pipeline_stages | tojson }};

        const TASK_TEMPLATES = {
            bs_missing_docs: JSON.stringify({
                note: "Runs full pipeline: ECW claims extraction, HCFA generation, Progress Notes capture, and BlueShield submission for Missing Documentation cases"
            }, null, 2),
            blueshield_submissions: JSON.stringify({
                note: "Uploads claim documentation (HCFA, Prog Notes, Encounter File) to Blue Shield via SympliSend."
            }, null, 2),
            ecw_status_update: JSON.stringify({
                note: "Updates claim status in ECW from 'Ready to Submit to Symplisend' to 'Claim sent via Symplisend' for all submitted claims."
            }, null, 2),
            fix_coding_ivs__all: JSON.stringify({
                note: "Runs Stage 1 (create encounter claims) → Stage 2 (delete test claims) → Stage 3 (generate Excel)."
            }, null, 2),
            fix_coding_ivs__1: JSON.stringify({
                note: "Stage 1 — Encounters page: filter Progress Notes Done/Locked, click Claims IPE All to create encounter claims."
            }, null, 2),
            fix_coding_ivs__2: JSON.stringify({
                note: "Stage 2 — Find claims whose patient name contains 'test' and delete them from ECW."
            }, null, 2),
            fix_coding_ivs__3: JSON.stringify({
                note: "Stage 3 — Pull Pending and Pending With Errors claims from ECW, combine into a single Excel, upload to S3."
            }, null, 2)
        };

        // Maps the dropdown value into the actual SQS payload {task_type, stage?}
        const TASK_DISPATCH = {
            fix_coding_ivs__all: { task_type: 'fix_coding_ivs', stage: 'all' },
            fix_coding_ivs__1:   { task_type: 'fix_coding_ivs', stage: '1' },
            fix_coding_ivs__2:   { task_type: 'fix_coding_ivs', stage: '2' },
            fix_coding_ivs__3:   { task_type: 'fix_coding_ivs', stage: '3' }
        };

        const TASK_DESCRIPTIONS = {
            bs_missing_docs: {
                title: 'ECW obtain claims documentation',
                desc: 'Runs the full pipeline: extract claims from ECW, generate HCFA forms, capture Progress Notes and submit Missing Documentation cases to Blue Shield.',
                steps: []
            },
            blueshield_submissions: {
                title: 'Blueshield Submissions',
                desc: 'Uploads claim documentation (HCFA, Progress Notes, Encounter File) to Blue Shield via SympliSend.',
                steps: []
            },
            ecw_status_update: {
                title: 'ECW Status Update',
                desc: "Updates claim status in ECW from 'Ready to Submit to Symplisend' to 'Claim sent via Symplisend' for all submitted claims.",
                steps: []
            },
            fix_coding_ivs__all: {
                title: 'Fix Coding IVs — All Stages',
                desc: 'Runs <strong>Stage 1</strong> (create encounter claims) → <strong>Stage 2</strong> (delete test claims) → <strong>Stage 3</strong> (generate Excel) in a single session.',
                steps: [
                    '<strong>Stage 1</strong>: Encounters → <em>Progress Notes Done/Locked</em> → <strong>Claims IPE All</strong> to create encounter claims.',
                    '<strong>Stage 2</strong>: scan the Claims table and delete every claim whose patient name contains <code>test</code>.',
                    '<strong>Stage 3</strong>: pull <em>Pending</em> + <em>Pending With Errors</em> Excels (now clean of test claims), combine, upload to S3.'
                ]
            },
            fix_coding_ivs__1: {
                title: 'Fix Coding IVs — Stage 1: Create Encounter Claims',
                desc: 'On the Encounters page, run <strong>Claims IPE All</strong> to create encounter claims for all Progress Notes that are Done/Locked.',
                steps: [
                    'Login to ECW.',
                    'Billing → <strong>Encounters</strong>: dates <code>07/01/2025</code> → today, select <em>Progress Notes Done/Locked</em>, click <code>#btnFilter</code>.',
                    'Click <strong>Claims IPE All</strong> (<code>#btnClaimIPEAll</code>). Click <em>Yes</em> on the confirm dialog, then close the result modal with the X (<code>#ClaimIPEBtn1</code>). If "No encounters selected" appears, click OK and continue.'
                ]
            },
            fix_coding_ivs__2: {
                title: 'Fix Coding IVs — Stage 2: Delete Test Claims',
                desc: 'Scan the Claims table for entries whose patient name contains <code>test</code> and delete them.',
                steps: [
                    'Navigate to Billing → Claims.',
                    'For each status in [<em>Pending With Errors</em>, <em>Pending</em>]: set status filter + dates <code>07/01/2025</code> → today, click <code>#btnclaimlookup</code>.',
                    'For every row whose patient name contains <code>test</code>: select the row checkbox → Claims dropdown (<code>#claimLookupBtn10</code>) → <strong>Delete Claim</strong> → confirm <strong>Yes</strong>.',
                    'Repeat until no more test claims appear in the table.'
                ]
            },
            fix_coding_ivs__3: {
                title: 'Fix Coding IVs — Stage 3: Generate Excel Documentation',
                desc: 'Pull <strong>Pending</strong> and <strong>Pending With Errors</strong> claims from ECW and combine them into a single Excel.',
                steps: [
                    'Billing → <strong>Claims</strong>: filter <em>Pending With Errors</em>, dates <code>07/01/2025</code> → today, click <code>#btnclaimlookup</code>. Then Billing dropdown → <strong>View Claims Report → EXCEL</strong>.',
                    'Repeat the previous step for <em>Pending</em> status.',
                    'Combine the two Excels and upload to S3 under <code>fix_coding_ivs/&lt;timestamp&gt;/</code>.'
                ]
            }
        };

        function updateTaskTemplate() {
            const type = document.getElementById('task-type').value;
            document.getElementById('task-payload').value = TASK_TEMPLATES[type] || '{}';
            renderTaskSteps(type);
        }

        function renderTaskSteps(type) {
            const el = document.getElementById('task-steps');
            const info = TASK_DESCRIPTIONS[type];
            if (!el || !info) { if (el) el.innerHTML = ''; return; }
            const stepsHtml = (info.steps && info.steps.length)
                ? `<ol>${info.steps.map(s => `<li>${s}</li>`).join('')}</ol>`
                : '';
            el.innerHTML = `
                <div class="task-steps-title">📋 ${info.title}</div>
                <div class="task-steps-desc">${info.desc}</div>
                ${stepsHtml}
            `;
        }

        // Active bot tab. Drives which SQS queue / systemd service the dashboard talks to.
        window.activeBot = 'submissions';

        const BOT_NOVNC = {submissions: 6080, resubmissions: 6081, iv_corrections: 6082};

        function setActiveBot(bot) {
            if (!(bot in BOT_NOVNC)) bot = 'submissions';
            window.activeBot = bot;
            // Each bot runs on its own X display behind its own noVNC port —
            // pointing every tab at 6080 would show the submissions bot's
            // screen while claiming to show another's.
            const vnc = document.getElementById('novnc-link');
            if (vnc) vnc.href = `http://54.189.175.233:${BOT_NOVNC[bot]}/vnc.html`;
            // Tab styling
            document.querySelectorAll('#bot-tabs .subtab').forEach(el => {
                el.classList.toggle('on', el.dataset.bot === bot);
            });
            // Show/hide task-type optgroups + options
            document.querySelectorAll('#task-type [data-bot]').forEach(el => {
                el.hidden = (el.dataset.bot !== bot);
                el.disabled = (el.dataset.bot !== bot);
            });
            // Pick the first visible option for this bot
            const sel = document.getElementById('task-type');
            const firstVisible = Array.from(sel.options).find(o => !o.hidden);
            if (firstVisible) { sel.value = firstVisible.value; updateTaskTemplate(); }
            // Toggle hero KPI + claims table per bot. Resubmissions reuse the
            // submissions views — same claims table, filtered to its own work.
            const isIv = (bot === 'iv_corrections');
            const heroSub = document.getElementById('hero-submissions');
            const heroIv  = document.getElementById('hero-iv');
            const tblSub  = document.getElementById('claims-section-submissions');
            const tblIv   = document.getElementById('claims-section-iv');
            if (heroSub) heroSub.style.display = isIv ? 'none' : '';
            if (heroIv)  heroIv.style.display  = isIv ? '' : 'none';
            if (tblSub)  tblSub.style.display  = isIv ? 'none' : '';
            if (tblIv)   tblIv.style.display   = isIv ? '' : 'none';
            // Load the right data set
            if (isIv) loadIvClaims();
            else if (typeof loadData === 'function') loadData();
            // Refresh logs for the new service
            clearLogs('switched to ' + bot);
            loadLogs();
        }

        // Pulls latest fix_coding_ivs run from S3 and renders the IV claims table + KPI.
        async function loadIvClaims() {
            try {
                const res = await fetch('/api/iv-claims');
                const data = await res.json();
                const body = document.getElementById('iv-claims-body');
                const k = data.kpi || {total:0, pending:0, pending_with_errors:0};
                // KPI
                const setText = (id, v) => { const el = document.getElementById(id); if (el) el.textContent = v; };
                setText('iv-hero-total', k.total);
                setText('iv-hero-pending', k.pending);
                setText('iv-hero-errors', k.pending_with_errors);
                const warn = document.getElementById('iv-not-enriched');
                if (warn) warn.style.display = (data.enriched === false) ? 'inline-block' : 'none';
                const fill = document.getElementById('iv-hero-progress-fill');
                if (fill) fill.style.width = (k.total === 0 ? '0' : '100') + '%';
                // Section subtitle
                const runEl = document.getElementById('iv-claims-run');
                const runHero = document.getElementById('iv-hero-run');
                if (data.run) {
                    const t = data.run; // YYYYMMDD_HHMMSS
                    const human = `${t.slice(0,4)}-${t.slice(4,6)}-${t.slice(6,8)} ${t.slice(9,11)}:${t.slice(11,13)}`;
                    if (runEl)   runEl.textContent   = '· Run ' + human;
                    if (runHero) runHero.textContent = 'Last run: ' + human;
                } else {
                    if (runEl)   runEl.textContent   = '· No runs yet';
                    if (runHero) runHero.textContent = 'Last run: —';
                }
                const meta = document.getElementById('iv-claims-meta');
                if (meta) meta.textContent = `${(data.claims||[]).length} claims`;
                // Table
                if (!body) return;
                if (!data.claims || data.claims.length === 0) {
                    body.innerHTML = '<tr><td colspan="8" class="empty-state">No IV claims found. Run Stage 3 to generate the latest Excel.</td></tr>';
                    return;
                }
                const esc = s => String(s == null ? '' : s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
                const rows = data.claims.map(c => {
                    const statusClass = (c.status === 'Pending With Errors') ? 'revision' : 'ready';
                    return `<tr>
                        <td>${esc(c.claim_no)}</td>
                        <td>${esc(c.patient)}</td>
                        <td>${esc(c.dos)}</td>
                        <td>${esc(c.payer)}</td>
                        <td>${esc(c.charges)}</td>
                        <td>${esc(c.balance)}</td>
                        <td><span class="state-pill ${statusClass}">${esc(c.status)}</span></td>
                        <td>${esc(c.provider)}</td>
                    </tr>`;
                }).join('');
                body.innerHTML = rows;
            } catch (e) {
                const body = document.getElementById('iv-claims-body');
                if (body) body.innerHTML = `<tr><td colspan="8" class="empty-state">Failed to load IV claims: ${e.message}</td></tr>`;
            }
        }

        async function sendTask() {
            const selected = document.getElementById('task-type').value;
            let payload;
            try {
                payload = JSON.parse(document.getElementById('task-payload').value);
            } catch (e) {
                showToast('Invalid JSON payload', 'error');
                return;
            }
            // Translate the dropdown value into {task_type, stage?} via TASK_DISPATCH if mapped,
            // otherwise the dropdown value IS the task_type.
            const dispatch = TASK_DISPATCH[selected];
            if (dispatch) {
                payload.task_type = dispatch.task_type;
                if (dispatch.stage !== undefined) payload.stage = dispatch.stage;
            } else {
                payload.task_type = selected;
            }
            // Add test claim ID if specified
            const testClaimId = (document.getElementById('test-claim-id').value || '').trim();
            if (testClaimId) {
                payload.test_claim_id = testClaimId;
                payload.testing_mode = true;
            }
            payload.bot = window.activeBot;

            const res = await fetch('/api/send-task', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(payload)
            });
            const data = await res.json();
            if (data.success) {
                showToast('Task sent to SQS ✓', 'success');
                // New run starting — wipe the panel so this run's logs are not
                // visually mixed with the previous run's output.
                clearLogs('new ' + (payload.task_type || 'task') + ' run');
                setTimeout(loadLogs, 3000);
            } else {
                showToast('Failed: ' + data.error, 'error');
            }
        }

        async function deleteAllClaims() {
            if (!confirm('⚠️ DELETE all claims from the database? This cannot be undone. The next pipeline run will rediscover claims from ECW.')) return;
            try {
                const res = await fetch('/api/delete-all-claims', { method: 'POST' });
                const data = await res.json();
                if (data.success) {
                    showToast(data.message + ' \\u2713', 'success');
                    setTimeout(loadData, 500);
                } else {
                    showToast('Delete failed: ' + data.error, 'error');
                }
            } catch (e) {
                showToast('Delete failed: ' + e.message, 'error');
            }
        }

        // ── Stage classification for each claim ──
        function getStageKey(state) {
            for (const [key, stage] of Object.entries(PIPELINE_STAGES)) {
                if (stage.states.includes(state)) return key;
            }
            return 'documentation';
        }

        function getStagePill(state) {
            const key = getStageKey(state);
            const label = PIPELINE_STAGES[key]?.label || 'Unknown';
            return `<span class="state-pill ${key}">${label}</span>`;
        }

        // Manual override: flag a claim as not needing a Progress Note (e.g.,
        // diagnostic-only CPT mix). Posts to /api/claim/<cid>/skip_progress_note.
        async function skipProgressNote(claimId, skip) {
            const action = skip ? 'mark as NOT needing a Progress Note' : 'undo the override';
            if (!confirm(`Claim ${claimId}: ${action}?`)) return;
            try {
                const res = await fetch(`/api/claim/${claimId}/skip_progress_note`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ skip: skip }),
                });
                const data = await res.json();
                if (!res.ok) { alert('Failed: ' + (data.error || res.status)); return; }
                loadData();
            } catch (e) {
                alert('Error: ' + e.message);
            }
        }

        // Manual upload of a progress note PDF for a claim that the bot couldn't
        // capture. Opens a file picker, POSTs to /api/encounter_file/<cid>/upload.
        function uploadEncounterFile(claimId) {
            const input = document.createElement('input');
            input.type = 'file';
            input.accept = 'application/pdf,.pdf';
            input.style.display = 'none';
            document.body.appendChild(input);
            input.onchange = async () => {
                const file = input.files && input.files[0];
                document.body.removeChild(input);
                if (!file) return;
                if (!file.name.toLowerCase().endsWith('.pdf')) {
                    alert('Please choose a PDF file.');
                    return;
                }
                const fd = new FormData();
                fd.append('file', file);
                try {
                    const res = await fetch(`/api/encounter_file/${claimId}/upload`, { method: 'POST', body: fd });
                    const data = await res.json();
                    if (!res.ok) { alert('Upload failed: ' + (data.error || res.status)); return; }
                    alert(`✅ Uploaded progress note for claim ${claimId} (${data.size} bytes)`);
                    loadData();  // refresh table
                } catch (e) {
                    alert('Upload error: ' + e.message);
                }
            };
            input.click();
        }

        // Progress Note Date — prefer the Rx Start Date parsed from the IV Note
        // (iv_note_rx_start_date, stored as YYYY-MM-DD). This is the date used to
        // look up the progress note in ECW, and it's available even when the
        // encounter PDF capture failed. Fall back to encounter_date (MM/DD/YYYY)
        // when no IV Note Rx date was parsed yet. Cell is clickable to edit.
        function formatProgNoteDate(c) {
            const rx = c.iv_note_rx_start_date;
            let shown = '—';
            if (rx) {
                const m = String(rx).match(/^(\d{4})-(\d{2})-(\d{2})/);
                shown = m ? `${m[2]}/${m[3]}/${m[1]}` : rx;
            } else if (c.encounter_date) {
                shown = c.encounter_date;
            }
            const cid = c.claim_id;
            return `<span style="cursor:pointer;" onclick="editProgNoteDate('${cid}', '${rx || ''}')" title="Click to edit">${shown} <span style="color:var(--info);font-size:9px;">✏️</span></span>`;
        }

        // Manual edit of Progress Note Date (iv_note_rx_start_date in DDB).
        // Prompts user for MM/DD/YYYY, POSTs to /api/claim/<cid>/prog_note_date.
        async function editProgNoteDate(claimId, currentIso) {
            // Show current as MM/DD/YYYY in the prompt for clarity.
            let prefill = '';
            if (currentIso) {
                const m = String(currentIso).match(/^(\d{4})-(\d{2})-(\d{2})/);
                if (m) prefill = `${m[2]}/${m[3]}/${m[1]}`;
                else prefill = currentIso;
            }
            const val = prompt(`Progress Note Date for claim ${claimId} (MM/DD/YYYY, blank to clear):`, prefill);
            if (val === null) return;  // cancel
            try {
                const res = await fetch(`/api/claim/${claimId}/prog_note_date`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ date: val.trim() }),
                });
                const data = await res.json();
                if (!res.ok) { alert('Update failed: ' + (data.error || res.status)); return; }
                loadData();
            } catch (e) {
                alert('Update error: ' + e.message);
            }
        }

        // New submissions and resubmissions are separate bots with separate
        // queues, displays and browser profiles, so each tab shows only its own
        // claims. Anything not explicitly marked a resubmission counts as a new
        // submission — an unclassified claim belongs with the first-time work.
        function claimsForActiveBot(claims) {
            const isResub = c => /resub/i.test(c.submission_type || '');
            if (window.activeBot === 'resubmissions') return claims.filter(isResub);
            if (window.activeBot === 'submissions') return claims.filter(c => !isResub(c));
            return claims;
        }

        async function loadData() {
            try {
                const res = await fetch('/api/claims');
                const data = await res.json();
                const claims = claimsForActiveBot(data.claims);
                renderStats(claims);
                renderPipeline(claims);
                renderClaims(claims);
                renderTasks(data.tasks);
            } catch (e) {
                console.error('Failed to load data:', e);
            }
        }

        // Incremental log rendering — only APPEND new lines instead of replacing
        // the whole panel on every poll. Stops the 5s flicker that wiped any text
        // the user was selecting / reading. State:
        //   _logSeen        : Set of recently-rendered lines (de-dupe)
        //   _logRunAnchor   : timestamp set when a new task is sent; lines older
        //                    than this are filtered out so a fresh run starts clean
        let _logSeen = new Set();
        let _logRunAnchor = 0;
        const _MAX_LOG_LINES = 800;
        const _MAX_SEEN = 1500;

        function clearLogs(reason) {
            const body = document.getElementById('logs-body');
            if (body) body.innerHTML = `<div class="log-line info">— logs cleared (${reason}) —</div>`;
            _logSeen = new Set();
            _logRunAnchor = Date.now();
        }

        // Best-effort parse of journalctl's "May 21 03:25:11" prefix → ms timestamp.
        function _parseLogTs(line) {
            const m = (line || '').match(/^([A-Z][a-z]{2})\s+(\d+)\s+(\d{2}):(\d{2}):(\d{2})/);
            if (!m) return 0;
            const months = {Jan:0,Feb:1,Mar:2,Apr:3,May:4,Jun:5,Jul:6,Aug:7,Sep:8,Oct:9,Nov:10,Dec:11};
            const now = new Date();
            const d = new Date(now.getFullYear(), months[m[1]] || 0, parseInt(m[2]),
                               parseInt(m[3]), parseInt(m[4]), parseInt(m[5]));
            return d.getTime();
        }

        async function loadLogs() {
            try {
                const res = await fetch('/api/logs?bot=' + encodeURIComponent(window.activeBot || 'submissions'));
                const data = await res.json();
                const body = document.getElementById('logs-body');
                if (!body) return;
                // First load: clear the "Loading..." placeholder.
                if (body.firstElementChild && body.firstElementChild.textContent === 'Loading logs...') {
                    body.innerHTML = '';
                }
                // Preserve user's scroll position unless they're pinned to the bottom.
                const atBottom = (body.scrollHeight - body.scrollTop - body.clientHeight) < 40;
                const frag = document.createDocumentFragment();
                let added = 0;
                for (const line of (data.logs || [])) {
                    if (_logSeen.has(line)) continue;
                    if (_logRunAnchor && _parseLogTs(line) > 0 && _parseLogTs(line) < _logRunAnchor) continue;
                    _logSeen.add(line);
                    let cls = 'info';
                    if (line.includes('ERROR')) cls = 'error';
                    else if (line.includes('WARNING')) cls = 'warning';
                    else if (line.includes('complete') || line.includes('SUCCESS')) cls = 'success';
                    const div = document.createElement('div');
                    div.className = 'log-line ' + cls;
                    div.textContent = line;
                    frag.appendChild(div);
                    added++;
                }
                if (added > 0) {
                    body.appendChild(frag);
                    // Trim the panel — keep newest _MAX_LOG_LINES only.
                    while (body.childElementCount > _MAX_LOG_LINES) body.removeChild(body.firstElementChild);
                    // Trim the seen-set so it doesn't grow forever.
                    if (_logSeen.size > _MAX_SEEN) {
                        const arr = Array.from(_logSeen);
                        _logSeen = new Set(arr.slice(arr.length - _MAX_SEEN / 2));
                    }
                    if (atBottom) body.scrollTop = body.scrollHeight;
                }
            } catch (e) {
                console.error('Failed to load logs:', e);
            }
        }

        function renderStats(claims) {
            const getState = c => parseInt(c.state || c.current_state || 0);
            const submittedStates = PIPELINE_STAGES.submitted.states;

            const total = claims.length;
            const submittedCount = claims.filter(c => submittedStates.includes(getState(c))).length;
            const pct = total > 0 ? Math.round((submittedCount / total) * 100) : 0;

            document.getElementById('hero-submitted').textContent = submittedCount;
            document.getElementById('hero-total').textContent = total;
            document.getElementById('hero-pct').textContent = `${pct}%`;
            document.getElementById('hero-remaining').textContent = total - submittedCount;
            document.getElementById('hero-progress-fill').style.width = `${pct}%`;
        }

        function renderPipeline(claims) {
            const grid = document.getElementById('pipeline-grid');
            if (!grid) return;
            const getState = c => parseInt(c.state || c.current_state || 0);
            grid.innerHTML = '';

            const stageKeys = Object.keys(PIPELINE_STAGES);
            stageKeys.forEach((key, idx) => {
                const stage = PIPELINE_STAGES[key];
                const stageClaims = claims.filter(c => stage.states.includes(getState(c)));
                const count = stageClaims.length;

                // Build detail chips showing internal state breakdown
                const breakdown = {};
                for (const c of stageClaims) {
                    const s = getState(c);
                    const label = STATE_LABELS[s] || `State ${s}`;
                    breakdown[label] = (breakdown[label] || 0) + 1;
                }
                const detailChips = Object.entries(breakdown).map(([label, cnt]) =>
                    `<span class="detail-chip has-count">${cnt} ${label}</span>`
                ).join('');

                const arrow = idx < stageKeys.length - 1
                    ? '<div class="stage-arrow">→</div>'
                    : '';

                grid.innerHTML += `
                    <div class="pipeline-stage" onclick="openBSDrawer('${key}', '${stage.label}')">
                        <div class="stage-count" style="color: ${stage.color};">${count}</div>
                        <div class="stage-name">${stage.label}</div>
                        <div class="stage-bar" style="background: ${stage.color};"></div>
                        ${detailChips ? `<div class="stage-details">${detailChips}</div>` : ''}
                        ${arrow}
                    </div>`;
            });
        }

        function renderClaims(claims) {
            const body = document.getElementById('claims-body');
            const meta = document.getElementById('claims-meta');
            if (!claims.length) {
                body.innerHTML = '<tr><td colspan="11" class="empty-state">No claims yet. Send a task to begin.</td></tr>';
                if (meta) meta.textContent = '';
                return;
            }
            // Show metadata about the last scrape
            if (meta) {
                const scraped = claims.find(c => c.scraped_at);
                meta.textContent = scraped ? `Last scraped: ${scraped.scraped_at}` : '';
            }
            // Find the single most-recent processing_at across ALL claims. The bot
            // processes claims serially, so only the claim with the LATEST timestamp
            // (and within the last 90s, with a non-empty current_step) is "live".
            let _liveProcessingAt = null;
            let _liveProcessingClaimId = null;
            for (const cc of claims) {
                if (!cc.processing_at) continue;
                if (!(cc.current_step || '').trim()) continue;
                try {
                    const _ts = new Date(String(cc.processing_at).replace(' UTC', 'Z').replace(' ', 'T')).getTime();
                    if (!isNaN(_ts) && (Date.now() - _ts) < 90000) {
                        if (_liveProcessingAt === null || _ts > _liveProcessingAt) {
                            _liveProcessingAt = _ts;
                            _liveProcessingClaimId = cc.claim_id;
                        }
                    }
                } catch(e) {}
            }
            // Sort claims by readiness so, top→bottom, the user sees:
            //   0. The claim being processed RIGHT NOW (live)
            //   1. Pending — still missing HCFA / IV / Progress note
            //   2. Ready to submit — all docs captured, not yet sent
            //   3. Submitted — done, sinks to the bottom (out of attention)
            // Within each group, newer service dates first.
            const _sortOfficeRe = /\b(9920[1-5]|9921[1-5])\b/;
            const _claimNeedsWork = (c) => {
                if (c.symplisend_submitted) return false;
                const isOffice = !!c.office_visit || _sortOfficeRe.test(String(c.cpt || ''));
                const hcfaOk = !!c.hcfa_s3_path;
                const ivOk = !!c.prog_notes_s3_path;
                const encOk = isOffice || !!c.encounter_file_s3_path;
                return !(hcfaOk && ivOk && encOk);
            };
            const _readinessRank = (c) => {
                if (_liveProcessingClaimId !== null && c.claim_id === _liveProcessingClaimId) return 0;
                if (c.symplisend_submitted) return 3;
                if (_claimNeedsWork(c)) return 1;
                return 2;
            };
            claims.sort((a, b) => {
                const ra = _readinessRank(a);
                const rb = _readinessRank(b);
                if (ra !== rb) return ra - rb;
                return (b.service_date || '').localeCompare(a.service_date || '');
            });
            // If every claim shares one payer, hide the column and surface the
            // payer name once in the section title — saves a wide column for a
            // value that never changes.
            const _uniquePayers = [...new Set(claims.map(c => (c.payer || '').trim()).filter(Boolean))];
            const _payerTitle = document.getElementById('claims-payer-title');
            const _hidePayer = _uniquePayers.length === 1;
            if (_payerTitle) _payerTitle.textContent = _hidePayer ? `· ${_uniquePayers[0]}` : '';
            const _tableWrap = document.querySelector('.claims-table-wrap');
            if (_tableWrap) _tableWrap.classList.toggle('hide-payer', _hidePayer);

            // Type counts — Office Visit vs IV Therapy, plus IVs missing their
            // progress note (encounter_file_s3_path empty) so reviewers see at a
            // glance how many still need a doc.
            const _officeRe = /\b(9920[1-5]|9921[1-5])\b/;
            let _ovCount = 0, _ivCount = 0, _ivNoProgCount = 0;
            for (const c of claims) {
                const isOfc = !!c.office_visit || _officeRe.test(String(c.cpt || ''));
                if (isOfc) {
                    _ovCount++;
                } else {
                    _ivCount++;
                    if (!c.encounter_file_s3_path) _ivNoProgCount++;
                }
            }
            const _typeCounts = document.getElementById('claims-type-counts');
            if (_typeCounts) {
                const missingPart = _ivNoProgCount > 0
                    ? ` <span style="color:var(--text-muted);margin:0 4px;">·</span>` +
                      `<span style="color:#ef4444;font-weight:600;" title="IV Therapy claims with no progress note captured yet">🚫 ${_ivNoProgCount} IV${_ivNoProgCount === 1 ? '' : 's'} without progress note</span>`
                    : '';
                _typeCounts.innerHTML = `<span style="color:#3b82f6;">🏥 ${_ovCount} Office Visit${_ovCount === 1 ? '' : 's'}</span>` +
                                        ` <span style="color:var(--text-muted);margin:0 4px;">·</span>` +
                                        `<span style="color:var(--text-muted);">💉 ${_ivCount} IV Therap${_ivCount === 1 ? 'y' : 'ies'}</span>` +
                                        missingPart;
            }

            body.innerHTML = claims.map(c => {
                const state = parseInt(c.state || c.current_state || 0);
                const stageKey = getStageKey(state);
                // HCFA status — clickable if generated (opens actual PDF)
                const hasHCFA = !!c.hcfa_s3_path || state >= 2;
                const hcfaTitle = c.hcfa_generated_at ? `Generated: ${c.hcfa_generated_at}` : '';
                const hcfaCell = hasHCFA
                    ? `<span class="hcfa-link" onclick="window.open('/api/hcfa_pdf/${c.claim_id}', '_blank')" title="${hcfaTitle}">📄 HCFA</span>`
                    : '—';
                // Submission type from Box 22
                const subType = c.submission_type || (state >= 2 ? 'Pending' : '—');
                const refNo = c.original_ref_no || '';
                const subscriberId = c.subscriber_id || '';
                const subColor = subType === 'First Time Submission' ? '#4ade80' 
                    : subType === 'Resubmission' ? '#f59e0b' : 'var(--text-muted)';
                let subCell;
                if (subType === 'Resubmission' && refNo) {
                    subCell = `<span style="color:${subColor};font-weight:600;font-size:11px;">${subType}</span><br><span style="font-size:10px;color:var(--text-muted);">Ref: ${refNo}</span>`;
                } else if (subType === 'First Time Submission') {
                    subCell = `<span style="color:${subColor};font-weight:600;font-size:11px;">${subType}</span>`
                        + (subscriberId ? `<br><span style="font-size:10px;color:var(--text-muted);">ID: ${subscriberId}</span>` : '');
                } else {
                    subCell = `<span style="color:${subColor};font-weight:500;font-size:11px;">${subType}</span>`;
                }
                // ECW claim-status indicator (post-submission status update in eCW).
                // Green ✓ once the bot set the claim to "Claim sent via Symplisend" and
                // verified it; amber ⏳ while submitted but the ECW status isn't updated yet.
                if (c.ecw_status_updated) {
                    const _ecwTip = 'ECW status set to "' + (c.ecw_status_code || 'Claim sent via Symplisend') + '"'
                        + (c.ecw_status_updated_at ? ' · ' + c.ecw_status_updated_at : '');
                    subCell += `<br><span style="font-size:10px;color:#4ade80;font-weight:600;" title="${_ecwTip}">✓ ECW status updated</span>`;
                } else if (c.symplisend_submitted) {
                    subCell += `<br><span style="font-size:10px;color:var(--warning);" title="Submitted to SympliSend — ECW claim status not updated yet">⏳ ECW status pending</span>`;
                }
                // Progress Notes — clickable if captured
                const hasProgNotes = !!c.prog_notes_s3_path;
                const progNotesCell = hasProgNotes
                    ? `<span class="hcfa-link" onclick="window.open('/api/prog_notes/${c.claim_id}', '_blank')" title="Captured: ${c.prog_notes_captured_at || ''}">📝 IV Note</span>`
                    : '—';
                // Office visit (E/M CPT 99201-99205 / 99211-99215) → only HCFA + IV Note,
                // no Progress Note required. Honor the persisted flag or detect from CPT.
                const isOffice = !!c.office_visit || /\b(9920[1-5]|9921[1-5])\b/.test(String(c.cpt || ''));
                const officeVisitCell = isOffice
                    ? `<span style="color:#3b82f6;font-size:11px;font-weight:600;" title="CPT ${c.cpt || ''} — Progress Note not required">🏥 Office Visit</span>`
                    : `<span style="color:var(--text-muted);font-size:11px;">💉 IV Therapy</span>`;
                // Encounter File — clickable if captured; N/A for office visits;
                // "Not Found" (red) when the Encounters tab had no row matching
                // the Rx Start Date + allowed visit type; "Needs Review" (yellow)
                // for any other failure (capture/PDF/click problems where the
                // encounter exists but couldn't be retrieved).
                // Manual-upload affordance for any failed state (Not Found /
                // Needs Review): user clicks "Upload" → file picker → POST PDF
                // to /api/encounter_file/<cid>/upload → DDB cleared + S3 written.
                // "Skip" button flags the claim as not needing a Progress Note
                // (manual override for diagnostic-only / mixed CPT cases).
                const uploadBtn = `<a href="#" onclick="event.stopPropagation();uploadEncounterFile('${c.claim_id}');return false;" style="margin-left:6px;color:var(--info);text-decoration:underline;font-size:10px;">📤 Upload</a>`;
                const skipBtn = `<a href="#" onclick="event.stopPropagation();skipProgressNote('${c.claim_id}', true);return false;" style="margin-left:6px;color:var(--success);text-decoration:underline;font-size:10px;" title="Mark this claim as not needing a Progress Note (submittable without one)">✓ Skip</a>`;
                let encFileCell = '—';
                if (c.encounter_file_s3_path) {
                    encFileCell = `<span class="hcfa-link" onclick="window.open('/api/encounter_file/${c.claim_id}', '_blank')" title="Captured: ${c.encounter_file_captured_at || ''}">📄 Progress Note</span>`;
                } else if (isOffice) {
                    encFileCell = `<span style="color:var(--text-muted);font-size:11px;">N/A · Office Visit</span>`;
                } else if (c.progress_note_not_required) {
                    // Manual override flag set — claim is treated like no-progress-note-needed.
                    encFileCell = `<span style="color:#4ade80;font-size:11px;font-weight:600;" title="Manually flagged: no Progress Note needed">✓ Not Required</span>` +
                                  ` <a href="#" onclick="event.stopPropagation();skipProgressNote('${c.claim_id}', false);return false;" style="margin-left:4px;color:var(--text-muted);text-decoration:underline;font-size:10px;" title="Undo manual override">undo</a>`;
                } else if (c.encounter_capture_failed === 'no_np_or_fu_visit_type'
                           || c.encounter_capture_failed === 'no_rx_start_date') {
                    const reason = c.encounter_capture_failed === 'no_rx_start_date'
                        ? 'No Rx Start Date parsed from IV Note'
                        : 'No encounter on target date with allowed visit type';
                    encFileCell = `<span style="color:#ef4444;font-size:11px;font-weight:600;" title="${reason}">🚫 Not Found</span>${uploadBtn}${skipBtn}`;
                } else if (c.encounter_revision_needed) {
                    encFileCell = `<span style="color:#f59e0b;font-size:11px;font-weight:600;" title="${c.encounter_capture_failed || 'capture failed'}">⚠️ Needs Review</span>${uploadBtn}${skipBtn}`;
                }
                // Currently-processing indicator: only the SINGLE claim with the
                // newest processing_at AND a non-empty current_step lights up. The
                // bot clears current_step at end of each iteration, so the badge
                // disappears as soon as work moves on, not 90s later.
                const _stepLive = (c.current_step || '').trim();
                const isProcessing = (_liveProcessingClaimId !== null
                                      && c.claim_id === _liveProcessingClaimId
                                      && _stepLive.length > 0);
                const rowClass = isProcessing ? 'processing-row' : '';
                const procBadge = isProcessing
                    ? `<br><span class="proc-badge"><span class="proc-spin"></span>${_stepLive}</span>`
                    : '';
                return `<tr class="${rowClass}">
                    <td style="font-weight:600;color:var(--accent);">${c.claim_id || '—'}${procBadge}</td>
                    <td>${c.patient_name || '—'}</td>
                    <td>${c.encounter_date || c.service_date || c.dos || '—'}</td>
                    <td class="col-payer">${c.payer || '—'}</td>
                    <td>${c.charges || '—'}</td>
                    <td>${officeVisitCell}</td>
                    <td>${hcfaCell}</td>
                    <td>${progNotesCell}</td>
                    <td style="font-size:11px;color:var(--text-muted);">${formatProgNoteDate(c)}</td>
                    <td>${encFileCell}</td>
                    <td>${subCell}</td>
                    <td>${getStagePill(state)}</td>
                </tr>`;
            }).join('');

            // Simple per-doc rate indicator. Avg time/claim from deltas between
            // consecutive recent processing_at stamps; drop outliers (<1s, >5min)
            // to ignore stuck claims. This is just informational — not a live
            // backlog countdown.
            const _parseTs = (x) => {
                if (!x) return NaN;
                try { return new Date(String(x).replace(' UTC', 'Z').replace(' ', 'T')).getTime(); }
                catch(e) { return NaN; }
            };
            const _tsAsc = claims
                .map(c => _parseTs(c.processing_at))
                .filter(t => !isNaN(t))
                .sort((a, b) => a - b);
            const _deltas = [];
            for (let i = 1; i < _tsAsc.length; i++) {
                const d = _tsAsc[i] - _tsAsc[i - 1];
                if (d > 1000 && d < 300000) _deltas.push(d);
            }
            const _avgMs = _deltas.length > 0 ? _deltas.reduce((a, b) => a + b, 0) / _deltas.length : null;
            const TYPICAL_UNITS_PER_CLAIM = 2.5;
            const _perUnitSec = _avgMs !== null ? Math.round((_avgMs / TYPICAL_UNITS_PER_CLAIM) / 1000) : null;

            window._etaState = _perUnitSec !== null ? { perUnitSec: _perUnitSec } : null;
            renderEta();
        }

        function renderEta() {
            const etaEl = document.getElementById('claims-eta');
            if (!etaEl) return;
            const s = window._etaState;
            if (!s || s.perUnitSec == null) { etaEl.style.display = 'none'; return; }
            etaEl.style.display = 'inline';
            etaEl.style.color = 'var(--text-muted)';
            etaEl.style.fontWeight = '500';
            etaEl.textContent = `⏱ ~${s.perUnitSec}s/doc avg`;
        }

        function renderTasks(tasks) {
            const list = document.getElementById('tasks-list');
            if (!tasks.length) {
                list.innerHTML = '<div class="empty-state" style="padding:20px;">No open tasks</div>';
                return;
            }
            list.innerHTML = tasks.map(t => `
                <div style="padding:10px;border-bottom:1px solid var(--border);font-size:13px;">
                    <strong style="color:var(--warning);">${t.task_type}</strong> — ${t.claim_id || ''}
                    <div style="color:var(--text-muted);font-size:11px;margin-top:4px;">${t.notes || ''}</div>
                </div>
            `).join('');
        }

        function showToast(msg, type) {
            const toast = document.getElementById('toast');
            toast.textContent = msg;
            toast.className = 'toast ' + type;
            setTimeout(() => { toast.className = 'toast'; }, 3000);
        }

        // ─── BS Claims Drawer ───
        let bsClaimsCache = null;

        async function loadBSClaims() {
            try {
                const res = await fetch('/api/bs-claims');
                bsClaimsCache = await res.json();
            } catch (e) {
                console.error('Failed to load BS claims:', e);
            }
        }

        function openBSDrawer(stageKey, stageName) {
            document.getElementById('bs-overlay').classList.add('active');
            document.getElementById('bs-drawer').classList.add('active');
            document.getElementById('bs-drawer-title').textContent = stageName;

            // Show claims for this stage
            const stage = PIPELINE_STAGES[stageKey];
            if (!stage) return;

            fetch('/api/claims').then(r => r.json()).then(data => {
                const claims = data.claims || [];
                const filtered = claims.filter(c => {
                    const s = parseInt(c.state || c.current_state || 0);
                    return stage.states.includes(s);
                });

                document.getElementById('bs-drawer-meta').innerHTML =
                    `<span>📊 <strong>${filtered.length}</strong> claims in this stage</span>`;

                if (filtered.length === 0) {
                    document.getElementById('bs-drawer-body').innerHTML =
                        `<div class="empty-state">No claims in <strong>${stageName}</strong> yet.</div>`;
                    return;
                }

                document.getElementById('bs-drawer-body').innerHTML = filtered.map(c => {
                    const state = parseInt(c.state || c.current_state || 0);
                    return `<div class="bs-claim-card">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;">
                            <strong style="color:var(--accent);">${c.patient_name || 'Unknown'}</strong>
                            <span class="state-pill documentation" style="font-size:10px;">${STATE_LABELS[state] || 'State ' + state}</span>
                        </div>
                        <div style="display:grid;grid-template-columns:1fr 1fr;gap:4px 16px;font-size:11px;color:var(--text-secondary);">
                            <span>Claim #: <strong style="color:var(--text-primary);">${c.claim_id}</strong></span>
                            <span>DOS: ${c.service_date || '—'}</span>
                            <span>DOS: <strong>${c.encounter_date || '—'}</strong></span>
                            <span>Charges: <strong style="color:var(--warning);">${c.charges || '—'}</strong></span>
                            <span>Subscriber: ${c.subscriber_id || '—'}</span>
                            <span>Payer: ${c.payer || '—'}</span>
                        </div>
                    </div>`;
                }).join('');
            });
        }

        function closeBSDrawer() {
            document.getElementById('bs-overlay').classList.remove('active');
            document.getElementById('bs-drawer').classList.remove('active');
        }

        // ─── HCFA Viewer ───
        let hcfaCurrentClaimId = null;
        let hcfaCurrentView = 'popup'; // 'popup' or 'after_click'

        async function openHCFAViewer(claimId, patientName) {
            hcfaCurrentClaimId = claimId;
            hcfaCurrentView = 'popup';

            document.getElementById('hcfa-lightbox').classList.add('active');
            document.getElementById('hcfa-lightbox-title').textContent =
                `HCFA — Claim ${claimId} (${patientName})`;

            // Build tabs
            const tabsEl = document.getElementById('hcfa-tabs');
            tabsEl.innerHTML = `
                <button class="hcfa-tab active" data-view="popup" onclick="switchHCFATab('popup', this)">📋 Claim Popup</button>
                <button class="hcfa-tab" data-view="after_click" onclick="switchHCFATab('after_click', this)">📄 After HCFA Click</button>
            `;

            loadHCFAImage(claimId, 'popup');
        }

        function switchHCFATab(view, tabEl) {
            // Update tab styles
            document.querySelectorAll('.hcfa-tab').forEach(t => t.classList.remove('active'));
            tabEl.classList.add('active');
            hcfaCurrentView = view;
            loadHCFAImage(hcfaCurrentClaimId, view);
        }

        function loadHCFAImage(claimId, view) {
            const body = document.getElementById('hcfa-lightbox-body');
            const imgUrl = `/api/hcfa/${claimId}/${view}?t=${Date.now()}`;
            body.innerHTML = `<div class="hcfa-placeholder">Loading screenshot...</div>`;

            const img = new Image();
            img.onload = () => {
                body.innerHTML = '';
                body.appendChild(img);
            };
            img.onerror = () => {
                body.innerHTML = `<div class="hcfa-placeholder">
                    <div style="font-size:48px;margin-bottom:16px;">📭</div>
                    <div style="font-size:14px;">No screenshot available for this view.</div>
                    <div style="font-size:12px;margin-top:8px;color:var(--text-muted);">Run the HCFA generation task to capture screenshots.</div>
                </div>`;
            };
            img.src = imgUrl;
        }

        function closeHCFAViewer() {
            document.getElementById('hcfa-lightbox').classList.remove('active');
            hcfaCurrentClaimId = null;
        }

        // Close on Escape
        document.addEventListener('keydown', e => {
            if (e.key === 'Escape') {
                closeHCFAViewer();
                closeBSDrawer();
            }
        });

        // Stop Agent
        async function stopAgent() {
            if (!confirm('⛔ Stop the agent? This will kill any running browser, purge the SQS queue, and stop the service.')) return;
            const btn = document.getElementById('stopBtn');
            btn.disabled = true;
            btn.classList.add('stopping');
            btn.innerHTML = '⏳ Stopping...';
            try {
                const resp = await fetch('/api/stop-agent?bot=' + encodeURIComponent(window.activeBot || 'submissions'), { method: 'POST' });
                const d = await resp.json();
                if (d.success) {
                    btn.innerHTML = '✅ Stopped';
                    btn.style.borderColor = 'var(--success)';
                    btn.style.background = 'rgba(16,185,129,0.15)';
                    setTimeout(() => {
                        btn.innerHTML = '⛔ Stop Agent';
                        btn.disabled = false;
                        btn.classList.remove('stopping');
                        btn.style.borderColor = '';
                        btn.style.background = '';
                        loadLogs();
                    }, 3000);
                } else {
                    btn.innerHTML = '❌ Error';
                    alert('Stop failed: ' + (d.error || 'Unknown'));
                    setTimeout(() => {
                        btn.innerHTML = '⛔ Stop Agent';
                        btn.disabled = false;
                        btn.classList.remove('stopping');
                    }, 2000);
                }
            } catch (e) {
                btn.innerHTML = '❌ Error';
                alert('Stop failed: ' + e);
                setTimeout(() => {
                    btn.innerHTML = '⛔ Stop Agent';
                    btn.disabled = false;
                    btn.classList.remove('stopping');
                }, 2000);
            }
        }

        // ── Fix Coding IVs saved reports ──
        function _fmtBytes(b) {
            if (b < 1024) return b + ' B';
            if (b < 1048576) return (b / 1024).toFixed(1) + ' KB';
            return (b / 1048576).toFixed(1) + ' MB';
        }
        function _fmtRunId(rid) {
            // YYYYMMDD_HHMMSS → YYYY-MM-DD HH:MM:SS
            const m = rid.match(/^(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})$/);
            if (!m) return rid;
            return `${m[1]}-${m[2]}-${m[3]} ${m[4]}:${m[5]}:${m[6]}`;
        }
        async function deleteAllFixIvsFiles() {
            if (!confirm('⚠️ Borrar TODOS los saved reports de Fix Coding IVs en S3? Esto no se puede deshacer.')) return;
            try {
                const res = await fetch('/api/delete-fix-coding-ivs-files', { method: 'POST' });
                const data = await res.json();
                if (data.success) {
                    showToast(`🗑️ ${data.deleted} archivo(s) eliminado(s)`, 'success');
                    loadFixIvsFiles();
                } else {
                    showToast(`Error: ${data.error || 'unknown'}`, 'error');
                }
            } catch (e) {
                showToast(`Error: ${e.message}`, 'error');
            }
        }

        async function loadFixIvsFiles() {
            const wrap = document.getElementById('fix-ivs-files');
            if (!wrap) return;
            try {
                const res = await fetch('/api/fix-coding-ivs-files');
                const data = await res.json();
                if (!data.success) {
                    wrap.innerHTML = `<div class="empty-state" style="padding:16px;">${data.error || 'Failed to load'}</div>`;
                    return;
                }
                if (!data.runs || data.runs.length === 0) {
                    wrap.innerHTML = `<div class="empty-state" style="padding:16px;">No reports yet. Run the "Fix Coding IVs" task to generate Excel files.</div>`;
                    return;
                }
                const html = data.runs.slice(0, 8).map(run => {
                    const files = run.files.map(f => `
                        <a href="${f.url}" target="_blank" class="fix-ivs-file" download>
                            <span>📄 ${f.name}</span>
                            <span class="fix-ivs-size">${_fmtBytes(f.size)}</span>
                        </a>`).join('');
                    return `
                        <div class="fix-ivs-run">
                            <div class="fix-ivs-run-id">🕒 ${_fmtRunId(run.run_id)}</div>
                            ${files}
                        </div>`;
                }).join('');
                wrap.innerHTML = html;
            } catch (e) {
                wrap.innerHTML = `<div class="empty-state" style="padding:16px;">Error: ${e.message}</div>`;
            }
        }

        // Auto-refresh
        updateTaskTemplate();
        loadData();
        loadLogs();
        loadBSClaims();
        loadFixIvsFiles();
        setInterval(loadData, 4000);  // 4s — fast enough for live "processing" row updates
        setInterval(loadLogs, 5000);
        setInterval(loadBSClaims, 30000);
        setInterval(loadFixIvsFiles, 30000);
    </script>
</body>
</html>
"""


@app.route('/')
def dashboard():
    return render_template_string(DASHBOARD_HTML,
        state_labels=STATE_LABELS,
        pipeline_stages=PIPELINE_STAGES)


@app.route('/client')
def client_dashboard():
    """Client-facing read-only dashboard showing claims submission status."""
    template_path = os.path.join(os.path.dirname(__file__), 'templates', 'client.html')
    with open(template_path, 'r') as f:
        html = f.read()
    return html


def _scan_all(table):
    """Every item in a table, following DynamoDB's pagination.

    A bare `table.scan()` returns at most 1MB and silently stops there. On the
    claims table that was cutting the dashboard off at roughly half the data —
    it reported 690 of 1198 claims and called the backlog "100% complete", so
    the headline number was wrong in the reassuring direction.
    """
    items, kwargs = [], {}
    while True:
        resp = table.scan(**kwargs)
        items.extend(resp.get('Items', []))
        if 'LastEvaluatedKey' not in resp:
            return items
        kwargs['ExclusiveStartKey'] = resp['LastEvaluatedKey']


@app.route('/api/claims')
def api_claims():
    try:
        claims_table = dynamodb.Table('helixona-claims')
        tasks_table = dynamodb.Table('helixona-tasks')

        claims = _scan_all(claims_table)
        tasks = [t for t in _scan_all(tasks_table) if t.get('status') == 'Open']

        # Convert Decimal to int/float for JSON
        for c in claims:
            for k, v in c.items():
                if hasattr(v, 'real'):
                    c[k] = int(v)

        return jsonify({'claims': claims, 'tasks': tasks})
    except Exception as e:
        return jsonify({'claims': [], 'tasks': [], 'error': str(e)})


@app.route('/api/logs')
def api_logs():
    """Live agent log tail.

    Query params:
      n      : max line count to return (default 50, capped at 5000).
      level  : 'problems' → only WARNING / ERROR / CRITICAL / Traceback /
                Exception / ❌ / ⛔ / ⚠️ lines (great for triage).
      since  : journalctl --since value (e.g. '2 hours ago'). Default '30 min ago'.
    """
    try:
        n = int(request.args.get('n', 50))
        n = max(1, min(n, 5000))
        level = request.args.get('level', '')
        since = request.args.get('since', '30 min ago')
        service = BOT_ROUTING[_bot_from_request()]['service']
        cmd = ['journalctl', '-u', service, '--since', since, '--no-pager', '--output=short']
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=20)
        lines = [l.strip() for l in result.stdout.strip().split('\n') if l.strip()]
        cleaned = []
        for line in lines:
            if 'python[' in line:
                parts = line.split('python[')
                if len(parts) > 1:
                    cleaned.append(parts[1].split(']: ', 1)[-1] if ']: ' in parts[1] else parts[1])
                else:
                    cleaned.append(line)
            else:
                cleaned.append(line)
        if level == 'problems':
            markers = ('WARNING', 'ERROR', 'CRITICAL', 'Traceback', 'Exception',
                       '❌', '⛔', '⚠️')
            cleaned = [l for l in cleaned if any(m in l for m in markers)]
        return jsonify({'logs': cleaned[-n:], 'total': len(cleaned)})
    except Exception as e:
        return jsonify({'logs': [f'Error fetching logs: {str(e)}']})


@app.route('/api/send-task', methods=['POST'])
def api_send_task():
    try:
        payload = request.get_json() or {}
        bot = (payload.pop('bot', None) or 'submissions')
        if bot not in BOT_ROUTING:
            bot = 'submissions'
        routing = BOT_ROUTING[bot]
        if not routing['queue_url']:
            return jsonify({'success': False, 'error': f'No queue URL configured for bot={bot}'})
        sqs.send_message(
            QueueUrl=routing['queue_url'],
            MessageBody=json.dumps(payload)
        )
        # Ensure the right agent service is running (in case it was stopped)
        subprocess.run(
            ['sudo', 'systemctl', 'start', routing['service']],
            capture_output=True, timeout=10
        )
        return jsonify({'success': True, 'bot': bot})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/bs-claims')
def api_bs_claims():
    """Serve scraped Blue Shield claims data from the agent."""
    try:
        bs_file = '/opt/helixona-agent/bs_claims.json'
        if os.path.exists(bs_file):
            with open(bs_file) as f:
                data = json.load(f)
            return jsonify(data)
        else:
            return jsonify({'raw_claims': [], 'total_found': 0, 'scraped_at': None})
    except Exception as e:
        return jsonify({'error': str(e), 'raw_claims': [], 'total_found': 0})


@app.route('/api/hcfa_pdf/<claim_id>')
def api_hcfa_pdf(claim_id):
    """Serve HCFA PDF for a claim from /tmp or S3."""
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400

    local_path = f'/tmp/hcfa_{claim_id}.pdf'
    if os.path.exists(local_path) and os.path.getsize(local_path) > 500:
        return send_file(local_path, mimetype='application/pdf',
                         download_name=f'hcfa_{claim_id}.pdf')

    try:
        import boto3
        s3 = boto3.client('s3', region_name='us-west-2')
        bucket = 'helixona-claims-docs-eb2f8e3c'
        s3_key = f'hcfa_forms/{claim_id}_hcfa.pdf'
        local_dl = f'/tmp/hcfa_{claim_id}_dl.pdf'
        s3.download_file(bucket, s3_key, local_dl)
        return send_file(local_dl, mimetype='application/pdf',
                         download_name=f'hcfa_{claim_id}.pdf')
    except Exception as e:
        return jsonify({'error': f'HCFA PDF not found: {str(e)}'}), 404


@app.route('/api/hcfa/<claim_id>/<view_type>')
def api_hcfa_screenshot(claim_id, view_type):
    """Serve HCFA screenshots from /tmp on the EC2.
    
    view_type: 'popup' → /tmp/hcfa_popup_{claim_id}.png
               'after_click' → /tmp/hcfa_after_click_{claim_id}.png
    """
    import re
    # Sanitize claim_id to prevent path traversal
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400

    file_map = {
        'popup': f'/tmp/hcfa_popup_{claim_id}.png',
        'after_click': f'/tmp/hcfa_after_click_{claim_id}.png',
    }

    file_path = file_map.get(view_type)
    if not file_path or not os.path.exists(file_path):
        return jsonify({'error': 'Screenshot not found'}), 404

    return send_file(file_path, mimetype='image/png')


@app.route('/api/prog_notes/<claim_id>')
def api_prog_notes_pdf(claim_id):
    """Serve Progress Notes PDF for a claim.
    
    First check /tmp for the local file, then fall back to S3.
    """
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400

    # Check local /tmp first
    local_path = f'/tmp/prog_notes_{claim_id}.pdf'
    if os.path.exists(local_path) and os.path.getsize(local_path) > 500:
        return send_file(local_path, mimetype='application/pdf',
                         download_name=f'prog_notes_{claim_id}.pdf')

    # Fall back to S3
    try:
        import boto3
        s3 = boto3.client('s3', region_name='us-west-2')
        bucket = 'helixona-claims-docs-eb2f8e3c'
        s3_key = f'prog_notes/{claim_id}_prog_notes.pdf'
        local_dl = f'/tmp/prog_notes_{claim_id}_dl.pdf'
        s3.download_file(bucket, s3_key, local_dl)
        return send_file(local_dl, mimetype='application/pdf',
                         download_name=f'prog_notes_{claim_id}.pdf')
    except Exception as e:
        return jsonify({'error': f'Progress Notes not found: {str(e)}'}), 404


@app.route('/api/encounter_file/<claim_id>')
@app.route('/api/encounter_file/<claim_id>/<int:idx>')
def api_encounter_file(claim_id, idx=None):
    """Serve Encounter File PDF for a claim.

    The agent uploads each matching encounter as a numbered file:
      encounter_files/{claim_id}_encounter_{n}.pdf
    DynamoDB stores the first one in `encounter_file_s3_path` and the full list
    in `encounter_files_s3_paths`. With no idx, returns the first; with idx=N
    (1-based), returns the N-th captured encounter for that claim.
    """
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400

    # Try local /tmp first — agent writes encounter_file_{cid}_{n}.pdf
    nth = idx if idx else 1
    for local_candidate in [
        f'/tmp/encounter_file_{claim_id}_{nth}.pdf',
        f'/tmp/encounter_file_{claim_id}.pdf',  # legacy single-file path
    ]:
        if os.path.exists(local_candidate) and os.path.getsize(local_candidate) > 500:
            return send_file(local_candidate, mimetype='application/pdf',
                             download_name=f'encounter_{claim_id}.pdf')

    # Read the actual S3 path from DynamoDB
    s3_uri = None
    try:
        item = dynamodb.Table('helixona-claims').get_item(Key={'claim_id': claim_id}).get('Item', {})
        if idx and item.get('encounter_files_s3_paths'):
            paths = item['encounter_files_s3_paths']
            if 1 <= idx <= len(paths):
                s3_uri = paths[idx - 1]
        if not s3_uri:
            s3_uri = item.get('encounter_file_s3_path')
    except Exception as ddb_err:
        return jsonify({'error': f'DynamoDB lookup failed: {ddb_err}'}), 500

    # Fall back to the legacy hardcoded key if DDB has no record
    if not s3_uri:
        s3_uri = f's3://helixona-claims-docs-eb2f8e3c/encounter_files/{claim_id}_encounter_1.pdf'

    try:
        import boto3
        s3 = boto3.client('s3', region_name='us-west-2')
        # Parse s3://bucket/key into bucket + key
        if s3_uri.startswith('s3://'):
            without_scheme = s3_uri[5:]
            bucket, s3_key = without_scheme.split('/', 1)
        else:
            bucket = 'helixona-claims-docs-eb2f8e3c'
            s3_key = s3_uri
        local_dl = f'/tmp/encounter_file_{claim_id}_{nth}_dl.pdf'
        s3.download_file(bucket, s3_key, local_dl)
        return send_file(local_dl, mimetype='application/pdf',
                         download_name=f'encounter_{claim_id}.pdf')
    except Exception as e:
        return jsonify({'error': f'Encounter file not found: {str(e)}'}), 404


@app.route('/api/encounter_files/<claim_id>/list')
def api_encounter_files_list(claim_id):
    """List all encounter file URLs + visit types for a claim (from DynamoDB)."""
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400
    try:
        item = dynamodb.Table('helixona-claims').get_item(Key={'claim_id': claim_id}).get('Item', {})
        paths = item.get('encounter_files_s3_paths', [])
        visit_types = item.get('encounter_files_visit_types', [])
        files = []
        for i, p in enumerate(paths, start=1):
            files.append({
                'idx': i,
                'visit_type': visit_types[i - 1] if i - 1 < len(visit_types) else '',
                'url': f'/api/encounter_file/{claim_id}/{i}',
                's3_path': p,
            })
        return jsonify({'claim_id': claim_id, 'count': len(files), 'files': files})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/encounter_file/<claim_id>/upload', methods=['POST'])
def api_encounter_file_upload(claim_id):
    """Manual upload of a Progress Note PDF for a claim the bot couldn't capture.
    Stores it at the same S3 key the bot would write (encounter_files/{cid}_encounter_1.pdf),
    updates DynamoDB so the claim is no longer flagged Not Found / Needs Review.
    """
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'No file in request'}), 400
    upload = request.files['file']
    if not upload.filename or not upload.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'File must be a PDF'}), 400
    # Save to /tmp, then upload to S3
    local_path = f'/tmp/manual_upload_{claim_id}.pdf'
    try:
        upload.save(local_path)
        size = os.path.getsize(local_path)
        if size < 500:
            return jsonify({'error': f'File too small ({size} bytes)'}), 400
        import boto3
        s3 = boto3.client('s3', region_name='us-west-2')
        s3_key = f'encounter_files/{claim_id}_encounter_1.pdf'
        bucket = 'helixona-claims-docs-eb2f8e3c'
        s3.upload_file(local_path, bucket, s3_key)
        s3_uri = f's3://{bucket}/{s3_key}'
        # Update DynamoDB — clear Not Found / Needs Review state, mark manual upload.
        import time as _t
        now = _t.strftime('%Y-%m-%d %H:%M:%S UTC', _t.gmtime())
        dynamodb.Table('helixona-claims').update_item(
            Key={'claim_id': claim_id},
            UpdateExpression=(
                'SET encounter_file_s3_path = :p, encounter_files_s3_paths = :pl, '
                'encounter_files_count = :one, encounter_files_visit_types = :vt, '
                'encounter_file_captured_at = :now, encounter_revision_needed = :f, '
                'encounter_date_mismatch = :f, encounter_pick_strategy = :ms '
                'REMOVE encounter_capture_failed, encounter_capture_failed_at'
            ),
            ExpressionAttributeValues={
                ':p': s3_uri, ':pl': [s3_uri], ':one': 1, ':vt': ['MANUAL'],
                ':now': now, ':f': False, ':ms': 'manual_upload',
            },
        )
        try: os.remove(local_path)
        except Exception: pass
        return jsonify({'ok': True, 's3_path': s3_uri, 'size': size})
    except Exception as e:
        try: os.remove(local_path)
        except Exception: pass
        return jsonify({'error': str(e)}), 500


@app.route('/api/claim/<claim_id>/skip_progress_note', methods=['POST'])
def api_skip_progress_note(claim_id):
    """Manual override: flag this claim as not requiring a Progress Note so it
    becomes submittable without one. Used for claims like diagnostic-only or
    other CPT codes the reviewer knows don't need an encounter file.
    Body: {"skip": true|false}. true sets the flag, false removes it.
    """
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400
    body = request.get_json(silent=True) or {}
    skip = bool(body.get('skip'))
    try:
        if skip:
            # Set flag + clear the Needs Review state so the claim is submittable.
            dynamodb.Table('helixona-claims').update_item(
                Key={'claim_id': claim_id},
                UpdateExpression=(
                    'SET progress_note_not_required = :t, '
                    'encounter_revision_needed = :f '
                    'REMOVE encounter_capture_failed, encounter_capture_failed_at'
                ),
                ExpressionAttributeValues={':t': True, ':f': False},
            )
        else:
            dynamodb.Table('helixona-claims').update_item(
                Key={'claim_id': claim_id},
                UpdateExpression='REMOVE progress_note_not_required',
            )
        return jsonify({'ok': True, 'skip': skip})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/claim/<claim_id>/prog_note_date', methods=['POST'])
def api_edit_prog_note_date(claim_id):
    """Manually set the Progress Note Date (iv_note_rx_start_date) for a claim.
    Accepts JSON {"date": "MM/DD/YYYY"} or {"date": "YYYY-MM-DD"} or empty string to clear.
    Stored canonically as YYYY-MM-DD in DDB to match what the bot writes.
    """
    import re
    if not re.match(r'^[a-zA-Z0-9_-]+$', claim_id):
        return jsonify({'error': 'Invalid claim ID'}), 400
    body = request.get_json(silent=True) or {}
    raw = (body.get('date') or '').strip()
    # Empty string → remove the field entirely.
    if not raw:
        try:
            dynamodb.Table('helixona-claims').update_item(
                Key={'claim_id': claim_id},
                UpdateExpression='REMOVE iv_note_rx_start_date',
            )
            return jsonify({'ok': True, 'cleared': True})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    # Normalize: accept MM/DD/YYYY or YYYY-MM-DD
    m_us = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', raw)
    m_iso = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', raw)
    if m_us:
        mm, dd, yyyy = m_us.groups()
        iso = f'{yyyy}-{mm}-{dd}'
    elif m_iso:
        iso = raw
    else:
        return jsonify({'error': 'Date must be MM/DD/YYYY or YYYY-MM-DD'}), 400
    try:
        dynamodb.Table('helixona-claims').update_item(
            Key={'claim_id': claim_id},
            UpdateExpression='SET iv_note_rx_start_date = :d',
            ExpressionAttributeValues={':d': iso},
        )
        return jsonify({'ok': True, 'date': iso})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stop-agent', methods=['POST'])
def api_stop_agent():
    """Emergency stop — kills browser processes and stops the agent."""
    try:
        # Kill any running chrome/chromium browsers spawned by the agent
        subprocess.run(
            ['pkill', '-f', 'chromium.*helixona'],
            capture_output=True, timeout=5
        )
        subprocess.run(
            ['pkill', '-f', 'chrome.*user-data-dir.*browser-profile'],
            capture_output=True, timeout=5
        )
        bot = _bot_from_request()
        routing = BOT_ROUTING[bot]
        # Purge the bot's SQS queue to prevent re-processing the same task
        if routing['queue_url']:
            try:
                sqs.purge_queue(QueueUrl=routing['queue_url'])
            except Exception:
                pass
        # STOP the bot's agent service (not restart)
        result = subprocess.run(
            ['sudo', 'systemctl', 'stop', routing['service']],
            capture_output=True, text=True, timeout=15
        )
        return jsonify({
            'success': result.returncode == 0,
            'message': 'Agent stopped. SQS queue purged.',
            'output': result.stdout + result.stderr
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/fix-coding-ivs-files')
def api_fix_coding_ivs_files():
    """List Excel files saved under fix_coding_ivs/ in S3, grouped by run timestamp."""
    try:
        s3 = boto3.client('s3', region_name='us-west-2')
        bucket = 'helixona-claims-docs-eb2f8e3c'
        prefix = 'fix_coding_ivs/'
        paginator = s3.get_paginator('list_objects_v2')
        runs = {}
        for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
            for obj in page.get('Contents', []):
                key = obj['Key']
                parts = key[len(prefix):].split('/', 1)
                if len(parts) != 2 or not parts[1]:
                    continue
                run_id, filename = parts
                presigned = s3.generate_presigned_url(
                    'get_object',
                    Params={'Bucket': bucket, 'Key': key, 'ResponseContentDisposition': f'attachment; filename="{filename}"'},
                    ExpiresIn=3600
                )
                runs.setdefault(run_id, []).append({
                    'name': filename,
                    'key': key,
                    'size': obj['Size'],
                    'last_modified': obj['LastModified'].isoformat(),
                    'url': presigned,
                })
        # Sort runs newest-first by run_id (which is a timestamp like YYYYMMDD_HHMMSS)
        ordered = sorted(runs.items(), key=lambda kv: kv[0], reverse=True)
        return jsonify({
            'success': True,
            'runs': [{'run_id': rid, 'files': sorted(files, key=lambda f: f['name'])} for rid, files in ordered]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'runs': []})


@app.route('/api/iv-claims')
def api_iv_claims():
    """Parses the most recent fix_coding_ivs combined.xlsx in S3 and returns
    each claim plus headline KPIs for the IV Fix Coding tab."""
    try:
        import io
        from openpyxl import load_workbook
        s3 = boto3.client('s3', region_name='us-west-2')
        bucket = 'helixona-claims-docs-eb2f8e3c'
        prefix = 'fix_coding_ivs/'
        paginator = s3.get_paginator('list_objects_v2')
        latest_run = None
        latest_combined_key = None
        latest_modified = None
        for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
            for obj in page.get('Contents', []):
                key = obj['Key']
                if not key.endswith('/combined.xlsx'):
                    continue
                parts = key[len(prefix):].split('/', 1)
                if len(parts) != 2:
                    continue
                run_id = parts[0]
                if latest_run is None or run_id > latest_run:
                    latest_run = run_id
                    latest_combined_key = key
                    latest_modified = obj['LastModified']
        if not latest_combined_key:
            return jsonify({'success': True, 'run': None, 'kpi': {'total': 0, 'pending': 0, 'pending_with_errors': 0}, 'claims': []})

        obj = s3.get_object(Bucket=bucket, Key=latest_combined_key)
        wb = load_workbook(io.BytesIO(obj['Body'].read()), data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [c.value for c in ws[1]]
        # Build header → column index (Excel has a leading None column; skip Nones)
        idx = {h: i for i, h in enumerate(headers) if h}

        def cell(row, key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        # CLAIM_TYPE column is added by Stage 3's enrichment step. If present we
        # filter to IVs only; if absent (older runs) we return all rows with an
        # `enriched: false` flag so the UI can warn the user.
        enriched = 'CLAIM_TYPE' in idx

        claims = []
        pending = 0
        pending_with_errors = 0
        unknown_count = 0
        other_count = 0
        for r in range(2, ws.max_row + 1):
            row = [c.value for c in ws[r]]
            claim_no = cell(row, 'CLAIMS#')
            if claim_no is None:
                continue
            ctype = (cell(row, 'CLAIM_TYPE') or '').strip() if enriched else ''
            if enriched:
                if ctype == 'IV':
                    pass  # keep
                else:
                    if ctype == 'Unknown':
                        unknown_count += 1
                    else:
                        other_count += 1
                    continue
            status = cell(row, 'STATUS') or ''
            if status == 'Pending With Errors':
                pending_with_errors += 1
            elif status == 'Pending':
                pending += 1
            claims.append({
                'claim_no': str(claim_no),
                'patient': cell(row, 'PATIENT') or '',
                'dos': cell(row, 'SERVICE DATE') or '',
                'charges': cell(row, 'CHARGES') or '',
                'payer': cell(row, 'PAYER') or '',
                'status': status,
                'balance': cell(row, 'Balance') or '',
                'provider': cell(row, 'Provider Name') or '',
            })
        return jsonify({
            'success': True,
            'run': latest_run,
            'run_at': latest_modified.isoformat() if latest_modified else None,
            'enriched': enriched,
            'kpi': {
                'total': len(claims),
                'pending': pending,
                'pending_with_errors': pending_with_errors,
                'excluded_other': other_count,
                'excluded_unknown': unknown_count,
            },
            'claims': claims,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'claims': []})


@app.route('/api/delete-fix-coding-ivs-files', methods=['POST'])
def api_delete_fix_coding_ivs_files():
    """Delete all objects under fix_coding_ivs/ in S3."""
    try:
        s3 = boto3.client('s3', region_name='us-west-2')
        bucket = 'helixona-claims-docs-eb2f8e3c'
        prefix = 'fix_coding_ivs/'
        paginator = s3.get_paginator('list_objects_v2')
        deleted = 0
        for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
            keys = [{'Key': obj['Key']} for obj in page.get('Contents', [])]
            if not keys:
                continue
            # delete_objects accepts up to 1000 keys per call
            for i in range(0, len(keys), 1000):
                batch = keys[i:i+1000]
                s3.delete_objects(Bucket=bucket, Delete={'Objects': batch, 'Quiet': True})
                deleted += len(batch)
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/delete-all-claims', methods=['POST'])
def api_delete_all_claims():
    """Delete all claims from DynamoDB for a fresh start."""
    try:
        table = dynamodb.Table('helixona-claims')
        scan = table.scan()
        delete_count = 0
        for item in scan.get('Items', []):
            cid = item['claim_id']
            table.delete_item(Key={'claim_id': cid})
            delete_count += 1
        return jsonify({'success': True, 'delete_count': delete_count, 'message': f'Deleted {delete_count} claims. Next pipeline run will rediscover from ECW.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ──────────────────────────────────────────────────────────
# SUBMISSION AUDIT LOG
# ──────────────────────────────────────────────────────────
# One row per submission attempt, append-only, written by
# src/audit/submission_log.py. This is the record shown to a payer who
# disputes having received documentation.

AUDIT_COLUMNS = [
    ('submitted_date_pt', 'Date'),
    ('submitted_time_pt', 'Time (PT)'),
    ('method', 'Method'),
    ('submission_form_type', 'Form type'),
    ('claim_id', 'Claim #'),
    ('blueshield_claim_number', 'BS claim #'),
    ('dos', 'DOS'),
    ('patient_name', 'Patient'),
    ('subscriber_id', 'Subscriber ID'),
    ('documents_label', 'Documents submitted'),
    ('outcome', 'Outcome'),
    ('fln', 'FLN'),
    ('linkage_label', 'Attached to prior claim?'),
    ('notes', 'Notes'),
]

DOC_LABELS = {
    'hcfa': 'HCFA-1500',
    'prog_notes': 'IV Note',
    'encounter': 'Progress Note',
}


def _linkage_risk(row):
    """True when a replacement claim was transmitted as a brand-new one.

    The claim itself is coded as a replacement, but the packet went to the
    payer under a first-submission form type — so the payer opens a fresh
    claim instead of attaching the documents to the one being disputed. This
    is the single most useful thing the log can surface: it marks exactly the
    submissions a payer is likely to report as never received.
    """
    return (
        str(row.get('claim_submission_type', '')).strip().lower() == 'resubmission'
        and 'first submission' in str(row.get('submission_form_type', '')).lower()
    )


def _load_audit_rows():
    """Fetch and normalise every audit row, newest first."""
    rows = _scan_all(dynamodb.Table('helixona-submissions'))

    out = []
    for r in rows:
        # Rows predating the audit log (written by the Stage 6 stub) lack the
        # timestamp fields. Skip them rather than render blank lines.
        if not r.get('submitted_at_utc'):
            continue
        for k, v in list(r.items()):
            if hasattr(v, 'real') and not isinstance(v, bool):
                r[k] = int(v)
        docs = r.get('documents') or []
        r['documents_label'] = ', '.join(
            DOC_LABELS.get(d.get('document', ''), d.get('document', '?')) for d in docs
        )
        r['linkage_risk'] = _linkage_risk(r)
        r['linkage_label'] = 'No — sent as a new claim' if r['linkage_risk'] else 'Yes'
        out.append(r)

    out.sort(key=lambda r: r.get('submitted_at_utc', ''), reverse=True)
    return out


def _filter_audit_rows(rows):
    """Apply the ?q= quick search plus ?from= / ?to= / ?outcome= / ?flag= filters.

    ?q= is one box over everything a payer would quote back at you — patient
    name, either claim number, date of service, subscriber ID. Asking staff to
    know which field a number belongs to is a step they should not have to take.
    The older per-field ?claim= / ?patient= params still work.
    """
    q = (request.args.get('q') or '').strip().lower()
    claim = (request.args.get('claim') or '').strip()
    patient = (request.args.get('patient') or '').strip().lower()
    date_from = (request.args.get('from') or '').strip()
    date_to = (request.args.get('to') or '').strip()
    outcome = (request.args.get('outcome') or '').strip().lower()
    flag = (request.args.get('flag') or '').strip().lower()
    sub_type = (request.args.get('type') or '').strip().lower()

    def keep(r):
        if sub_type and sub_type not in str(r.get('claim_submission_type', '')).lower():
            return False
        if q:
            haystack = ' '.join(str(r.get(k, '')) for k in (
                'patient_name', 'claim_id', 'blueshield_claim_number',
                'dos', 'subscriber_id', 'fln')).lower()
            if q not in haystack:
                return False
        if claim and claim not in (str(r.get('claim_id', '')),
                                   str(r.get('blueshield_claim_number', ''))):
            return False
        if patient and patient not in str(r.get('patient_name', '')).lower():
            return False
        if date_from and str(r.get('submitted_date_pt', '')) < date_from:
            return False
        if date_to and str(r.get('submitted_date_pt', '')) > date_to:
            return False
        if outcome and str(r.get('outcome', '')).lower() != outcome:
            return False
        if flag == 'unlinked' and not r.get('linkage_risk'):
            return False
        if flag == 'no-fln' and str(r.get('fln', '')):
            return False
        return True

    return [r for r in rows if keep(r)]


@app.route('/api/audit-log')
def api_audit_log():
    try:
        rows = _filter_audit_rows(_load_audit_rows())
        return jsonify({'rows': rows, 'count': len(rows)})
    except Exception as e:
        return jsonify({'rows': [], 'count': 0, 'error': str(e)})


@app.route('/api/audit-log.csv')
def api_audit_log_csv():
    """CSV export — the artefact you actually hand to a payer or auditor."""
    import csv
    import io

    try:
        rows = _filter_audit_rows(_load_audit_rows())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in AUDIT_COLUMNS])
    for r in rows:
        writer.writerow([r.get(key, '') for key, _ in AUDIT_COLUMNS])

    stamp = datetime.utcnow().strftime('%Y%m%d')
    return app.response_class(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition':
                 f'attachment; filename=helixona-submission-audit-{stamp}.csv'},
    )

AUDIT_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Submission Audit Log — Helixona</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;500;600;700&family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#07070a;--bg2:#0c0c12;--panel:#0f0f16;--panel2:#15151e;--card:#14141c;--card2:#1a1a24;
  --bdr:#1f1f2b;--bdr2:#2a2a38;
  --accent:#CDB486;--accent2:#b09968;--accent-glow:rgba(205,180,134,.18);
  --success:#10b981;--warning:#f59e0b;--bad:#ef4444;--info:#3b82f6;
  --text-primary:#f1f5f9;--text-secondary:#94a3b8;--text-muted:#64748b;--text-dim:#475569;
}
*{margin:0;padding:0;box-sizing:border-box}
html,body{background:var(--bg);color:var(--text-primary);font-family:'Inter',sans-serif;font-size:13px;-webkit-font-smoothing:antialiased}
a{color:inherit;text-decoration:none}
input,select,button{font-family:'Inter',sans-serif}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--bdr2);border-radius:3px}

.wrap{max-width:1500px;margin:0 auto;padding:22px 26px 70px}

/* ---------- header ---------- */
.top{display:flex;align-items:flex-start;justify-content:space-between;gap:20px;flex-wrap:wrap;margin-bottom:20px}
h1{font-family:'Playfair Display',serif;font-size:26px;font-weight:600;color:var(--accent);letter-spacing:-.3px}
.sub{color:var(--text-secondary);font-size:13px;margin-top:5px;max-width:640px;line-height:1.5}
.back{font-size:12px;color:var(--text-muted);border:1px solid var(--bdr);padding:8px 13px;border-radius:8px;transition:all .15s;white-space:nowrap}
.back:hover{border-color:var(--accent);color:var(--accent)}

/* ---------- stat tiles ---------- */
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(178px,1fr));gap:12px;margin-bottom:20px}
.tile{background:linear-gradient(180deg,var(--panel) 0%,var(--bg2) 100%);border:1px solid var(--bdr);border-radius:13px;padding:15px 17px;position:relative;overflow:hidden}
.tile .k{font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--text-muted);font-weight:600;margin-bottom:7px;display:flex;align-items:center;gap:5px}
.tile .v{font-size:29px;font-weight:700;letter-spacing:-1px;line-height:1;color:var(--text-primary)}
.tile .v.sm{font-size:15px;font-weight:600;letter-spacing:-.2px;padding-top:7px}
.tile .n{font-size:11px;color:var(--text-muted);margin-top:6px;line-height:1.4}
.tile.hero .v{color:var(--accent);text-shadow:0 0 24px var(--accent-glow)}
.tile.warn{border-color:rgba(245,158,11,.32)}
.tile.warn .v{color:var(--warning)}
.tile.crit{border-color:rgba(239,68,68,.34)}
.tile.crit .v{color:var(--bad)}
.tile.act{cursor:pointer;transition:all .15s}
.tile.act:hover{border-color:var(--accent);transform:translateY(-1px)}
.tile.on{border-color:var(--accent);box-shadow:0 0 0 1px var(--accent-glow)}
.dot{width:7px;height:7px;border-radius:50%;flex:none}
.dot.w{background:var(--warning)} .dot.c{background:var(--bad)} .dot.g{background:var(--success)}

/* ---------- controls ---------- */
.bar{display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap;background:var(--panel);border:1px solid var(--bdr);border-radius:13px;padding:14px 16px;margin-bottom:16px}
.fld{display:flex;flex-direction:column;gap:5px}
.fld label{font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:var(--text-muted);font-weight:600}
.fld input,.fld select{background:var(--bg2);border:1px solid var(--bdr);color:var(--text-primary);padding:9px 11px;border-radius:8px;font-size:12.5px;outline:none;transition:border-color .15s}
.fld input:focus,.fld select:focus{border-color:var(--accent)}
/* Native date controls default to the light theme: a near-black calendar icon
   on a near-black field, and a white popup. color-scheme:dark hands the whole
   native widget — icon and calendar — to the browser's dark rendering. */
.fld input[type=date]{color-scheme:dark;cursor:pointer}
.fld input[type=date]::-webkit-calendar-picker-indicator{cursor:pointer;opacity:.65;transition:opacity .15s}
.fld input[type=date]:hover::-webkit-calendar-picker-indicator{opacity:1}
.fld.grow{flex:1;min-width:250px}
.fld.grow input{width:100%}
.fld input::placeholder{color:var(--text-dim)}
.btn{padding:9px 15px;border-radius:8px;font-size:12px;font-weight:600;cursor:pointer;border:1px solid var(--bdr);background:var(--bg2);color:var(--text-secondary);transition:all .15s;white-space:nowrap}
.btn:hover{border-color:var(--accent);color:var(--accent)}
.btn.gold{background:linear-gradient(135deg,var(--accent),var(--accent2));color:#000;border-color:var(--accent)}
.btn.gold:hover{filter:brightness(1.08);color:#000}

/* ---------- table ---------- */
.count{color:var(--text-secondary);font-size:12px;margin-bottom:10px;display:flex;align-items:center;gap:9px;flex-wrap:wrap}
.count b{color:var(--text-primary)}
.chip{background:var(--accent-glow);color:var(--accent);padding:3px 9px;border-radius:20px;font-size:11px;font-weight:600;cursor:pointer}
.chip:hover{background:rgba(205,180,134,.28)}
.panel{background:var(--panel);border:1px solid var(--bdr);border-radius:13px;overflow:hidden}
.scroll{overflow-x:auto}
table{border-collapse:collapse;width:100%;min-width:900px}
th{background:var(--bg2);font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:var(--text-muted);font-weight:600;text-align:left;padding:11px 14px;border-bottom:1px solid var(--bdr);position:sticky;top:0;z-index:2}
td{padding:11px 14px;border-bottom:1px solid var(--bdr);vertical-align:top;font-size:12.5px}
tr.row{cursor:pointer;transition:background .12s}
tr.row:hover{background:var(--card)}
tr.row.open{background:var(--card2)}
.mono{font-family:'JetBrains Mono','Monaco',monospace;font-size:11.5px;font-variant-numeric:tabular-nums}
.pt{color:var(--text-primary);font-weight:500}
.dim{color:var(--text-muted)}
.docs{color:var(--text-secondary);line-height:1.5}
.nowrap{white-space:nowrap}
.pill{display:inline-flex;align-items:center;gap:5px;padding:3px 9px;border-radius:20px;font-size:10.5px;font-weight:600;white-space:nowrap}
.pill.submitted{background:rgba(16,185,129,.13);color:var(--success)}
.pill.failed{background:rgba(239,68,68,.13);color:var(--bad)}
.pill.blocked{background:rgba(245,158,11,.13);color:var(--warning)}
.ty{display:inline-block;padding:2px 8px;border-radius:6px;font-size:10.5px;font-weight:600;white-space:nowrap}
.ty.first{background:rgba(74,222,128,.12);color:#4ade80}
.ty.resub{background:rgba(245,158,11,.13);color:var(--warning)}
.caret{color:var(--text-dim);font-size:10px;transition:transform .15s;display:inline-block;width:11px}
tr.row.open .caret{transform:rotate(90deg);color:var(--accent)}

/* ---------- detail ---------- */
.detail td{background:var(--bg2);padding:0;border-bottom:1px solid var(--bdr)}
.dwrap{padding:17px 20px 19px;display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:17px}
.dsec h4{font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--accent);font-weight:600;margin-bottom:9px}
.dl{display:flex;gap:9px;font-size:12px;padding:3px 0;line-height:1.5}
.dl .dt{color:var(--text-muted);min-width:96px;flex:none}
.dl .dd{color:var(--text-primary);word-break:break-word}
.doc{background:var(--card);border:1px solid var(--bdr);border-radius:8px;padding:9px 11px;margin-bottom:7px}
.doc .dn{font-weight:600;font-size:12px;color:var(--text-primary);margin-bottom:3px}
.doc .dm{font-size:10.5px;color:var(--text-muted);font-family:'JetBrains Mono','Monaco',monospace;word-break:break-all;line-height:1.5}
.callout{grid-column:1/-1;background:rgba(239,68,68,.07);border:1px solid rgba(239,68,68,.28);border-radius:9px;padding:11px 14px;font-size:12px;color:var(--text-secondary);line-height:1.55}
.callout b{color:var(--bad)}
.empty{padding:52px 20px;text-align:center;color:var(--text-muted)}
.empty .big{font-size:15px;color:var(--text-secondary);margin-bottom:6px}
.more{padding:15px;text-align:center;border-top:1px solid var(--bdr)}
.loading{padding:52px;text-align:center;color:var(--text-muted)}
</style>
</head>
<body>
<div class="wrap">

  <div class="top">
    <div>
      <h1>Submission Audit Log</h1>
      <div class="sub">Every documentation packet sent to a payer — when it went, how, for which
      claim and date of service, and exactly which documents were included.</div>
    </div>
    <a class="back" href="/">&larr; Dashboard</a>
  </div>

  <div class="tiles" id="tiles"></div>

  <div class="bar">
    <div class="fld grow">
      <label>Search</label>
      <input id="q" placeholder="Patient, claim #, Blue Shield claim #, date of service, subscriber ID…" autocomplete="off">
    </div>
    <div class="fld"><label>From</label><input id="from" type="date"></div>
    <div class="fld"><label>To</label><input id="to" type="date"></div>
    <div class="fld"><label>Type</label>
      <select id="type">
        <option value="">All types</option>
        <option value="first time">New submission</option>
        <option value="resubmission">Resubmission</option>
      </select>
    </div>
    <div class="fld"><label>Outcome</label>
      <select id="outcome">
        <option value="">All</option>
        <option value="submitted">Submitted</option>
        <option value="failed">Failed</option>
        <option value="blocked">Blocked</option>
      </select>
    </div>
    <button class="btn" onclick="reset()">Clear</button>
    <a class="btn gold" id="csv" href="/api/audit-log.csv">Export CSV</a>
  </div>

  <div class="count" id="count">Loading…</div>
  <div class="panel">
    <div class="scroll">
      <table>
        <thead><tr>
          <th style="width:26px"></th>
          <th class="nowrap">Sent</th>
          <th>Patient</th>
          <th class="nowrap">Claim #</th>
          <th class="nowrap">Date of service</th>
          <th class="nowrap">Type</th>
          <th>Documents</th>
          <th class="nowrap">Status</th>
        </tr></thead>
        <tbody id="body"><tr><td colspan="8" class="loading">Loading submissions…</td></tr></tbody>
      </table>
    </div>
    <div class="more" id="more" style="display:none">
      <button class="btn" onclick="showMore()">Show more</button>
    </div>
  </div>
</div>

<script>
const PAGE = 60;
let ALL = [], SHOWN = [], limit = PAGE, flag = '';

const esc = s => String(s == null ? '' : s).replace(/[&<>"]/g,
  c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));

const DOCN = {hcfa:'HCFA-1500', prog_notes:'IV Note', encounter:'Progress Note'};

function qs() {
  const p = new URLSearchParams();
  for (const id of ['q','from','to','type','outcome']) {
    const v = document.getElementById(id).value.trim();
    if (v) p.set(id, v);
  }
  if (flag) p.set('flag', flag);
  return p.toString();
}

function tiles() {
  // Tiles describe what is currently on screen, not a fixed global total —
  // filtering to one patient should answer "how many of THEIRS went out
  // unattached", which is the question being asked at that moment.
  const view = SHOWN;
  const n = view.length;
  const filtered = n !== ALL.length;
  const bad = view.filter(r => r.outcome !== 'submitted').length;
  const dates = view.map(r => r.submitted_date_pt).filter(Boolean).sort();
  const range = dates.length ? dates[0] + ' → ' + dates[dates.length - 1] : '—';

  document.getElementById('tiles').innerHTML = `
    <div class="tile hero">
      <div class="k">${filtered ? 'Submissions in this view' : 'Submissions on record'}</div>
      <div class="v">${n.toLocaleString()}</div>
      <div class="n">${filtered ? 'of ' + ALL.length.toLocaleString() + ' on record' : 'packets sent to a payer'}</div>
    </div>
    <div class="tile">
      <div class="k">Period covered</div>
      <div class="v sm">${esc(range)}</div>
      <div class="n">first to most recent</div>
    </div>
    <div class="tile ${bad?'warn':''}">
      <div class="k"><span class="dot ${bad?'w':'g'}"></span>Failed or blocked</div>
      <div class="v">${bad.toLocaleString()}</div>
      <div class="n">never reached the payer</div>
    </div>`;
}

function setFlag(f) { flag = (flag === f) ? '' : f; limit = PAGE; apply(); }
function reset() {
  for (const id of ['q','from','to','type','outcome']) document.getElementById(id).value = '';
  flag = ''; limit = PAGE; apply();
}

function rowHtml(r, i) {
  const docs = (r.documents || []).length
    ? (r.documents || []).map(d => DOCN[d.document] || d.document).join(', ')
    : '<span class="dim">none</span>';
  const bs = r.blueshield_claim_number
    ? `<div class="dim mono" style="font-size:10.5px">BS ${esc(r.blueshield_claim_number)}</div>` : '';
  const t = String(r.claim_submission_type || '');
  const typeCell = /resub/i.test(t) ? '<span class="ty resub">Resubmission</span>'
                 : /first/i.test(t) ? '<span class="ty first">New submission</span>'
                 : '<span class="dim">—</span>';
  return `<tr class="row" data-i="${i}" onclick="toggle(${i})">
    <td><span class="caret">&#9654;</span></td>
    <td class="nowrap mono">${esc(r.submitted_date_pt)}<div class="dim" style="font-size:10.5px">${esc(r.submitted_time_pt)} PT</div></td>
    <td class="pt">${esc(r.patient_name) || '<span class="dim">—</span>'}</td>
    <td class="nowrap mono">${esc(r.claim_id)}${bs}</td>
    <td class="nowrap mono">${esc(r.dos) || '<span class="dim">—</span>'}</td>
    <td class="nowrap">${typeCell}</td>
    <td class="docs">${docs}</td>
    <td><span class="pill ${esc(r.outcome)}">${esc(r.outcome)}</span></td>
  </tr>
  <tr class="detail" id="d${i}" style="display:none"><td colspan="8"></td></tr>`;
}

function detailHtml(r) {
  const docs = (r.documents || []).length ? (r.documents || []).map(d => `
    <div class="doc">
      <div class="dn">${esc(DOCN[d.document] || d.document)}</div>
      <div class="dm">${d.filename ? esc(d.filename) + (d.bytes ? ' · ' + Number(d.bytes).toLocaleString() + ' bytes' : '') + '<br>' : ''}${d.sha256 ? 'sha256 ' + esc(d.sha256).slice(0,32) + '…<br>' : ''}${esc(d.s3_path || '')}${d.note ? '<br>' + esc(d.note) : ''}</div>
    </div>`).join('') : '<div class="dim">No documents were attached to this attempt.</div>';

  const warn = r.linkage_risk ? `<div class="callout">
      <b>Not attached to the prior claim.</b> Our records classify this as a
      <b>${esc(r.claim_submission_type)}</b>, but it went to the payer as
      &ldquo;${esc(r.submission_form_type)}&rdquo;${r.blueshield_claim_number ? ' and the prior claim number (' + esc(r.blueshield_claim_number) + ') was never entered' : ''}.
      The payer opens a new claim instead of adding these documents to the one under dispute —
      which is how a packet we did send gets reported as never received.
    </div>` : '';

  return `<div class="dwrap">
    <div class="dsec">
      <h4>Submission</h4>
      <div class="dl"><span class="dt">Date &amp; time</span><span class="dd">${esc(r.submitted_date_pt)} ${esc(r.submitted_time_pt)} PT</span></div>
      <div class="dl"><span class="dt">UTC</span><span class="dd mono">${esc(r.submitted_at_utc)}</span></div>
      <div class="dl"><span class="dt">Method</span><span class="dd">${esc(r.method)}</span></div>
      <div class="dl"><span class="dt">Form type</span><span class="dd">${esc(r.submission_form_type) || '<span class="dim">not recorded</span>'}</span></div>
      <div class="dl"><span class="dt">Outcome</span><span class="dd">${esc(r.outcome)}</span></div>
      <div class="dl"><span class="dt">Payer ack (FLN)</span><span class="dd">${r.fln ? esc(r.fln) : '<span class="dim">not captured</span>'}</span></div>
    </div>
    <div class="dsec">
      <h4>Claim</h4>
      <div class="dl"><span class="dt">Patient</span><span class="dd">${esc(r.patient_name)}</span></div>
      <div class="dl"><span class="dt">Claim # (ECW)</span><span class="dd mono">${esc(r.claim_id)}</span></div>
      <div class="dl"><span class="dt">Prior BS claim #</span><span class="dd mono">${esc(r.blueshield_claim_number) || '<span class="dim">none on file</span>'}</span></div>
      <div class="dl"><span class="dt">Date of service</span><span class="dd mono">${esc(r.dos)}</span></div>
      <div class="dl"><span class="dt">Subscriber ID</span><span class="dd mono">${esc(r.subscriber_id)}</span></div>
      <div class="dl"><span class="dt">Our class.</span><span class="dd">${esc(r.claim_submission_type) || '<span class="dim">—</span>'}</span></div>
    </div>
    <div class="dsec">
      <h4>Documents (${(r.documents || []).length})</h4>
      ${docs}
    </div>
    ${r.error ? `
    <div class="dsec">
      <h4>Error</h4>
      <div class="dl"><span class="dd">${esc(r.error)}</span></div>
    </div>` : ''}
    ${warn}
  </div>`;
}

function toggle(i) {
  const d = document.getElementById('d' + i);
  const row = document.querySelector(`tr.row[data-i="${i}"]`);
  const open = d.style.display !== 'none';
  if (open) { d.style.display = 'none'; row.classList.remove('open'); return; }
  d.querySelector('td').innerHTML = detailHtml(SHOWN[i]);
  d.style.display = '';
  row.classList.add('open');
}

function render() {
  const body = document.getElementById('body');
  if (!SHOWN.length) {
    body.innerHTML = `<tr><td colspan="8" class="empty">
      <div class="big">No submissions match this search.</div>
      <div>Try a surname, a claim number, or a date of service like 07/23/2026.</div></td></tr>`;
    document.getElementById('more').style.display = 'none';
    return;
  }
  const slice = SHOWN.slice(0, limit);
  body.innerHTML = slice.map((r, i) => rowHtml(r, i)).join('');
  document.getElementById('more').style.display = SHOWN.length > limit ? '' : 'none';

  const active = flag === 'unlinked' ? ' <span class="chip" onclick="setFlag(\\'unlinked\\')">not attached to prior claim &times;</span>'
              : flag === 'no-fln' ? ' <span class="chip" onclick="setFlag(\\'no-fln\\')">no payer acknowledgement &times;</span>' : '';
  document.getElementById('count').innerHTML =
    `Showing <b>${slice.length.toLocaleString()}</b> of <b>${SHOWN.length.toLocaleString()}</b> submissions${active}`;
}

function showMore() { limit += PAGE; render(); }

// Filtering happens in the browser over the rows already in memory.
//
// It used to re-query the server on every keystroke, which was both slow and
// wrong. Slow because each keystroke re-scanned the whole table and shipped up
// to 1.7MB back. Wrong because the responses could land out of order: clearing
// the box asked for all 1188 rows while a one-row response for the text just
// deleted was still in flight, and the small one arrived last and won — so an
// empty search box showed a single result.
//
// The server-side filters stay exactly as they were; the CSV export and the
// API still use them.
function matches(r) {
  const q = document.getElementById('q').value.trim().toLowerCase();
  if (q) {
    const hay = [r.patient_name, r.claim_id, r.blueshield_claim_number,
                 r.dos, r.subscriber_id, r.fln].join(' ').toLowerCase();
    if (!hay.includes(q)) return false;
  }
  const from = document.getElementById('from').value.trim();
  const to = document.getElementById('to').value.trim();
  const d = r.submitted_date_pt || '';
  if (from && d < from) return false;
  if (to && d > to) return false;

  const ty = document.getElementById('type').value.trim().toLowerCase();
  if (ty && !String(r.claim_submission_type || '').toLowerCase().includes(ty)) return false;

  const oc = document.getElementById('outcome').value.trim().toLowerCase();
  if (oc && String(r.outcome || '').toLowerCase() !== oc) return false;

  if (flag === 'unlinked' && !r.linkage_risk) return false;
  if (flag === 'no-fln' && r.fln) return false;
  return true;
}

function apply() {
  const q = qs();
  document.getElementById('csv').href = '/api/audit-log.csv' + (q ? '?' + q : '');
  SHOWN = ALL.filter(matches);
  limit = PAGE;
  tiles();
  render();
}

for (const id of ['q','from','to','type','outcome']) {
  document.getElementById(id).addEventListener('input', apply);
}

// Clicking a date field anywhere — not just the small calendar glyph — opens
// the picker. Typing dd.mm.yyyy by hand is the slow path; the calendar is the
// one people actually want.
for (const id of ['from','to']) {
  const el = document.getElementById(id);
  el.addEventListener('click', () => {
    if (typeof el.showPicker === 'function') {
      try { el.showPicker(); } catch (e) { /* not user-activated; ignore */ }
    }
  });
}

// One fetch of the whole log, then everything is local. Reload the page to
// pick up submissions made since.
(async function init() {
  try {
    const res = await fetch('/api/audit-log');
    const data = await res.json();
    ALL = data.rows || [];
    if (data.error) {
      document.getElementById('body').innerHTML =
        `<tr><td colspan="8" class="empty"><div class="big">Could not load the audit log.</div>
         <div>${esc(data.error)}</div></td></tr>`;
      document.getElementById('count').textContent = '';
      return;
    }
  } catch (e) {
    document.getElementById('body').innerHTML =
      `<tr><td colspan="8" class="empty"><div class="big">Could not reach the server.</div>
       <div>${esc(e.message || e)}</div></td></tr>`;
    document.getElementById('count').textContent = '';
    return;
  }
  apply();
})();
</script>
</body>
</html>
"""
@app.route('/audit')
def audit_page():
    return AUDIT_HTML


if __name__ == '__main__':
    print("\n" + "="*50)
    print("  Helixona Dashboard — http://localhost:5050")
    print("  Audit log      — http://localhost:5050/audit")
    print("="*50 + "\n")
    app.run(host='0.0.0.0', port=5050, debug=True)
